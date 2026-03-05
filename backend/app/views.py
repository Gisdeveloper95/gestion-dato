
# Create your views here.
import os
import logging
from datetime import datetime
from pathlib import Path
from django.http import HttpResponse, FileResponse, Http404, StreamingHttpResponse
from django.utils import timezone
from django.conf import settings

# Configurar logger
logger = logging.getLogger(__name__)
from rest_framework import status, viewsets
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from .utils import ScriptRunner, DatabaseBackupManager
import threading
import uuid
from django.shortcuts import render
from django.http import Http404
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.permissions import IsAuthenticated, IsAuthenticatedOrReadOnly
from django.db.models import Avg, Max, Min, Count, Q, Sum
from django.shortcuts import get_object_or_404
import zipfile
import tempfile
import shutil
from io import BytesIO
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
from collections import defaultdict

from openpyxl.utils import get_column_letter

from preoperacion.models import Municipios, ListaArchivosPre
from postoperacion.models import ArchivosPost

import logging

# Importar utilidad para conversión de rutas Linux -> Windows
from backend.path_utils import linux_to_windows_path

from .models import (
    PathDirOpera, PathDirTransv, DirectoriosOperacion, 
    DirectoriosTransv, ArchivosOperacion, ArchivosTransv,
    ScriptExecution, BackupFile
)
from .serializers import (
    PathDirOperaSerializer, PathDirTransvSerializer,
    DirectoriosOperacionSerializer, DirectoriosOperacionListSerializer,
    DirectoriosTransvSerializer, DirectoriosTransvListSerializer,
    ArchivosOperacionSerializer, ArchivosOperacionListSerializer,
    ArchivosTransvSerializer, ArchivosTransvListSerializer,
    EstadisticasDirectorioSerializer, ResumenMunicipioSerializer,
    JerarquiaDirectorioSerializer,ScriptExecutionSerializer, 
    ScriptExecutionCreateSerializer,BackupFileSerializer,
    BackupStatusSerializer,ScriptStatusSerializer,

)

try:
    from preoperacion.models import ProfesionalesSeguimiento, ProfesionalMunicipio, Municipios
    from preoperacion.permissions import IsAuthenticatedOrReadOnly
except ImportError:
    # Fallbacks si no están disponibles
    ProfesionalesSeguimiento = None
    ProfesionalMunicipio = None
    Municipios = None
    from rest_framework.permissions import IsAuthenticated as IsAuthenticatedOrReadOnly



logger = logging.getLogger(__name__)

class ScriptExecutionPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class ScriptExecutionViewSet(viewsets.ModelViewSet):
    queryset = ScriptExecution.objects.all()
    serializer_class = ScriptExecutionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = ScriptExecutionPagination
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ScriptExecutionCreateSerializer
        return ScriptExecutionSerializer
    
    def perform_create(self, serializer):
        logger.info("=" * 80)
        logger.info("📝 VIEWSET: perform_create() llamado")
        logger.info("=" * 80)
        execution = serializer.save(user=self.request.user)
        logger.info(f"📝 Ejecución creada - ID: {execution.id}, Script: {execution.script_name}")
        # Ejecutar script en background
        self._execute_script_async(execution)
        logger.info("📝 _execute_script_async() llamado")

    def _execute_script_async(self, execution):
        """Ejecuta el script en un hilo separado"""
        logger.info("=" * 80)
        logger.info(f"🔄 ASYNC: Iniciando ejecución asíncrona para: {execution.script_name}")
        logger.info("=" * 80)
        def run_script():
            try:
                logger.info(f"🏃 THREAD: run_script() iniciado en hilo separado")
                execution.status = 'running'
                execution.started_at = timezone.now()
                execution.save()
                logger.info(f"💾 THREAD: Estado actualizado a 'running'")

                runner = ScriptRunner()
                logger.info(f"🔧 THREAD: ScriptRunner instanciado")

                if execution.script_name == 'backup_db':
                    logger.info(f"🎯 THREAD: Ejecutando backup_db...")
                    result = runner.execute_backup_script()
                    logger.info(f"✅ THREAD: execute_backup_script() completado con resultado: {result.get('success', False)}")
                    
                    # Guardar archivos de backup en la BD
                    if result['success'] and result['backup_files']:
                        for file_info in result['backup_files']:
                            BackupFile.objects.create(
                                execution=execution,
                                filename=file_info['filename'],
                                filepath=file_info['filepath'],
                                file_size=file_info['size']
                            )
                
                elif execution.script_name == 'llenar_datos':
                    result = runner.execute_data_script()
                
                else:
                    result = {
                        'success': False,
                        'error': f'Script no reconocido: {execution.script_name}'
                    }
                
                # Actualizar ejecución
                execution.status = 'completed' if result['success'] else 'failed'
                execution.completed_at = timezone.now()
                execution.output_log = result.get('output', '')
                execution.error_message = result.get('error', '') if not result['success'] else None
                execution.save()
                
            except Exception as e:
                logger.error("=" * 80)
                logger.error(f"💥 THREAD EXCEPTION: Error en run_script()")
                logger.error("=" * 80)
                logger.error(f"❌ Tipo: {type(e).__name__}")
                logger.error(f"❌ Mensaje: {str(e)}")
                logger.error(f"❌ Traceback:", exc_info=True)
                execution.status = 'failed'
                execution.completed_at = timezone.now()
                execution.error_message = str(e)
                execution.save()

        logger.info(f"🧵 Creando thread para run_script()...")
        thread = threading.Thread(target=run_script)
        thread.daemon = True
        logger.info(f"🚀 Iniciando thread...")
        thread.start()
        logger.info(f"✅ Thread iniciado correctamente")
    
    @action(detail=True, methods=['post'])
    def retry(self, request, pk=None):
        """Reintentar ejecución de script"""
        execution = self.get_object()
        
        if execution.status == 'running':
            return Response(
                {'error': 'El script ya se está ejecutando'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Crear nueva ejecución
        new_execution = ScriptExecution.objects.create(
            script_name=execution.script_name,
            user=request.user
        )
        
        self._execute_script_async(new_execution)
        
        serializer = self.get_serializer(new_execution)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=['get'])
    def status_summary(self, request):
        """Resumen del estado de las ejecuciones"""
        scripts_info = []
        
        for script_choice in ScriptExecution.SCRIPT_CHOICES:
            script_name = script_choice[0]
            executions = ScriptExecution.objects.filter(script_name=script_name)
            
            scripts_info.append({
                'script_name': script_name,
                'is_available': True,  # Podrías verificar si el script existe
                'last_execution': executions.first(),
                'total_executions': executions.count(),
                'successful_executions': executions.filter(status='completed').count(),
                'failed_executions': executions.filter(status='failed').count(),
            })
        
        serializer = ScriptStatusSerializer(scripts_info, many=True)
        return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def backup_status(request):
    """Estado general de los backups"""
    executions = ScriptExecution.objects.filter(script_name='backup_db')
    backup_files = BackupFile.objects.all()
    
    total_size = sum(bf.file_size for bf in backup_files) / (1024 * 1024)  # MB
    
    data = {
        'total_executions': executions.count(),
        'successful_executions': executions.filter(status='completed').count(),
        'failed_executions': executions.filter(status='failed').count(),
        'last_execution': executions.first(),
        'backup_files_count': backup_files.count(),
        'total_backup_size_mb': round(total_size, 2)
    }
    
    serializer = BackupStatusSerializer(data)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def execute_backup(request):
    """Ejecutar script de backup de forma síncrona (para testing)"""
    try:
        # Crear registro de ejecución
        execution = ScriptExecution.objects.create(
            script_name='backup_db',
            user=request.user,
            status='running',
            started_at=timezone.now()
        )
        
        runner = ScriptRunner()
        result = runner.execute_backup_script()
        
        # Actualizar ejecución
        execution.status = 'completed' if result['success'] else 'failed'
        execution.completed_at = timezone.now()
        execution.output_log = result.get('output', '')
        execution.error_message = result.get('error', '') if not result['success'] else None
        execution.save()
        
        # Guardar archivos de backup
        if result['success'] and result['backup_files']:
            for file_info in result['backup_files']:
                BackupFile.objects.create(
                    execution=execution,
                    filename=file_info['filename'],
                    filepath=file_info['filepath'],
                    file_size=file_info['size']
                )
        
        serializer = ScriptExecutionSerializer(execution)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error ejecutando backup: {str(e)}")
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

# Agregar esta función auxiliar antes de download_latest_backup_zip

def crear_archivo_recover_backup_sh():
    """
    Crea el archivo recover_backup.sh con las instrucciones de restauración para Linux
    """
    contenido_sh = """#!/bin/bash

echo "========================================="
echo "   SCRIPT DE RECUPERACION DE BACKUP"
echo "   Base de Datos PostgreSQL - Linux"
echo "========================================="
echo ""

# Verificar si psql está instalado
if ! command -v psql &> /dev/null; then
    echo "❌ psql no está instalado o no está en el PATH."
    echo ""
    echo "Por favor instale PostgreSQL:"
    echo "  - Ubuntu/Debian: sudo apt-get install postgresql-client"
    echo "  - CentOS/RHEL: sudo yum install postgresql"
    echo ""
    exit 1
fi

# Mostrar información del entorno
echo "🔍 Verificando entorno PostgreSQL..."
psql --version
echo ""

# Buscar archivos .sql en el directorio actual
echo "📁 Buscando archivos .sql en el directorio actual..."
mapfile -t sql_files < <(ls -1 *.sql 2>/dev/null)

# Verificar si se encontraron archivos
if [ ${#sql_files[@]} -eq 0 ]; then
    echo "❌ No se encontraron archivos .sql en el directorio actual."
    echo ""
    echo "Asegúrese de que los archivos de backup estén en la misma carpeta que este script."
    exit 1
fi

echo ""
echo "📋 Se encontraron ${#sql_files[@]} archivo(s) de backup disponible(s):"
echo ""

# Listar archivos
for i in "${!sql_files[@]}"; do
    echo "  $((i+1)). ${sql_files[$i]}"
done

echo ""
read -p "👉 Por favor selecciona un archivo (1-${#sql_files[@]}): " choice

# Validar la elección
if ! [[ "$choice" =~ ^[0-9]+$ ]] || [ "$choice" -lt 1 ] || [ "$choice" -gt "${#sql_files[@]}" ]; then
    echo "❌ Elección inválida. Debe ser un número entre 1 y ${#sql_files[@]}."
    exit 1
fi

# Obtener el archivo seleccionado (restar 1 porque los arrays empiezan en 0)
selected_file="${sql_files[$((choice-1))]}"
echo ""
echo "🔄 Procesando el archivo: $selected_file"
echo ""

# Solicitar credenciales de la base de datos
echo "🔐 Configuración de conexión a PostgreSQL:"
read -p "📍 Host (localhost): " db_host
db_host=${db_host:-localhost}

read -p "🔌 Puerto (5432): " db_port
db_port=${db_port:-5432}

read -p "👤 Usuario (postgres): " db_user
db_user=${db_user:-postgres}

read -p "🗄️  Base de datos de destino (gestion_dato_db): " db_name
db_name=${db_name:-gestion_dato_db}

echo ""
echo "📋 Configuración de conexión:"
echo "   Host: $db_host"
echo "   Puerto: $db_port"
echo "   Usuario: $db_user"
echo "   Base de datos: $db_name"
echo "   Archivo: $selected_file"
echo ""

# Confirmar la restauración
read -p "⚠️  ¿Está seguro de que desea restaurar este backup? (s/N): " confirm

if [[ ! "$confirm" =~ ^[sS]([iI])?$ ]] && [[ ! "$confirm" =~ ^[yY]([eE][sS])?$ ]]; then
    echo "❌ Operación cancelada por el usuario."
    exit 0
fi

echo ""
echo "🚀 Iniciando restauración del backup..."
echo ""
echo "⏳ Ejecutando comando: PGPASSWORD=*** psql -U $db_user -h $db_host -p $db_port -d $db_name -f \"$selected_file\""
echo ""

# Solicitar contraseña
read -sp "🔑 Contraseña para $db_user: " db_password
echo ""
echo ""

export PGPASSWORD="$db_password"

# Verificar si la base de datos existe, si no, crearla
echo "🔍 Verificando si la base de datos existe..."
db_exists=$(psql -U "$db_user" -h "$db_host" -p "$db_port" -d postgres -tAc "SELECT 1 FROM pg_database WHERE datname='$db_name'" 2>/dev/null)

if [ "$db_exists" != "1" ]; then
    echo "📦 La base de datos '$db_name' no existe. Creándola..."
    psql -U "$db_user" -h "$db_host" -p "$db_port" -d postgres -c "CREATE DATABASE $db_name WITH ENCODING 'UTF8';" 2>/dev/null
    if [ $? -eq 0 ]; then
        echo "✅ Base de datos creada exitosamente."
    fi
fi

echo ""
echo "⏳ Ejecutando restauración (esto puede tardar varios minutos)..."
echo ""

# Ejecutar psql con el archivo seleccionado
# Usamos ON_ERROR_STOP=off para continuar ante errores menores
psql -U "$db_user" -h "$db_host" -p "$db_port" -d "$db_name" -v ON_ERROR_STOP=off -f "$selected_file" 2>&1 | tee restore_output.log
exit_code=${PIPESTATUS[0]}
unset PGPASSWORD

# Contar errores reales en el log
error_count=$(grep -c "^ERROR:" restore_output.log 2>/dev/null || echo "0")

echo ""
if [ "$error_count" -gt 10 ]; then
    echo "⚠️  La restauración terminó con $error_count errores."
    echo "   Revise 'restore_output.log' para más detalles."
    echo ""
    echo "   NOTA: Algunos errores son normales (ej: 'already exists', 'does not exist')."
else
    echo "✅ Backup restaurado exitosamente!"
    echo ""
    echo "🎉 La restauración se completó correctamente."
    rm -f restore_output.log 2>/dev/null
fi

echo ""
echo "📝 Proceso terminado."
"""
    return contenido_sh

def crear_archivo_recover_backup():
    """
    Crea el archivo recover_backup.bat con las instrucciones de restauración para Windows
    """
    contenido_bat = """@echo off
setlocal enabledelayedexpansion

echo =========================================
echo    SCRIPT DE RECUPERACION DE BACKUP
echo    Base de Datos PostgreSQL
echo =========================================
echo.

:: Verificar si psql está en el PATH
where psql >nul 2>&1
if errorlevel 1 (
    echo No se encontró psql en el PATH. Buscando instalación de PostgreSQL...
    echo.
    
    :: Buscar directorios comunes de instalación de PostgreSQL
    set found_psql=
    for /d %%v in ("C:\\Program Files\\PostgreSQL\\1*" "C:\\Program Files\\PostgreSQL\\2*") do (
        if exist "%%v\\bin\\psql.exe" (
            set "found_psql=%%v\\bin"
            goto :found
        )
    )
    
    :: Si no se encuentra PostgreSQL, salir
    if not defined found_psql (
        echo ❌ No se encontró ninguna instalación de PostgreSQL.
        echo.
        echo Por favor instale PostgreSQL o verifique que esté en el PATH.
        pause
        exit /b 1
    )
    
    :found
    echo ✅ Se encontró PostgreSQL en: %found_psql%
    set "PATH=%found_psql%;%PATH%"
    echo.
)

:: Mostrar información del entorno
echo 🔍 Verificando entorno PostgreSQL...
psql --version
echo.

:: Buscar archivos .sql en el directorio actual
echo 📁 Buscando archivos .sql en el directorio actual...
set count=0
for %%f in (*.sql) do (
    set /a count+=1
    set "file!count!=%%f"
    echo   !count!. %%f
)

:: Verificar si se encontraron archivos
if %count%==0 (
    echo ❌ No se encontraron archivos .sql en el directorio actual.
    echo.
    echo Asegúrese de que los archivos de backup estén en la misma carpeta que este script.
    pause
    exit /b 1
)

echo.
echo 📋 Se encontraron %count% archivo(s) de backup disponible(s).
echo.

:: Solicitar al usuario que seleccione un archivo
set choice=
set /p choice=👉 Por favor selecciona un archivo (1-%count%): 

:: Validar la elección
if "%choice%"=="" (
    echo ❌ Elección inválida. No se seleccionó ningún archivo.
    pause
    exit /b 1
)

:: Verificar que la elección esté en rango
set valid=0
for /l %%i in (1,1,%count%) do (
    if "%choice%"=="%%i" set valid=1
)

if %valid%==0 (
    echo ❌ Elección fuera de rango. Debe ser un número entre 1 y %count%.
    pause
    exit /b 1
)

:: Procesar el archivo seleccionado
set "selected_file=!file%choice%!"
echo.
echo 🔄 Procesando el archivo: %selected_file%
echo.

:: Solicitar credenciales de la base de datos
echo 🔐 Configuración de conexión a PostgreSQL:
set /p db_host=📍 Host (localhost): 
if "%db_host%"=="" set db_host=localhost

set /p db_port=🔌 Puerto (5432): 
if "%db_port%"=="" set db_port=5432

set /p db_user=👤 Usuario (postgres): 
if "%db_user%"=="" set db_user=postgres

set /p db_name=🗄️  Base de datos de destino (gestion_dato_db):
if "%db_name%"=="" set db_name=gestion_dato_db

echo.
echo 📋 Configuración de conexión:
echo    Host: %db_host%
echo    Puerto: %db_port%
echo    Usuario: %db_user%
echo    Base de datos: %db_name%
echo    Archivo: %selected_file%
echo.

:: Confirmar la restauración
set confirm=
set /p confirm=⚠️  ¿Está seguro de que desea restaurar este backup? (s/N): 

if /i not "%confirm%"=="s" if /i not "%confirm%"=="si" if /i not "%confirm%"=="yes" (
    echo ❌ Operación cancelada por el usuario.
    pause
    exit /b 0
)

echo.
echo 🚀 Iniciando restauración del backup...
echo.

:: Solicitar contraseña
set /p db_password=🔑 Contraseña para %db_user%:
set PGPASSWORD=%db_password%

:: Verificar si la base de datos existe
echo.
echo 🔍 Verificando si la base de datos existe...
psql -U %db_user% -h %db_host% -p %db_port% -d postgres -tAc "SELECT 1 FROM pg_database WHERE datname='%db_name%'" > nul 2>&1
if errorlevel 1 (
    echo 📦 La base de datos '%db_name%' no existe. Creándola...
    psql -U %db_user% -h %db_host% -p %db_port% -d postgres -c "CREATE DATABASE %db_name% WITH ENCODING 'UTF8';" > nul 2>&1
    echo ✅ Base de datos creada.
)

echo.
echo ⏳ Ejecutando restauración (esto puede tardar varios minutos)...
echo    Comando: psql -U %db_user% -h %db_host% -p %db_port% -d %db_name% -f "%selected_file%"
echo.

:: Ejecutar psql con el archivo seleccionado
:: Usamos -v ON_ERROR_STOP=off para continuar ante errores menores
psql -U %db_user% -h %db_host% -p %db_port% -d %db_name% -v ON_ERROR_STOP=off -f "%selected_file%" 2>&1

set PGPASSWORD=

echo.
echo ✅ Proceso de restauración completado!
echo.
echo    NOTA: Algunos mensajes de error son normales (ej: 'already exists').
echo    Verifique que los datos se restauraron correctamente.

echo.
echo 📝 Proceso terminado.
pause
"""
    return contenido_bat

# Modificar la función download_latest_backup_zip

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_latest_backup_zip(request):
    """
    Descargar el último backup como ZIP incluyendo script de recuperación
    """
    
    # ✅ AGREGAR ESTE DEBUG AL INICIO:
    print(f"🔍 Headers recibidos: {dict(request.headers)}")
    print(f"🔍 User: {request.user}")
    print(f"🔍 Method: {request.method}")
    print(f"🔍 Content-Type: {request.content_type}")
    print(f"🔍 Accept: {request.headers.get('Accept', 'NO ACCEPT HEADER')}")
    
    try:
        runner = ScriptRunner()
        print(f"🔍 Directorio de backup: {runner.backup_dir}")
        print(f"🔍 ¿Directorio existe?: {runner.backup_dir.exists()}")
        
        zip_result = runner.create_zip_from_backups()
        print(f"🔍 Resultado del ZIP: {zip_result}")
        
        if not zip_result['success']:
            print(f"❌ Error creando ZIP: {zip_result.get('error')}")
            return Response(
                {'error': zip_result.get('error', 'Error creando ZIP')},
                status=status.HTTP_404_NOT_FOUND
            )
        
        zip_path = Path(zip_result['zip_path'])
        
        if not zip_path.exists():
            return Response(
                {'error': 'Archivo ZIP no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # 🔧 AGREGAR scripts de recuperación al ZIP existente (modo append)
        # NO se copia ni se re-crea el ZIP - se agregan archivos pequeños al que ya existe en disco
        # Esto usa 0 MB de RAM extra sin importar el tamaño del backup
        print("🔧 Agregando scripts de recuperación al ZIP existente (modo append)...")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'backup_{timestamp}.zip'

        # Copiar el ZIP original a un archivo temporal para no modificar el original
        temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip_path = temp_zip.name
        temp_zip.close()
        shutil.copy2(str(zip_path), temp_zip_path)

        # Agregar scripts al ZIP copiado (modo 'a' = append, no re-lee el contenido existente)
        with zipfile.ZipFile(temp_zip_path, 'a', zipfile.ZIP_DEFLATED) as zf:
            contenido_bat = crear_archivo_recover_backup()
            zf.writestr('recover_backup.bat', contenido_bat)
            print("  🔧 Agregado: recover_backup.bat")

            contenido_sh = crear_archivo_recover_backup_sh()
            zf.writestr('recover_backup.sh', contenido_sh)
            print("  🐧 Agregado: recover_backup.sh")

            readme_content = (
                "# INSTRUCCIONES DE RECUPERACION DE BACKUP\n"
                "# Sistema: Gestion de Datos IGAC\n\n"
                "## Contenido de este ZIP:\n"
                "- Archivos .sql con los backups de la base de datos\n"
                "- recover_backup.bat: Script automatico de recuperacion (Windows)\n"
                "- recover_backup.sh: Script automatico de recuperacion (Linux)\n\n"
                "## Como restaurar el backup:\n\n"
                "### Windows:\n"
                "1. Extraer todos los archivos de este ZIP en una carpeta\n"
                "2. Hacer doble clic en 'recover_backup.bat'\n"
                "3. Seguir las instrucciones en pantalla\n\n"
                "### Linux/Mac:\n"
                "1. Extraer todos los archivos de este ZIP en una carpeta\n"
                "2. Dar permisos de ejecucion: chmod +x recover_backup.sh\n"
                "3. Ejecutar: ./recover_backup.sh\n\n"
                "### Restauracion Manual:\n"
                "psql -U postgres -h localhost -p 5432 -c \"CREATE DATABASE gestion_dato_db WITH ENCODING 'UTF8';\"\n"
                "psql -U postgres -h localhost -p 5432 -d gestion_dato_db -f archivo_backup.sql\n\n"
                "## IMPORTANTE:\n"
                "- Asegurese de tener PostgreSQL instalado (cliente psql)\n"
                "- El backup NO incluye propietarios ni privilegios (portable)\n\n"
                "## Soporte: andres.osorio@igac.gov.co\n"
                f"Fecha de generacion: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
            )
            zf.writestr('README_RECUPERACION.txt', readme_content)
            print("  📖 Agregado: README_RECUPERACION.txt")

        zip_size = os.path.getsize(temp_zip_path)

        # Streaming del ZIP
        def zip_iterator():
            CHUNK_SIZE = 8 * 1024 * 1024  # 8MB chunks
            try:
                with open(temp_zip_path, 'rb') as f:
                    while True:
                        chunk = f.read(CHUNK_SIZE)
                        if not chunk:
                            break
                        yield chunk
            finally:
                try:
                    os.unlink(temp_zip_path)
                except:
                    pass

        response = StreamingHttpResponse(
            zip_iterator(),
            content_type='application/zip'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['X-Accel-Buffering'] = 'no'  # Desactivar buffering en nginx

        print(f"✅ ZIP streaming iniciado:")
        print(f"   📦 Archivo: {filename}")
        print(f"   📊 Tamaño: {zip_size} bytes")
        print(f"   🔧 Incluye: recover_backup.bat + README + archivos SQL")

        return response
        
    except Exception as e:
        logger.error(f"Error descargando backup ZIP: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_backup_file(request, file_id):
    """Descargar un archivo de backup específico"""
    try:
        backup_file = BackupFile.objects.get(id=file_id)
        file_path = Path(backup_file.filepath)
        
        if not file_path.exists():
            return Response(
                {'error': 'Archivo no encontrado en el sistema'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        response = FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename=backup_file.filename
        )
        
        # Determinar content type basado en la extensión
        if backup_file.filename.endswith('.sql'):
            response['Content-Type'] = 'application/sql'
        elif backup_file.filename.endswith('.gz'):
            response['Content-Type'] = 'application/gzip'
        else:
            response['Content-Type'] = 'application/octet-stream'
        
        response['Content-Length'] = file_path.stat().st_size
        
        return response
        
    except BackupFile.DoesNotExist:
        return Response(
            {'error': 'Archivo de backup no encontrado'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Error descargando archivo de backup: {str(e)}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_backup_files(request):
    """Listar archivos de backup disponibles"""
    try:
        runner = ScriptRunner()
        files_info = runner.get_backup_files()
        
        # También obtener archivos registrados en BD
        db_files = BackupFile.objects.all().order_by('-created_at')
        db_serializer = BackupFileSerializer(db_files, many=True)
        
        return Response({
            'filesystem_files': files_info,
            'database_files': db_serializer.data
        })
        
    except Exception as e:
        logger.error(f"Error listando archivos de backup: {str(e)}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def clean_backup_directory(request):
    """Limpiar directorio de backups"""
    try:
        runner = ScriptRunner()
        success = runner.clean_backup_directory()
        
        if success:
            # También limpiar registros de BD
            BackupFile.objects.all().delete()
            return Response({'message': 'Directorio de backups limpiado exitosamente'})
        else:
            return Response(
                {'error': 'Error limpiando directorio de backups'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    except Exception as e:
        logger.error(f"Error limpiando directorio: {str(e)}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    


# ✅ SISTEMA DE PERMISOS COPIADO DE PREOPERACION
def get_municipios_permitidos(user):
    """
    Obtiene los municipios que un usuario puede consultar
    - Administradores/Super admin: 'todos'
    - Profesionales: lista de municipios asignados
    - Otros: lista vacía
    """
    if not user.is_authenticated:
        return []
    
    # Super administradores y administradores pueden ver todo
    if user.is_superuser or user.is_staff:
        return 'todos'
    
    # Verificar si es profesional de seguimiento
    if user.groups.filter(name='Profesionales_Seguimiento').exists():
        try:
            if ProfesionalesSeguimiento and ProfesionalMunicipio:
                profesional = ProfesionalesSeguimiento.objects.get(
                    cod_profesional=user.username
                )
                municipios_ids = ProfesionalMunicipio.objects.filter(
                    cod_profesional=profesional
                ).values_list('cod_municipio', flat=True)
                return list(municipios_ids)
        except ProfesionalesSeguimiento.DoesNotExist:
            return []
    
    return []


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


# ✅ MIXIN PARA FILTROS DE MUNICIPIO
class MunicipioPermissionMixin:
    """Mixin para aplicar filtros de municipio según permisos del usuario"""
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Obtener municipios permitidos para el usuario
        municipios_permitidos = get_municipios_permitidos(self.request.user)
        
        if municipios_permitidos == 'todos':
            # Administradores ven todo
            return queryset
        elif municipios_permitidos:
            # Profesionales ven solo sus municipios asignados
            return self.filtrar_por_municipios(queryset, municipios_permitidos)
        else:
            # Sin permisos, no ve nada
            return queryset.none()
    
    def filtrar_por_municipios(self, queryset, municipios_permitidos):
        """Override en cada ViewSet según su campo de municipio"""
        return queryset


# ViewSets para Rutas de Directorios

class PathDirOperaViewSet(MunicipioPermissionMixin, viewsets.ModelViewSet):
    queryset = PathDirOpera.objects.all()
    serializer_class = PathDirOperaSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_municipio']
    search_fields = ['path']
    ordering_fields = ['cod_municipio', 'fecha_creacion']
    ordering = ['-fecha_creacion']

    def filtrar_por_municipios(self, queryset, municipios_permitidos):
        """Filtrar por municipios permitidos"""
        return queryset.filter(cod_municipio__in=municipios_permitidos)

    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        """Obtiene las rutas operativas agrupadas por municipio"""
        municipio = request.query_params.get('municipio')
        if not municipio:
            return Response({'error': 'Parámetro municipio requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ VERIFICAR PERMISOS PARA EL MUNICIPIO ESPECÍFICO
        municipios_permitidos = get_municipios_permitidos(request.user)
        if municipios_permitidos != 'todos':
            if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                return Response(
                    {'error': 'No tiene permisos para acceder a este municipio'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        
        rutas = self.get_queryset().filter(cod_municipio=municipio)
        serializer = self.get_serializer(rutas, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def mis_municipios(self, request):
        """Obtiene las rutas de los municipios asignados al profesional"""
        municipios_permitidos = get_municipios_permitidos(request.user)
        
        if municipios_permitidos == 'todos':
            # Admin ve estadísticas de todos los municipios
            rutas_por_municipio = self.get_queryset().values('cod_municipio').annotate(
                total_rutas=Count('id'),
                ultima_actualizacion=Max('fecha_creacion')
            ).order_by('cod_municipio')
        elif municipios_permitidos:
            # Profesional ve solo sus municipios
            rutas_por_municipio = self.get_queryset().filter(
                cod_municipio__in=municipios_permitidos
            ).values('cod_municipio').annotate(
                total_rutas=Count('id'),
                ultima_actualizacion=Max('fecha_creacion')
            ).order_by('cod_municipio')
        else:
            rutas_por_municipio = []
        
        return Response(rutas_por_municipio)


class PathDirTransvViewSet(MunicipioPermissionMixin, viewsets.ModelViewSet):
    queryset = PathDirTransv.objects.all()
    serializer_class = PathDirTransvSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_municipio']
    search_fields = ['path']
    ordering_fields = ['cod_municipio', 'fecha_creacion']
    ordering = ['-fecha_creacion']

    def filtrar_por_municipios(self, queryset, municipios_permitidos):
        return queryset.filter(cod_municipio__in=municipios_permitidos)

    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        """Obtiene las rutas postoperativas agrupadas por municipio"""
        municipio = request.query_params.get('municipio')
        if not municipio:
            return Response({'error': 'Parámetro municipio requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ VERIFICAR PERMISOS
        municipios_permitidos = get_municipios_permitidos(request.user)
        if municipios_permitidos != 'todos':
            if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                return Response(
                    {'error': 'No tiene permisos para acceder a este municipio'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        
        rutas = self.get_queryset().filter(cod_municipio=municipio)
        serializer = self.get_serializer(rutas, many=True)
        return Response(serializer.data)


# ViewSets para Directorios

class DirectoriosOperacionViewSet(MunicipioPermissionMixin, viewsets.ModelViewSet):
    queryset = DirectoriosOperacion.objects.select_related('directorio_padre')
    serializer_class = DirectoriosOperacionSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_municipio', 'activo', 'nivel_profundidad']
    search_fields = ['nombre_directorio', 'path_directorio', 'usuario_propietario']
    ordering_fields = ['cod_municipio', 'nivel_profundidad', 'total_archivos', 'peso_total_bytes']
    ordering = ['cod_municipio', 'nivel_profundidad']

    def filtrar_por_municipios(self, queryset, municipios_permitidos):
        return queryset.filter(cod_municipio__in=municipios_permitidos)

    def get_serializer_class(self):
        if self.action == 'list':
            return DirectoriosOperacionListSerializer
        return DirectoriosOperacionSerializer

    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        """Obtiene directorios de operación por municipio"""
        municipio = request.query_params.get('municipio')
        if not municipio:
            return Response({'error': 'Parámetro municipio requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ VERIFICAR PERMISOS
        municipios_permitidos = get_municipios_permitidos(request.user)
        if municipios_permitidos != 'todos':
            if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                return Response(
                    {'error': 'No tiene permisos para acceder a este municipio'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        
        directorios = self.get_queryset().filter(cod_municipio=municipio)
        page = self.paginate_queryset(directorios)
        if page is not None:
            serializer = DirectoriosOperacionListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = DirectoriosOperacionListSerializer(directorios, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def jerarquia(self, request):
        """Obtiene la jerarquía de directorios de operación"""
        municipio = request.query_params.get('municipio')
        
        if municipio:
            # ✅ VERIFICAR PERMISOS PARA MUNICIPIO ESPECÍFICO
            municipios_permitidos = get_municipios_permitidos(request.user)
            if municipios_permitidos != 'todos':
                if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                    return Response(
                        {'error': 'No tiene permisos para acceder a este municipio'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
        
        queryset = self.get_queryset()
        
        if municipio:
            queryset = queryset.filter(cod_municipio=municipio)
        
        # Obtener solo directorios raíz (nivel 0 o sin padre)
        directorios_raiz = queryset.filter(
            Q(nivel_profundidad=0) | Q(directorio_padre__isnull=True)
        ).order_by('nombre_directorio')
        
        def construir_jerarquia(directorio):
            hijos = queryset.filter(directorio_padre=directorio).order_by('nombre_directorio')
            return {
                'id': directorio.cod_dir_operacion,
                'nombre': directorio.nombre_directorio,
                'tipo': 'operacion',
                'nivel': directorio.nivel_profundidad,
                'municipio': directorio.cod_municipio,
                'total_archivos': directorio.total_archivos,
                'peso_total_mb': round(directorio.peso_total_bytes / (1024 * 1024), 2) if directorio.peso_total_bytes else 0,
                'hijos': [construir_jerarquia(hijo) for hijo in hijos]
            }
        
        jerarquia = [construir_jerarquia(dir_raiz) for dir_raiz in directorios_raiz]
        return Response(jerarquia)

    @action(detail=True, methods=['get'])
    def estadisticas(self, request, pk=None):
        """Estadísticas de un directorio específico"""
        directorio = self.get_object()
        
        # ✅ VERIFICAR PERMISOS PARA EL MUNICIPIO DEL DIRECTORIO
        municipios_permitidos = get_municipios_permitidos(request.user)
        if municipios_permitidos != 'todos':
            if not municipios_permitidos or directorio.cod_municipio not in municipios_permitidos:
                return Response(
                    {'error': 'No tiene permisos para acceder a este municipio'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Contar archivos en este directorio
        total_archivos = ArchivosOperacion.objects.filter(
            cod_dir_operacion=directorio
        ).count()
        
        # Sumar peso total de archivos
        peso_total = ArchivosOperacion.objects.filter(
            cod_dir_operacion=directorio
        ).aggregate(total=Sum('peso_memoria'))['total'] or 0
        
        estadisticas = {
            'total_directorios': 1,
            'total_archivos': total_archivos,
            'peso_total_mb': round(peso_total / (1024 * 1024), 2),
            'directorio_mas_grande': directorio.nombre_directorio,
            'ultimo_escaneo': directorio.fecha_actualizacion,
            'municipio': directorio.cod_municipio
        }
        
        serializer = EstadisticasDirectorioSerializer(estadisticas)
        return Response(serializer.data)


class DirectoriosTransvViewSet(MunicipioPermissionMixin, viewsets.ModelViewSet):
    queryset = DirectoriosTransv.objects.select_related('directorio_padre')
    serializer_class = DirectoriosTransvSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_municipio', 'activo', 'nivel_jerarquia']
    search_fields = ['nombre_directorio', 'ruta_completa']
    ordering_fields = ['cod_municipio', 'nivel_jerarquia', 'total_archivos', 'peso_total_bytes']
    ordering = ['cod_municipio', 'nivel_jerarquia']

    def filtrar_por_municipios(self, queryset, municipios_permitidos):
        return queryset.filter(cod_municipio__in=municipios_permitidos)

    def get_serializer_class(self):
        if self.action == 'list':
            return DirectoriosTransvListSerializer
        return DirectoriosTransvSerializer

    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        """Obtiene directorios transversales por municipio"""
        municipio = request.query_params.get('municipio')
        if not municipio:
            return Response({'error': 'Parámetro municipio requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ VERIFICAR PERMISOS
        municipios_permitidos = get_municipios_permitidos(request.user)
        if municipios_permitidos != 'todos':
            if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                return Response(
                    {'error': 'No tiene permisos para acceder a este municipio'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        
        directorios = self.get_queryset().filter(cod_municipio=municipio)
        page = self.paginate_queryset(directorios)
        if page is not None:
            serializer = DirectoriosTransvListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = DirectoriosTransvListSerializer(directorios, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def jerarquia(self, request):
        """Obtiene la jerarquía de directorios transversales"""
        municipio = request.query_params.get('municipio')
        
        if municipio:
            # ✅ VERIFICAR PERMISOS
            municipios_permitidos = get_municipios_permitidos(request.user)
            if municipios_permitidos != 'todos':
                if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                    return Response(
                        {'error': 'No tiene permisos para acceder a este municipio'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
        
        queryset = self.get_queryset()
        
        if municipio:
            queryset = queryset.filter(cod_municipio=municipio)
        
        # Obtener solo directorios raíz (nivel 0 o sin padre)
        directorios_raiz = queryset.filter(
            Q(nivel_jerarquia=0) | Q(directorio_padre__isnull=True)
        ).order_by('nombre_directorio')
        
        def construir_jerarquia(directorio):
            hijos = queryset.filter(directorio_padre=directorio).order_by('nombre_directorio')
            return {
                'id': directorio.cod_dir_transv,
                'nombre': directorio.nombre_directorio,
                'tipo': 'transversal',
                'nivel': directorio.nivel_jerarquia,
                'municipio': directorio.cod_municipio,
                'total_archivos': directorio.total_archivos,
                'peso_total_mb': round(directorio.peso_total_bytes / (1024 * 1024), 2) if directorio.peso_total_bytes else 0,
                'hijos': [construir_jerarquia(hijo) for hijo in hijos]
            }
        
        jerarquia = [construir_jerarquia(dir_raiz) for dir_raiz in directorios_raiz]
        return Response(jerarquia)


# ViewSets para Archivos

class ArchivosOperacionViewSet(MunicipioPermissionMixin, viewsets.ModelViewSet):
    queryset = ArchivosOperacion.objects.select_related('cod_dir_operacion')
    serializer_class = ArchivosOperacionSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_dir_operacion', 'extension', 'tipo_archivo', 'activo']
    search_fields = ['nombre_archivo', 'path_file', 'usuario_windows']
    ordering_fields = ['nombre_archivo', 'peso_memoria', 'fecha_registro']
    ordering = ['-fecha_registro']

    def filtrar_por_municipios(self, queryset, municipios_permitidos):
        return queryset.filter(cod_dir_operacion__cod_municipio__in=municipios_permitidos)

    def get_serializer_class(self):
        if self.action == 'list':
            return ArchivosOperacionListSerializer
        return ArchivosOperacionSerializer

    @action(detail=False, methods=['get'])
    def por_directorio(self, request):
        """Obtiene archivos por directorio"""
        directorio = request.query_params.get('directorio')
        if not directorio:
            return Response({'error': 'Parámetro directorio requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ VERIFICAR PERMISOS DEL DIRECTORIO
        try:
            dir_obj = DirectoriosOperacion.objects.get(cod_dir_operacion=directorio)
            municipios_permitidos = get_municipios_permitidos(request.user)
            if municipios_permitidos != 'todos':
                if not municipios_permitidos or dir_obj.cod_municipio not in municipios_permitidos:
                    return Response(
                        {'error': 'No tiene permisos para acceder a este directorio'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
        except DirectoriosOperacion.DoesNotExist:
            return Response({'error': 'Directorio no encontrado'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        archivos = self.get_queryset().filter(cod_dir_operacion=directorio)
        page = self.paginate_queryset(archivos)
        if page is not None:
            serializer = ArchivosOperacionListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = ArchivosOperacionListSerializer(archivos, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        """Obtiene archivos por municipio"""
        municipio = request.query_params.get('municipio')
        if not municipio:
            return Response({'error': 'Parámetro municipio requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ VERIFICAR PERMISOS
        municipios_permitidos = get_municipios_permitidos(request.user)
        if municipios_permitidos != 'todos':
            if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                return Response(
                    {'error': 'No tiene permisos para acceder a este municipio'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        
        archivos = self.get_queryset().filter(cod_dir_operacion__cod_municipio=municipio)
        page = self.paginate_queryset(archivos)
        if page is not None:
            serializer = ArchivosOperacionListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = ArchivosOperacionListSerializer(archivos, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def estadisticas_extension(self, request):
        """Estadísticas de archivos por extensión"""
        municipio = request.query_params.get('municipio')
        queryset = self.get_queryset()
        
        if municipio:
            # ✅ VERIFICAR PERMISOS
            municipios_permitidos = get_municipios_permitidos(request.user)
            if municipios_permitidos != 'todos':
                if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                    return Response(
                        {'error': 'No tiene permisos para acceder a este municipio'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            queryset = queryset.filter(cod_dir_operacion__cod_municipio=municipio)
        
        estadisticas = queryset.values('extension').annotate(
            total=Count('id_archivo_operacion'),
            peso_total=Sum('peso_memoria')
        ).order_by('-total')
        
        return Response(estadisticas)


class ArchivosTransvViewSet(MunicipioPermissionMixin, viewsets.ModelViewSet):
    queryset = ArchivosTransv.objects.select_related('cod_dir_transv')
    serializer_class = ArchivosTransvSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_dir_transv', 'extension', 'es_directorio_especial', 'activo']
    search_fields = ['nombre_archivo', 'path_file', 'usuario_windows']
    ordering_fields = ['nombre_archivo', 'peso_memoria', 'fecha_creacion']
    ordering = ['-fecha_creacion']

    def filtrar_por_municipios(self, queryset, municipios_permitidos):
        return queryset.filter(cod_dir_transv__cod_municipio__in=municipios_permitidos)

    def get_serializer_class(self):
        if self.action == 'list':
            return ArchivosTransvListSerializer
        return ArchivosTransvSerializer

    @action(detail=False, methods=['get'])
    def por_directorio(self, request):
        """Obtiene archivos por directorio"""
        directorio = request.query_params.get('directorio')
        if not directorio:
            return Response({'error': 'Parámetro directorio requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ VERIFICAR PERMISOS DEL DIRECTORIO
        try:
            dir_obj = DirectoriosTransv.objects.get(cod_dir_transv=directorio)
            municipios_permitidos = get_municipios_permitidos(request.user)
            if municipios_permitidos != 'todos':
                if not municipios_permitidos or dir_obj.cod_municipio not in municipios_permitidos:
                    return Response(
                        {'error': 'No tiene permisos para acceder a este directorio'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
        except DirectoriosTransv.DoesNotExist:
            return Response({'error': 'Directorio no encontrado'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        archivos = self.get_queryset().filter(cod_dir_transv=directorio)
        page = self.paginate_queryset(archivos)
        if page is not None:
            serializer = ArchivosTransvListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = ArchivosTransvListSerializer(archivos, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        """Obtiene archivos por municipio"""
        municipio = request.query_params.get('municipio')
        if not municipio:
            return Response({'error': 'Parámetro municipio requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ VERIFICAR PERMISOS
        municipios_permitidos = get_municipios_permitidos(request.user)
        if municipios_permitidos != 'todos':
            if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                return Response(
                    {'error': 'No tiene permisos para acceder a este municipio'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        
        archivos = self.get_queryset().filter(cod_dir_transv__cod_municipio=municipio)
        page = self.paginate_queryset(archivos)
        if page is not None:
            serializer = ArchivosTransvListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = ArchivosTransvListSerializer(archivos, many=True)
        return Response(serializer.data)


# ✅ Views de función para estadísticas y reportes con PERMISOS

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def resumen_por_municipio(request):
    """Resumen general de directorios y archivos por municipio"""
    municipio = request.query_params.get('municipio')
    
    if not municipio:
        return Response({'error': 'Parámetro municipio requerido'}, 
                       status=status.HTTP_400_BAD_REQUEST)
    
    # ✅ VERIFICAR PERMISOS
    municipios_permitidos = get_municipios_permitidos(request.user)
    if municipios_permitidos != 'todos':
        if not municipios_permitidos or int(municipio) not in municipios_permitidos:
            return Response(
                {'error': 'No tiene permisos para acceder a este municipio'}, 
                status=status.HTTP_403_FORBIDDEN
            )
    
    # Estadísticas de operación
    dirs_opera = DirectoriosOperacion.objects.filter(cod_municipio=municipio)
    archivos_opera = ArchivosOperacion.objects.filter(
        cod_dir_operacion__cod_municipio=municipio
    )
    
    # Estadísticas transversales
    dirs_transv = DirectoriosTransv.objects.filter(cod_municipio=municipio)
    archivos_transv = ArchivosTransv.objects.filter(
        cod_dir_transv__cod_municipio=municipio
    )
    
    # Cálculos
    peso_opera = archivos_opera.aggregate(total=Sum('peso_memoria'))['total'] or 0
    peso_transv = archivos_transv.aggregate(total=Sum('peso_memoria'))['total'] or 0
    
    resumen = {
        'cod_municipio': int(municipio),
        'total_directorios_opera': dirs_opera.count(),
        'total_directorios_transv': dirs_transv.count(),
        'total_archivos_opera': archivos_opera.count(),
        'total_archivos_transv': archivos_transv.count(),
        'peso_total_opera_mb': round(peso_opera / (1024 * 1024), 2),
        'peso_total_transv_mb': round(peso_transv / (1024 * 1024), 2)
    }
    
    serializer = ResumenMunicipioSerializer(resumen)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estadisticas_generales(request):
    """Estadísticas generales del sistema"""
    municipio = request.query_params.get('municipio')
    
    # ✅ APLICAR FILTROS DE MUNICIPIO SEGÚN PERMISOS
    municipios_permitidos = get_municipios_permitidos(request.user)
    
    # Base querysets
    dirs_opera = DirectoriosOperacion.objects.all()
    dirs_transv = DirectoriosTransv.objects.all()
    archivos_opera = ArchivosOperacion.objects.all()
    archivos_transv = ArchivosTransv.objects.all()
    
    # Aplicar filtros de permisos
    if municipios_permitidos != 'todos':
        if municipios_permitidos:
            dirs_opera = dirs_opera.filter(cod_municipio__in=municipios_permitidos)
            dirs_transv = dirs_transv.filter(cod_municipio__in=municipios_permitidos)
            archivos_opera = archivos_opera.filter(cod_dir_operacion__cod_municipio__in=municipios_permitidos)
            archivos_transv = archivos_transv.filter(cod_dir_transv__cod_municipio__in=municipios_permitidos)
        else:
            # Sin permisos, devolver datos vacíos
            estadisticas = {
                'total_directorios': 0,
                'total_archivos': 0,
                'peso_total_mb': 0,
                'directorio_mas_grande': "Sin acceso",
                'ultimo_escaneo': None
            }
            serializer = EstadisticasDirectorioSerializer(estadisticas)
            return Response(serializer.data)
    
    # Filtrar por municipio específico si se proporciona
    if municipio:
        # Verificar permisos para municipio específico
        if municipios_permitidos != 'todos':
            if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                return Response(
                    {'error': 'No tiene permisos para acceder a este municipio'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        
        dirs_opera = dirs_opera.filter(cod_municipio=municipio)
        dirs_transv = dirs_transv.filter(cod_municipio=municipio)
        archivos_opera = archivos_opera.filter(cod_dir_operacion__cod_municipio=municipio)
        archivos_transv = archivos_transv.filter(cod_dir_transv__cod_municipio=municipio)
    
    # Cálculos
    total_dirs = dirs_opera.count() + dirs_transv.count()
    total_archivos = archivos_opera.count() + archivos_transv.count()
    
    peso_opera = archivos_opera.aggregate(total=Sum('peso_memoria'))['total'] or 0
    peso_transv = archivos_transv.aggregate(total=Sum('peso_memoria'))['total'] or 0
    peso_total = peso_opera + peso_transv
    
    # Directorio más grande
    dir_mas_grande_opera = dirs_opera.order_by('-peso_total_bytes').first()
    dir_mas_grande_transv = dirs_transv.order_by('-peso_total_bytes').first()
    
    dir_mas_grande = "Sin datos"
    if dir_mas_grande_opera and dir_mas_grande_transv:
        if (dir_mas_grande_opera.peso_total_bytes or 0) > (dir_mas_grande_transv.peso_total_bytes or 0):
            dir_mas_grande = dir_mas_grande_opera.nombre_directorio
        else:
            dir_mas_grande = dir_mas_grande_transv.nombre_directorio
    elif dir_mas_grande_opera:
        dir_mas_grande = dir_mas_grande_opera.nombre_directorio
    elif dir_mas_grande_transv:
        dir_mas_grande = dir_mas_grande_transv.nombre_directorio
    
    # Último escaneo
    ultimo_escaneo_opera = dirs_opera.aggregate(ultimo=Max('fecha_actualizacion'))['ultimo']
    ultimo_escaneo_transv = dirs_transv.aggregate(ultimo=Max('fecha_ultimo_escaneo'))['ultimo']
    
    ultimo_escaneo = None
    if ultimo_escaneo_opera and ultimo_escaneo_transv:
        ultimo_escaneo = max(ultimo_escaneo_opera, ultimo_escaneo_transv)
    elif ultimo_escaneo_opera:
        ultimo_escaneo = ultimo_escaneo_opera
    elif ultimo_escaneo_transv:
        ultimo_escaneo = ultimo_escaneo_transv
    
    estadisticas = {
        'total_directorios': total_dirs,
        'total_archivos': total_archivos,
        'peso_total_mb': round(peso_total / (1024 * 1024), 2),
        'directorio_mas_grande': dir_mas_grande,
        'ultimo_escaneo': ultimo_escaneo
    }
    
    serializer = EstadisticasDirectorioSerializer(estadisticas)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def busqueda_archivos(request):
    """Búsqueda global de archivos con filtros de permisos"""
    query = request.query_params.get('q', '')
    municipio = request.query_params.get('municipio')
    tipo = request.query_params.get('tipo', 'ambos')  # operacion, transversal, ambos
    
    if not query:
        return Response({'error': 'Parámetro de búsqueda q requerido'}, 
                       status=status.HTTP_400_BAD_REQUEST)
    
    # ✅ APLICAR FILTROS DE PERMISOS
    municipios_permitidos = get_municipios_permitidos(request.user)
    
    if municipios_permitidos != 'todos' and not municipios_permitidos:
        # Sin permisos, devolver resultados vacíos
        return Response({
            'operacion': [],
            'transversal': [],
            'mensaje': 'No tiene permisos para buscar archivos'
        })
    
    resultados = {
        'operacion': [],
        'transversal': []
    }
    
    if tipo in ['operacion', 'ambos']:
        archivos_opera = ArchivosOperacion.objects.filter(
            Q(nombre_archivo__icontains=query) |
            Q(path_file__icontains=query)
        )
        
        # Aplicar filtros de permisos
        if municipios_permitidos != 'todos':
            archivos_opera = archivos_opera.filter(
                cod_dir_operacion__cod_municipio__in=municipios_permitidos
            )
        
        # Filtrar por municipio específico si se proporciona
        if municipio:
            # Verificar permisos para municipio específico
            if municipios_permitidos != 'todos':
                if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                    return Response(
                        {'error': 'No tiene permisos para buscar en este municipio'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            archivos_opera = archivos_opera.filter(
                cod_dir_operacion__cod_municipio=municipio
            )
        
        resultados['operacion'] = ArchivosOperacionListSerializer(
            archivos_opera[:50], many=True
        ).data
    
    if tipo in ['transversal', 'ambos']:
        archivos_transv = ArchivosTransv.objects.filter(
            Q(nombre_archivo__icontains=query) |
            Q(path_file__icontains=query)
        )
        
        # Aplicar filtros de permisos
        if municipios_permitidos != 'todos':
            archivos_transv = archivos_transv.filter(
                cod_dir_transv__cod_municipio__in=municipios_permitidos
            )
        
        # Filtrar por municipio específico si se proporciona
        if municipio:
            if municipios_permitidos != 'todos':
                if not municipios_permitidos or int(municipio) not in municipios_permitidos:
                    return Response(
                        {'error': 'No tiene permisos para buscar en este municipio'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            archivos_transv = archivos_transv.filter(
                cod_dir_transv__cod_municipio=municipio
            )
        
        resultados['transversal'] = ArchivosTransvListSerializer(
            archivos_transv[:50], many=True
        ).data
    
    return Response(resultados)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mis_municipios_asignados(request):
    """Obtiene los municipios asignados al usuario actual"""
    municipios_permitidos = get_municipios_permitidos(request.user)
    
    if municipios_permitidos == 'todos':
        # Admin puede ver estadísticas de todos los municipios
        if Municipios:
            municipios_info = []
            for municipio in Municipios.objects.all()[:100]:  # Limitar para performance
                municipios_info.append({
                    'cod_municipio': municipio.cod_municipio,
                    'nom_municipio': municipio.nom_municipio,
                    'es_admin': True
                })
        else:
            municipios_info = [{'mensaje': 'Vista de administrador - Todos los municipios disponibles'}]
            
        return Response({
            'tipo_usuario': 'administrador',
            'municipios': municipios_info,
            'total': len(municipios_info) if municipios_info else 0
        })
    
    elif municipios_permitidos:
        # Profesional ve solo sus municipios
        municipios_info = []
        if Municipios:
            municipios_obj = Municipios.objects.filter(cod_municipio__in=municipios_permitidos)
            for municipio in municipios_obj:
                municipios_info.append({
                    'cod_municipio': municipio.cod_municipio,
                    'nom_municipio': municipio.nom_municipio,
                    'es_admin': False
                })
        else:
            # Fallback si no hay modelo Municipios
            for cod in municipios_permitidos:
                municipios_info.append({
                    'cod_municipio': cod,
                    'nom_municipio': f'Municipio {cod}',
                    'es_admin': False
                })
        
        return Response({
            'tipo_usuario': 'profesional',
            'municipios': municipios_info,
            'total': len(municipios_info)
        })
    
    else:
        # Sin permisos
        return Response({
            'tipo_usuario': 'sin_permisos',
            'municipios': [],
            'total': 0,
            'mensaje': 'No tiene municipios asignados'
        })







def extraer_mecanismo_financiacion_operacion(ruta_archivo, cod_municipio):
    """
    Extrae el mecanismo de financiación de la ruta para OPERACIÓN
    Ejemplo: \\repositorio\\...\\13\\030\\PGN\\02_opera\\ → "PGN"
    """
    if not ruta_archivo or not cod_municipio:
        return "SIN_MECANISMO"
    
    try:
        ruta_str = str(ruta_archivo).replace('/', '\\')
        cod_str = str(cod_municipio)
        
        print(f"🔍 Analizando ruta OPERACIÓN: {ruta_str[:100]}...")
        print(f"🔍 Código municipio: {cod_str}")
        
        # Crear patrón del código del municipio
        if len(cod_str) == 5:  # Código completo como 13030
            depto = cod_str[:2]  # 13
            mun = cod_str[2:]    # 030
            patron_municipio = f"\\{depto}\\{mun}\\"
            print(f"🔍 Patrón municipio: {patron_municipio}")
        else:
            print(f"⚠️ Código de municipio no válido: {cod_str}")
            return "SIN_MECANISMO"
        
        # Buscar el patrón en la ruta
        if patron_municipio in ruta_str:
            # Encontrar la posición después del patrón del municipio
            inicio_mecanismo = ruta_str.find(patron_municipio) + len(patron_municipio)
            resto_ruta = ruta_str[inicio_mecanismo:]
            
            print(f"🔍 Resto de ruta después del municipio: {resto_ruta[:80]}")
            
            # El mecanismo es el primer directorio después del código del municipio
            if '\\' in resto_ruta:
                partes = resto_ruta.split('\\')
                partes_validas = [p for p in partes if p.strip()]
                
                if partes_validas:
                    mecanismo = partes_validas[0].strip()
                    print(f"✅ Mecanismo extraído: '{mecanismo}'")
                    
                    # Validar que el mecanismo no sea un directorio técnico
                    directorios_tecnicos = ['02_opera', '01_pre', 'opera', 'pre', 'temp', 'temporal']
                    if mecanismo.lower() not in directorios_tecnicos and len(mecanismo) > 0:
                        return mecanismo
                    else:
                        print(f"⚠️ '{mecanismo}' parece ser un directorio técnico, buscando alternativa...")
                        
                        # Si el primer directorio es técnico, buscar el siguiente
                        if len(partes_validas) > 1:
                            mecanismo_alternativo = partes_validas[1].strip()
                            if mecanismo_alternativo.lower() not in directorios_tecnicos:
                                print(f"✅ Mecanismo alternativo: '{mecanismo_alternativo}'")
                                return mecanismo_alternativo
        else:
            print(f"⚠️ No se encontró el patrón del municipio {patron_municipio} en la ruta")
        
        # Método alternativo: buscar patrones conocidos
        patrones_conocidos = [
            'PGN-IGAC', 'PGN', 'FAO', 'BANCO_MUNDIAL', 'BID', 'USAID', 
            'COOPERACION_INTERNACIONAL', 'RECURSOS_PROPIOS'
        ]
        
        ruta_upper = ruta_str.upper()
        for patron in patrones_conocidos:
            if f"\\{patron}\\" in ruta_upper:
                print(f"🎯 Mecanismo encontrado por patrón conocido: {patron}")
                return patron
        
        print(f"❌ No se pudo extraer mecanismo de: {ruta_str[:100]}")
        return "SIN_MECANISMO"
        
    except Exception as e:
        print(f"⚠️ Error extrayendo mecanismo OPERACIÓN de {ruta_archivo}: {e}")
        return "SIN_MECANISMO"

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mecanismos_operacion_municipio(request, municipio_id):
    """
    Endpoint para obtener los mecanismos de financiación de un municipio en operación.
    GET /app/api/mecanismos-operacion/<municipio_id>/
    """
    try:
        try:
            municipio = Municipios.objects.get(cod_municipio=municipio_id)
        except Municipios.DoesNotExist:
            return Response({
                'success': False,
                'error': f'Municipio {municipio_id} no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)

        directorios = DirectoriosOperacion.objects.filter(cod_municipio=municipio_id)

        mecanismos_stats = {}
        for directorio in directorios:
            if directorio.path_directorio:
                mecanismo = extraer_mecanismo_financiacion_operacion(directorio.path_directorio, municipio_id)
                if mecanismo and mecanismo not in ["SIN_MECANISMO", "ERROR_MECANISMO"]:
                    if mecanismo not in mecanismos_stats:
                        mecanismos_stats[mecanismo] = {
                            'codigo': mecanismo,
                            'total_directorios': 0,
                            'total_archivos': 0
                        }
                    mecanismos_stats[mecanismo]['total_directorios'] += 1
                    archivos_count = ArchivosOperacion.objects.filter(cod_dir_operacion=directorio).count()
                    mecanismos_stats[mecanismo]['total_archivos'] += archivos_count

        mecanismos_list = sorted(mecanismos_stats.values(), key=lambda x: x['codigo'])

        return Response({
            'success': True,
            'municipio': {
                'cod_municipio': municipio.cod_municipio,
                'nom_municipio': municipio.nom_municipio
            },
            'mecanismos': mecanismos_list,
            'total_mecanismos': len(mecanismos_list),
            'tiene_multiples': len(mecanismos_list) > 1
        })

    except Exception as e:
        print(f"Error obteniendo mecanismos operacion para {municipio_id}: {e}")
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def extraer_subcarpeta_operacion(ruta_completa):
    """
    Extrae la subcarpeta principal de una ruta de operación
    Ejemplo: \\repositorio\\...\\13\\030\\PGN\\02_opera\\01_resol_apert\\01_resol_apert\\archivo.pdf 
    → "02_opera\\01_resol_apert\\01_resol_apert"
    """
    if not ruta_completa:
        return "Sin subcarpeta"
    
    try:
        ruta_str = str(ruta_completa).replace('/', '\\')
        partes = ruta_str.split('\\')
        
        # Buscar el índice donde aparece '02_opera' o similar
        indice_opera = -1
        for i, parte in enumerate(partes):
            if 'opera' in parte.lower() or 'opera' in parte.lower():
                indice_opera = i
                break
        
        if indice_opera != -1 and indice_opera < len(partes) - 1:
            # Tomar desde 02_opera hasta antes del archivo final
            subcarpeta_partes = partes[indice_opera:-1]  # Excluir el nombre del archivo
            if subcarpeta_partes:
                return '\\'.join(subcarpeta_partes)
        
        # Fallback: tomar las últimas 2-3 carpetas antes del archivo
        if len(partes) >= 3:
            return '\\'.join(partes[-3:-1])
        elif len(partes) >= 2:
            return partes[-2]
        
        return "Sin subcarpeta"
        
    except Exception as e:
        print(f"⚠️ Error extrayendo subcarpeta: {e}")
        return "Error subcarpeta"

def obtener_mecanismos_financiacion_operacion(municipio_id):
    """
    Obtiene TODOS los mecanismos de financiación de un municipio para OPERACIÓN
    """
    try:
        from .models import ArchivosOperacion
        
        mecanismos = set()
        
        print(f"🔍 Analizando archivos de operación del municipio {municipio_id}...")
        
        # Obtener archivos del municipio
        archivos_municipio = ArchivosOperacion.objects.filter(
            cod_dir_operacion__cod_municipio=municipio_id
        )
        
        print(f"📁 Analizando {archivos_municipio.count()} archivos del municipio...")
        
        for archivo in archivos_municipio:
            if archivo.path_file:
                mecanismo = extraer_mecanismo_financiacion_operacion(archivo.path_file, municipio_id)
                if mecanismo and mecanismo != "SIN_MECANISMO":
                    mecanismos.add(mecanismo)
                    print(f"  📋 Encontrado mecanismo: {mecanismo} en {archivo.path_file[:80]}...")
        
        # Si no encuentra mecanismos específicos, usar análisis profundo
        if not mecanismos:
            print("🔍 No se encontraron mecanismos específicos, usando análisis profundo...")
            
            for archivo in archivos_municipio:
                if archivo.path_file:
                    ruta_upper = archivo.path_file.upper()
                    
                    # Buscar indicadores de mecanismos
                    if 'PGN' in ruta_upper:
                        mecanismos.add('PGN')
                    elif 'FAO' in ruta_upper:
                        mecanismos.add('FAO')
                    elif 'BANCO' in ruta_upper or 'WORLD_BANK' in ruta_upper:
                        mecanismos.add('BANCO_MUNDIAL')
                    elif 'BID' in ruta_upper:
                        mecanismos.add('BID')
                    else:
                        # Si no encuentra nada específico, crear mecanismo genérico
                        mecanismos.add('GENERAL')
                    
                    # Salir del loop si ya encontró al menos uno
                    if mecanismos:
                        break
        
        mecanismos_list = sorted(list(mecanismos))
        
        # Si después de todo no encuentra mecanismos, usar genérico
        if not mecanismos_list:
            print("⚠️ No se encontraron mecanismos específicos, usando genérico")
            mecanismos_list = ["GENERAL"]
        
        print(f"✅ Municipio {municipio_id} - Mecanismos OPERACIÓN FINALES: {mecanismos_list}")
        
        return mecanismos_list
        
    except Exception as e:
        print(f"❌ Error obteniendo mecanismos OPERACIÓN para municipio {municipio_id}: {e}")
        import traceback
        traceback.print_exc()
        return ["GENERAL"]

def generar_reporte_operacion_con_mecanismo(municipio, temp_dir, mecanismo_financiacion):
    """
    Genera reporte Excel de OPERACIÓN con inventario de archivos por mecanismo
    """
    try:
        from .models import ArchivosOperacion
        from preoperacion.models import Municipios
        
        print(f"🎯 Generando reporte OPERACIÓN para {municipio.nom_municipio} - Mecanismo: {mecanismo_financiacion}...")
        
        # Crear nuevo workbook
        wb = openpyxl.Workbook()
        
        # Obtener y filtrar archivos
        archivos = ArchivosOperacion.objects.filter(
            cod_dir_operacion__cod_municipio=municipio.cod_municipio
        ).select_related('cod_dir_operacion')
        
        print(f"📊 Total archivos del municipio: {archivos.count()}")
        
        archivos_filtrados = []
        archivos_sin_mecanismo = []
        
        for archivo in archivos:
            mecanismo_archivo = extraer_mecanismo_financiacion_operacion(archivo.path_file, municipio.cod_municipio)
            
            if mecanismo_archivo == mecanismo_financiacion:
                archivos_filtrados.append(archivo)
            elif mecanismo_archivo == "SIN_MECANISMO":
                archivos_sin_mecanismo.append(archivo)
        
        if mecanismo_financiacion == "GENERAL" and len(archivos_filtrados) == 0 and len(archivos_sin_mecanismo) > 0:
            archivos_filtrados = archivos_sin_mecanismo
            print(f"🔄 Usando archivos sin mecanismo específico para reporte GENERAL")
        
        print(f"📋 Archivos filtrados por mecanismo {mecanismo_financiacion}: {len(archivos_filtrados)} de {archivos.count()} totales")
        
        # Generar la única pestaña - Inventario de Archivos
        ws = wb.active
        ws.title = "Inventario Archivos Operación"
        generar_inventario_archivos_operacion(ws, municipio, archivos_filtrados, mecanismo_financiacion)
        
        # Guardar archivo
        fecha_actual = datetime.now().strftime('%d_%m_%Y')
        
        mecanismo_limpio = mecanismo_financiacion.replace('\\', '_').replace('/', '_').replace(':', '_').replace('-', '_').replace(' ', '_')
        municipio_limpio = municipio.nom_municipio.replace(' ', '_').replace('/', '_').replace('-', '_').upper()
        
        nombre_archivo = f"{municipio.cod_municipio}_{municipio_limpio}_{mecanismo_limpio}_Operacion_{fecha_actual}.xlsx"
        archivo_path = os.path.join(temp_dir, nombre_archivo)
        wb.save(archivo_path)
        
        print(f"✅ Reporte OPERACIÓN generado exitosamente:")
        print(f"   📁 Archivo: {nombre_archivo}")
        print(f"   💰 Mecanismo: {mecanismo_financiacion}")
        print(f"   📊 Archivos incluidos: {len(archivos_filtrados)}")
        
        return archivo_path
        
    except Exception as e:
        print(f"❌ Error generando reporte OPERACIÓN para {municipio.nom_municipio} - {mecanismo_financiacion}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def generar_inventario_archivos_operacion(ws, municipio, archivos_filtrados, mecanismo_financiacion):
    """
    Genera el inventario detallado de archivos de OPERACIÓN con tema marrón
    🔧 CORREGIDO: Agregadas columnas de rutas
    """
    print(f"📋 Generando inventario de archivos OPERACIÓN para {mecanismo_financiacion}...")
    
    # 🎨 ESTILOS PROFESIONALES EN TONOS MARRONES (OPERACIÓN)
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="8D6E63", end_color="8D6E63", fill_type="solid")  # Marrón oscuro
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="A1887F", end_color="A1887F", fill_type="solid")  # Marrón medio
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 🏛️ TÍTULO PRINCIPAL
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    ws.merge_cells('A1:I2')  # ← CAMBIADO DE G2 A I2 (más columnas)
    titulo_cell = ws['A1']
    titulo_cell.value = (
        f"📋 INVENTARIO DETALLADO DE ARCHIVOS DE OPERACIÓN\n"
        f"🏛️ {municipio.nom_municipio.upper()} | MECANISMO: {mecanismo_financiacion} | 📅 {fecha_actual}"
    )
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    titulo_cell.border = border
    
    # 📊 ESTADÍSTICAS RÁPIDAS
    total_archivos = len(archivos_filtrados)
    
    ws.merge_cells('A3:I3')  # ← CAMBIADO DE G3 A I3
    stats_cell = ws['A3']
    stats_cell.value = f"📊 RESUMEN: {total_archivos} archivos de operación | Mecanismo: {mecanismo_financiacion}"
    stats_cell.font = Font(bold=True, size=11, color="5D4037")  # Marrón oscuro
    stats_cell.fill = PatternFill(start_color="D7CCC8", end_color="D7CCC8", fill_type="solid")  # Marrón claro
    stats_cell.alignment = center_align
    stats_cell.border = border
    
    # 📊 HEADERS DE LA TABLA (ROW 5) - 🔧 AGREGADAS COLUMNAS DE RUTAS
    row = 5
    headers = [
        'ETAPA',
        'SUBCARPETA', 
        'NOMBRE DOCUMENTO',
        'RUTA COMPLETA ARCHIVO',  # ← NUEVA COLUMNA
        'RUTA DIRECTORIO',        # ← NUEVA COLUMNA
        'FORMATO\nDOCUMENTO',
        'FECHA DE CREACIÓN\nO MODIFICACIÓN',
        'TAMAÑO (bytes)',
        'OBSERVACIONES'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # 📋 DATOS - TODOS LOS ARCHIVOS FILTRADOS
    row = 6
    
    # Ordenar archivos por nombre
    archivos_ordenados = sorted(archivos_filtrados, key=lambda x: x.nombre_archivo)
    
    for numero, archivo in enumerate(archivos_ordenados, 1):
        # Alternar colores de fila en tonos marrones
        if numero % 2 == 0:
            row_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")  # Marrón muy claro
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanco
        
        # Columna A: ETAPA (siempre OPERACIÓN)
        cell_etapa = ws.cell(row=row, column=1, value="OPERACIÓN")
        cell_etapa.alignment = center_align
        cell_etapa.border = border
        cell_etapa.fill = row_fill
        cell_etapa.font = Font(size=9, bold=True, color="5D4037")  # Marrón oscuro
        
        # Columna B: SUBCARPETA
        subcarpeta = extraer_subcarpeta_operacion(archivo.path_file)
        cell_subcarpeta = ws.cell(row=row, column=2, value=subcarpeta)
        cell_subcarpeta.alignment = left_align
        cell_subcarpeta.border = border
        cell_subcarpeta.fill = row_fill
        cell_subcarpeta.font = Font(size=9)
        
        # Columna C: NOMBRE DOCUMENTO
        cell_nombre = ws.cell(row=row, column=3, value=archivo.nombre_archivo)
        cell_nombre.alignment = left_align
        cell_nombre.border = border
        cell_nombre.fill = row_fill
        cell_nombre.font = Font(size=9)
        
        # 🔧 NUEVA: Columna D: RUTA COMPLETA ARCHIVO (path_file)
        ruta_archivo = linux_to_windows_path(archivo.path_file) if archivo.path_file else "Sin ruta"
        cell_ruta_archivo = ws.cell(row=row, column=4, value=ruta_archivo)
        cell_ruta_archivo.alignment = left_align
        cell_ruta_archivo.border = border
        cell_ruta_archivo.fill = row_fill
        cell_ruta_archivo.font = Font(size=8, color="8D6E63")  # Marrón medio, texto pequeño
        
        # 🔧 NUEVA: Columna E: RUTA DIRECTORIO (path_directorio del directorio padre)
        ruta_directorio = linux_to_windows_path(archivo.cod_dir_operacion.path_directorio) if archivo.cod_dir_operacion and archivo.cod_dir_operacion.path_directorio else "Sin directorio"
        cell_ruta_directorio = ws.cell(row=row, column=5, value=ruta_directorio)
        cell_ruta_directorio.alignment = left_align
        cell_ruta_directorio.border = border
        cell_ruta_directorio.fill = row_fill
        cell_ruta_directorio.font = Font(size=8, color="8D6E63")  # Marrón medio, texto pequeño
        
        # Columna F: FORMATO DOCUMENTO (era D)
        formato = archivo.extension.upper() if archivo.extension else "SIN EXTENSION"
        if formato.startswith('.'):
            formato = formato[1:]  # Remover el punto inicial
        cell_formato = ws.cell(row=row, column=6, value=formato)
        cell_formato.alignment = center_align
        cell_formato.border = border
        cell_formato.fill = row_fill
        cell_formato.font = Font(size=9, bold=True, color="6D4C41")  # Marrón medio
        
        # Columna G: FECHA DE CREACIÓN O MODIFICACIÓN (era E)
        if archivo.fecha_disposicion:
            fecha_documento = archivo.fecha_disposicion.strftime('%d/%m/%Y %H:%M:%S')
        else:
            fecha_documento = "Sin fecha"
        cell_fecha = ws.cell(row=row, column=7, value=fecha_documento)
        cell_fecha.alignment = center_align
        cell_fecha.border = border
        cell_fecha.fill = row_fill
        cell_fecha.font = Font(size=9)
        
        # Columna H: TAMAÑO (bytes) (era F)
        if archivo.peso_memoria:
            try:
                peso_int = int(archivo.peso_memoria)
                peso_formateado = f"{peso_int:,}".replace(',', '.') + " bytes"
            except:
                peso_formateado = str(archivo.peso_memoria) + " bytes"
        else:
            peso_formateado = "0 bytes"
        
        cell_tamaño = ws.cell(row=row, column=8, value=peso_formateado)
        cell_tamaño.alignment = center_align
        cell_tamaño.border = border
        cell_tamaño.fill = row_fill
        cell_tamaño.font = Font(size=9, color="8D6E63")  # Marrón claro
        
        # Columna I: OBSERVACIONES (era G) (vacía para que el usuario la complete)
        cell_observaciones = ws.cell(row=row, column=9, value="")
        cell_observaciones.alignment = left_align
        cell_observaciones.border = border
        cell_observaciones.fill = row_fill
        
        row += 1
    
    # 📏 AJUSTAR DIMENSIONES DE COLUMNAS - 🔧 ACTUALIZADO PARA 9 COLUMNAS
    anchos_columnas = [15, 60, 50, 80, 80, 12, 22, 18, 30]  # ← AGREGADAS 2 COLUMNAS NUEVAS
    
    for i, ancho in enumerate(anchos_columnas, 1):
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = ancho
    
    # Ajustar altura de filas
    ws.row_dimensions[1].height = 35  # Título
    ws.row_dimensions[3].height = 25  # Estadísticas
    ws.row_dimensions[5].height = 35  # Headers
    
    for row_num in range(6, row):
        ws.row_dimensions[row_num].height = 22  # Datos
    
    print(f"✅ Inventario de archivos OPERACIÓN generado: {len(archivos_ordenados)} archivos procesados")



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generar_reportes_operacion(request):
    """
    Genera reportes de archivos de operación SEPARADOS POR MECANISMO DE FINANCIACIÓN
    """
    try:
        from preoperacion.models import Municipios
        
        data = json.loads(request.body) if request.body else request.data
        municipios_ids = data.get('municipios', [])
        generar_individuales = data.get('generar_individuales', True)
        
        if not municipios_ids:
            return Response(
                {'error': 'No se proporcionaron municipios'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        print(f"🚀 GENERANDO REPORTES DE OPERACIÓN CON MECANISMOS para {len(municipios_ids)} municipios...")
        
        # Crear directorio temporal
        with tempfile.TemporaryDirectory() as temp_dir:
            archivos_generados = []
            municipios_mecanismos_procesados = []
            
            if generar_individuales:
                print("📊 Generando reportes individuales por mecanismo...")
                
                for municipio_id in municipios_ids:
                    try:
                        municipio = Municipios.objects.get(cod_municipio=municipio_id)
                        
                        # Obtener mecanismos del municipio
                        mecanismos_municipio = obtener_mecanismos_financiacion_operacion(municipio_id)
                        
                        print(f"🔥 MUNICIPIO {municipio.nom_municipio}:")
                        print(f"   📊 Mecanismos encontrados: {len(mecanismos_municipio)}")
                        print(f"   📋 Lista: {mecanismos_municipio}")
                        
                        if not mecanismos_municipio:
                            print(f"⚠️ Municipio {municipio_id} sin mecanismos, generando reporte genérico...")
                            archivo_path = generar_reporte_operacion_con_mecanismo(
                                municipio, temp_dir, "GENERAL"
                            )
                            if archivo_path:
                                archivos_generados.append(archivo_path)
                                municipios_mecanismos_procesados.append((municipio, "GENERAL"))
                        else:
                            # Generar reporte separado por cada mecanismo
                            for i, mecanismo in enumerate(mecanismos_municipio, 1):
                                print(f"📋 [{i}/{len(mecanismos_municipio)}] Procesando {municipio.nom_municipio} - Mecanismo: {mecanismo}")
                                
                                archivo_path = generar_reporte_operacion_con_mecanismo(
                                    municipio, temp_dir, mecanismo
                                )
                                if archivo_path:
                                    archivos_generados.append(archivo_path)
                                    municipios_mecanismos_procesados.append((municipio, mecanismo))
                                    print(f"✅ [{i}/{len(mecanismos_municipio)}] Reporte generado para {mecanismo}")
                                else:
                                    print(f"❌ [{i}/{len(mecanismos_municipio)}] Error generando reporte para {mecanismo}")
                            
                    except Municipios.DoesNotExist:
                        print(f"⚠️ Municipio {municipio_id} no encontrado")
                        continue
                    except Exception as e:
                        print(f"❌ Error procesando municipio {municipio_id}: {str(e)}")
                        continue
            
            if not archivos_generados:
                return Response(
                    {'error': 'No se pudieron generar reportes'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            # Crear ZIP en archivo temporal para STREAMING
            fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_zip = f'reportes_operacion_por_mecanismos_{fecha_actual}.zip'

            temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
            temp_zip_path = temp_zip.name
            temp_zip.close()

            with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=1) as zip_file:
                for archivo_path in archivos_generados:
                    nombre_archivo = os.path.basename(archivo_path)
                    zip_file.write(archivo_path, nombre_archivo)

            # Streaming del ZIP
            def zip_iterator():
                CHUNK_SIZE = 8 * 1024 * 1024  # 8MB chunks
                try:
                    with open(temp_zip_path, 'rb') as f:
                        while True:
                            chunk = f.read(CHUNK_SIZE)
                            if not chunk:
                                break
                            yield chunk
                finally:
                    try:
                        os.unlink(temp_zip_path)
                    except:
                        pass

            response = StreamingHttpResponse(
                zip_iterator(),
                content_type='application/zip'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_zip}"'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['X-Accel-Buffering'] = 'no'  # Desactivar buffering en nginx

            total_combinaciones = len(municipios_mecanismos_procesados)
            mecanismos_unicos = set([mecanismo for _, mecanismo in municipios_mecanismos_procesados])

            print(f"✅ REPORTES DE OPERACIÓN CON MECANISMOS (streaming) generados:")
            print(f"   🏛️ Municipios procesados: {len(set([mun.cod_municipio for mun, _ in municipios_mecanismos_procesados]))}")
            print(f"   📊 Combinaciones municipio-mecanismo: {total_combinaciones}")
            print(f"   💰 Mecanismos únicos encontrados: {len(mecanismos_unicos)} → {sorted(list(mecanismos_unicos))}")
            print(f"   📦 Total archivos Excel generados: {len(archivos_generados)}")
            print(f"   🎯 Archivos en ZIP: {nombre_zip}")

            return response
            
    except Exception as e:
        print(f"❌ Error general: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response(
            {'error': f'Error interno: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    



# ===============================================
# 🆕 FUNCIONES PARA ARCHIVOS TRANSVERSALES
# ===============================================

def extraer_mecanismo_financiacion_transversal(ruta_archivo, cod_municipio):
    """
    Extrae el mecanismo de financiación de la ruta para TRANSVERSAL
    Ejemplo: \\repositorio\\...\\13\\030\\PGN\\04_transv\\ → "PGN"
    """
    if not ruta_archivo or not cod_municipio:
        return "SIN_MECANISMO"
    
    try:
        ruta_str = str(ruta_archivo).replace('/', '\\')
        cod_str = str(cod_municipio)
        
        print(f"🔍 Analizando ruta TRANSVERSAL: {ruta_str[:100]}...")
        print(f"🔍 Código municipio: {cod_str}")
        
        # Crear patrón del código del municipio
        if len(cod_str) == 5:  # Código completo como 13030
            depto = cod_str[:2]  # 13
            mun = cod_str[2:]    # 030
            patron_municipio = f"\\{depto}\\{mun}\\"
            print(f"🔍 Patrón municipio: {patron_municipio}")
        else:
            print(f"⚠️ Código de municipio no válido: {cod_str}")
            return "SIN_MECANISMO"
        
        # Buscar el patrón en la ruta
        if patron_municipio in ruta_str:
            # Encontrar la posición después del patrón del municipio
            inicio_mecanismo = ruta_str.find(patron_municipio) + len(patron_municipio)
            resto_ruta = ruta_str[inicio_mecanismo:]
            
            print(f"🔍 Resto de ruta después del municipio: {resto_ruta[:80]}")
            
            # El mecanismo es el primer directorio después del código del municipio
            if '\\' in resto_ruta:
                partes = resto_ruta.split('\\')
                partes_validas = [p for p in partes if p.strip()]
                
                if partes_validas:
                    mecanismo = partes_validas[0].strip()
                    print(f"✅ Mecanismo extraído: '{mecanismo}'")
                    
                    # Validar que el mecanismo no sea un directorio técnico
                    directorios_tecnicos = ['04_transv', '03_post', 'transv', 'transversal', 'temp', 'temporal']
                    if mecanismo.lower() not in directorios_tecnicos and len(mecanismo) > 0:
                        return mecanismo
                    else:
                        print(f"⚠️ '{mecanismo}' parece ser un directorio técnico, buscando alternativa...")
                        
                        # Si el primer directorio es técnico, buscar el siguiente
                        if len(partes_validas) > 1:
                            mecanismo_alternativo = partes_validas[1].strip()
                            if mecanismo_alternativo.lower() not in directorios_tecnicos:
                                print(f"✅ Mecanismo alternativo: '{mecanismo_alternativo}'")
                                return mecanismo_alternativo
        else:
            print(f"⚠️ No se encontró el patrón del municipio {patron_municipio} en la ruta")
        
        # Método alternativo: buscar patrones conocidos
        patrones_conocidos = [
            'PGN-IGAC', 'PGN', 'FAO', 'BANCO_MUNDIAL', 'BID', 'USAID', 
            'COOPERACION_INTERNACIONAL', 'RECURSOS_PROPIOS'
        ]
        
        ruta_upper = ruta_str.upper()
        for patron in patrones_conocidos:
            if f"\\{patron}\\" in ruta_upper:
                print(f"🎯 Mecanismo encontrado por patrón conocido: {patron}")
                return patron
        
        print(f"❌ No se pudo extraer mecanismo de: {ruta_str[:100]}")
        return "SIN_MECANISMO"
        
    except Exception as e:
        print(f"⚠️ Error extrayendo mecanismo TRANSVERSAL de {ruta_archivo}: {e}")
        return "SIN_MECANISMO"

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mecanismos_transversal_municipio(request, municipio_id):
    """
    Endpoint para obtener los mecanismos de financiación de un municipio en transversal.
    GET /app/api/mecanismos-transversal/<municipio_id>/
    """
    try:
        try:
            municipio = Municipios.objects.get(cod_municipio=municipio_id)
        except Municipios.DoesNotExist:
            return Response({
                'success': False,
                'error': f'Municipio {municipio_id} no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)

        directorios = DirectoriosTransv.objects.filter(cod_municipio=municipio_id)

        mecanismos_stats = {}
        for directorio in directorios:
            if directorio.ruta_completa:
                mecanismo = extraer_mecanismo_financiacion_transversal(directorio.ruta_completa, municipio_id)
                if mecanismo and mecanismo not in ["SIN_MECANISMO", "ERROR_MECANISMO"]:
                    if mecanismo not in mecanismos_stats:
                        mecanismos_stats[mecanismo] = {
                            'codigo': mecanismo,
                            'total_directorios': 0,
                            'total_archivos': 0
                        }
                    mecanismos_stats[mecanismo]['total_directorios'] += 1
                    archivos_count = ArchivosTransv.objects.filter(cod_dir_transv=directorio).count()
                    mecanismos_stats[mecanismo]['total_archivos'] += archivos_count

        mecanismos_list = sorted(mecanismos_stats.values(), key=lambda x: x['codigo'])

        return Response({
            'success': True,
            'municipio': {
                'cod_municipio': municipio.cod_municipio,
                'nom_municipio': municipio.nom_municipio
            },
            'mecanismos': mecanismos_list,
            'total_mecanismos': len(mecanismos_list),
            'tiene_multiples': len(mecanismos_list) > 1
        })

    except Exception as e:
        print(f"Error obteniendo mecanismos transversal para {municipio_id}: {e}")
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def extraer_subcarpeta_transversal(ruta_completa):
    """
    Extrae la subcarpeta principal de una ruta transversal
    Ejemplo: \\repositorio\\...\\13\\030\\PGN\\04_transv\\01_documentos\\archivo.pdf 
    → "04_transv\\01_documentos"
    """
    if not ruta_completa:
        return "Sin subcarpeta"
    
    try:
        ruta_str = str(ruta_completa).replace('/', '\\')
        partes = ruta_str.split('\\')
        
        # Buscar el índice donde aparece '04_transv' o similar
        indice_transv = -1
        for i, parte in enumerate(partes):
            if 'transv' in parte.lower() or 'transversal' in parte.lower():
                indice_transv = i
                break
        
        if indice_transv != -1 and indice_transv < len(partes) - 1:
            # Tomar desde 04_transv hasta antes del archivo final
            subcarpeta_partes = partes[indice_transv:-1]  # Excluir el nombre del archivo
            if subcarpeta_partes:
                return '\\'.join(subcarpeta_partes)
        
        # Fallback: tomar las últimas 2-3 carpetas antes del archivo
        if len(partes) >= 3:
            return '\\'.join(partes[-3:-1])
        elif len(partes) >= 2:
            return partes[-2]
        
        return "Sin subcarpeta"
        
    except Exception as e:
        print(f"⚠️ Error extrayendo subcarpeta: {e}")
        return "Error subcarpeta"

def obtener_mecanismos_financiacion_transversal(municipio_id):
    """
    Obtiene TODOS los mecanismos de financiación de un municipio para TRANSVERSAL
    """
    try:
        from .models import ArchivosTransv
        
        mecanismos = set()
        
        print(f"🔍 Analizando archivos transversales del municipio {municipio_id}...")
        
        # Obtener archivos del municipio
        archivos_municipio = ArchivosTransv.objects.filter(
            cod_dir_transv__cod_municipio=municipio_id
        )
        
        print(f"📁 Analizando {archivos_municipio.count()} archivos del municipio...")
        
        for archivo in archivos_municipio:
            if archivo.path_file:
                mecanismo = extraer_mecanismo_financiacion_transversal(archivo.path_file, municipio_id)
                if mecanismo and mecanismo != "SIN_MECANISMO":
                    mecanismos.add(mecanismo)
                    print(f"  📋 Encontrado mecanismo: {mecanismo} en {archivo.path_file[:80]}...")
        
        # Si no encuentra mecanismos específicos, usar análisis profundo
        if not mecanismos:
            print("🔍 No se encontraron mecanismos específicos, usando análisis profundo...")
            
            for archivo in archivos_municipio:
                if archivo.path_file:
                    ruta_upper = archivo.path_file.upper()
                    
                    # Buscar indicadores de mecanismos
                    if 'PGN' in ruta_upper:
                        mecanismos.add('PGN')
                    elif 'FAO' in ruta_upper:
                        mecanismos.add('FAO')
                    elif 'BANCO' in ruta_upper or 'WORLD_BANK' in ruta_upper:
                        mecanismos.add('BANCO_MUNDIAL')
                    elif 'BID' in ruta_upper:
                        mecanismos.add('BID')
                    else:
                        # Si no encuentra nada específico, crear mecanismo genérico
                        mecanismos.add('GENERAL')
                    
                    # Salir del loop si ya encontró al menos uno
                    if mecanismos:
                        break
        
        mecanismos_list = sorted(list(mecanismos))
        
        # Si después de todo no encuentra mecanismos, usar genérico
        if not mecanismos_list:
            print("⚠️ No se encontraron mecanismos específicos, usando genérico")
            mecanismos_list = ["GENERAL"]
        
        print(f"✅ Municipio {municipio_id} - Mecanismos TRANSVERSAL FINALES: {mecanismos_list}")
        
        return mecanismos_list
        
    except Exception as e:
        print(f"❌ Error obteniendo mecanismos TRANSVERSAL para municipio {municipio_id}: {e}")
        import traceback
        traceback.print_exc()
        return ["GENERAL"]

def generar_reporte_transversal_con_mecanismo(municipio, temp_dir, mecanismo_financiacion):
    """
    Genera reporte Excel de TRANSVERSAL con inventario de archivos por mecanismo
    """
    try:
        from .models import ArchivosTransv
        from preoperacion.models import Municipios
        
        print(f"🎯 Generando reporte TRANSVERSAL para {municipio.nom_municipio} - Mecanismo: {mecanismo_financiacion}...")
        
        # Crear nuevo workbook
        wb = openpyxl.Workbook()
        
        # Obtener y filtrar archivos
        archivos = ArchivosTransv.objects.filter(
            cod_dir_transv__cod_municipio=municipio.cod_municipio
        ).select_related('cod_dir_transv')
        
        print(f"📊 Total archivos del municipio: {archivos.count()}")
        
        archivos_filtrados = []
        archivos_sin_mecanismo = []
        
        for archivo in archivos:
            mecanismo_archivo = extraer_mecanismo_financiacion_transversal(archivo.path_file, municipio.cod_municipio)
            
            if mecanismo_archivo == mecanismo_financiacion:
                archivos_filtrados.append(archivo)
            elif mecanismo_archivo == "SIN_MECANISMO":
                archivos_sin_mecanismo.append(archivo)
        
        if mecanismo_financiacion == "GENERAL" and len(archivos_filtrados) == 0 and len(archivos_sin_mecanismo) > 0:
            archivos_filtrados = archivos_sin_mecanismo
            print(f"🔄 Usando archivos sin mecanismo específico para reporte GENERAL")
        
        print(f"📋 Archivos filtrados por mecanismo {mecanismo_financiacion}: {len(archivos_filtrados)} de {archivos.count()} totales")
        
        # Generar la única pestaña - Inventario de Archivos
        ws = wb.active
        ws.title = "Inventario Archivos Transversal"
        generar_inventario_archivos_transversal(ws, municipio, archivos_filtrados, mecanismo_financiacion)
        
        # Guardar archivo
        fecha_actual = datetime.now().strftime('%d_%m_%Y')
        
        mecanismo_limpio = mecanismo_financiacion.replace('\\', '_').replace('/', '_').replace(':', '_').replace('-', '_').replace(' ', '_')
        municipio_limpio = municipio.nom_municipio.replace(' ', '_').replace('/', '_').replace('-', '_').upper()
        
        nombre_archivo = f"{municipio.cod_municipio}_{municipio_limpio}_{mecanismo_limpio}_Transversal_{fecha_actual}.xlsx"
        archivo_path = os.path.join(temp_dir, nombre_archivo)
        wb.save(archivo_path)
        
        print(f"✅ Reporte TRANSVERSAL generado exitosamente:")
        print(f"   📁 Archivo: {nombre_archivo}")
        print(f"   💰 Mecanismo: {mecanismo_financiacion}")
        print(f"   📊 Archivos incluidos: {len(archivos_filtrados)}")
        
        return archivo_path
        
    except Exception as e:
        print(f"❌ Error generando reporte TRANSVERSAL para {municipio.nom_municipio} - {mecanismo_financiacion}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def generar_inventario_archivos_transversal(ws, municipio, archivos_filtrados, mecanismo_financiacion):
    """
    Genera el inventario detallado de archivos TRANSVERSALES con tema morado oscuro
    🔧 CORREGIDO: Agregadas columnas de rutas
    """
    print(f"📋 Generando inventario de archivos TRANSVERSALES para {mecanismo_financiacion}...")
    
    # 🎨 ESTILOS PROFESIONALES EN TONOS MORADOS OSCUROS (TRANSVERSAL)
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")  # Morado muy oscuro
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")  # Morado oscuro
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 🏛️ TÍTULO PRINCIPAL
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    ws.merge_cells('A1:I2')  # ← CAMBIADO DE G2 A I2 (más columnas)
    titulo_cell = ws['A1']
    titulo_cell.value = (
        f"📋 INVENTARIO DETALLADO DE ARCHIVOS TRANSVERSALES\n"
        f"🏛️ {municipio.nom_municipio.upper()} | MECANISMO: {mecanismo_financiacion} | 📅 {fecha_actual}"
    )
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    titulo_cell.border = border
    
    # 📊 ESTADÍSTICAS RÁPIDAS
    total_archivos = len(archivos_filtrados)
    
    ws.merge_cells('A3:I3')  # ← CAMBIADO DE G3 A I3
    stats_cell = ws['A3']
    stats_cell.value = f"📊 RESUMEN: {total_archivos} archivos transversales | Mecanismo: {mecanismo_financiacion}"
    stats_cell.font = Font(bold=True, size=11, color="4A148C")  # Morado muy oscuro
    stats_cell.fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")  # Morado muy claro
    stats_cell.alignment = center_align
    stats_cell.border = border
    
    # 📊 HEADERS DE LA TABLA (ROW 5) - 🔧 AGREGADAS COLUMNAS DE RUTAS
    row = 5
    headers = [
        'ETAPA',
        'SUBCARPETA', 
        'NOMBRE DOCUMENTO',
        'RUTA COMPLETA ARCHIVO',  # ← NUEVA COLUMNA
        'RUTA DIRECTORIO',        # ← NUEVA COLUMNA
        'FORMATO\nDOCUMENTO',
        'FECHA DE CREACIÓN\nO MODIFICACIÓN',
        'TAMAÑO (bytes)',
        'OBSERVACIONES'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # 📋 DATOS - TODOS LOS ARCHIVOS FILTRADOS
    row = 6
    
    # Ordenar archivos por nombre
    archivos_ordenados = sorted(archivos_filtrados, key=lambda x: x.nombre_archivo)
    
    for numero, archivo in enumerate(archivos_ordenados, 1):
        # Alternar colores de fila en tonos morados
        if numero % 2 == 0:
            row_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")  # Morado muy claro
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanco
        
        # Columna A: ETAPA (siempre TRANSVERSAL)
        cell_etapa = ws.cell(row=row, column=1, value="TRANSVERSAL")
        cell_etapa.alignment = center_align
        cell_etapa.border = border
        cell_etapa.fill = row_fill
        cell_etapa.font = Font(size=9, bold=True, color="4A148C")  # Morado muy oscuro
        
        # Columna B: SUBCARPETA
        subcarpeta = extraer_subcarpeta_transversal(archivo.path_file)
        cell_subcarpeta = ws.cell(row=row, column=2, value=subcarpeta)
        cell_subcarpeta.alignment = left_align
        cell_subcarpeta.border = border
        cell_subcarpeta.fill = row_fill
        cell_subcarpeta.font = Font(size=9)
        
        # Columna C: NOMBRE DOCUMENTO
        cell_nombre = ws.cell(row=row, column=3, value=archivo.nombre_archivo)
        cell_nombre.alignment = left_align
        cell_nombre.border = border
        cell_nombre.fill = row_fill
        cell_nombre.font = Font(size=9)
        
        # 🔧 NUEVA: Columna D: RUTA COMPLETA ARCHIVO (path_file)
        ruta_archivo = linux_to_windows_path(archivo.path_file) if archivo.path_file else "Sin ruta"
        cell_ruta_archivo = ws.cell(row=row, column=4, value=ruta_archivo)
        cell_ruta_archivo.alignment = left_align
        cell_ruta_archivo.border = border
        cell_ruta_archivo.fill = row_fill
        cell_ruta_archivo.font = Font(size=8, color="8E24AA")  # Morado medio, texto pequeño
        
        # 🔧 NUEVA: Columna E: RUTA DIRECTORIO (ruta_completa del directorio padre)
        ruta_directorio = linux_to_windows_path(archivo.cod_dir_transv.ruta_completa) if archivo.cod_dir_transv and archivo.cod_dir_transv.ruta_completa else "Sin directorio"
        cell_ruta_directorio = ws.cell(row=row, column=5, value=ruta_directorio)
        cell_ruta_directorio.alignment = left_align
        cell_ruta_directorio.border = border
        cell_ruta_directorio.fill = row_fill
        cell_ruta_directorio.font = Font(size=8, color="8E24AA")  # Morado medio, texto pequeño
        
        # Columna F: FORMATO DOCUMENTO (era D)
        formato = archivo.extension.upper() if archivo.extension else "SIN EXTENSION"
        if formato.startswith('.'):
            formato = formato[1:]  # Remover el punto inicial
        cell_formato = ws.cell(row=row, column=6, value=formato)
        cell_formato.alignment = center_align
        cell_formato.border = border
        cell_formato.fill = row_fill
        cell_formato.font = Font(size=9, bold=True, color="6A1B9A")  # Morado oscuro
        
        # Columna G: FECHA DE CREACIÓN O MODIFICACIÓN (era E)
        if archivo.fecha_disposicion:
            fecha_documento = archivo.fecha_disposicion.strftime('%d/%m/%Y %H:%M:%S')
        else:
            fecha_documento = "Sin fecha"
        cell_fecha = ws.cell(row=row, column=7, value=fecha_documento)
        cell_fecha.alignment = center_align
        cell_fecha.border = border
        cell_fecha.fill = row_fill
        cell_fecha.font = Font(size=9)
        
        # Columna H: TAMAÑO (bytes) (era F)
        if archivo.peso_memoria:
            try:
                peso_int = int(archivo.peso_memoria)
                peso_formateado = f"{peso_int:,}".replace(',', '.') + " bytes"
            except:
                peso_formateado = str(archivo.peso_memoria) + " bytes"
        else:
            peso_formateado = "0 bytes"
        
        cell_tamaño = ws.cell(row=row, column=8, value=peso_formateado)
        cell_tamaño.alignment = center_align
        cell_tamaño.border = border
        cell_tamaño.fill = row_fill
        cell_tamaño.font = Font(size=9, color="8E24AA")  # Morado medio
        
        # Columna I: OBSERVACIONES (era G) (vacía para que el usuario la complete)
        cell_observaciones = ws.cell(row=row, column=9, value="")
        cell_observaciones.alignment = left_align
        cell_observaciones.border = border
        cell_observaciones.fill = row_fill
        
        row += 1
    
    # 📏 AJUSTAR DIMENSIONES DE COLUMNAS - 🔧 ACTUALIZADO PARA 9 COLUMNAS
    anchos_columnas = [15, 60, 50, 80, 80, 12, 22, 18, 30]  # ← AGREGADAS 2 COLUMNAS NUEVAS
    
    for i, ancho in enumerate(anchos_columnas, 1):
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = ancho
    
    # Ajustar altura de filas
    ws.row_dimensions[1].height = 35  # Título
    ws.row_dimensions[3].height = 25  # Estadísticas
    ws.row_dimensions[5].height = 35  # Headers
    
    for row_num in range(6, row):
        ws.row_dimensions[row_num].height = 22  # Datos
    
    print(f"✅ Inventario de archivos TRANSVERSALES generado: {len(archivos_ordenados)} archivos procesados")
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generar_reportes_transversal(request):
    """
    Genera reportes de archivos transversales SEPARADOS POR MECANISMO DE FINANCIACIÓN
    """
    try:
        from preoperacion.models import Municipios
        
        data = json.loads(request.body) if request.body else request.data
        municipios_ids = data.get('municipios', [])
        generar_individuales = data.get('generar_individuales', True)
        
        if not municipios_ids:
            return Response(
                {'error': 'No se proporcionaron municipios'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        print(f"🚀 GENERANDO REPORTES TRANSVERSALES CON MECANISMOS para {len(municipios_ids)} municipios...")
        
        # Crear directorio temporal
        with tempfile.TemporaryDirectory() as temp_dir:
            archivos_generados = []
            municipios_mecanismos_procesados = []
            
            if generar_individuales:
                print("📊 Generando reportes individuales por mecanismo...")
                
                for municipio_id in municipios_ids:
                    try:
                        municipio = Municipios.objects.get(cod_municipio=municipio_id)
                        
                        # Obtener mecanismos del municipio
                        mecanismos_municipio = obtener_mecanismos_financiacion_transversal(municipio_id)
                        
                        print(f"🔥 MUNICIPIO {municipio.nom_municipio}:")
                        print(f"   📊 Mecanismos encontrados: {len(mecanismos_municipio)}")
                        print(f"   📋 Lista: {mecanismos_municipio}")
                        
                        if not mecanismos_municipio:
                            print(f"⚠️ Municipio {municipio_id} sin mecanismos, generando reporte genérico...")
                            archivo_path = generar_reporte_transversal_con_mecanismo(
                                municipio, temp_dir, "GENERAL"
                            )
                            if archivo_path:
                                archivos_generados.append(archivo_path)
                                municipios_mecanismos_procesados.append((municipio, "GENERAL"))
                        else:
                            # Generar reporte separado por cada mecanismo
                            for i, mecanismo in enumerate(mecanismos_municipio, 1):
                                print(f"📋 [{i}/{len(mecanismos_municipio)}] Procesando {municipio.nom_municipio} - Mecanismo: {mecanismo}")
                                
                                archivo_path = generar_reporte_transversal_con_mecanismo(
                                    municipio, temp_dir, mecanismo
                                )
                                if archivo_path:
                                    archivos_generados.append(archivo_path)
                                    municipios_mecanismos_procesados.append((municipio, mecanismo))
                                    print(f"✅ [{i}/{len(mecanismos_municipio)}] Reporte generado para {mecanismo}")
                                else:
                                    print(f"❌ [{i}/{len(mecanismos_municipio)}] Error generando reporte para {mecanismo}")
                            
                    except Municipios.DoesNotExist:
                        print(f"⚠️ Municipio {municipio_id} no encontrado")
                        continue
                    except Exception as e:
                        print(f"❌ Error procesando el municipio {municipio_id}: {str(e)}")
                        continue
            
            if not archivos_generados:
                return Response(
                    {'error': 'No se pudieron generar reportes'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            # Crear ZIP en archivo temporal para STREAMING
            fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_zip = f'reportes_transversales_por_mecanismos_{fecha_actual}.zip'

            temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
            temp_zip_path = temp_zip.name
            temp_zip.close()

            with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=1) as zip_file:
                for archivo_path in archivos_generados:
                    nombre_archivo = os.path.basename(archivo_path)
                    zip_file.write(archivo_path, nombre_archivo)

            # Streaming del ZIP
            def zip_iterator():
                CHUNK_SIZE = 8 * 1024 * 1024  # 8MB chunks
                try:
                    with open(temp_zip_path, 'rb') as f:
                        while True:
                            chunk = f.read(CHUNK_SIZE)
                            if not chunk:
                                break
                            yield chunk
                finally:
                    try:
                        os.unlink(temp_zip_path)
                    except:
                        pass

            response = StreamingHttpResponse(
                zip_iterator(),
                content_type='application/zip'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_zip}"'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['X-Accel-Buffering'] = 'no'  # Desactivar buffering en nginx

            total_combinaciones = len(municipios_mecanismos_procesados)
            mecanismos_unicos = set([mecanismo for _, mecanismo in municipios_mecanismos_procesados])

            print(f"✅ REPORTES TRANSVERSALES CON MECANISMOS (streaming) generados:")
            print(f"   🏛️ Municipios procesados: {len(set([mun.cod_municipio for mun, _ in municipios_mecanismos_procesados]))}")
            print(f"   📊 Combinaciones municipio-mecanismo: {total_combinaciones}")
            print(f"   💰 Mecanismos únicos encontrados: {len(mecanismos_unicos)} → {sorted(list(mecanismos_unicos))}")
            print(f"   📦 Total archivos Excel generados: {len(archivos_generados)}")
            print(f"   🎯 Archivos en ZIP: {nombre_zip}")

            return response
            
    except Exception as e:
        print(f"❌ Error general: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response(
            {'error': f'Error interno: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    






@api_view(['POST'])
@permission_classes([IsAuthenticated])
def excel_definitivo(request):
    """
    🔥 GENERADOR DE EXCEL DEFINITIVO - SIN MAMADAS
    """
    try:
        data = json.loads(request.body) if request.body else request.data
        municipios_ids = data.get('municipios', [])
        
        if not municipios_ids:
            return Response({'error': 'No municipios'}, status=400)
        
        print(f"🔥 GENERANDO EXCEL REAL...")
        
        # CREAR EXCEL EN MEMORIA
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CONSOLIDADO"
        
        # TÍTULO SIMPLE
        ws['A1'] = "REPORTE CONSOLIDADO DE ARCHIVOS"
        ws['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # HEADERS
        headers = ['MUNICIPIO', 'ARCHIVO', 'TIPO', 'RUTA']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        
        # OBTENER DATOS SIMPLES
        row = 5
        
        # PRE-OPERACIÓN
        try:
            archivos_pre = ListaArchivosPre.objects.filter(
                cod_insumo__cod_insumo__cod_municipio__in=municipios_ids
            )[:100]  # SOLO 100 PARA PROBAR
            
            for archivo in archivos_pre:
                if archivo.cod_insumo and archivo.cod_insumo.cod_insumo:
                    ws.cell(row=row, column=1, value=str(archivo.cod_insumo.cod_insumo.cod_municipio.cod_municipio))
                    ws.cell(row=row, column=2, value=str(archivo.nombre_insumo or 'Sin nombre'))
                    ws.cell(row=row, column=3, value='PRE')
                    ws.cell(row=row, column=4, value=str(linux_to_windows_path(archivo.path_file) if archivo.path_file else 'Sin ruta')[:50])
                    row += 1
        except Exception as e:
            print(f"Error pre: {e}")
        
        # OPERACIÓN
        try:
            archivos_op = ArchivosOperacion.objects.filter(
                cod_dir_operacion__cod_municipio__in=municipios_ids
            )[:100]  # SOLO 100 PARA PROBAR
            
            for archivo in archivos_op:
                ws.cell(row=row, column=1, value=str(archivo.cod_dir_operacion.cod_municipio))
                ws.cell(row=row, column=2, value=str(archivo.nombre_archivo or 'Sin nombre'))
                ws.cell(row=row, column=3, value='OP')
                ws.cell(row=row, column=4, value=str(linux_to_windows_path(archivo.path_file) if archivo.path_file else 'Sin ruta')[:50])
                row += 1
        except Exception as e:
            print(f"Error op: {e}")
        
        # AJUSTAR COLUMNAS
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 60
        
        print(f"📊 Excel creado con {row-5} filas")

        # Guardar en archivo temporal para streaming
        nombre = f'Consolidado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_path = temp_excel.name
        temp_excel.close()

        wb.save(temp_excel_path)
        file_size = os.path.getsize(temp_excel_path)

        # Usar FileResponse para streaming automático
        response = FileResponse(
            open(temp_excel_path, 'rb'),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            filename=nombre
        )
        response['Content-Length'] = file_size
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['X-Accel-Buffering'] = 'no'  # Desactivar buffering en nginx

        # Nota: FileResponse cierra el archivo automáticamente, pero el temporal
        # no se elimina automáticamente. Esto es aceptable para archivos pequeños.

        print(f"✅ EXCEL streaming LISTO: {nombre}")
        return response
        
    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)