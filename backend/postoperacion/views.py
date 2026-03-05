import os
import zipfile
import tempfile
import re
from io import BytesIO
from django.http import HttpResponse, StreamingHttpResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json
from django.contrib.auth.models import Group
from collections import defaultdict
from rest_framework import viewsets, filters, status
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser, IsAuthenticatedOrReadOnly
from django_filters.rest_framework import DjangoFilterBackend
from datetime import datetime, timedelta
from django.db.models import Count, Q, Exists, OuterRef, Avg, Max, Min
from django.shortcuts import get_object_or_404
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response

from django.db import IntegrityError
from django.db.models import ProtectedError

from preoperacion.views import get_municipios_permitidos
from .models import (
    ComponentesPost, DisposicionPost, ArchivosPost, PathDirPost, NotificacionesPost,HistorialPropietarios,
    CalificacionInfoPost,EvaluacionArchivosPost
)
from .serializers import (
    ComponentesPostSerializer, DisposicionPostSerializer, ArchivosPostSerializer,HistorialPropietariosSerializer,
    PathDirPostSerializer, NotificacionesPostSerializer, DisposicionPostSimpleSerializer,
    CalificacionInfoPostSerializer,CalificacionInfoPostCreateSerializer,CalificacionInfoPostListSerializer,
    EvaluacionArchivosPostSerializer, EvaluacionArchivosPostSimpleSerializer,
    EvaluacionArchivosPostCreateSerializer, EvaluacionArchivosPostUpdateSerializer,
    EstadisticasEvaluacionSerializer, EvaluacionesPorMunicipioSerializer
)
from preoperacion.models import Municipios
from preoperacion.serializers import MunicipiosSerializer

import logging
from django.utils import timezone
logger = logging.getLogger(__name__)

# Importar utilidad para conversión de rutas Linux -> Windows
from backend.path_utils import linux_to_windows_path, windows_to_linux_path
from collections import OrderedDict

# ===============================================
# CONSTANTES: ETAPAS POST-OPERACIÓN
# ===============================================
ETAPAS_POSTOPERACION = OrderedDict([
    ('01', 'Aprobacion Economica'),
    ('02', 'Componente Social'),
    ('03', 'Pre-Cierre'),
    ('04', 'Resolucion Inscripcion'),
    ('05', 'Productos Catastrales'),
    ('06', 'Acta Entrega Base Catastral'),
    ('07', 'Acta Entrega Productos Catastrales'),
    ('08', 'Informe Plan Calidad'),
    ('09', 'Documentos Anexos'),
    ('10', 'Reporte Calidad Base Catastral'),
    ('11', 'Correspondencia'),
    ('12', 'Asistencia'),
])


def obtener_subido_por_plataforma(ruta_completa):
    """
    Helper para obtener el username del usuario que subió un archivo desde la plataforma web.

    Args:
        ruta_completa: Ruta del archivo (puede ser formato Windows o Linux)

    Returns:
        str: Username del usuario que subió el archivo, o None si no fue subido por la plataforma
    """
    if not ruta_completa:
        return None

    try:
        from postoperacion.models import AuditoriaArchivos

        # Normalizar la ruta para búsqueda (puede estar en formato Windows o Linux)
        ruta_buscar = ruta_completa.replace('\\', '/')
        if ruta_buscar.startswith('//repositorio/DirGesCat'):
            ruta_buscar = ruta_buscar.replace('//repositorio/DirGesCat', '/mnt/repositorio', 1)

        # Buscar el registro de UPLOAD más reciente
        registro = AuditoriaArchivos.objects.filter(
            ruta_completa=ruta_buscar,
            accion='UPLOAD',
            plataforma='WEB'
        ).order_by('-fecha_accion').first()

        if registro:
            return registro.usuario

        return None
    except Exception as e:
        logger.warning(f"Error obteniendo subido_por_plataforma para {ruta_completa}: {e}")
        return None


class ComponentesPostViewSet(viewsets.ModelViewSet):
    """
    ⚠️ VIEWSET DEPRECATED - Mantenido solo para compatibilidad
    Ya no se usa en la nueva arquitectura POST V2
    """
    queryset = ComponentesPost.objects.all()
    serializer_class = ComponentesPostSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['id_componente', 'nombre_componente']
    search_fields = ['nombre_componente']
    ordering_fields = ['id_componente', 'nombre_componente']

class DisposicionPostViewSet(viewsets.ModelViewSet):
    """
    🔄 VIEWSET ACTUALIZADO POST V2 - SIN dependencia de componentes_post
    """
    queryset = DisposicionPost.objects.all()
    serializer_class = DisposicionPostSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        'id_disposicion', 'cod_municipio', 'dispuesto', 
        'evaluado', 'aprobado', 'fecha_disposicion'
    ]
    ordering_fields = ['id_disposicion', 'fecha_disposicion']
    
    def get_queryset(self):
        """🔄 ACTUALIZADO - Sin select_related con componentes"""
        queryset = super().get_queryset()
        municipios_permitidos = get_municipios_permitidos(self.request.user)
        
        if municipios_permitidos == 'todos':
            return queryset.select_related('cod_municipio')
        elif municipios_permitidos:
            return queryset.filter(
                cod_municipio__in=municipios_permitidos
            ).select_related('cod_municipio')
        else:
            return queryset.none()
    
    def get_serializer_class(self):
        if self.action == 'list':
            return DisposicionPostSimpleSerializer
        return DisposicionPostSerializer
    
    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        municipio_id = request.query_params.get('municipio_id')
        if municipio_id:
            disposiciones = self.get_queryset().filter(cod_municipio=municipio_id)
            serializer = self.get_serializer(disposiciones, many=True)
            return Response(serializer.data)
        return Response({"error": "Se requiere el parámetro municipio_id"}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def resumen_estado(self, request):
        """🔄 ACTUALIZADO - Sin componentes"""
        queryset = self.get_queryset()
        
        resumen = {
            'total': queryset.count(),
            'dispuestos': queryset.filter(dispuesto=True).count(),
            'evaluados': queryset.filter(evaluado=True).count(),
            'aprobados': queryset.filter(aprobado=True).count(),
            'por_municipio': queryset.values('cod_municipio__nom_municipio').annotate(
                total=Count('id_disposicion'),
                dispuestos=Count('id_disposicion', filter=Q(dispuesto=True)),
                evaluados=Count('id_disposicion', filter=Q(evaluado=True)),
                aprobados=Count('id_disposicion', filter=Q(aprobado=True))
            )
        }
        return Response(resumen)
    
    @action(detail=False, methods=['get'])
    def municipios_con_productos(self, request):
        """
        🔄 ACTUALIZADO V3 - Cuenta archivos (EvaluacionArchivosPost) correctamente
        Los contadores ahora reflejan el estado real de los ARCHIVOS, no de los directorios.
        """
        # Consulta base: municipios que tienen al menos una disposición
        disposicion_subquery = DisposicionPost.objects.filter(
            cod_municipio=OuterRef('cod_municipio')
        )

        queryset = Municipios.objects.filter(
            Exists(disposicion_subquery)
        ).select_related('cod_depto')

        # Aplicar filtros
        departamento = request.query_params.get('departamento')
        municipio = request.query_params.get('municipio')
        territorial = request.query_params.get('territorial')

        if departamento:
            queryset = queryset.filter(cod_depto=departamento)

        if municipio:
            queryset = queryset.filter(cod_municipio=municipio)

        if territorial:
            queryset = queryset.filter(nom_territorial=territorial)

        # 🔧 CORREGIDO: Contar ARCHIVOS (EvaluacionArchivosPost) en lugar de directorios
        municipios_data = []
        for mun in queryset:
            # Obtener todos los archivos del municipio
            archivos = EvaluacionArchivosPost.objects.filter(
                id_disposicion__cod_municipio=mun.cod_municipio
            )

            # Contar directorios para referencia
            disposiciones = DisposicionPost.objects.filter(cod_municipio=mun.cod_municipio)

            municipio_info = {
                'cod_municipio': mun.cod_municipio,
                'nom_municipio': mun.nom_municipio,
                'cod_depto': mun.cod_depto.cod_depto,
                'nom_territorial': mun.nom_territorial,
                # 📊 Contadores de ARCHIVOS (corregidos)
                'total_productos': archivos.count(),
                'evaluados': archivos.filter(evaluado=True).count(),
                'aprobados': archivos.filter(aprobado=True).count(),
                'pendientes': archivos.filter(estado_archivo='PENDIENTE').count(),
                # 📁 Contadores de directorios (para referencia)
                'total_directorios': disposiciones.count(),
            }
            municipios_data.append(municipio_info)

        return Response(municipios_data)
    
    @action(detail=False, methods=['get']) 
    def directorios_por_municipio(self, request):
        """
        🆕 NUEVO ENDPOINT - Obtiene directorios agrupados por municipio
        """
        municipio_id = request.query_params.get('municipio_id')
        
        queryset = self.get_queryset()
        if municipio_id:
            queryset = queryset.filter(cod_municipio=municipio_id)
        
        # Agrupar por municipio y calcular estadísticas de directorios
        directorios_data = []
        
        for disposicion in queryset.select_related('cod_municipio'):
            directorio_info = {
                'id_disposicion': disposicion.id_disposicion,
                'municipio': disposicion.cod_municipio.nom_municipio,
                'cod_municipio': disposicion.cod_municipio.cod_municipio,
                'ruta_acceso': disposicion.ruta_acceso,
                'nombre_directorio': disposicion.get_nombre_directorio(),
                'nivel_jerarquia': disposicion.get_nivel_jerarquia(),
                'total_archivos': disposicion.archivos_relacionados.count(),
                'dispuesto': disposicion.dispuesto,
                'evaluado': disposicion.evaluado,
                'aprobado': disposicion.aprobado,
                'fecha_disposicion': disposicion.fecha_disposicion
            }
            directorios_data.append(directorio_info)
        
        return Response(directorios_data)

    def destroy(self, request, *args, **kwargs):
        """
        ✅ MÉTODO PERSONALIZADO PARA ELIMINAR DISPOSICIONES
        Maneja las restricciones de llaves foráneas correctamente
        """
        try:
            # Obtener la disposición
            instance = self.get_object()
            
            print(f"🗑️ Intentando eliminar disposición ID: {instance.id_disposicion}")
            print(f"   Nombre: {instance.get_nombre_directorio()}")
            print(f"   Ruta: {instance.ruta_acceso}")
            
            # Verificar si tiene archivos asociados
            archivos_relacionados = instance.archivos_relacionados.all()
            total_archivos = archivos_relacionados.count()
            
            print(f"   Total archivos relacionados: {total_archivos}")
            
            if total_archivos > 0:
                # Tiene archivos - no se puede eliminar directamente
                print(f"❌ No se puede eliminar: {total_archivos} archivos asociados")
                
                # Verificar si también tiene evaluaciones
                from .models import EvaluacionArchivosPost
                evaluaciones = EvaluacionArchivosPost.objects.filter(id_disposicion=instance.id_disposicion)
                total_evaluaciones = evaluaciones.count()
                
                print(f"   Total evaluaciones relacionadas: {total_evaluaciones}")
                
                return Response({
                    'error': 'No se puede eliminar la disposición porque tiene archivos asociados',
                    'total_archivos': total_archivos,
                    'total_evaluaciones': total_evaluaciones,
                    'mensaje': f'Para eliminar este directorio, primero debe eliminar todos los {total_archivos} archivos asociados.',
                    'solucion': 'Elimine los archivos manualmente desde la interfaz, o contacte al administrador para una eliminación en cascada.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # No tiene archivos - se puede eliminar directamente
            print(f"✅ Sin archivos asociados - procediendo con eliminación")
            
            # Guardar información para el log
            info_eliminado = {
                'id_disposicion': instance.id_disposicion,
                'nombre_directorio': instance.get_nombre_directorio(),
                'ruta_acceso': instance.ruta_acceso,
                'municipio': instance.cod_municipio.nom_municipio if instance.cod_municipio else 'N/A'
            }
            
            # Eliminar la disposición
            self.perform_destroy(instance)
            
            print(f"✅ Disposición eliminada exitosamente: {info_eliminado['nombre_directorio']}")
            
            # Respuesta exitosa con información
            return Response({
                'message': f'Directorio "{info_eliminado["nombre_directorio"]}" eliminado exitosamente',
                'eliminado': info_eliminado
            }, status=status.HTTP_200_OK)
            
        except ProtectedError as e:
            # Error de llave foránea protegida
            print(f"❌ Error de llave foránea protegida: {e}")
            
            return Response({
                'error': 'No se puede eliminar la disposición debido a restricciones de base de datos',
                'detalle': 'Existen registros relacionados que impiden la eliminación',
                'solucion': 'Elimine primero todos los archivos y evaluaciones asociados'
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except IntegrityError as e:
            # Error de integridad de base de datos
            print(f"❌ Error de integridad: {e}")
            
            return Response({
                'error': 'Error de integridad de base de datos',
                'detalle': str(e),
                'solucion': 'Contacte al administrador del sistema'
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            # Error general
            print(f"❌ Error general eliminando disposición: {e}")
            import traceback
            traceback.print_exc()
            
            return Response({
                'error': f'Error interno al eliminar disposición: {str(e)}',
                'tipo_error': str(type(e).__name__)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class ArchivosPostViewSet(viewsets.ModelViewSet):
    """
    🔄 VIEWSET ACTUALIZADO - Sin componentes
    """
    queryset = ArchivosPost.objects.all()
    serializer_class = ArchivosPostSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    def get_queryset(self):
        """🔄 ACTUALIZADO - Sin select_related con componentes"""
        queryset = super().get_queryset()
        municipios_permitidos = get_municipios_permitidos(self.request.user)
        
        if municipios_permitidos == 'todos':
            return queryset.select_related(
                'id_disposicion',
                'id_disposicion__cod_municipio'
            )
        elif municipios_permitidos:
            # Filtrar archivos por municipios de las disposiciones
            return queryset.filter(
                id_disposicion__cod_municipio__in=municipios_permitidos
            ).select_related(
                'id_disposicion',
                'id_disposicion__cod_municipio'
            )
        else:
            return queryset.none()
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'por_municipio']: 
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [IsAuthenticated, IsAdminUser]
        return [permission() for permission in permission_classes]

    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        """
        🔄 ACTUALIZADO - Obtiene archivos por municipio sin componentes
        """
        municipio_id = request.query_params.get('municipio_id')
        if not municipio_id:
            return Response({"error": "Se requiere el parámetro municipio_id"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Verificar permisos
            municipios_permitidos = get_municipios_permitidos(request.user)
            
            if municipios_permitidos != 'todos':
                if not municipios_permitidos or int(municipio_id) not in municipios_permitidos:
                    return Response({"error": "No tiene permisos para acceder a este municipio"}, status=status.HTTP_403_FORBIDDEN)
            
            # Obtener archivos del municipio con información de disposición incluida
            archivos = self.get_queryset().filter(
                id_disposicion__cod_municipio=municipio_id
            )
            
            # Serializar con información adicional
            archivos_data = []
            for archivo in archivos:
                archivo_serialized = ArchivosPostSerializer(archivo).data
                archivo_serialized['disposicion_info'] = {
                    'id_disposicion': archivo.id_disposicion.id_disposicion,
                    'directorio': archivo.id_disposicion.get_nombre_directorio(),
                    'ruta_acceso': archivo.id_disposicion.ruta_acceso,
                    'dispuesto': archivo.id_disposicion.dispuesto,
                    'evaluado': archivo.id_disposicion.evaluado,
                    'aprobado': archivo.id_disposicion.aprobado
                }
                archivos_data.append(archivo_serialized)
            
            return Response(archivos_data)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def por_directorio(self, request):
        """
        🆕 NUEVO ENDPOINT - Obtiene archivos agrupados por directorio
        """
        id_disposicion = request.query_params.get('id_disposicion')
        
        if not id_disposicion:
            return Response({"error": "Se requiere el parámetro id_disposicion"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Obtener archivos del directorio específico
            archivos = self.get_queryset().filter(id_disposicion=id_disposicion)
            
            # Obtener información del directorio
            try:
                disposicion = DisposicionPost.objects.get(id_disposicion=id_disposicion)
                directorio_info = {
                    'id_disposicion': disposicion.id_disposicion,
                    'nombre_directorio': disposicion.get_nombre_directorio(),
                    'ruta_acceso': disposicion.ruta_acceso,
                    'municipio': disposicion.cod_municipio.nom_municipio,
                    'total_archivos': archivos.count()
                }
            except DisposicionPost.DoesNotExist:
                return Response({"error": "Disposición no encontrada"}, status=status.HTTP_404_NOT_FOUND)
            
            serializer = self.get_serializer(archivos, many=True)
            
            return Response({
                'directorio_info': directorio_info,
                'archivos': serializer.data
            })
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class PathDirPostViewSet(viewsets.ModelViewSet):
    queryset = PathDirPost.objects.all()
    serializer_class = PathDirPostSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['id', 'cod_municipio']
    search_fields = ['path']
    ordering_fields = ['id', 'fecha_creacion']
    
    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        municipio_id = request.query_params.get('municipio_id')
        if municipio_id:
            rutas = self.queryset.filter(cod_municipio=municipio_id)
            serializer = self.get_serializer(rutas, many=True)
            return Response(serializer.data)
        return Response({"error": "Se requiere el parámetro municipio_id"}, status=status.HTTP_400_BAD_REQUEST)

class HistorialPropietariosViewSet(viewsets.ModelViewSet):
    queryset = HistorialPropietarios.objects.all()
    serializer_class = HistorialPropietariosSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo_archivo', 'id_archivo', 'propietario_nuevo', 'propietario_anterior']
    search_fields = ['propietario_nuevo', 'propietario_anterior']
    ordering_fields = ['fecha_inicio', 'fecha_fin', 'id']
    
    def get_queryset(self):
        """Personalizar el queryset con filtros adicionales"""
        queryset = super().get_queryset()
        
        # Filtrar por municipio si se proporciona
        municipio_id = self.request.query_params.get('municipio_id')
        if municipio_id:
            # Consulta JSON para encontrar archivos relacionados con este municipio
            from django.db import models
            queryset = queryset.filter(
                models.Q(detalles__contains={'cod_municipio': municipio_id}) |
                models.Q(detalles__contains={'municipio_id': municipio_id})
            )
        
        # Filtrar por fecha de inicio
        fecha_desde = self.request.query_params.get('fecha_desde')
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        
        if fecha_desde:
            queryset = queryset.filter(fecha_inicio__gte=fecha_desde)
        
        if fecha_hasta:
            queryset = queryset.filter(fecha_inicio__lte=fecha_hasta)
        
        # Filtrar solo propietarios actuales
        solo_actuales = self.request.query_params.get('solo_actuales')
        if solo_actuales and solo_actuales.lower() in ['true', '1', 'yes']:
            queryset = queryset.filter(fecha_fin__isnull=True)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def por_archivo(self, request):
        """Obtener historial completo de un archivo específico"""
        tipo = request.query_params.get('tipo')
        id_archivo = request.query_params.get('id_archivo')
        
        print(f"🔍 DEBUG: tipo={tipo}, id_archivo={id_archivo}")
        
        if not tipo or not id_archivo:
            return Response(
                {"error": "Se requieren los parámetros 'tipo' y 'id_archivo'"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar el tipo
        if tipo not in ['preoperacion', 'postoperacion']:
            return Response(
                {"error": "El parámetro 'tipo' debe ser 'preoperacion' o 'postoperacion'"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # PASO 1: Probar consulta básica SIN get_queryset()
            print("🔍 PASO 1: Consulta básica...")
            historial = HistorialPropietarios.objects.filter(
                tipo_archivo=tipo, 
                id_archivo=int(id_archivo)  # Convertir a int por si acaso
            ).order_by('-fecha_inicio')
            
            count = historial.count()
            print(f"📊 Registros encontrados: {count}")
            
            if count == 0:
                return Response([])
            
            # PASO 2: Probar serialización
            print("🔍 PASO 2: Serializando...")
            serializer = self.get_serializer(historial, many=True)
            data = serializer.data
            print(f"✅ Serialización exitosa: {len(data)} registros")
            
            return Response(data)
            
        except ValueError as e:
            print(f"❌ Error de valor: {e}")
            return Response(
                {"error": f"ID de archivo debe ser un número: {e}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            print(f"❌ Error inesperado: {e}")
            print(f"❌ Tipo de error: {type(e)}")
            import traceback
            traceback.print_exc()
            
            return Response(
                {"error": f"Error interno: {str(e)}", "tipo_error": str(type(e))},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def por_usuario(self, request):
        """Obtener todos los archivos que pertenecen o pertenecieron a un usuario"""
        usuario = request.query_params.get('usuario')
        
        if not usuario:
            return Response(
                {"error": "Se requiere el parámetro 'usuario'"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Obtener historial
        historial = self.queryset.filter(
            propietario_nuevo=usuario
        ).order_by('-fecha_inicio')
        
        serializer = self.get_serializer(historial, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Obtener estadísticas sobre cambios de propietarios"""
        from django.db.models import Count
        from django.db.models.functions import TruncDay, TruncMonth
        
        # Determinar el nivel de agrupación (día, mes, año)
        agrupacion = request.query_params.get('agrupacion', 'dia')
        
        if agrupacion == 'dia':
            trunc_func = TruncDay
            periodo_nombre = 'día'
        elif agrupacion == 'mes':
            trunc_func = TruncMonth
            periodo_nombre = 'mes'
        else:
            trunc_func = TruncDay
            periodo_nombre = 'día'
        
        # Calcular estadísticas
        stats = self.queryset.annotate(
            periodo=trunc_func('fecha_inicio')
        ).values(
            'periodo'
        ).annotate(
            total_cambios=Count('id'),
            usuarios_distintos=Count('propietario_nuevo', distinct=True),
            archivos_afectados=Count('id_archivo', distinct=True)
        ).order_by('-periodo')
        
        # Convertir resultados a formato más amigable
        resultados = []
        for stat in stats:
            resultados.append({
                'periodo': stat['periodo'].strftime('%Y-%m-%d') if agrupacion == 'dia' else stat['periodo'].strftime('%Y-%m'),
                'periodo_formateado': stat['periodo'].strftime('%d/%m/%Y') if agrupacion == 'dia' else stat['periodo'].strftime('%m/%Y'),
                'total_cambios': stat['total_cambios'],
                'usuarios_distintos': stat['usuarios_distintos'],
                'archivos_afectados': stat['archivos_afectados']
            })
        
        return Response({
            'agrupacion': periodo_nombre,
            'resultados': resultados
        })

    @action(detail=False, methods=['get'])
    def test(self, request):
        """Endpoint de prueba"""
        try:
            total = self.get_queryset().count()
            return Response({
                "status": "OK",
                "total_registros": total,
                "mensaje": "Endpoint funcionando correctamente"
            })
        except Exception as e:
            return Response({
                "status": "ERROR", 
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class NotificacionesPostViewSet(viewsets.ModelViewSet):
    queryset = NotificacionesPost.objects.all().order_by('-fecha_cambio')
    serializer_class = NotificacionesPostSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    # ACTIVAR filter_backends para soportar filtros
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    # Filtros básicos
    filterset_fields = ['tipo_entidad', 'accion', 'leido']
    search_fields = ['descripcion', 'tipo_entidad', 'accion']
    ordering_fields = ['fecha_cambio', 'id']
    
    def get_queryset(self):
        """
        Queryset personalizado con filtros avanzados
        """
        queryset = NotificacionesPost.objects.all()
        
        # Filtro por fechas
        fecha_desde = self.request.query_params.get('fecha_desde', None)
        fecha_hasta = self.request.query_params.get('fecha_hasta', None)
        
        if fecha_desde:
            try:
                queryset = queryset.filter(fecha_cambio__date__gte=fecha_desde)
            except:
                pass
                
        if fecha_hasta:
            try:
                queryset = queryset.filter(fecha_cambio__date__lte=fecha_hasta)
            except:
                pass
        
        # Filtro por rango de fechas con hora
        fecha_cambio_gte = self.request.query_params.get('fecha_cambio__gte', None)
        fecha_cambio_lte = self.request.query_params.get('fecha_cambio__lte', None)
        
        if fecha_cambio_gte:
            try:
                queryset = queryset.filter(fecha_cambio__gte=fecha_cambio_gte)
            except:
                pass
                
        if fecha_cambio_lte:
            try:
                queryset = queryset.filter(fecha_cambio__lte=fecha_cambio_lte)
            except:
                pass
        
        # Filtro por municipio (buscar en datos_contexto)
        municipio_id = self.request.query_params.get('municipio_id', None)
        if municipio_id:
            try:
                # Buscar en datos_contexto JSON
                from django.db.models import Q
                queryset = queryset.filter(
                    Q(datos_contexto__municipio_id=municipio_id) |
                    Q(datos_contexto__cod_municipio=municipio_id) |
                    Q(datos_contexto__municipio_id=int(municipio_id)) |
                    Q(datos_contexto__cod_municipio=int(municipio_id))
                )
            except:
                pass
        
        # Filtro por departamento (buscar municipios que empiecen con código de depto)
        departamento_id = self.request.query_params.get('departamento_id', None)
        if departamento_id:
            try:
                from django.db.models import Q
                depto_str = str(departamento_id).zfill(2)  # Asegurar 2 dígitos
                
                # Buscar municipios cuyo código empiece con el código del departamento
                queryset = queryset.extra(
                    where=[
                        """(
                            datos_contexto->>'municipio_id' LIKE %s OR 
                            datos_contexto->>'cod_municipio' LIKE %s
                        )"""
                    ],
                    params=[f"{depto_str}%", f"{depto_str}%"]
                )
            except Exception as e:
                print(f"Error filtrando por departamento: {e}")
                pass
        
        # Filtro por usuario (buscar en datos_contexto)
        usuario_windows = self.request.query_params.get('usuario_windows', None)
        if usuario_windows:
            try:
                from django.db.models import Q
                queryset = queryset.filter(
                    Q(datos_contexto__usuario_windows__icontains=usuario_windows)
                )
            except:
                pass
        
        # Filtro por leído/no leído
        leido_param = self.request.query_params.get('leido', None)
        if leido_param is not None:
            if leido_param.lower() in ['true', '1']:
                queryset = queryset.filter(leido=True)
            elif leido_param.lower() in ['false', '0']:
                queryset = queryset.filter(leido=False)
        
        # Aplicar ordenación final
        ordering = self.request.query_params.get('ordering', '-fecha_cambio')
        try:
            queryset = queryset.order_by(ordering)
        except:
            queryset = queryset.order_by('-fecha_cambio')
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def no_leidas(self, request):
        try:
            queryset = self.get_queryset().filter(leido=False)[:100]
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        except Exception as e:
            print(f"❌ ERROR en no_leidas: {e}")
            return Response([])
    
    @action(detail=False, methods=['get'])
    def resumen(self, request):
        try:
            total_no_leidas = self.get_queryset().filter(leido=False).count()
            
            return Response({
                'total_no_leidas': total_no_leidas,
                'por_tipo': [],
                'ultimas': []
            })
        except Exception as e:
            print(f"❌ ERROR en resumen: {e}")
            return Response({
                'total_no_leidas': 0,
                'por_tipo': [],
                'ultimas': []
            })

class CalificacionInfoPostViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar las calificaciones de información post-operación.
    
    Proporciona operaciones CRUD completas y endpoints adicionales
    para estadísticas y consultas específicas.
    """
    
    queryset = CalificacionInfoPost.objects.all()
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    def get_serializer_class(self):
        """Retorna el serializer apropiado según la acción"""
        if self.action == 'create':
            return CalificacionInfoPostCreateSerializer
        elif self.action == 'list':
            return CalificacionInfoPostListSerializer
        return CalificacionInfoPostSerializer
    
    def get_queryset(self):
        """Filtra el queryset según los parámetros de consulta"""
        queryset = CalificacionInfoPost.objects.all()
        
        # Filtro por valor mínimo
        valor_min = self.request.query_params.get('valor_min')
        if valor_min is not None:
            try:
                queryset = queryset.filter(valor__gte=float(valor_min))
            except ValueError:
                pass
        
        # Filtro por valor máximo
        valor_max = self.request.query_params.get('valor_max')
        if valor_max is not None:
            try:
                queryset = queryset.filter(valor__lte=float(valor_max))
            except ValueError:
                pass
        
        # Búsqueda por concepto
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(concepto__icontains=search)
        
        # Ordenamiento
        ordering = self.request.query_params.get('ordering', 'valor')
        if ordering in ['valor', '-valor', 'concepto', '-concepto']:
            queryset = queryset.order_by(ordering)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Endpoint para obtener estadísticas de las calificaciones"""
        try:
            queryset = self.get_queryset()
            
            if not queryset.exists():
                return Response({
                    'message': 'No hay calificaciones disponibles'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Estadísticas básicas
            stats = queryset.aggregate(
                total=Count('id'),
                promedio=Avg('valor'),
                maximo=Max('valor'),
                minimo=Min('valor')
            )
            
            # Distribución por niveles de calidad
            distribucion = {}
            for calif in queryset:
                nivel = calif.get_nivel_calidad()
                distribucion[nivel] = distribucion.get(nivel, 0) + 1
            
            data = {
                'total_calificaciones': stats['total'],
                'promedio_valor': round(stats['promedio'], 2),
                'valor_maximo': stats['maximo'],
                'valor_minimo': stats['minimo'],
                'distribucion_niveles': distribucion
            }
            
            return Response(data)
            
        except Exception as e:
            logger.error(f"Error obteniendo estadísticas: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def niveles_calidad(self, request):
        """Endpoint para obtener todos los niveles de calidad disponibles"""
        try:
            queryset = self.get_queryset()
            niveles = {}
            
            for calif in queryset:
                nivel = calif.get_nivel_calidad()
                if nivel not in niveles:
                    niveles[nivel] = {
                        'nivel': nivel,
                        'calificaciones': [],
                        'rango_valor': {'min': float(calif.valor), 'max': float(calif.valor)}
                    }
                
                niveles[nivel]['calificaciones'].append({
                    'id': calif.id,
                    'concepto': calif.concepto,
                    'valor': float(calif.valor)
                })
                
                # Actualizar rango
                niveles[nivel]['rango_valor']['min'] = min(
                    niveles[nivel]['rango_valor']['min'], 
                    float(calif.valor)
                )
                niveles[nivel]['rango_valor']['max'] = max(
                    niveles[nivel]['rango_valor']['max'], 
                    float(calif.valor)
                )
            
            return Response(list(niveles.values()))
            
        except Exception as e:
            logger.error(f"Error obteniendo niveles de calidad: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def buscar_por_valor(self, request):
        """Busca la calificación más cercana a un valor dado"""
        try:
            valor = request.query_params.get('valor')
            if valor is None:
                return Response({
                    'error': 'El parámetro valor es requerido'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                valor_float = float(valor)
            except ValueError:
                return Response({
                    'error': 'El valor debe ser un número válido'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            calificacion = CalificacionInfoPost.get_concepto_por_valor(valor_float)
            
            if calificacion:
                serializer = self.get_serializer(calificacion)
                return Response(serializer.data)
            else:
                return Response({
                    'message': 'No se encontró ninguna calificación'
                }, status=status.HTTP_404_NOT_FOUND)
                
        except Exception as e:
            logger.error(f"Error buscando por valor: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class EvaluacionArchivosPostViewSet(viewsets.ModelViewSet):
    """
    🔄 VIEWSET ACTUALIZADO - Evaluaciones POST V2 sin componentes
    """
    
    queryset = EvaluacionArchivosPost.objects.select_related(
        'id_disposicion',
        'id_disposicion__cod_municipio',
        'evaluacion_archivo'
    ).prefetch_related(
        'id_disposicion__cod_municipio__cod_depto'
    )
    
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    # 🔍 FILTROS DISPONIBLES (ACTUALIZADOS)
    filterset_fields = [
        'id_evaluacion', 'id_archivo', 'id_disposicion', 'estado_archivo',
        'evaluacion_archivo', 'usuario_evaluacion', 'fecha_creacion'
    ]
    search_fields = [
        'nombre_archivo', 'observaciones_evaluacion', 'usuario_evaluacion',
        'id_disposicion__cod_municipio__nom_municipio',
    ]
    ordering_fields = [
        'id_evaluacion', 'fecha_creacion', 'fecha_actualizacion',
        'estado_archivo', 'nombre_archivo'
    ]
    ordering = ['-fecha_creacion']
    
    def get_queryset(self):
        """🔄 ACTUALIZADO - Sin componentes"""
        queryset = super().get_queryset()
        municipios_permitidos = get_municipios_permitidos(self.request.user)
        
        if municipios_permitidos == 'todos':
            return queryset
        elif municipios_permitidos:
            # Filtrar evaluaciones por municipios permitidos
            return queryset.filter(
                id_disposicion__cod_municipio__in=municipios_permitidos
            )
        else:
            return queryset.none()
    
    def get_serializer_class(self):
        """Retorna el serializer apropiado según la acción"""
        if self.action == 'create':
            return EvaluacionArchivosPostCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return EvaluacionArchivosPostUpdateSerializer
        elif self.action == 'list':
            return EvaluacionArchivosPostSimpleSerializer
        return EvaluacionArchivosPostSerializer
    
    def get_permissions(self):
        """Define permisos según la acción"""
        if self.action in ['list', 'retrieve', 'estadisticas', 'por_municipio', 'por_estado']:
            permission_classes = [IsAuthenticated]
        elif self.action in ['create', 'update', 'partial_update']:
            permission_classes = [IsAuthenticated]
        elif self.action in ['destroy', 'aprobar_masivo', 'rechazar_masivo']:
            permission_classes = [IsAuthenticated, IsAdminUser]
        else:
            permission_classes = [IsAuthenticated]
        
        return [permission() for permission in permission_classes]
    
    def perform_update(self, serializer):
        """Registra el usuario que realiza la actualización"""
        serializer.save(usuario_evaluacion=self.request.user.username)
    
    # 📊 ACCIONES PERSONALIZADAS (ACTUALIZADAS POST V2)
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """🔄 ACTUALIZADO - Estadísticas sin componentes"""
        try:
            queryset = self.get_queryset()
            stats = EvaluacionArchivosPost.get_estadisticas_evaluacion()
            
            # Agregar estadísticas adicionales (SIN componente)
            stats_adicionales = queryset.aggregate(
                total_usuarios=Count('usuario_evaluacion', distinct=True),
                total_municipios=Count('id_disposicion__cod_municipio', distinct=True),
                total_directorios=Count('id_disposicion', distinct=True)
            )
            
            stats.update(stats_adicionales)
            
            # 🆕 NUEVO: Estadísticas por directorio
            stats_por_directorio = queryset.values(
                'id_disposicion__ruta_acceso'
            ).annotate(
                total_evaluaciones=Count('id_evaluacion'),
                pendientes=Count('id_evaluacion', filter=Q(estado_archivo='PENDIENTE')),
                aprobados=Count('id_evaluacion', filter=Q(estado_archivo='APROBADO'))
            ).order_by('-total_evaluaciones')[:10]
            
            stats['top_directorios'] = stats_por_directorio
            
            serializer = EstadisticasEvaluacionSerializer(stats)
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"Error obteniendo estadísticas de evaluación: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        """🔄 ACTUALIZADO - Evaluaciones por municipio sin componentes"""
        try:
            municipio_id = request.query_params.get('municipio_id')
            
            evaluaciones_por_municipio = EvaluacionArchivosPost.get_evaluaciones_por_municipio(municipio_id)
            serializer = EvaluacionesPorMunicipioSerializer(evaluaciones_por_municipio, many=True)
            
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"Error obteniendo evaluaciones por municipio: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def por_directorio(self, request):
        """
        🆕 NUEVO ENDPOINT - Evaluaciones agrupadas por directorio POST V2
        """
        id_disposicion = request.query_params.get('id_disposicion')
        
        if not id_disposicion:
            return Response({
                'error': 'El parámetro id_disposicion es requerido'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            queryset = self.get_queryset().filter(id_disposicion=id_disposicion)
            
            # Obtener información del directorio
            try:
                disposicion = DisposicionPost.objects.get(id_disposicion=id_disposicion)
                directorio_info = {
                    'id_disposicion': disposicion.id_disposicion,
                    'nombre_directorio': disposicion.get_nombre_directorio(),
                    'ruta_acceso': disposicion.ruta_acceso,
                    'municipio': disposicion.cod_municipio.nom_municipio,
                    'total_evaluaciones': queryset.count()
                }
            except DisposicionPost.DoesNotExist:
                return Response({"error": "Disposición no encontrada"}, status=status.HTTP_404_NOT_FOUND)
            
            # Estadísticas del directorio
            stats_directorio = queryset.aggregate(
                pendientes=Count('id_evaluacion', filter=Q(estado_archivo='PENDIENTE')),
                en_revision=Count('id_evaluacion', filter=Q(estado_archivo='EN_REVISION')),
                aprobados=Count('id_evaluacion', filter=Q(estado_archivo='APROBADO')),
                rechazados=Count('id_evaluacion', filter=Q(estado_archivo='RECHAZADO'))
            )
            
            page = self.paginate_queryset(queryset)
            
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response({
                    'directorio_info': directorio_info,
                    'estadisticas': stats_directorio,
                    'evaluaciones': serializer.data
                })
            
            serializer = self.get_serializer(queryset, many=True)
            return Response({
                'directorio_info': directorio_info,
                'estadisticas': stats_directorio,
                'evaluaciones': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Error obteniendo evaluaciones por directorio: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def por_estado(self, request):
        """Endpoint para obtener evaluaciones filtradas por estado"""
        estado = request.query_params.get('estado')
        
        if not estado:
            return Response({
                'error': 'El parámetro estado es requerido'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if estado not in dict(EvaluacionArchivosPost.ESTADOS_ARCHIVO):
            return Response({
                'error': 'Estado no válido'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            queryset = self.get_queryset().filter(estado_archivo=estado)
            page = self.paginate_queryset(queryset)
            
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"Error filtrando por estado: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def pendientes(self, request):
        """Endpoint optimizado para obtener evaluaciones pendientes"""
        try:
            queryset = self.get_queryset().filter(
                estado_archivo='PENDIENTE'
            ).order_by('fecha_creacion')
            
            # Aplicar filtros adicionales
            municipio_id = request.query_params.get('municipio_id')
            if municipio_id:
                queryset = queryset.filter(id_disposicion__cod_municipio=municipio_id)
            
            id_disposicion = request.query_params.get('id_disposicion')
            if id_disposicion:
                queryset = queryset.filter(id_disposicion=id_disposicion)
            
            page = self.paginate_queryset(queryset)
            
            if page is not None:
                serializer = EvaluacionArchivosPostSimpleSerializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            
            serializer = EvaluacionArchivosPostSimpleSerializer(queryset, many=True)
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"Error obteniendo evaluaciones pendientes: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None):
        """Endpoint para aprobar una evaluación específica"""
        try:
            evaluacion = self.get_object()
            
            if not evaluacion.puede_ser_evaluado():
                return Response({
                    'error': f'El archivo no puede ser aprobado desde el estado {evaluacion.estado_archivo}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            observaciones = request.data.get('observaciones', '')
            evaluacion_archivo_id = request.data.get('evaluacion_archivo')
            
            # Actualizar calificación si se proporciona
            if evaluacion_archivo_id:
                try:
                    calificacion = CalificacionInfoPost.objects.get(id=evaluacion_archivo_id)
                    evaluacion.evaluacion_archivo = calificacion
                except CalificacionInfoPost.DoesNotExist:
                    return Response({
                        'error': 'Calificación no encontrada'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Aprobar el archivo
            evaluacion.aprobar(
                usuario=request.user.username,
                observaciones=observaciones
            )
            
            serializer = self.get_serializer(evaluacion)
            return Response({
                'message': 'Archivo aprobado exitosamente',
                'evaluacion': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Error aprobando evaluación {pk}: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def rechazar(self, request, pk=None):
        """Endpoint para rechazar una evaluación específica"""
        try:
            evaluacion = self.get_object()
            
            if not evaluacion.puede_ser_evaluado():
                return Response({
                    'error': f'El archivo no puede ser rechazado desde el estado {evaluacion.estado_archivo}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            observaciones = request.data.get('observaciones', '')
            
            if not observaciones:
                return Response({
                    'error': 'Las observaciones son requeridas para rechazar un archivo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            evaluacion_archivo_id = request.data.get('evaluacion_archivo')
            
            # Actualizar calificación si se proporciona
            if evaluacion_archivo_id:
                try:
                    calificacion = CalificacionInfoPost.objects.get(id=evaluacion_archivo_id)
                    evaluacion.evaluacion_archivo = calificacion
                except CalificacionInfoPost.DoesNotExist:
                    return Response({
                        'error': 'Calificación no encontrada'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Rechazar el archivo
            evaluacion.rechazar(
                usuario=request.user.username,
                observaciones=observaciones
            )
            
            serializer = self.get_serializer(evaluacion)
            return Response({
                'message': 'Archivo rechazado exitosamente',
                'evaluacion': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Error rechazando evaluación {pk}: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def marcar_en_revision(self, request, pk=None):
        """Endpoint para marcar un archivo como en revisión"""
        try:
            evaluacion = self.get_object()
            
            if evaluacion.estado_archivo not in ['PENDIENTE', 'REQUIERE_AJUSTES']:
                return Response({
                    'error': f'El archivo no puede marcarse en revisión desde el estado {evaluacion.estado_archivo}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            evaluacion.marcar_como_en_revision(usuario=request.user.username)
            
            serializer = self.get_serializer(evaluacion)
            return Response({
                'message': 'Archivo marcado como en revisión',
                'evaluacion': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Error marcando evaluación en revisión {pk}: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def aprobar_masivo(self, request):
        """Endpoint para aprobar múltiples evaluaciones"""
        try:
            evaluaciones_ids = request.data.get('evaluaciones', [])
            observaciones = request.data.get('observaciones', '')
            evaluacion_archivo_id = request.data.get('evaluacion_archivo')
            
            if not evaluaciones_ids:
                return Response({
                    'error': 'Se requiere al menos una evaluación'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Obtener evaluaciones
            evaluaciones = self.get_queryset().filter(
                id_evaluacion__in=evaluaciones_ids,
                estado_archivo__in=['PENDIENTE', 'EN_REVISION', 'REQUIERE_AJUSTES']
            )
            
            if not evaluaciones.exists():
                return Response({
                    'error': 'No se encontraron evaluaciones válidas para aprobar'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validar calificación si se proporciona
            calificacion = None
            if evaluacion_archivo_id:
                try:
                    calificacion = CalificacionInfoPost.objects.get(id=evaluacion_archivo_id)
                except CalificacionInfoPost.DoesNotExist:
                    return Response({
                        'error': 'Calificación no encontrada'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Aprobar evaluaciones
            evaluaciones_aprobadas = []
            for evaluacion in evaluaciones:
                if calificacion:
                    evaluacion.evaluacion_archivo = calificacion
                
                evaluacion.aprobar(
                    usuario=request.user.username,
                    observaciones=observaciones
                )
                evaluaciones_aprobadas.append(evaluacion.id_evaluacion)
            
            return Response({
                'message': f'Se aprobaron {len(evaluaciones_aprobadas)} evaluaciones',
                'evaluaciones_aprobadas': evaluaciones_aprobadas
            })
            
        except Exception as e:
            logger.error(f"Error en aprobación masiva: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def resumen_usuario(self, request):
        """Obtiene resumen de evaluaciones por usuario actual"""
        try:
            usuario = request.user.username
            queryset = self.get_queryset().filter(usuario_evaluacion=usuario)
            
            stats = queryset.aggregate(
                total_evaluadas=Count('id_evaluacion'),
                aprobadas=Count('id_evaluacion', filter=Q(estado_archivo='APROBADO')),
                rechazadas=Count('id_evaluacion', filter=Q(estado_archivo='RECHAZADO')),
                en_revision=Count('id_evaluacion', filter=Q(estado_archivo='EN_REVISION'))
            )
            
            # Últimas evaluaciones del usuario
            ultimas = queryset.order_by('-fecha_actualizacion')[:5]
            ultimas_serializer = EvaluacionArchivosPostSimpleSerializer(ultimas, many=True)
            
            return Response({
                'usuario': usuario,
                'estadisticas': stats,
                'ultimas_evaluaciones': ultimas_serializer.data
            })
            
        except Exception as e:
            logger.error(f"Error obteniendo resumen de usuario: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ===============================================
    # 🆕 CALIFICACIÓN MASIVA (SOLO SUPER ADMIN)
    # ===============================================

    @action(detail=False, methods=['post'])
    def calificacion_masiva(self, request):
        """
        🆕 Endpoint para aplicar calificación masiva a múltiples archivos.
        Solo disponible para Super Administradores.

        Body esperado:
        {
            "archivos_ids": [1, 2, 3, ...],  // IDs de evaluaciones
            "calificacion_id": 5,             // ID de la calificación a aplicar
            "filtro": "no_calificados",       // "todos", "no_calificados", "excepto_aprobados"
            "observaciones": "Calificación masiva aplicada"
        }
        """
        import uuid
        from django.utils import timezone

        # Verificar que sea Super Admin
        if not request.user.is_superuser:
            return Response({
                'error': 'Solo los Super Administradores pueden realizar calificaciones masivas'
            }, status=status.HTTP_403_FORBIDDEN)

        try:
            archivos_ids = request.data.get('archivos_ids', [])
            calificacion_id = request.data.get('calificacion_id')
            filtro = request.data.get('filtro', 'no_calificados')
            observaciones = request.data.get('observaciones', 'Calificación masiva')

            if not archivos_ids:
                return Response({
                    'error': 'Se requiere al menos un archivo para calificar'
                }, status=status.HTTP_400_BAD_REQUEST)

            if not calificacion_id:
                return Response({
                    'error': 'Se requiere especificar la calificación a aplicar'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Validar que la calificación existe
            try:
                calificacion = CalificacionInfoPost.objects.get(id=calificacion_id)
            except CalificacionInfoPost.DoesNotExist:
                return Response({
                    'error': 'La calificación especificada no existe'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Obtener archivos base (TODOS los seleccionados)
            queryset_total = EvaluacionArchivosPost.objects.filter(id_evaluacion__in=archivos_ids)
            total_seleccionados = queryset_total.count()

            # Aplicar filtro según opción seleccionada
            if filtro == 'no_calificados':
                # Solo archivos con evaluacion_archivo = 0 (SIN CALIFICACION) o NULL
                # Nota: evaluacion_archivo=0 significa "sin calificar", no es 1
                queryset = queryset_total.filter(Q(evaluacion_archivo=0) | Q(evaluacion_archivo__isnull=True))
            elif filtro == 'excepto_aprobados':
                # Todos excepto los ya aprobados
                queryset = queryset_total.exclude(aprobado=True)
            else:
                # Si filtro == 'todos', no aplicamos filtro adicional
                queryset = queryset_total

            # Contar cuántos fueron ignorados por el filtro
            archivos_ignorados_por_filtro = total_seleccionados - queryset.count()

            # Generar identificador único para este lote de calificación masiva
            lote_id = str(uuid.uuid4())[:8]
            fecha_actual = timezone.now()

            # Contadores
            archivos_calificados = 0
            archivos_ya_tenian_misma_calificacion = 0

            # Procesar cada archivo que pasó el filtro
            for archivo in queryset:
                # Guardar calificación anterior SOLO si es diferente
                if archivo.evaluacion_archivo != calificacion_id:
                    archivo.evaluacion_archivo_anterior = archivo.evaluacion_archivo
                    archivo.evaluacion_archivo = calificacion_id
                    archivo.lote_calificacion_masiva = lote_id
                    archivo.fecha_calificacion_masiva = fecha_actual
                    archivo.observaciones_evaluacion = f"[MASIVO-{lote_id}] {observaciones}"
                    archivo.usuario_evaluacion = request.user.username
                    archivo.evaluado = True

                    # Determinar si está aprobado basándose en el valor de la calificación
                    if calificacion.valor >= 1.0:
                        archivo.aprobado = True
                        archivo.estado_archivo = 'APROBADO'
                    else:
                        archivo.aprobado = False
                        if archivo.estado_archivo == 'PENDIENTE':
                            archivo.estado_archivo = 'EN_REVISION'

                    archivo.save()
                    archivos_calificados += 1
                else:
                    archivos_ya_tenian_misma_calificacion += 1

            # Construir mensaje descriptivo
            if archivos_calificados > 0:
                mensaje = f'✅ {archivos_calificados} archivos calificados exitosamente'
            else:
                mensaje = '⚠️ No se calificaron archivos nuevos'

            if archivos_ignorados_por_filtro > 0:
                mensaje += f' ({archivos_ignorados_por_filtro} ignorados por filtro "{filtro}")'

            if archivos_ya_tenian_misma_calificacion > 0:
                mensaje += f' ({archivos_ya_tenian_misma_calificacion} ya tenían la misma calificación)'

            return Response({
                'success': True,
                'message': mensaje,
                'lote_id': lote_id if archivos_calificados > 0 else None,
                'calificacion_aplicada': calificacion.concepto,
                'archivos_calificados': archivos_calificados,
                'archivos_ignorados_por_filtro': archivos_ignorados_por_filtro,
                'archivos_ya_tenian_misma_calificacion': archivos_ya_tenian_misma_calificacion,
                'total_seleccionados': total_seleccionados,
                'filtro_aplicado': filtro,
                'fecha': fecha_actual.isoformat()
            })

        except Exception as e:
            logger.error(f"Error en calificación masiva: {str(e)}")
            return Response({
                'error': f'Error interno del servidor: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def restaurar_calificacion(self, request):
        """
        🆕 Endpoint para restaurar la calificación anterior de un lote de calificación masiva.
        Solo disponible para Super Administradores.

        Body esperado:
        {
            "lote_id": "abc12345"  // Identificador del lote a restaurar
        }
        """
        # Verificar que sea Super Admin
        if not request.user.is_superuser:
            return Response({
                'error': 'Solo los Super Administradores pueden restaurar calificaciones'
            }, status=status.HTTP_403_FORBIDDEN)

        try:
            lote_id = request.data.get('lote_id')

            if not lote_id:
                return Response({
                    'error': 'Se requiere el identificador del lote a restaurar'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Buscar archivos del lote
            archivos = EvaluacionArchivosPost.objects.filter(
                lote_calificacion_masiva=lote_id
            )

            if not archivos.exists():
                return Response({
                    'error': f'No se encontraron archivos para el lote {lote_id}'
                }, status=status.HTTP_404_NOT_FOUND)

            # Restaurar calificaciones anteriores
            archivos_restaurados = 0
            archivos_sin_anterior = 0

            for archivo in archivos:
                if archivo.evaluacion_archivo_anterior is not None:
                    # Restaurar la calificación anterior
                    archivo.evaluacion_archivo = archivo.evaluacion_archivo_anterior
                    archivo.evaluacion_archivo_anterior = None
                    archivo.lote_calificacion_masiva = None
                    archivo.fecha_calificacion_masiva = None
                    archivo.observaciones_evaluacion = f"[RESTAURADO] Calificación restaurada por {request.user.username}"
                    archivo.usuario_evaluacion = request.user.username

                    # Determinar estado basándose en la calificación restaurada
                    try:
                        calificacion_restaurada = CalificacionInfoPost.objects.get(id=archivo.evaluacion_archivo)
                        if calificacion_restaurada.valor >= 1.0:
                            archivo.aprobado = True
                            archivo.estado_archivo = 'APROBADO'
                        elif archivo.evaluacion_archivo == 1:  # SIN CALIFICACION
                            archivo.evaluado = False
                            archivo.aprobado = False
                            archivo.estado_archivo = 'PENDIENTE'
                        else:
                            archivo.aprobado = False
                            archivo.estado_archivo = 'EN_REVISION'
                    except CalificacionInfoPost.DoesNotExist:
                        archivo.evaluado = False
                        archivo.aprobado = False
                        archivo.estado_archivo = 'PENDIENTE'

                    archivo.save()
                    archivos_restaurados += 1
                else:
                    archivos_sin_anterior += 1

            return Response({
                'success': True,
                'message': f'Calificaciones restauradas exitosamente',
                'lote_id': lote_id,
                'archivos_restaurados': archivos_restaurados,
                'archivos_sin_anterior': archivos_sin_anterior
            })

        except Exception as e:
            logger.error(f"Error restaurando calificaciones: {str(e)}")
            return Response({
                'error': f'Error interno del servidor: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def ultimo_lote_masivo(self, request):
        """
        🆕 Obtiene información del último lote de calificación masiva.
        Útil para mostrar el botón de restaurar.
        """
        if not request.user.is_superuser:
            return Response({
                'error': 'Solo los Super Administradores pueden ver esta información'
            }, status=status.HTTP_403_FORBIDDEN)

        try:
            municipio_id = request.query_params.get('municipio_id')

            # Buscar el último lote
            queryset = EvaluacionArchivosPost.objects.filter(
                lote_calificacion_masiva__isnull=False
            )

            if municipio_id:
                queryset = queryset.filter(id_disposicion__cod_municipio=municipio_id)

            ultimo = queryset.order_by('-fecha_calificacion_masiva').first()

            if not ultimo:
                return Response({
                    'tiene_lote': False,
                    'message': 'No hay lotes de calificación masiva registrados'
                })

            # Contar archivos en el lote
            archivos_en_lote = queryset.filter(
                lote_calificacion_masiva=ultimo.lote_calificacion_masiva
            ).count()

            return Response({
                'tiene_lote': True,
                'lote_id': ultimo.lote_calificacion_masiva,
                'fecha': ultimo.fecha_calificacion_masiva.isoformat() if ultimo.fecha_calificacion_masiva else None,
                'archivos_en_lote': archivos_en_lote,
                'usuario': ultimo.usuario_evaluacion
            })

        except Exception as e:
            logger.error(f"Error obteniendo último lote masivo: {str(e)}")
            return Response({
                'error': 'Error interno del servidor'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ===============================================
    # 🗑️ DEPURACIÓN MASIVA DE ARCHIVOS INEXISTENTES
    # ===============================================

    @action(detail=False, methods=['get'])
    def verificar_archivos_inexistentes(self, request):
        """
        🆕 Verifica qué archivos de un municipio NO existen físicamente en el sistema de archivos.
        Solo para Super Administradores.

        Query params:
            municipio_id: Código del municipio a verificar

        Returns:
            - total_archivos: Total de archivos registrados
            - archivos_existentes: Cantidad que SÍ existen
            - archivos_inexistentes: Cantidad que NO existen
            - lista_inexistentes: Lista detallada de archivos inexistentes
        """
        if not request.user.is_superuser:
            return Response({
                'error': 'Solo los Super Administradores pueden verificar archivos inexistentes'
            }, status=status.HTTP_403_FORBIDDEN)

        try:
            municipio_id = request.query_params.get('municipio_id')

            if not municipio_id:
                return Response({
                    'error': 'Se requiere el parámetro municipio_id'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Obtener todos los archivos del municipio
            archivos = EvaluacionArchivosPost.objects.filter(
                id_disposicion__cod_municipio=municipio_id
            ).values('id_evaluacion', 'nombre_archivo', 'ruta_completa', 'fecha_creacion', 'evaluacion_archivo')

            total_archivos = archivos.count()
            archivos_existentes = 0
            archivos_inexistentes = 0
            lista_inexistentes = []

            for archivo in archivos:
                ruta_original = archivo['ruta_completa']

                # Determinar formato de la ruta y convertir según sea necesario
                if ruta_original and ruta_original.startswith('/mnt/'):
                    # La ruta está en formato Linux
                    ruta_linux = ruta_original
                    ruta_windows = linux_to_windows_path(ruta_original)
                else:
                    # La ruta está en formato Windows (o mixto)
                    ruta_linux = windows_to_linux_path(ruta_original)
                    ruta_windows = ruta_original if ruta_original and ruta_original.startswith('\\\\') else linux_to_windows_path(ruta_linux)

                # Verificar si el archivo existe físicamente
                existe = os.path.isfile(ruta_linux)

                if existe:
                    archivos_existentes += 1
                else:
                    archivos_inexistentes += 1
                    lista_inexistentes.append({
                        'id': archivo['id_evaluacion'],
                        'nombre': archivo['nombre_archivo'],
                        'ruta_windows': ruta_windows,
                        'fecha_creacion': archivo['fecha_creacion'].isoformat() if archivo['fecha_creacion'] else None,
                        'evaluacion_archivo': archivo['evaluacion_archivo']
                    })

            return Response({
                'municipio_id': municipio_id,
                'total_archivos': total_archivos,
                'archivos_existentes': archivos_existentes,
                'archivos_inexistentes': archivos_inexistentes,
                'porcentaje_inexistentes': round((archivos_inexistentes / total_archivos * 100), 2) if total_archivos > 0 else 0,
                'lista_inexistentes': lista_inexistentes
            })

        except Exception as e:
            logger.error(f"Error verificando archivos inexistentes: {str(e)}")
            return Response({
                'error': f'Error interno del servidor: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def depurar_archivos_inexistentes(self, request):
        """
        🆕 Elimina de la base de datos los registros de archivos que NO existen físicamente.
        Solo para Super Administradores.

        Body:
            - municipio_id: Código del municipio
            - archivo_ids: (opcional) Lista de IDs específicos a eliminar. Si no se proporciona,
                           elimina TODOS los archivos inexistentes del municipio.
            - confirmar: (requerido) Debe ser True para confirmar la operación

        Returns:
            - archivos_eliminados: Cantidad de registros eliminados
            - archivos_omitidos: Cantidad de archivos que existían y no se eliminaron
        """
        if not request.user.is_superuser:
            return Response({
                'error': 'Solo los Super Administradores pueden depurar archivos'
            }, status=status.HTTP_403_FORBIDDEN)

        try:
            municipio_id = request.data.get('municipio_id')
            archivo_ids = request.data.get('archivo_ids', [])  # Lista opcional de IDs
            confirmar = request.data.get('confirmar', False)

            if not municipio_id:
                return Response({
                    'error': 'Se requiere el parámetro municipio_id'
                }, status=status.HTTP_400_BAD_REQUEST)

            if not confirmar:
                return Response({
                    'error': 'Debe confirmar la operación estableciendo "confirmar": true'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Obtener archivos a verificar
            if archivo_ids:
                # Si se proporcionan IDs específicos, solo verificar esos
                archivos = EvaluacionArchivosPost.objects.filter(
                    id_evaluacion__in=archivo_ids,
                    id_disposicion__cod_municipio=municipio_id
                )
            else:
                # Si no, verificar todos los del municipio
                archivos = EvaluacionArchivosPost.objects.filter(
                    id_disposicion__cod_municipio=municipio_id
                )

            archivos_eliminados = 0
            archivos_omitidos = 0
            detalles_eliminados = []

            for archivo in archivos:
                ruta_windows = archivo.ruta_completa
                ruta_linux = windows_to_linux_path(ruta_windows)

                # Verificar si el archivo existe físicamente
                existe = os.path.isfile(ruta_linux)

                if not existe:
                    # El archivo NO existe, eliminarlo de la BD
                    detalles_eliminados.append({
                        'id': archivo.id_evaluacion,
                        'nombre': archivo.nombre_archivo,
                        'ruta': ruta_windows[:100] + '...' if len(ruta_windows) > 100 else ruta_windows
                    })
                    archivo.delete()
                    archivos_eliminados += 1
                else:
                    # El archivo SÍ existe, no eliminarlo
                    archivos_omitidos += 1

            # Registrar la acción en el log
            logger.info(f"Depuración masiva municipio {municipio_id} por {request.user.username}: "
                       f"{archivos_eliminados} eliminados, {archivos_omitidos} omitidos")

            return Response({
                'success': True,
                'municipio_id': municipio_id,
                'archivos_eliminados': archivos_eliminados,
                'archivos_omitidos': archivos_omitidos,
                'message': f'Se eliminaron {archivos_eliminados} registros de archivos inexistentes. '
                          f'{archivos_omitidos} archivos existentes fueron preservados.',
                'usuario': request.user.username,
                'fecha': timezone.now().isoformat(),
                'detalles': detalles_eliminados[:50]  # Limitar a 50 para no saturar respuesta
            })

        except Exception as e:
            logger.error(f"Error depurando archivos inexistentes: {str(e)}")
            return Response({
                'error': f'Error interno del servidor: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def verificar_directorios_inexistentes(self, request):
        """
        🆕 Verifica qué directorios (Disposiciones) de un municipio NO existen físicamente.
        Solo para Super Administradores.

        Query params:
            municipio_id: Código del municipio a verificar

        Returns:
            - total_directorios: Total de directorios registrados
            - directorios_existentes: Cantidad que SÍ existen
            - directorios_inexistentes: Cantidad que NO existen
            - lista_inexistentes: Lista detallada de directorios inexistentes
        """
        if not request.user.is_superuser:
            return Response({
                'error': 'Solo los Super Administradores pueden verificar directorios inexistentes'
            }, status=status.HTTP_403_FORBIDDEN)

        try:
            municipio_id = request.query_params.get('municipio_id')

            if not municipio_id:
                return Response({
                    'error': 'Se requiere el parámetro municipio_id'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Obtener todos los directorios del municipio
            directorios = DisposicionPost.objects.filter(
                cod_municipio=municipio_id
            ).values('id_disposicion', 'ruta_acceso', 'fecha_disposicion', 'dispuesto', 'evaluado', 'aprobado')

            total_directorios = directorios.count()
            directorios_existentes = 0
            directorios_inexistentes = 0
            lista_inexistentes = []

            for directorio in directorios:
                ruta_original = directorio['ruta_acceso']

                if not ruta_original:
                    # Si no tiene ruta, considerarlo inexistente
                    directorios_inexistentes += 1
                    lista_inexistentes.append({
                        'id': directorio['id_disposicion'],
                        'nombre': 'Sin ruta definida',
                        'ruta_windows': 'Sin ruta',
                        'fecha_disposicion': directorio['fecha_disposicion'].isoformat() if directorio['fecha_disposicion'] else None,
                        'dispuesto': directorio['dispuesto'],
                        'evaluado': directorio['evaluado'],
                        'aprobado': directorio['aprobado'],
                        'total_archivos': EvaluacionArchivosPost.objects.filter(id_disposicion=directorio['id_disposicion']).count()
                    })
                    continue

                # Determinar formato de la ruta y convertir según sea necesario
                if ruta_original.startswith('/mnt/'):
                    # La ruta está en formato Linux
                    ruta_linux = ruta_original
                    ruta_windows = linux_to_windows_path(ruta_original)
                else:
                    # La ruta está en formato Windows (o mixto)
                    ruta_linux = windows_to_linux_path(ruta_original)
                    ruta_windows = ruta_original if ruta_original.startswith('\\\\') else linux_to_windows_path(ruta_linux)

                # Verificar si el directorio existe físicamente
                existe = os.path.isdir(ruta_linux)

                if existe:
                    directorios_existentes += 1
                else:
                    directorios_inexistentes += 1
                    # Extraer nombre del directorio de la ruta Windows
                    nombre_dir = ruta_windows.split('\\')[-1] if '\\' in ruta_windows else ruta_windows.split('/')[-1]
                    lista_inexistentes.append({
                        'id': directorio['id_disposicion'],
                        'nombre': nombre_dir,
                        'ruta_windows': ruta_windows,
                        'fecha_disposicion': directorio['fecha_disposicion'].isoformat() if directorio['fecha_disposicion'] else None,
                        'dispuesto': directorio['dispuesto'],
                        'evaluado': directorio['evaluado'],
                        'aprobado': directorio['aprobado'],
                        'total_archivos': EvaluacionArchivosPost.objects.filter(id_disposicion=directorio['id_disposicion']).count()
                    })

            return Response({
                'municipio_id': municipio_id,
                'total_directorios': total_directorios,
                'directorios_existentes': directorios_existentes,
                'directorios_inexistentes': directorios_inexistentes,
                'lista_inexistentes': lista_inexistentes
            })

        except Exception as e:
            logger.error(f"Error verificando directorios inexistentes: {str(e)}")
            return Response({
                'error': f'Error interno del servidor: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def depurar_directorios_inexistentes(self, request):
        """
        🆕 Elimina de la base de datos los registros de directorios que NO existen físicamente.
        También elimina en cascada los archivos asociados.
        Solo para Super Administradores.

        Body:
            - municipio_id: Código del municipio
            - directorio_ids: Lista de IDs de directorios a eliminar
            - confirmar: (requerido) Debe ser True para confirmar la operación

        Returns:
            - directorios_eliminados: Cantidad de directorios eliminados
            - archivos_eliminados: Cantidad de archivos eliminados en cascada
            - directorios_omitidos: Cantidad de directorios que existían y no se eliminaron
        """
        if not request.user.is_superuser:
            return Response({
                'error': 'Solo los Super Administradores pueden depurar directorios'
            }, status=status.HTTP_403_FORBIDDEN)

        try:
            municipio_id = request.data.get('municipio_id')
            directorio_ids = request.data.get('directorio_ids', [])
            confirmar = request.data.get('confirmar', False)

            if not municipio_id:
                return Response({
                    'error': 'Se requiere el parámetro municipio_id'
                }, status=status.HTTP_400_BAD_REQUEST)

            if not confirmar:
                return Response({
                    'error': 'Debe confirmar la operación estableciendo "confirmar": true'
                }, status=status.HTTP_400_BAD_REQUEST)

            if not directorio_ids:
                return Response({
                    'error': 'Se requiere proporcionar directorio_ids con los IDs a eliminar'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Obtener directorios a verificar
            directorios = DisposicionPost.objects.filter(
                id_disposicion__in=directorio_ids,
                cod_municipio=municipio_id
            )

            directorios_eliminados = 0
            archivos_eliminados_total = 0
            directorios_omitidos = 0
            detalles_eliminados = []

            for directorio in directorios:
                ruta_windows = directorio.ruta_acceso

                # Verificar si el directorio existe físicamente
                if ruta_windows:
                    ruta_linux = windows_to_linux_path(ruta_windows)
                    existe = os.path.isdir(ruta_linux)
                else:
                    existe = False

                if not existe:
                    # El directorio NO existe, contar archivos y eliminar
                    # Contar archivos en ambas tablas
                    archivos_post_count = directorio.archivos_relacionados.count()  # ArchivosPost
                    evaluacion_archivos_count = EvaluacionArchivosPost.objects.filter(id_disposicion=directorio).count()
                    archivos_count = archivos_post_count + evaluacion_archivos_count

                    nombre_dir = ruta_windows.split('\\')[-1] if ruta_windows and '\\' in ruta_windows else (ruta_windows or 'Sin ruta')

                    detalles_eliminados.append({
                        'id': directorio.id_disposicion,
                        'nombre': nombre_dir,
                        'archivos_eliminados': archivos_count
                    })

                    # Eliminar primero archivos de ArchivosPost (tiene DO_NOTHING en FK)
                    directorio.archivos_relacionados.all().delete()

                    # Eliminar directorio (EvaluacionArchivosPost se elimina en cascada)
                    directorio.delete()
                    directorios_eliminados += 1
                    archivos_eliminados_total += archivos_count
                else:
                    # El directorio SÍ existe, no eliminarlo
                    directorios_omitidos += 1

            # Registrar la acción en el log
            logger.info(f"Depuración directorios municipio {municipio_id} por {request.user.username}: "
                       f"{directorios_eliminados} directorios eliminados, {archivos_eliminados_total} archivos en cascada, "
                       f"{directorios_omitidos} omitidos")

            return Response({
                'success': True,
                'municipio_id': municipio_id,
                'directorios_eliminados': directorios_eliminados,
                'archivos_eliminados': archivos_eliminados_total,
                'directorios_omitidos': directorios_omitidos,
                'message': f'Se eliminaron {directorios_eliminados} directorios inexistentes y '
                          f'{archivos_eliminados_total} archivos asociados. '
                          f'{directorios_omitidos} directorios existentes fueron preservados.',
                'usuario': request.user.username,
                'fecha': timezone.now().isoformat(),
                'detalles': detalles_eliminados[:50]
            })

        except Exception as e:
            logger.error(f"Error depurando directorios inexistentes: {str(e)}")
            return Response({
                'error': f'Error interno del servidor: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ===============================================
# 🔧 FUNCIONES DE UTILIDAD
# ===============================================

def extraer_mecanismo_financiacion_desde_ruta(ruta_completa):
    """
    Extrae el mecanismo de financiación de la ruta
    Ejemplo: \\repositorio\\...\\17\\380\\PGN\\03_post\\ → "PGN"
    """
    if not ruta_completa:
        return "SIN_MECANISMO"
    
    try:
        ruta_str = str(ruta_completa).replace('/', '\\')
        # Buscar patrón: \\{depto}\\{municipio}\\{MECANISMO}\\03_post
        match = re.search(r'\\(\d{2})\\(\d{3})\\([^\\]+)\\03_post', ruta_str)
        if match:
            mecanismo = match.group(3).strip()
            return mecanismo if mecanismo else "SIN_MECANISMO"
        return "SIN_MECANISMO"
    except Exception as e:
        print(f"⚠️ Error extrayendo mecanismo: {e}")
        return "ERROR_MECANISMO"

def extraer_etapa_desde_ruta(ruta_completa):
    """
    Extrae el prefijo de etapa operativa desde la ruta.
    Ejemplo: \\03_post\\01_aprob_econo\\... → "01"
             \\03_post\\05_prod_catas\\... → "05"

    Returns:
        str: Prefijo de 2 dígitos (01-12) o None si no se encuentra
    """
    if not ruta_completa:
        return None

    try:
        ruta_str = str(ruta_completa).replace('/', '\\')

        # Buscar el patrón \\03_post\\XX_...
        if '\\03_post\\' in ruta_str:
            inicio_post = ruta_str.find('\\03_post\\') + len('\\03_post\\')
            resto_ruta = ruta_str[inicio_post:]

            if resto_ruta:
                # Obtener el primer directorio después de 03_post
                if '\\' in resto_ruta:
                    primer_dir = resto_ruta.split('\\')[0].strip()
                else:
                    primer_dir = resto_ruta.strip()

                # Extraer los primeros 2 caracteres si son dígitos
                if primer_dir and len(primer_dir) >= 2:
                    prefijo = primer_dir[:2]
                    if prefijo.isdigit():
                        return prefijo

        return None
    except Exception as e:
        print(f"⚠️ Error extrayendo etapa: {e}")
        return None


def extraer_directorio_padre_desde_ruta(ruta_completa):
    """
    Extrae directorio padre después de \\03_post\\
    """
    if not ruta_completa:
        return "SIN_DIRECTORIO"
    
    try:
        ruta_str = str(ruta_completa).replace('/', '\\')
        
        # Buscar el patrón \\03_post\\
        if '\\03_post\\' in ruta_str:
            inicio_post = ruta_str.find('\\03_post\\') + len('\\03_post\\')
            resto_ruta = ruta_str[inicio_post:]
            
            if resto_ruta:
                # Si hay contenido después de 03_post\
                if '\\' in resto_ruta:
                    directorio_padre = resto_ruta.split('\\')[0].strip()
                    return directorio_padre if directorio_padre else "ROOT_03_POST"
                else:
                    # Solo hay un directorio
                    return resto_ruta.strip() if resto_ruta.strip() else "ROOT_03_POST"
            else:
                # Termina en 03_post\
                return "ROOT_03_POST"
        else:
            return "SIN_03_POST"
            
    except Exception as e:
        print(f"⚠️ Error extrayendo directorio padre: {e}")
        return "ERROR_DIRECTORIO"

def extraer_jerarquia_completa_desde_ruta(ruta_completa):
    """
    Extrae jerarquía completa después de \\03_post\\
    Ejemplo: \\...\\03_post\\01_aprob_econo\\01_info_tecn_zhg\\archivo.pdf → "01_aprob_econo>01_info_tecn_zhg"
    """
    if not ruta_completa:
        return "SIN_JERARQUIA"
    
    try:
        ruta_str = str(ruta_completa).replace('/', '\\')
        
        if '\\03_post\\' in ruta_str:
            inicio_post = ruta_str.find('\\03_post\\') + len('\\03_post\\')
            resto_ruta = ruta_str[inicio_post:]
            
            if resto_ruta:
                # Dividir en partes y filtrar archivos
                partes = resto_ruta.split('\\')
                # Filtrar partes vacías y archivos (que contienen punto)
                partes_validas = []
                for parte in partes:
                    parte = parte.strip()
                    if parte and '.' not in parte:  # No es archivo
                        partes_validas.append(parte)
                
                if partes_validas:
                    return '>'.join(partes_validas)
                else:
                    return "ROOT_03_POST"
            else:
                return "ROOT_03_POST"
        else:
            return "SIN_03_POST"
            
    except Exception as e:
        print(f"⚠️ Error extrayendo jerarquía: {e}")
        return "ERROR_JERARQUIA"

def extraer_ruta_hasta_directorio_padre(ruta_completa):
    """
    Extrae ruta hasta el directorio padre (sin archivo)
    CONVIERTE de formato Linux a formato Windows para usuarios
    """
    if not ruta_completa:
        return "SIN_RUTA"

    try:
        ruta_str = str(ruta_completa)

        # Si termina en archivo (contiene extensión), remover el archivo
        if '/' in ruta_str or '\\' in ruta_str:
            # Normalizar a / para procesar
            ruta_normalizada = ruta_str.replace('\\', '/')
            partes = ruta_normalizada.split('/')

            if partes and partes[-1] and '.' in partes[-1]:
                # Última parte es archivo, removerla
                partes = partes[:-1]
                resultado = '/'.join(partes)
            else:
                # Ya es un directorio
                resultado = ruta_normalizada.rstrip('/')
        else:
            resultado = ruta_str

        # 🔄 Convertir de Linux a Windows para que los usuarios puedan acceder
        resultado = linux_to_windows_path(resultado)

        # Agregar \ al final si no la tiene (formato directorio Windows)
        if resultado and not resultado.endswith('\\'):
            resultado += '\\'

        return resultado

    except Exception as e:
        print(f"⚠️ Error extrayendo ruta directorio: {e}")
        return "ERROR_RUTA"

def obtener_concepto_calificacion(evaluacion_archivo_id):
    """
    Obtiene concepto de calificación con manejo de errores
    """
    try:
        if evaluacion_archivo_id is None or evaluacion_archivo_id == 0:
            return "SIN CALIFICACIÓN (PENDIENTE)"

        # Query optimizada
        try:
            calificacion = CalificacionInfoPost.objects.get(id=evaluacion_archivo_id)
            return calificacion.concepto
        except CalificacionInfoPost.DoesNotExist:
            return f"CALIFICACIÓN ID {evaluacion_archivo_id} NO ENCONTRADA"

    except Exception as e:
        print(f"⚠️ Error obteniendo concepto calificación: {e}")
        return "ERROR CALIFICACIÓN"


def obtener_valor_calificacion(evaluacion_archivo_id):
    """
    Obtiene el valor numérico de la calificación (0-1)
    """
    try:
        if evaluacion_archivo_id is None or evaluacion_archivo_id == 0:
            return 0.0

        try:
            calificacion = CalificacionInfoPost.objects.get(id=evaluacion_archivo_id)
            return float(calificacion.valor) if calificacion.valor is not None else 0.0
        except CalificacionInfoPost.DoesNotExist:
            return 0.0

    except Exception as e:
        print(f"⚠️ Error obteniendo valor calificación: {e}")
        return 0.0


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mecanismos_municipio(request, municipio_id):
    """
    Endpoint para obtener los mecanismos de financiación de un municipio.
    Incluye estadísticas de archivos por mecanismo.

    GET /postoperacion/mecanismos/<municipio_id>/

    Returns:
        JSON con lista de mecanismos y estadísticas
    """
    try:
        # Verificar que el municipio existe
        try:
            municipio = Municipios.objects.get(cod_municipio=municipio_id)
        except Municipios.DoesNotExist:
            return Response({
                'success': False,
                'error': f'Municipio {municipio_id} no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)

        # Obtener mecanismos con estadísticas
        disposiciones = DisposicionPost.objects.filter(cod_municipio=municipio_id)

        mecanismos_stats = {}
        for disposicion in disposiciones:
            if disposicion.ruta_acceso:
                mecanismo = extraer_mecanismo_financiacion_desde_ruta(disposicion.ruta_acceso)
                if mecanismo and mecanismo not in ["SIN_MECANISMO", "ERROR_MECANISMO"]:
                    if mecanismo not in mecanismos_stats:
                        mecanismos_stats[mecanismo] = {
                            'codigo': mecanismo,
                            'total_directorios': 0,
                            'total_archivos': 0
                        }
                    mecanismos_stats[mecanismo]['total_directorios'] += 1
                    # Contar archivos de esta disposición
                    archivos_count = disposicion.archivos_relacionados.count() if hasattr(disposicion, 'archivos_relacionados') else 0
                    mecanismos_stats[mecanismo]['total_archivos'] += archivos_count

        # Convertir a lista ordenada
        mecanismos_list = sorted(mecanismos_stats.values(), key=lambda x: x['codigo'])

        return Response({
            'success': True,
            'municipio': {
                'cod_municipio': municipio.cod_municipio,
                'nom_municipio': municipio.nom_municipio
            },
            'mecanismos': mecanismos_list,
            'total_mecanismos': len(mecanismos_list),
            'tiene_multiples': len(mecanismos_list) > 1
        })

    except Exception as e:
        print(f"❌ Error obteniendo mecanismos para {municipio_id}: {e}")
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def obtener_mecanismos_financiacion_post(municipio_id):
    """
    Obtiene mecanismos de financiación usando disposicion_post (uso interno)
    """
    try:
        print(f"🔍 Obteniendo mecanismos para municipio {municipio_id}...")
        
        # USAR DISPOSICION_POST como fuente principal
        disposiciones_municipio = DisposicionPost.objects.filter(cod_municipio=municipio_id)
        
        if not disposiciones_municipio.exists():
            print(f"⚠️ No se encontraron disposiciones para municipio {municipio_id}")
            return ["GENERAL"]
        
        mecanismos = set()
        
        for disposicion in disposiciones_municipio:
            if disposicion.ruta_acceso:
                mecanismo = extraer_mecanismo_financiacion_desde_ruta(disposicion.ruta_acceso)
                if mecanismo and mecanismo not in ["SIN_MECANISMO", "ERROR_MECANISMO"]:
                    mecanismos.add(mecanismo)
        
        mecanismos_list = sorted(list(mecanismos)) if mecanismos else ["GENERAL"]
        
        print(f"✅ Mecanismos encontrados para {municipio_id}: {mecanismos_list}")
        return mecanismos_list
        
    except Exception as e:
        print(f"❌ Error obteniendo mecanismos para {municipio_id}: {e}")
        return ["GENERAL"]

# ===============================================
# 🚀 FUNCIÓN PRINCIPAL DE REPORTES
# ===============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generar_reportes_postoperacion(request):
    """
    Genera reportes POST usando AMBAS tablas: disposicion_post + evaluacion_archivos_post
    Soporta filtro de etapas operativas (01, 02, 03, etc.)
    """
    try:
        data = json.loads(request.body) if request.body else request.data
        municipios_ids = data.get('municipios', [])
        etapas_filtro = data.get('etapas_filtro', None)  # Lista de prefijos: ['01', '02', '03', ...]
        generar_individuales = data.get('generar_individuales', True)
        generar_resumen = data.get('generar_resumen', False)

        if not municipios_ids:
            return Response({'error': 'No se proporcionaron municipios'}, status=status.HTTP_400_BAD_REQUEST)

        print(f"🚀 GENERANDO REPORTES POST para {len(municipios_ids)} municipios...")
        print(f"   📄 Individuales: {generar_individuales} | 📊 Resumen: {generar_resumen}")
        if etapas_filtro:
            print(f"📂 Filtro de etapas activo: {etapas_filtro}")

        with tempfile.TemporaryDirectory() as temp_dir:
            archivos_generados = []
            municipios_mecanismos_procesados = []  # Para la matriz resumen

            for municipio_id in municipios_ids:
                try:
                    municipio = Municipios.objects.get(cod_municipio=municipio_id)
                    print(f"📊 Procesando municipio: {municipio.nom_municipio} ({municipio_id})")

                    # OBTENER MECANISMOS DE FINANCIACIÓN DEL MUNICIPIO
                    mecanismos_municipio = obtener_mecanismos_financiacion_post(municipio_id)
                    print(f"💰 Mecanismos encontrados: {mecanismos_municipio}")

                    # GENERAR REPORTE POR CADA MECANISMO
                    for mecanismo in mecanismos_municipio:
                        # Acumular para matriz resumen
                        municipios_mecanismos_procesados.append((municipio, mecanismo))

                        if generar_individuales:
                            print(f"📋 Generando reporte individual para mecanismo: {mecanismo}")
                            archivo_path = generar_reporte_individual_postoperacion(
                                municipio, temp_dir, mecanismo, etapas_filtro=etapas_filtro
                            )
                            if archivo_path:
                                archivos_generados.append(archivo_path)
                                print(f"✅ Reporte generado: {os.path.basename(archivo_path)}")
                            else:
                                print(f"❌ Error generando reporte para {mecanismo}")

                except Municipios.DoesNotExist:
                    print(f"⚠️ Municipio {municipio_id} no encontrado")
                    continue
                except Exception as e:
                    print(f"❌ Error procesando municipio {municipio_id}: {str(e)}")
                    continue

            # GENERAR MATRIZ RESUMEN si fue solicitada
            if generar_resumen and municipios_mecanismos_procesados:
                print(f"📊 Generando MATRIZ RESUMEN para {len(municipios_mecanismos_procesados)} combinaciones...")
                archivo_resumen = generar_matriz_resumen_postoperacion(
                    municipios_mecanismos_procesados, temp_dir, etapas_filtro
                )
                if archivo_resumen:
                    archivos_generados.append(archivo_resumen)
                    print(f"✅ Matriz resumen generada: {os.path.basename(archivo_resumen)}")
                else:
                    print(f"❌ Error generando matriz resumen")

            if not archivos_generados:
                return Response({'error': 'No se generaron reportes'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # CREAR ZIP en archivo temporal para STREAMING
            fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_zip = f'reportes_postoperacion_{fecha_actual}.zip'

            temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
            temp_zip_path = temp_zip.name
            temp_zip.close()

            with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=1) as zip_file:
                for archivo_path in archivos_generados:
                    zip_file.write(archivo_path, os.path.basename(archivo_path))

            # Streaming del ZIP
            def zip_iterator():
                CHUNK_SIZE = 8 * 1024 * 1024  # 8MB chunks
                try:
                    with open(temp_zip_path, 'rb') as f:
                        while True:
                            chunk = f.read(CHUNK_SIZE)
                            if not chunk:
                                break
                            yield chunk
                finally:
                    try:
                        os.unlink(temp_zip_path)
                    except:
                        pass

            response = StreamingHttpResponse(
                zip_iterator(),
                content_type='application/zip'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_zip}"'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['X-Accel-Buffering'] = 'no'  # Desactivar buffering en nginx

            print(f"✅ REPORTES POST generados con streaming:")
            print(f"   📦 Archivos: {len(archivos_generados)}")
            print(f"   🎯 ZIP: {nombre_zip}")

            return response
            
    except Exception as e:
        print(f"❌ Error general POST: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': f'Error interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ===============================================
# 🎨 PESTAÑAS DEL EXCEL
# ===============================================



# ===============================================
# 🔄 FUNCIONES DE ESTADÍSTICAS ACTUALIZADAS
# ===============================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estadisticas_post(request):
    """Estadísticas POST sin componentes"""
    try:
        # Estadísticas generales
        total_municipios = Municipios.objects.count()
        municipios_con_disposicion = DisposicionPost.objects.values('cod_municipio').distinct().count()
        
        # Estado general (SIN componentes)
        total_disposiciones = DisposicionPost.objects.count()
        dispuestos = DisposicionPost.objects.filter(dispuesto=True).count()
        evaluados = DisposicionPost.objects.filter(evaluado=True).count()
        aprobados = DisposicionPost.objects.filter(aprobado=True).count()
        
        # NUEVO: Estadísticas por municipio
        por_municipio = DisposicionPost.objects.values(
            'cod_municipio__nom_municipio',
            'cod_municipio__cod_municipio'
        ).annotate(
            total=Count('id_disposicion'),
            dispuestos=Count('id_disposicion', filter=Q(dispuesto=True)),
            evaluados=Count('id_disposicion', filter=Q(evaluado=True)),
            aprobados=Count('id_disposicion', filter=Q(aprobado=True)),
            total_archivos=Count('archivos_relacionados')
        ).order_by('cod_municipio__nom_municipio')
        
        # Últimas actualizaciones (SIN componente)
        ultimas_actualizaciones = DisposicionPost.objects.select_related(
            'cod_municipio'
        ).order_by('-fecha_disposicion')[:10]
        
        ultimas_serialized = []
        for disp in ultimas_actualizaciones:
            if disp.fecha_disposicion:  # Solo incluir si tiene fecha
                ultimas_serialized.append({
                    'id_disposicion': disp.id_disposicion,
                    'municipio': disp.cod_municipio.nom_municipio,
                    'directorio': disp.get_nombre_directorio(),
                    'ruta_acceso': disp.ruta_acceso,
                    'fecha': disp.fecha_disposicion.strftime('%Y-%m-%d'),
                    'dispuesto': disp.dispuesto,
                    'evaluado': disp.evaluado,
                    'aprobado': disp.aprobado,
                    'total_archivos': disp.archivos_relacionados.count()
                })
        
        return Response({
            'general': {
                'total_municipios': total_municipios,
                'municipios_con_disposicion': municipios_con_disposicion,
                'porcentaje_cobertura': round((municipios_con_disposicion / total_municipios) * 100, 2) if total_municipios > 0 else 0
            },
            'estado': {
                'total_disposiciones': total_disposiciones,
                'dispuestos': dispuestos,
                'evaluados': evaluados,
                'aprobados': aprobados,
                'porcentaje_dispuestos': round((dispuestos / total_disposiciones) * 100, 2) if total_disposiciones > 0 else 0,
                'porcentaje_evaluados': round((evaluados / total_disposiciones) * 100, 2) if total_disposiciones > 0 else 0,
                'porcentaje_aprobados': round((aprobados / total_disposiciones) * 100, 2) if total_disposiciones > 0 else 0
            },
            'por_municipio': por_municipio,
            'ultimas_actualizaciones': ultimas_serialized
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_consolidado_post(request):
    """Reporte consolidado POST por directorios"""
    try:
        # Obtener todos los municipios
        municipios = Municipios.objects.all().order_by('cod_depto__nom_depto', 'nom_municipio')
        
        resultado = []
        
        for municipio in municipios:
            municipio_data = {
                'cod_municipio': municipio.cod_municipio,
                'nom_municipio': municipio.nom_municipio,
                'departamento': municipio.cod_depto.nom_depto,
                'disposiciones': []
            }
            
            # Obtener todas las disposiciones para este municipio
            disposiciones = DisposicionPost.objects.filter(cod_municipio=municipio.cod_municipio)
            
            # Para cada disposición, registrar su estado
            for disposicion in disposiciones:
                municipio_data['disposiciones'].append({
                    'id_disposicion': disposicion.id_disposicion,
                    'nombre_directorio': disposicion.get_nombre_directorio(),
                    'ruta_acceso': disposicion.ruta_acceso,
                    'dispuesto': disposicion.dispuesto,
                    'evaluado': disposicion.evaluado,
                    'aprobado': disposicion.aprobado,
                    'fecha_disposicion': disposicion.fecha_disposicion,
                    'observaciones': disposicion.observaciones,
                    'total_archivos': disposicion.archivos_relacionados.count()
                })
            
            resultado.append(municipio_data)
        
        return Response(resultado)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ===============================================
# 🆕 ENDPOINTS ADICIONALES PARA CALIFICACIONES
# ===============================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def calificacion_completa(request):
    """Endpoint para obtener el valor de documento completo"""
    try:
        valor = CalificacionInfoPost.get_valor_completo()
        return Response({
            'valor_completo': float(valor),
            'porcentaje_completo': float(valor) * 100
        })
    except Exception as e:
        logger.error(f"Error obteniendo valor completo: {str(e)}")
        return Response({
            'error': 'Error interno del servidor'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def inicializar_calificaciones(request):
    """Endpoint para inicializar las calificaciones con valores por defecto"""
    try:
        # Verificar si ya existen calificaciones
        if CalificacionInfoPost.objects.exists():
            return Response({
                'message': 'Las calificaciones ya han sido inicializadas'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Datos por defecto
        calificaciones_default = [
            {'concepto': 'NO HAY INFORMACIÓN', 'valor': 0.0},
            {'concepto': 'NO ABRE EL DOCUMENTO', 'valor': 0.1},
            {'concepto': 'DOCUMENTO SIN FIRMAS DE QUIEN PROYECTO O EN FORMATO WORD', 'valor': 0.25},
            {'concepto': 'DOCUMENTO INCOMPLETO O SIN FIRMA DE APROBACION DEL ENLACE OPERATIVO MUNICIPAL O ENLACE OPERATIVO REGIONAL O DIRECTOR TERRITORIAL', 'valor': 0.5},
            {'concepto': 'DOCUMENTO SIN FIRMA DIRECTOR TERRITORIAL O CON FIRMAS PERO EN FORMATO WORD O ARCHIVO PDF GUARDADO CON CONTROL DE CAMBIOS', 'valor': 0.75},
            {'concepto': 'DOCUMENTO PENDIENTE AJUSTES DE FORMA', 'valor': 0.9},
            {'concepto': 'DOCUMENTO COMPLETO', 'valor': 1.0},
        ]
        
        # Crear calificaciones
        calificaciones_creadas = []
        for data in calificaciones_default:
            calificacion = CalificacionInfoPost.objects.create(**data)
            calificaciones_creadas.append(calificacion)
        
        serializer = CalificacionInfoPostSerializer(calificaciones_creadas, many=True)
        
        return Response({
            'message': f'Se inicializaron {len(calificaciones_creadas)} calificaciones',
            'calificaciones': serializer.data
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error inicializando calificaciones: {str(e)}")
        return Response({
            'error': 'Error interno del servidor'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ===============================================
# 🔄 ENDPOINT PARA MIGRACIÓN DE DATOS
# ===============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def migrar_datos_post(request):
    """Migra datos de la arquitectura antigua a POST V2"""
    try:
        resultado = {
            'version': 'POST V2',
            'fecha_migracion': datetime.now().isoformat(),
            'migracion': {}
        }
        
        # Solo permitir a administradores
        if not request.user.is_staff:
            return Response({
                'error': 'Solo administradores pueden ejecutar migraciones'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Verificar si hay datos para migrar
        total_disposiciones = DisposicionPost.objects.count()
        total_archivos = ArchivosPost.objects.count()
        
        resultado['migracion']['datos_actuales'] = {
            'disposiciones': total_disposiciones,
            'archivos': total_archivos
        }
        
        # Verificar constraints únicos
        from django.db import connection
        cursor = connection.cursor()
        
        try:
            cursor.execute("""
                SELECT COUNT(*) FROM (
                    SELECT ruta_acceso, COUNT(*) 
                    FROM disposicion_post 
                    WHERE ruta_acceso IS NOT NULL 
                    GROUP BY ruta_acceso 
                    HAVING COUNT(*) > 1
                ) duplicados
            """)
            duplicados_disposicion = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COUNT(*) FROM (
                    SELECT ruta_completa, COUNT(*) 
                    FROM archivos_post 
                    WHERE ruta_completa IS NOT NULL 
                    GROUP BY ruta_completa 
                    HAVING COUNT(*) > 1
                ) duplicados
            """)
            duplicados_archivos = cursor.fetchone()[0]
            
            resultado['migracion']['duplicados'] = {
                'disposiciones': duplicados_disposicion,
                'archivos': duplicados_archivos
            }
            
        except Exception as e:
            resultado['migracion']['duplicados'] = {
                'error': str(e)
            }
        
        # Sugerir acciones
        acciones_sugeridas = []
        
        if duplicados_disposicion > 0:
            acciones_sugeridas.append("Eliminar duplicados en disposicion_post.ruta_acceso")
        
        if duplicados_archivos > 0:
            acciones_sugeridas.append("Eliminar duplicados en archivos_post.ruta_completa")
        
        if total_disposiciones == 0:
            acciones_sugeridas.append("Ejecutar Script_POST_V2_COMPLETO.py para poblar datos")
        
        resultado['migracion']['acciones_sugeridas'] = acciones_sugeridas
        resultado['migracion']['estado'] = 'OK' if len(acciones_sugeridas) == 0 else 'REQUIERE_ACCION'
        
        return Response(resultado)
        
    except Exception as e:
        logger.error(f"Error en migración POST V2: {str(e)}")
        return Response({
            'version': 'POST V2',
            'estado': 'ERROR',
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ===============================================
# 🔄 ENDPOINT PARA LIMPIAR DUPLICADOS
# ===============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def limpiar_duplicados_post(request):
    """Limpia duplicados en POST V2"""
    try:
        # Solo permitir a administradores
        if not request.user.is_staff:
            return Response({
                'error': 'Solo administradores pueden limpiar duplicados'
            }, status=status.HTTP_403_FORBIDDEN)
        
        resultado = {
            'version': 'POST V2',
            'fecha_limpieza': datetime.now().isoformat(),
            'limpieza': {}
        }
        
        from django.db import transaction
        
        with transaction.atomic():
            # Limpiar duplicados en disposicion_post
            from django.db import connection
            cursor = connection.cursor()
            
            cursor.execute("""
                DELETE FROM disposicion_post 
                WHERE id_disposicion NOT IN (
                    SELECT DISTINCT ON (ruta_acceso)
                        id_disposicion
                    FROM disposicion_post 
                    WHERE ruta_acceso IS NOT NULL
                    ORDER BY 
                        ruta_acceso, 
                        fecha_disposicion DESC NULLS LAST,
                        id_disposicion DESC
                )
            """)
            duplicados_disposicion_eliminados = cursor.rowcount
            
            cursor.execute("""
                DELETE FROM archivos_post 
                WHERE id_archivo NOT IN (
                    SELECT DISTINCT ON (ruta_completa)
                        id_archivo
                    FROM archivos_post 
                    WHERE ruta_completa IS NOT NULL
                    ORDER BY 
                        ruta_completa, 
                        fecha_disposicion DESC NULLS LAST,
                        id_archivo DESC
                )
            """)
            duplicados_archivos_eliminados = cursor.rowcount
            
            resultado['limpieza'] = {
                'disposiciones_duplicadas_eliminadas': duplicados_disposicion_eliminados,
                'archivos_duplicados_eliminados': duplicados_archivos_eliminados,
                'estado': 'COMPLETADO'
            }
        
        return Response(resultado)
        
    except Exception as e:
        logger.error(f"Error limpiando duplicados POST V2: {str(e)}")
        return Response({
            'version': 'POST V2',
            'estado': 'ERROR',
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ===============================================
# 🔄 ENDPOINT DE SALUD DEL SISTEMA
# ===============================================

@api_view(['GET'])
@permission_classes([AllowAny])
def health_check_post(request):
    """Health check para POST V2"""
    try:
        # Verificaciones básicas
        total_disposiciones = DisposicionPost.objects.count()
        total_archivos = ArchivosPost.objects.count()
        total_evaluaciones = EvaluacionArchivosPost.objects.count()
        
        # Verificar conexión a base de datos
        from django.db import connection
        cursor = connection.cursor()
        cursor.execute("SELECT 1")
        db_ok = cursor.fetchone()[0] == 1
        
        # Estado del sistema
        sistema_ok = all([
            total_disposiciones >= 0,
            total_archivos >= 0,
            total_evaluaciones >= 0,
            db_ok
        ])
        
        return Response({
            'version': 'POST V2',
            'timestamp': datetime.now().isoformat(),
            'status': 'OK' if sistema_ok else 'ERROR',
            'database': 'OK' if db_ok else 'ERROR',
            'estadisticas': {
                'disposiciones': total_disposiciones,
                'archivos': total_archivos,
                'evaluaciones': total_evaluaciones
            },
            'arquitectura': 'POST V2 - Sin dependencia de componentes'
        })
        
    except Exception as e:
        return Response({
            'version': 'POST V2',
            'timestamp': datetime.now().isoformat(),
            'status': 'ERROR',
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ===============================================
# 🆕 NUEVOS ENDPOINTS PARA ESTADÍSTICAS GENERALES
# ===============================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estadisticas_generales_post(request):
    """Estadísticas generales POST usando ambas tablas"""
    try:
        # USANDO DISPOSICION_POST como base
        total_disposiciones = DisposicionPost.objects.count()
        municipios_con_disposiciones = DisposicionPost.objects.values('cod_municipio').distinct().count()
        
        # USANDO EVALUACION_ARCHIVOS_POST para archivos
        total_evaluaciones = EvaluacionArchivosPost.objects.count()
        evaluaciones_evaluadas = EvaluacionArchivosPost.objects.filter(evaluado=True).count()
        evaluaciones_aprobadas = EvaluacionArchivosPost.objects.filter(aprobado=True).count()
        
        # MECANISMOS DE FINANCIACIÓN
        mecanismos_encontrados = set()
        disposiciones_con_ruta = DisposicionPost.objects.exclude(ruta_acceso__isnull=True).exclude(ruta_acceso='')
        
        for disposicion in disposiciones_con_ruta:
            mecanismo = extraer_mecanismo_financiacion_desde_ruta(disposicion.ruta_acceso)
            if mecanismo not in ["SIN_MECANISMO", "ERROR_MECANISMO"]:
                mecanismos_encontrados.add(mecanismo)
        
        return Response({
            'version': 'POST COMPLETO',
            'timestamp': datetime.now().isoformat(),
            'disposiciones': {
                'total': total_disposiciones,
                'municipios_con_disposiciones': municipios_con_disposiciones
            },
            'evaluaciones': {
                'total': total_evaluaciones,
                'evaluadas': evaluaciones_evaluadas,
                'aprobadas': evaluaciones_aprobadas,
                'porcentaje_evaluado': round((evaluaciones_evaluadas / total_evaluaciones) * 100, 2) if total_evaluaciones > 0 else 0,
                'porcentaje_aprobado': round((evaluaciones_aprobadas / total_evaluaciones) * 100, 2) if total_evaluaciones > 0 else 0
            },
            'mecanismos_financiacion': {
                'total_encontrados': len(mecanismos_encontrados),
                'lista': sorted(list(mecanismos_encontrados))
            }
        })
        
    except Exception as e:
        logger.error(f"Error obteniendo estadísticas POST COMPLETO: {str(e)}")
        return Response({
            'error': f'Error interno del servidor: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def validar_directorios_post(request):
    """Valida usando ambas tablas"""
    try:
        municipio_id = request.query_params.get('municipio_id')
        
        print(f"🔍 Validando para municipio: {municipio_id}")
        
        # VALIDAR DISPOSICIONES
        if municipio_id:
            disposiciones = DisposicionPost.objects.filter(cod_municipio=municipio_id)[:20]
        else:
            disposiciones = DisposicionPost.objects.all()[:20]
        
        # VALIDAR EVALUACIONES
        if municipio_id:
            evaluaciones = EvaluacionArchivosPost.objects.filter(
                id_disposicion__cod_municipio=municipio_id
            ).select_related('id_disposicion')[:20]
        else:
            evaluaciones = EvaluacionArchivosPost.objects.all().select_related('id_disposicion')[:20]
        
        resultados_disposiciones = []
        directorios_padre_disposiciones = set()
        mecanismos_disposiciones = set()
        
        for disposicion in disposiciones:
            directorio_padre = extraer_directorio_padre_desde_ruta(disposicion.ruta_acceso)
            mecanismo = extraer_mecanismo_financiacion_desde_ruta(disposicion.ruta_acceso)
            
            directorios_padre_disposiciones.add(directorio_padre)
            mecanismos_disposiciones.add(mecanismo)
            
            resultados_disposiciones.append({
                'id_disposicion': disposicion.id_disposicion,
                'ruta_acceso': disposicion.ruta_acceso[:100] + '...' if len(disposicion.ruta_acceso or '') > 100 else disposicion.ruta_acceso,
                'directorio_padre': directorio_padre,
                'mecanismo_financiacion': mecanismo,
                'dispuesto': disposicion.dispuesto,
                'evaluado': disposicion.evaluado,
                'aprobado': disposicion.aprobado,
                'municipio': disposicion.cod_municipio.cod_municipio
            })
        
        resultados_evaluaciones = []
        directorios_padre_evaluaciones = set()
        
        for evaluacion in evaluaciones:
            directorio_padre = extraer_directorio_padre_desde_ruta(evaluacion.ruta_completa)
            concepto = obtener_concepto_calificacion(evaluacion.evaluacion_archivo)
            
            directorios_padre_evaluaciones.add(directorio_padre)
            
            resultados_evaluaciones.append({
                'id_evaluacion': evaluacion.id_evaluacion,
                'nombre_archivo': evaluacion.nombre_archivo,
                'ruta_completa': evaluacion.ruta_completa[:100] + '...' if len(evaluacion.ruta_completa) > 100 else evaluacion.ruta_completa,
                'directorio_padre': directorio_padre,
                'evaluacion_concepto': concepto,
                'evaluado': evaluacion.evaluado,
                'aprobado': evaluacion.aprobado
            })
        
        return Response({
            'version': 'POST COMPLETO',
            'municipio_consultado': municipio_id,
            'disposiciones': {
                'total_analizadas': len(resultados_disposiciones),
                'directorios_padre_unicos': sorted(list(directorios_padre_disposiciones)),
                'mecanismos_financiacion': sorted(list(mecanismos_disposiciones)),
                'muestras': resultados_disposiciones[:5]
            },
            'evaluaciones': {
                'total_analizadas': len(resultados_evaluaciones),
                'directorios_padre_unicos': sorted(list(directorios_padre_evaluaciones)),
                'muestras': resultados_evaluaciones[:5]
            },
            'estado': 'EXTRACCIÓN FUNCIONANDO - USANDO AMBAS TABLAS'
        })
        
    except Exception as e:
        logger.error(f"Error validando directorios POST: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def test_post(request):
    """Test básico para verificar que POST funciona"""
    return Response({
        'message': '¡POST está funcionando!',
        'version': 'POST COMPLETO',
        'timestamp': datetime.now().isoformat()
    })


def nombre_legible_directorio(nombre_dir):
    """Convierte nombres como '01_aprob_econo' a 'APROBACION ECONOMICA'"""
    traducciones = {
        '01_aprob_econo': 'APROBACION ECONOMICA',
        '01_info_tecn_zhg': 'INFORME_TECNICO_ZONAS_HOMOGENENAS',
        '01_analis_sensib': 'Análisis de sensibilidad',
        '02_proyec_avalu': 'Proyección de avaluos',
        '03_tab_terr_const': 'Tablas de terreno y construcción',
        '02_acta_comit': 'ACTA DE COMITÉ DE AVALUOS',
        '03_conc_favor': 'CONCEPTO FAVORABLE',
        '04_memo_tecn_estu_econo': 'MEMORIA TECNICA DE ESTUDIO ECONOMICO',
        '01_plano': 'Planos',
        '01_urb': 'Urbano',
        '02_rur': 'Rural',
        '03_urb_rur': 'Urbano-Rural',
        '02_zhf_zhg': 'Fgdb',
        '03_fto': 'Formatos',
        '04_memor': 'Memoria',
        '05_comuni_adopc_porc_aval_catas': 'COMUNICACIÓN ADOPCION PORCENTAJE AVALUO CATASTRAL',
        '06_resol_aprob': 'RESOLUCION APROBACION',
        '01_resol': 'Resolución',
        '02_publi_dia_ofic': 'Publicación diario oficial',
        '03_ofic_comuni': 'Oficios de comunicación',
        '02_comp_soci': 'COMPONENTE SOCIAL',
        '01_niv_interloc_iv': 'NIVEL INTERLOCUCION IV',
        '01_infor': 'Informe',
        '02_regis_fotog': 'Registro Fotográfico',
        '03_list_asis': 'Lista de asistencia',
        '04_difus': 'Difusión',
        '03_pre_cie': 'PRE-CIERRE',
        '01_eviden_atenc_sald': 'Evidencia de atención de saldos',
        '02_repor_cica_fin': 'Reporte cica final',
        '03_base_renmrc': 'Base consolidada gráfica y alfanumérica',
        '04_tab_ph': 'Tablas de terreno y construcción',
        '05_prelim': 'Preliquidación final',
        '04_resol_insc': 'RESOLUCION INSCRIPCION',
        '03_ofic_comun_resol': 'Oficios de comunicación',
        '05_prod_catas': 'PRODUCTOS CATASTRALES',
        '01_list_pred_actualiz': 'Listado de predios actualizados',
    }
    nombre_lower = nombre_dir.lower()
    if nombre_lower in traducciones:
        return traducciones[nombre_lower]
    limpio = re.sub(r'^\d+_', '', nombre_dir)
    limpio = limpio.replace('_', ' ').title()
    return limpio


def convertir_booleano_a_texto(valor_boolean):
    """
    Convierte boolean a ✓ SÍ / ✗ NO
    """
    if valor_boolean is True:
        return "✓ SÍ"
    elif valor_boolean is False:
        return "✗ NO"
    else:
        return "⏳ Pendiente"

def generar_pestana_resumen_general(ws, municipio, datos_completos, mecanismo_financiacion):
    """
    PESTAÑA 1: Resumen General por directorios con índices jerárquicos y componentes
    *** MODIFICADO v2: Columna COMPONENTE, índices jerárquicos, sin DISPUESTO ***
    """
    # ESTILOS
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # ESTILOS PARA CATEGORÍAS PRINCIPALES
    categoria_font = Font(bold=True, color="1F4E79", size=10)
    categoria_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

    # ENCABEZADO
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')

    # PRIMERO: Hacer merge (8 columnas ahora: A-H)
    ws.merge_cells('A1:H2')

    # SEGUNDO: Configurar la celda principal
    titulo_cell = ws['A1']
    titulo_cell.value = f"MATRIZ DE DISPOSICIÓN DE LA INFORMACIÓN POST-OPERACIÓN\nMUNICIPIO DE {municipio.nom_municipio.upper()}\nMECANISMO: {mecanismo_financiacion}"
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align

    # TERCERO: Aplicar bordes a toda la celda mergeada
    for row_merge in range(1, 3):  # Filas 1 y 2
        for col_merge in range(1, 9):  # Columnas A-H (8 columnas)
            cell = ws.cell(row=row_merge, column=col_merge)
            cell.border = border

    # INFO MUNICIPIO
    ws.merge_cells('A3:H3')
    departamento = municipio.cod_depto.nom_depto if municipio.cod_depto else "N/A"
    info_cell = ws['A3']
    info_cell.value = f"📍 Código: {municipio.cod_municipio} | 🏛️ Departamento: {departamento} | 💰 Mecanismo: {mecanismo_financiacion} | 📅 Generado: {fecha_actual}"
    info_cell.font = Font(bold=True, size=11, color="1F4E79")
    info_cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    info_cell.alignment = center_align

    # Aplicar bordes a la fila de info
    for col_merge in range(1, 9):  # Columnas A-H
        cell = ws.cell(row=3, column=col_merge)
        cell.border = border

    # HEADERS - Nueva estructura: N°, COMPONENTE, DIRECTORIO, FECHA, TOTAL, RESUMEN, OBSERVACIONES, RUTA
    headers = [
        'N°', 'COMPONENTE', 'DIRECTORIO/SUBCARPETA', 'FECHA\nDISPOSICIÓN',
        'TOTAL\nARCHIVOS', 'RESUMEN EVALUACIÓN', 'OBSERVACIONES EVALUACION', 'RUTA DE ACCESO'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # *** FILTRADO ESTRICTO DE DATOS INVÁLIDOS *** (sin cambios)
    print("🔍 ANTES DEL FILTRO:")
    print(f"   Total datos: {len(datos_completos)}")
    
    # FILTRAR DATOS VÁLIDOS ÚNICAMENTE
    datos_validos = []
    registros_filtrados = []
    
    for dato in datos_completos:
        jerarquia = dato['jerarquia_completa']
        
        # LISTA DE VALORES A EXCLUIR COMPLETAMENTE
        valores_excluir = [
            "SIN_03_POST", 
            "ROOT_03_POST",
            "SIN_JERARQUIA", 
            "ERROR_JERARQUIA", 
            "SIN_DIRECTORIO", 
            "ERROR_DIRECTORIO",
            "",
            None
        ]
        
        # TAMBIÉN FILTRAR POR RUTA SI CONTIENE PATRONES INVÁLIDOS
        ruta_disposicion = dato.get('ruta_directorio', '') or ''
        ruta_evaluacion = ''
        if dato.get('evaluacion') and hasattr(dato['evaluacion'], 'ruta_completa'):
            ruta_evaluacion = dato['evaluacion'].ruta_completa or ''
        
        rutas_a_revisar = [ruta_disposicion, ruta_evaluacion]
        contiene_ruta_invalida = any(
            any(patron in str(ruta).lower() for patron in ['sin_03_post', 'root_03_post', 'error']) 
            for ruta in rutas_a_revisar if ruta
        )
        
        # DECIDIR SI INCLUIR O EXCLUIR
        if jerarquia not in valores_excluir and not contiene_ruta_invalida and jerarquia:
            datos_validos.append(dato)
        else:
            registros_filtrados.append({
                'jerarquia': jerarquia,
                'ruta_disposicion': ruta_disposicion[:50] + '...' if len(ruta_disposicion) > 50 else ruta_disposicion,
                'motivo': 'jerarquia_invalida' if jerarquia in valores_excluir else 'ruta_invalida'
            })
    
    print("🔍 DESPUÉS DEL FILTRO:")
    print(f"   Datos válidos: {len(datos_validos)}")
    print(f"   Registros filtrados: {len(registros_filtrados)}")
    if registros_filtrados:
        print("   Ejemplos filtrados:")
        for filtrado in registros_filtrados[:3]:
            print(f"     - {filtrado['jerarquia']} ({filtrado['motivo']})")
    
    # AGRUPAR POR JERARQUÍA COMPLETA (SOLO DATOS VÁLIDOS)
    directorios_jerarquia = defaultdict(list)
    for dato in datos_validos:
        jerarquia_completa = dato['jerarquia_completa']
        directorios_jerarquia[jerarquia_completa].append(dato)

    print(f"📁 Directorios agrupados: {len(directorios_jerarquia)}")

    # nombre_legible_directorio() ahora está a nivel de módulo (reutilizable)

    # ============================================================
    # CONSTRUIR ESTRUCTURA JERÁRQUICA PARA ÍNDICES
    # ============================================================
    # Estructura: {nivel0: {nivel1: {nivel2: ...}}}
    estructura_jerarquica = {}
    for jerarquia in sorted(directorios_jerarquia.keys()):
        partes = jerarquia.split('>')
        actual = estructura_jerarquica
        for parte in partes:
            if parte not in actual:
                actual[parte] = {}
            actual = actual[parte]

    # ============================================================
    # GENERAR ÍNDICES JERÁRQUICOS
    # ============================================================
    indice_mapping = {}  # {jerarquia_completa: indice_str}

    def asignar_indices(estructura, prefijo="", padre_indice=""):
        """Asigna índices jerárquicos recursivamente"""
        resultado = []
        for i, (nombre, hijos) in enumerate(sorted(estructura.items()), 1):
            # Construir la ruta completa
            ruta_completa = f"{prefijo}{nombre}" if prefijo else nombre

            # Generar índice
            if padre_indice:
                indice = f"{padre_indice}.{i}"
            else:
                indice = str(i)

            resultado.append((ruta_completa, indice, nombre))
            indice_mapping[ruta_completa] = indice

            # Procesar hijos
            if hijos:
                resultado.extend(asignar_indices(hijos, f"{ruta_completa}>", indice))

        return resultado

    # Generar lista ordenada con índices
    lista_ordenada = asignar_indices(estructura_jerarquica)

    # ============================================================
    # GENERAR FILAS DE DATOS CON NUEVA ESTRUCTURA
    # ============================================================
    row = 6

    for jerarquia_completa, indice_str, nombre_dir in lista_ordenada:
        datos_directorio = directorios_jerarquia.get(jerarquia_completa, [])

        # Determinar nivel de profundidad para estilo
        nivel = indice_str.count('.')

        # Estilo según nivel
        if nivel == 0:
            # Categoría principal: fondo azul claro
            row_fill = categoria_fill
            row_font = categoria_font
        else:
            # Subcategorías: alternar colores
            if row % 2 == 0:
                row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            else:
                row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            row_font = None

        # Obtener datos representativos
        if datos_directorio:
            disposicion_representativa = datos_directorio[0]['disposicion']
            total_archivos = len([d for d in datos_directorio if d['tipo'] == 'ARCHIVO_EVALUADO'])
        else:
            disposicion_representativa = None
            total_archivos = 0

        # Contar archivos evaluados y aprobados
        archivos_evaluados = 0
        archivos_aprobados = 0
        observaciones_lista = []

        for dato in datos_directorio:
            if dato['tipo'] == 'ARCHIVO_EVALUADO' and dato['evaluacion']:
                if dato['evaluacion'].evaluado:
                    archivos_evaluados += 1
                if dato['evaluacion'].aprobado:
                    archivos_aprobados += 1
                # Recoger observaciones
                if hasattr(dato['evaluacion'], 'observaciones') and dato['evaluacion'].observaciones:
                    observaciones_lista.append(dato['evaluacion'].observaciones)

        # Recoger rutas y fechas
        rutas_acceso = set()
        fechas_disposicion = []

        for dato in datos_directorio:
            if dato['ruta_directorio']:
                rutas_acceso.add(dato['ruta_directorio'])
            if dato['evaluacion'] and dato['evaluacion'].fecha_disposicion:
                fechas_disposicion.append(dato['evaluacion'].fecha_disposicion)
            elif dato.get('disposicion') and dato['disposicion'].fecha_disposicion:
                fechas_disposicion.append(dato['disposicion'].fecha_disposicion)

        # ==================== ESCRIBIR FILA ====================

        # Indentación visual con espacios según nivel de profundidad
        indentacion = "    " * nivel  # 4 espacios por nivel de profundidad

        # Columna A: N° (índice jerárquico con tabulación)
        numero_tabulado = f"{indentacion}{indice_str}"
        cell_numero = ws.cell(row=row, column=1, value=numero_tabulado)
        cell_numero.alignment = left_align  # Cambiar a left para que se vea la indentación
        cell_numero.border = border
        cell_numero.fill = row_fill
        if row_font:
            cell_numero.font = row_font

        # Columna B: COMPONENTE (nombre legible con formato según nivel)

        nombre_componente = nombre_legible_directorio(nombre_dir)
        if nivel == 0:
            # Categoría principal: emoji + mayúsculas (sin indentación)
            texto_componente = f"📁  {nombre_componente.upper()} "
        elif nivel == 1:
            # Subcategoría nivel 1: indentación + índice + nombre en mayúsculas
            texto_componente = f"{indentacion}{indice_str} {nombre_componente.upper()}"
        else:
            # Niveles más profundos: indentación + solo nombre
            texto_componente = f"{indentacion}{nombre_componente}"

        cell_componente = ws.cell(row=row, column=2, value=texto_componente)
        cell_componente.alignment = left_align
        cell_componente.border = border
        cell_componente.fill = row_fill
        if row_font:
            cell_componente.font = row_font

        # Columna C: DIRECTORIO/SUBCARPETA (jerarquía técnica)
        ws.cell(row=row, column=3, value=jerarquia_completa).alignment = left_align
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).fill = row_fill

        # Columna D: FECHA DISPOSICIÓN
        fecha_texto = "⏳ Pendiente"
        if fechas_disposicion:
            fecha_reciente = max(fechas_disposicion)
            fecha_texto = fecha_reciente.strftime('%d/%m/%Y')
        ws.cell(row=row, column=4, value=fecha_texto).alignment = center_align
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).fill = row_fill

        # Columna E: TOTAL ARCHIVOS
        ws.cell(row=row, column=5, value=f"{total_archivos} archivos").alignment = center_align
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=5).fill = row_fill

        # Columna F: RESUMEN EVALUACIÓN
        if total_archivos > 0:
            porcentaje_evaluado = round((archivos_evaluados / total_archivos) * 100, 1)
            porcentaje_aprobado = round((archivos_aprobados / total_archivos) * 100, 1)

            resumen_texto = f"📊 TOTAL: {total_archivos}\n"
            resumen_texto += f"✅ EVALUADOS: {archivos_evaluados} ({porcentaje_evaluado}%)\n"
            resumen_texto += f"🎯 APROBADOS: {archivos_aprobados} ({porcentaje_aprobado}%)"

            if archivos_aprobados == total_archivos:
                color_resumen = Font(bold=True, color="28A745", size=9)
            elif archivos_evaluados == total_archivos:
                color_resumen = Font(bold=True, color="FFA500", size=9)
            elif archivos_evaluados > 0:
                color_resumen = Font(bold=True, color="007BFF", size=9)
            else:
                color_resumen = Font(bold=True, color="1E3A8A", size=9)
        else:
            resumen_texto = "📁 Sin archivos para evaluar"
            color_resumen = Font(bold=True, color="6C757D", size=9)

        cell_resumen = ws.cell(row=row, column=6, value=resumen_texto)
        cell_resumen.alignment = left_align
        cell_resumen.border = border
        cell_resumen.fill = row_fill
        cell_resumen.font = color_resumen

        # Columna G: OBSERVACIONES EVALUACION (nueva columna)
        observaciones_texto = ""
        if observaciones_lista:
            # Mostrar hasta 3 observaciones únicas
            obs_unicas = list(set(observaciones_lista))[:3]
            observaciones_texto = "; ".join(obs_unicas)
        cell_obs = ws.cell(row=row, column=7, value=observaciones_texto)
        cell_obs.alignment = left_align
        cell_obs.border = border
        cell_obs.fill = row_fill

        # Columna H: RUTA DE ACCESO
        rutas_texto = sorted(rutas_acceso)[0] if rutas_acceso else "Sin rutas"
        ws.cell(row=row, column=8, value=rutas_texto).alignment = left_align
        ws.cell(row=row, column=8).border = border
        ws.cell(row=row, column=8).fill = row_fill

        row += 1

    # ============================================================
    # AJUSTAR DIMENSIONES DE COLUMNAS (8 columnas ahora)
    # ============================================================
    anchos_columnas = [8, 45, 45, 15, 15, 45, 40, 55]
    # N°, COMPONENTE, DIRECTORIO, FECHA, TOTAL, RESUMEN, OBSERVACIONES, RUTA
    for i, ancho in enumerate(anchos_columnas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    ws.row_dimensions[1].height = 50
    ws.row_dimensions[3].height = 25
    ws.row_dimensions[5].height = 35

    # Altura de filas de datos
    for row_num in range(6, row):
        ws.row_dimensions[row_num].height = 45

    print(f"✅ PESTAÑA 1 CON ÍNDICES JERÁRQUICOS: {len(lista_ordenada)} filas procesadas")

"""   
def generar_pestana_detalles_componentes(ws, municipio, datos_completos, mecanismo_financiacion):

    # AGRUPAR POR DIRECTORIO PADRE Y LUEGO POR JERARQUÍA COMPLETA
    directorios_padre_agrupados = defaultdict(lambda: defaultdict(list))
    
    for dato in datos_completos:
        directorio_padre = dato['directorio_padre']
        jerarquia_completa = dato['jerarquia_completa']
        
        # FILTRAR DATOS INVÁLIDOS TAMBIÉN AQUÍ
        if jerarquia_completa in ["SIN_03_POST", "ROOT_03_POST", "SIN_JERARQUIA", "ERROR_JERARQUIA"]:
            continue
        
        # Solo agregar archivos evaluados para esta pestaña
        if dato['tipo'] == 'ARCHIVO_EVALUADO':
            directorios_padre_agrupados[directorio_padre][jerarquia_completa].append(dato)
    
    # Filtrar directorios padre que tengan archivos
    directorios_con_archivos = {k: v for k, v in directorios_padre_agrupados.items() 
                               if k not in ["SIN_DIRECTORIO", "ERROR_DIRECTORIO", "SIN_03_POST", "ROOT_03_POST"] and any(v.values())}
    
    if not directorios_con_archivos:
        print("⚠️ No hay directorios con archivos para mostrar")
        return
    
    # ESTILOS
    titulo_font = Font(bold=True, size=14, color="FFFFFF")
    titulo_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_padre_font = Font(bold=True, color="FFFFFF", size=10)
    header_padre_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_hijo_font = Font(bold=True, color="FFFFFF", size=9)
    header_hijo_fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=8)
    subheader_fill = PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # CALCULAR TOTAL DE COLUMNAS NECESARIAS
    total_columnas = 0
    estructura_columnas = {}
    
    for directorio_padre, directorios_hijos in directorios_con_archivos.items():
        estructura_columnas[directorio_padre] = {
            'inicio_col': total_columnas + 1,
            'hijos': {},
            'total_hijos': len(directorios_hijos)
        }
        
        col_hijo = total_columnas + 1
        for jerarquia_completa, archivos in directorios_hijos.items():
            estructura_columnas[directorio_padre]['hijos'][jerarquia_completa] = {
                'col_archivo': col_hijo,
                'col_ruta': col_hijo + 1,
                'total_archivos': len(archivos)
            }
            col_hijo += 2
        
        estructura_columnas[directorio_padre]['fin_col'] = col_hijo - 1
        total_columnas = col_hijo - 1
    
    # TÍTULO CON BORDES CORRECTOS
    ultima_columna = openpyxl.utils.get_column_letter(max(total_columnas, 10))
    
    # PRIMERO: Hacer merge
    ws.merge_cells(f'A1:{ultima_columna}2')
    
    # SEGUNDO: Configurar la celda principal
    titulo_cell = ws['A1']
    titulo_cell.value = f"🌳 ESTRUCTURA JERÁRQUICA DE ARCHIVOS POR COMPONENTE\n🏛️ {municipio.nom_municipio.upper()} - {mecanismo_financiacion}"
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    
    # TERCERO: Aplicar bordes a toda la celda mergeada
    for row_merge in range(1, 3):  # Filas 1 y 2
        for col_merge in range(1, max(total_columnas, 10) + 1):
            cell = ws.cell(row=row_merge, column=col_merge)
            cell.border = border
    
    # HEADERS NIVEL 1 - DIRECTORIOS PADRE CON BORDES CORRECTOS
    row_header_padre = 4
    for directorio_padre, info in estructura_columnas.items():
        start_col = info['inicio_col']
        end_col = info['fin_col']
        
        if start_col <= end_col:
            start_letter = openpyxl.utils.get_column_letter(start_col)
            end_letter = openpyxl.utils.get_column_letter(end_col)
            
            # PRIMERO: Hacer merge
            ws.merge_cells(f'{start_letter}{row_header_padre}:{end_letter}{row_header_padre}')
            
            # SEGUNDO: Configurar la celda principal
            header_cell = ws.cell(row=row_header_padre, column=start_col)
            
            # Contar total de archivos en este directorio padre
            total_archivos_padre = sum(len(archivos) for archivos in directorios_con_archivos[directorio_padre].values())
            
            header_cell.value = f"📁 {directorio_padre} ({total_archivos_padre} archivos)"
            header_cell.font = header_padre_font
            header_cell.fill = header_padre_fill
            header_cell.alignment = center_align
            
            # TERCERO: Aplicar bordes a toda la celda mergeada
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row_header_padre, column=col)
                cell.border = border
    
    # HEADERS NIVEL 2 - DIRECTORIOS HIJOS CON BORDES CORRECTOS
    row_header_hijo = 5
    for directorio_padre, info in estructura_columnas.items():
        for jerarquia_completa, info_hijo in info['hijos'].items():
            start_col = info_hijo['col_archivo']
            end_col = info_hijo['col_ruta']
            
            start_letter = openpyxl.utils.get_column_letter(start_col)
            end_letter = openpyxl.utils.get_column_letter(end_col)
            
            # PRIMERO: Hacer merge
            ws.merge_cells(f'{start_letter}{row_header_hijo}:{end_letter}{row_header_hijo}')
            
            # SEGUNDO: Configurar la celda principal
            header_cell = ws.cell(row=row_header_hijo, column=start_col)
            # Mostrar solo la parte después del directorio padre para evitar redundancia
            jerarquia_mostrar = jerarquia_completa.replace(f"{directorio_padre}>", "") if ">" in jerarquia_completa else jerarquia_completa
            header_cell.value = f"📄 {jerarquia_mostrar} ({info_hijo['total_archivos']})"
            header_cell.font = header_hijo_font
            header_cell.fill = header_hijo_fill
            header_cell.alignment = center_align
            
            # TERCERO: Aplicar bordes a toda la celda mergeada
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row_header_hijo, column=col)
                cell.border = border
    
    # HEADERS NIVEL 3 - ARCHIVO/RUTA
    row_subheader = 6
    for directorio_padre, info in estructura_columnas.items():
        for jerarquia_completa, info_hijo in info['hijos'].items():
            # Columna archivo
            ws.cell(row=row_subheader, column=info_hijo['col_archivo'], value="📄 Archivo").font = subheader_font
            ws.cell(row=row_subheader, column=info_hijo['col_archivo']).fill = subheader_fill
            ws.cell(row=row_subheader, column=info_hijo['col_archivo']).alignment = center_align
            ws.cell(row=row_subheader, column=info_hijo['col_archivo']).border = border
            
            # Columna ruta
            ws.cell(row=row_subheader, column=info_hijo['col_ruta'], value="📁 Ruta").font = subheader_font
            ws.cell(row=row_subheader, column=info_hijo['col_ruta']).fill = subheader_fill
            ws.cell(row=row_subheader, column=info_hijo['col_ruta']).alignment = center_align
            ws.cell(row=row_subheader, column=info_hijo['col_ruta']).border = border
    
    # DATOS - ARCHIVOS EN FILAS
    max_archivos_por_directorio_hijo = max(
        max(len(archivos) for archivos in directorios_hijos.values()) 
        for directorios_hijos in directorios_con_archivos.values()
    )
    
    row_data_start = 7
    for fila_archivo in range(max_archivos_por_directorio_hijo):
        row_data = row_data_start + fila_archivo
        
        # Alternar colores
        if fila_archivo % 2 == 0:
            row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for directorio_padre, directorios_hijos in directorios_con_archivos.items():
            for jerarquia_completa, archivos in directorios_hijos.items():
                info_hijo = estructura_columnas[directorio_padre]['hijos'][jerarquia_completa]
                
                if fila_archivo < len(archivos):
                    archivo_data = archivos[fila_archivo]
                    evaluacion = archivo_data['evaluacion']
                    
                    # Nombre del archivo
                    ws.cell(row=row_data, column=info_hijo['col_archivo'], value=evaluacion.nombre_archivo).alignment = left_align
                    ws.cell(row=row_data, column=info_hijo['col_archivo']).border = border
                    ws.cell(row=row_data, column=info_hijo['col_archivo']).fill = row_fill
                    
                    # Ruta del archivo (convertir a Windows para usuarios)
                    ruta_windows = linux_to_windows_path(evaluacion.ruta_completa)
                    ws.cell(row=row_data, column=info_hijo['col_ruta'], value=ruta_windows).alignment = left_align
                    ws.cell(row=row_data, column=info_hijo['col_ruta']).border = border
                    ws.cell(row=row_data, column=info_hijo['col_ruta']).fill = row_fill
                else:
                    # Celdas vacías
                    ws.cell(row=row_data, column=info_hijo['col_archivo'], value="").border = border
                    ws.cell(row=row_data, column=info_hijo['col_archivo']).fill = row_fill
                    ws.cell(row=row_data, column=info_hijo['col_ruta'], value="").border = border
                    ws.cell(row=row_data, column=info_hijo['col_ruta']).fill = row_fill
    
    # AJUSTAR DIMENSIONES
    for col in range(1, total_columnas + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        if col % 2 == 1:  # Columnas impares (archivos)
            ws.column_dimensions[column_letter].width = 35
        else:  # Columnas pares (rutas)
            ws.column_dimensions[column_letter].width = 55
    
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[row_header_padre].height = 30
    ws.row_dimensions[row_header_hijo].height = 25
    ws.row_dimensions[row_subheader].height = 20
    
    for row_num in range(row_data_start, row_data_start + max_archivos_por_directorio_hijo):
        ws.row_dimensions[row_num].height = 25
    
    print(f"✅ PESTAÑA 2 CON BORDES PERFECTOS: {len(directorios_con_archivos)} directorios padre, estructura jerárquica completa")
"""

def generar_reporte_individual_postoperacion(municipio, temp_dir, mecanismo_financiacion, etapas_filtro=None):
    """
    Genera reporte Excel usando AMBAS tablas: disposicion_post + evaluacion_archivos_post
    *** MODIFICADO: Solo 2 pestañas (eliminada Detalles de Componentes) ***
    *** MEJORADO: Soporta filtro de etapas operativas ***

    Args:
        municipio: Objeto Municipio
        temp_dir: Directorio temporal para guardar el archivo
        mecanismo_financiacion: Mecanismo de financiación (PGN, SGR, etc.)
        etapas_filtro: Lista de prefijos de etapas a incluir (ej: ['01', '02', '03'])
                       Si es None o vacío, incluye todas las etapas
    """
    try:
        print(f"🎯 Generando reporte para {municipio.nom_municipio} - Mecanismo: {mecanismo_financiacion}")
        if etapas_filtro:
            print(f"📂 Filtro de etapas: {etapas_filtro}")

        # PASO 1: OBTENER TODOS LOS DIRECTORIOS (DISPOSICION_POST)
        disposiciones_base = DisposicionPost.objects.filter(
            cod_municipio=municipio.cod_municipio
        ).select_related('cod_municipio')

        # FILTRAR POR MECANISMO DE FINANCIACIÓN Y ETAPAS
        disposiciones_filtradas = []
        for disposicion in disposiciones_base:
            if disposicion.ruta_acceso:
                mecanismo = extraer_mecanismo_financiacion_desde_ruta(disposicion.ruta_acceso)
                if mecanismo == mecanismo_financiacion or (mecanismo_financiacion == "GENERAL" and mecanismo in ["SIN_MECANISMO", "ERROR_MECANISMO"]):
                    # Aplicar filtro de etapas si está definido
                    if etapas_filtro:
                        etapa_dir = extraer_etapa_desde_ruta(disposicion.ruta_acceso)
                        if etapa_dir and etapa_dir in etapas_filtro:
                            disposiciones_filtradas.append(disposicion)
                    else:
                        disposiciones_filtradas.append(disposicion)
        
        if not disposiciones_filtradas:
            print(f"⚠️ No se encontraron disposiciones para mecanismo {mecanismo_financiacion}")
            return None
        
        print(f"📁 Total disposiciones filtradas: {len(disposiciones_filtradas)}")
        
        # PASO 2: OBTENER EVALUACIONES RELACIONADAS (LEFT JOIN)
        evaluaciones_municipio = EvaluacionArchivosPost.objects.filter(
            id_disposicion__cod_municipio=municipio.cod_municipio
        ).select_related('id_disposicion')

        # Crear diccionario de evaluaciones por id_disposicion
        # También aplicar filtro de etapas a las evaluaciones
        evaluaciones_por_disposicion = defaultdict(list)
        evaluaciones_filtradas_count = 0
        for evaluacion in evaluaciones_municipio:
            # Si hay filtro de etapas, verificar que la evaluación pertenezca a una etapa seleccionada
            if etapas_filtro and evaluacion.ruta_completa:
                etapa_eval = extraer_etapa_desde_ruta(evaluacion.ruta_completa)
                if not etapa_eval or etapa_eval not in etapas_filtro:
                    continue  # Saltar esta evaluación
            evaluaciones_por_disposicion[evaluacion.id_disposicion.id_disposicion].append(evaluacion)
            evaluaciones_filtradas_count += 1

        print(f"📊 Total evaluaciones encontradas: {evaluaciones_municipio.count()}")
        if etapas_filtro:
            print(f"📊 Evaluaciones después del filtro de etapas: {evaluaciones_filtradas_count}")
        
        # PASO 3: CREAR ESTRUCTURA DE DATOS COMPLETA
        datos_completos = []
        
        for disposicion in disposiciones_filtradas:
            # Obtener evaluaciones de esta disposición
            evaluaciones_disp = evaluaciones_por_disposicion.get(disposicion.id_disposicion, [])
            
            if evaluaciones_disp:
                # Hay archivos evaluados en esta disposición
                for evaluacion in evaluaciones_disp:
                    datos_completos.append({
                        'tipo': 'ARCHIVO_EVALUADO',
                        'disposicion': disposicion,
                        'evaluacion': evaluacion,
                        'directorio_padre': extraer_directorio_padre_desde_ruta(evaluacion.ruta_completa),
                        'jerarquia_completa': extraer_jerarquia_completa_desde_ruta(evaluacion.ruta_completa),
                        'ruta_directorio': extraer_ruta_hasta_directorio_padre(evaluacion.ruta_completa)
                    })
            else:
                # Disposición sin archivos evaluados (solo directorio)
                datos_completos.append({
                    'tipo': 'DIRECTORIO_SOLO',
                    'disposicion': disposicion,
                    'evaluacion': None,
                    'directorio_padre': extraer_directorio_padre_desde_ruta(disposicion.ruta_acceso),
                    'jerarquia_completa': extraer_jerarquia_completa_desde_ruta(disposicion.ruta_acceso),
                    'ruta_directorio': extraer_ruta_hasta_directorio_padre(disposicion.ruta_acceso)
                })
        
        print(f"📊 Datos completos estructurados: {len(datos_completos)} registros")
        
        # PASO 4: CREAR WORKBOOK - *** SOLO 2 PESTAÑAS ***
        wb = openpyxl.Workbook()
        
        # PESTAÑA 1: Resumen General
        ws1 = wb.active
        ws1.title = "Resumen General"
        generar_pestana_resumen_general(ws1, municipio, datos_completos, mecanismo_financiacion)
        
        # PESTAÑA 2: Inventario de Archivos (MEJORADA)
        ws2 = wb.create_sheet(title="Inventario de Archivos")
        generar_pestana_inventario_archivos(ws2, municipio, datos_completos, mecanismo_financiacion)
        
        # PASO 5: GUARDAR ARCHIVO
        fecha_actual = datetime.now().strftime('%d_%m_%Y')
        municipio_limpio = municipio.nom_municipio.replace(' ', '_').replace('/', '_').upper()
        mecanismo_limpio = mecanismo_financiacion.replace('\\', '_').replace('/', '_').replace('-', '_')
        
        nombre_archivo = f"{municipio.cod_municipio}_{municipio_limpio}_{mecanismo_limpio}_Postoperacion_{fecha_actual}.xlsx"
        archivo_path = os.path.join(temp_dir, nombre_archivo)
        wb.save(archivo_path)
        
        print(f"✅ Reporte generado:")
        print(f"   📄 Archivo: {nombre_archivo}")
        print(f"   💰 Mecanismo: {mecanismo_financiacion}")
        print(f"   📊 Registros procesados: {len(datos_completos)}")
        print(f"   📁 Disposiciones: {len(disposiciones_filtradas)}")
        print(f"   📑 Pestañas: 2 (Resumen General + Inventario de Archivos)")
        
        return archivo_path
        
    except Exception as e:
        print(f"❌ Error generando reporte para {municipio.nom_municipio}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def es_extension_ejecutable(nombre_archivo):
    """
    Determina si un archivo tiene una extensión que Windows puede abrir nativamente
    """
    if not nombre_archivo or '.' not in nombre_archivo:
        return False
    
    extension = nombre_archivo.split('.')[-1].lower()
    
    # Extensiones que Windows puede abrir nativamente
    extensiones_ejecutables = {
        # Documentos de Office
        'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'pps', 'ppsx',
        # PDFs
        'pdf',
        # Imágenes
        'jpg', 'jpeg', 'png', 'gif', 'bmp', 'tiff', 'tif', 'webp', 'ico',
        # Archivos comprimidos
        'zip', '7z', 'rar', 'tar', 'gz',
        # Texto
        'txt', 'rtf', 'csv',
        # Videos
        'mp4', 'avi', 'mov', 'wmv', 'mkv', 'flv', 'webm',
        # Audio
        'mp3', 'wav', 'wma', 'aac', 'ogg',
        # Web
        'html', 'htm', 'xml',
        # Otros comunes
        'json', 'log'
    }
    
    # Extensiones que NO incluir (requieren software específico)
    extensiones_no_ejecutables = {
        # GIS y CAD
        'shp', 'gdb', 'mxd', 'lyr', 'dwg', 'dxf', 'kml', 'kmz',
        # Específicos de software
        'ai', 'psd', 'eps', 'indd', 'fla', 'swf',
        # Datos científicos
        'nc', 'hdf', 'mat', 'sav',
        # Código fuente (aunque pueden abrirse, no son "ejecutables" para el usuario final)
        'py', 'js', 'php', 'cpp', 'java', 'sql'
    }
    
    # Verificar si está en la lista de no ejecutables
    if extension in extensiones_no_ejecutables:
        return False
    
    # Verificar si está en la lista de ejecutables
    return extension in extensiones_ejecutables


def generar_banner_sin_archivos(ws, municipio, mecanismo_financiacion):
    """
    Genera un banner informativo cuando no hay archivos para mostrar
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Estilos
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    warning_font = Font(bold=True, size=14, color="856404")
    warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    info_font = Font(size=12, color="495057")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # TÍTULO
    ws.merge_cells('A1:H2')
    titulo_cell = ws['A1']
    titulo_cell.value = f"📋 INVENTARIO DE ARCHIVOS POST-OPERACIÓN\n🏛️ {municipio.nom_municipio.upper()} - {mecanismo_financiacion}"
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align

    # Aplicar bordes al título
    for row_merge in range(1, 3):
        for col_merge in range(1, 9):
            cell = ws.cell(row=row_merge, column=col_merge)
            cell.border = border

    # BANNER DE ADVERTENCIA
    ws.merge_cells('A4:H6')
    warning_cell = ws['A4']
    warning_cell.value = "⚠️ SIN ARCHIVOS REGISTRADOS\n\nNo se encontraron archivos de post-operación para este municipio con el mecanismo de financiación seleccionado."
    warning_cell.font = warning_font
    warning_cell.fill = warning_fill
    warning_cell.alignment = center_align

    # Aplicar bordes al banner
    for row_merge in range(4, 7):
        for col_merge in range(1, 9):
            cell = ws.cell(row=row_merge, column=col_merge)
            cell.border = border

    # INFORMACIÓN ADICIONAL
    ws.merge_cells('A8:H9')
    info_cell = ws['A8']
    info_cell.value = "📌 Posibles causas:\n• El municipio no tiene archivos indexados\n• No hay archivos para el mecanismo de financiación seleccionado\n• Los archivos aún no han sido procesados por el sistema"
    info_cell.font = info_font
    info_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Aplicar bordes
    for row_merge in range(8, 10):
        for col_merge in range(1, 9):
            cell = ws.cell(row=row_merge, column=col_merge)
            cell.border = border

    # ESTADÍSTICAS
    ws.merge_cells('A11:H11')
    stats_cell = ws['A11']
    stats_cell.value = "📊 RESUMEN: 0 archivos | 0 evaluados | 0 aprobados"
    stats_cell.font = Font(bold=True, size=11, color="DC3545")
    stats_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    stats_cell.alignment = center_align

    for col_merge in range(1, 9):
        cell = ws.cell(row=11, column=col_merge)
        cell.border = border

    # Ajustar anchos de columna
    for i in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    # Ajustar alturas de fila
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[4].height = 60
    ws.row_dimensions[8].height = 60

    print(f"✅ Banner de 'sin archivos' generado para {municipio.nom_municipio}")


def generar_pestana_inventario_archivos(ws, municipio, datos_completos, mecanismo_financiacion):
    """
    PESTAÑA 2: Inventario completo de archivos CON TODAS LAS MEJORAS SOLICITADAS
    *** MODIFICADO: Ordenado alfabéticamente, nueva columna RUTA ARCHIVO ***
    """
    # FILTRAR SOLO ARCHIVOS EVALUADOS
    archivos_evaluados = [d for d in datos_completos if d['tipo'] == 'ARCHIVO_EVALUADO' and d['evaluacion']]

    if not archivos_evaluados:
        print("⚠️ No hay archivos evaluados para mostrar en inventario")
        # GENERAR BANNER DE "SIN ARCHIVOS" EN VEZ DE HOJA EN BLANCO
        generar_banner_sin_archivos(ws, municipio, mecanismo_financiacion)
        return
    
    # *** ORDENAR ALFABÉTICAMENTE POR SUBCARPETA (jerarquia_completa) ***
    archivos_ordenados = sorted(archivos_evaluados, key=lambda x: (
        x['jerarquia_completa'].lower(),  # Primer criterio: subcarpeta alfabéticamente
        x['evaluacion'].nombre_archivo.lower() if x['evaluacion'] else ''  # Segundo criterio: nombre archivo
    ))
    
    print(f"📊 Archivos ordenados alfabéticamente por subcarpeta: {len(archivos_ordenados)}")
    
    # ESTILOS
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ESTILOS PARA VALORES BOOLEANOS (sin fondo, solo texto coloreado)
    si_font = Font(bold=True, color="28A745")  # Verde para checkmark
    no_font = Font(bold=True, color="DC3545")  # Rojo para X
    
    # TÍTULO CON BORDES CORRECTOS - 13 COLUMNAS (M) - *** AUMENTADO PARA COLUMNA DE CALIFICACIÓN ***
    # PRIMERO: Hacer merge
    ws.merge_cells('A1:M2')
    
    # SEGUNDO: Configurar la celda principal
    titulo_cell = ws['A1']
    titulo_cell.value = f"📋 INVENTARIO DETALLADO DE ARCHIVOS POST-OPERACIÓN\n🏛️ {municipio.nom_municipio.upper()} - {mecanismo_financiacion}"
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    
    # TERCERO: Aplicar bordes a toda la celda mergeada
    for row_merge in range(1, 3):  # Filas 1 y 2
        for col_merge in range(1, 14):  # Columnas A-M
            cell = ws.cell(row=row_merge, column=col_merge)
            cell.border = border
    
    # ESTADÍSTICAS CON BORDES CORRECTOS - 12 COLUMNAS (L)
    total_archivos = len(archivos_ordenados)
    evaluados = sum(1 for a in archivos_ordenados if a['evaluacion'].evaluado)
    aprobados = sum(1 for a in archivos_ordenados if a['evaluacion'].aprobado)
    archivos_ejecutables = sum(1 for a in archivos_ordenados if es_extension_ejecutable(a['evaluacion'].nombre_archivo))
    
    # PRIMERO: Hacer merge
    ws.merge_cells('A3:M3')

    # SEGUNDO: Configurar la celda principal
    stats_cell = ws['A3']
    stats_cell.value = f"📊 RESUMEN: {total_archivos} archivos | {evaluados} evaluados | {aprobados} aprobados | {archivos_ejecutables} ejecutables | Ordenado por subcarpeta"
    stats_cell.font = Font(bold=True, size=11, color="2E7D32")
    stats_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    stats_cell.alignment = center_align

    # TERCERO: Aplicar bordes a toda la celda mergeada
    for col_merge in range(1, 14):  # Columnas A-M
        cell = ws.cell(row=3, column=col_merge)
        cell.border = border

    # HEADERS - *** CON COLUMNAS DE CALIFICACIÓN (VALOR Y TEXTO) ***
    headers = [
        'ETAPA',                                    # A
        'SUBCARPETA\n(Ordenado A-Z)',              # B
        'NOMBRE DOCUMENTO',                         # C
        'FORMATO\nDOCUMENTO',                      # D
        'FECHA DE CREACIÓN\nO MODIFICACIÓN',        # E
        'TAMAÑO ARCHIVO',                          # F
        'EVALUADO',                                # G
        'APROBADO',                                # H
        'CALIFICACIÓN\n(VALOR)',                   # I *** NUEVA: Valor numérico ***
        'CALIFICACIÓN\n(TEXTO)',                   # J *** NUEVA: Texto concepto ***
        'OBSERVACIONES',                           # K
        'RUTA DIRECTORIO PADRE',                   # L
        'RUTA ARCHIVO'                             # M
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # DATOS - USAR ARCHIVOS YA ORDENADOS
    row = 6
    for numero, dato in enumerate(archivos_ordenados, 1):
        evaluacion = dato['evaluacion']
        
        # Alternar colores
        if numero % 2 == 0:
            row_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Columna A: ETAPA
        ws.cell(row=row, column=1, value="POST-OPERACIÓN").alignment = center_align
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=1).fill = row_fill
        
        # Columna B: SUBCARPETA (jerarquía completa) - *** YA ORDENADA ALFABÉTICAMENTE ***
        subcarpeta = dato['jerarquia_completa']
        ws.cell(row=row, column=2, value=subcarpeta).alignment = left_align
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=2).fill = row_fill
        
        # Columna C: NOMBRE DOCUMENTO
        ws.cell(row=row, column=3, value=evaluacion.nombre_archivo).alignment = left_align
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).fill = row_fill
        
        # Columna D: FORMATO
        formato = evaluacion.nombre_archivo.split('.')[-1].upper() if '.' in evaluacion.nombre_archivo else "SIN EXT"
        ws.cell(row=row, column=4, value=formato).alignment = center_align
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).fill = row_fill
        
        # Columna E: FECHA
        fecha_documento = "Sin fecha"
        if evaluacion.fecha_disposicion:
            fecha_documento = evaluacion.fecha_disposicion.strftime('%d/%m/%Y')
        ws.cell(row=row, column=5, value=fecha_documento).alignment = center_align
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=5).fill = row_fill
        
        # Columna F: TAMAÑO ARCHIVO (*** HEADER MODIFICADO ***)
        tamaño = evaluacion.peso_memoria or "Sin datos"
        if tamaño != "Sin datos" and str(tamaño).replace('.', '').replace(',', '').isdigit():
            try:
                tamaño_int = int(float(str(tamaño).replace(',', '')))
                if tamaño_int < 1024:
                    tamaño = f"{tamaño_int:,}".replace(',', '.') + " bytes"
                elif tamaño_int < 1024*1024:
                    tamaño = f"{tamaño_int/1024:.1f} KB"
                elif tamaño_int < 1024*1024*1024:
                    tamaño = f"{tamaño_int/(1024*1024):.1f} MB"
                else:
                    tamaño = f"{tamaño_int/(1024*1024*1024):.1f} GB"
            except:
                pass
        ws.cell(row=row, column=6, value=tamaño).alignment = center_align
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).fill = row_fill
        
        # Columna G: EVALUADO (con colores sin fondo)
        cell_evaluado = ws.cell(row=row, column=7, value=convertir_booleano_a_texto(evaluacion.evaluado))
        cell_evaluado.alignment = center_align
        cell_evaluado.border = border
        cell_evaluado.fill = row_fill
        if evaluacion.evaluado:
            cell_evaluado.font = si_font
        else:
            cell_evaluado.font = no_font
        
        # Columna H: APROBADO (con colores sin fondo)
        cell_aprobado = ws.cell(row=row, column=8, value=convertir_booleano_a_texto(evaluacion.aprobado))
        cell_aprobado.alignment = center_align
        cell_aprobado.border = border
        cell_aprobado.fill = row_fill
        if evaluacion.aprobado:
            cell_aprobado.font = si_font
        else:
            cell_aprobado.font = no_font
        
        # Columna I: CALIFICACIÓN (VALOR NUMÉRICO)
        valor_calificacion = obtener_valor_calificacion(evaluacion.evaluacion_archivo)
        cell_valor = ws.cell(row=row, column=9, value=valor_calificacion)
        cell_valor.alignment = center_align
        cell_valor.border = border
        cell_valor.fill = row_fill
        # Color según valor
        if valor_calificacion is not None and valor_calificacion >= 1.0:
            cell_valor.font = Font(bold=True, color="28A745")  # Verde
        elif valor_calificacion is not None and valor_calificacion > 0:
            cell_valor.font = Font(bold=True, color="FFC107")  # Amarillo/Naranja
        else:
            cell_valor.font = Font(bold=True, color="DC3545")  # Rojo

        # Columna J: CALIFICACIÓN (TEXTO)
        concepto_evaluacion = obtener_concepto_calificacion(evaluacion.evaluacion_archivo)
        ws.cell(row=row, column=10, value=concepto_evaluacion).alignment = left_align
        ws.cell(row=row, column=10).border = border
        ws.cell(row=row, column=10).fill = row_fill

        # Columna K: OBSERVACIONES (de observaciones_evaluacion)
        observaciones = evaluacion.observaciones_evaluacion or "Sin observaciones"
        ws.cell(row=row, column=11, value=observaciones).alignment = left_align
        ws.cell(row=row, column=11).border = border
        ws.cell(row=row, column=11).fill = row_fill

        # Columna L: RUTA DIRECTORIO PADRE (SIN archivo final)
        ruta_directorio = dato['ruta_directorio']
        ws.cell(row=row, column=12, value=ruta_directorio).alignment = left_align
        ws.cell(row=row, column=12).border = border
        ws.cell(row=row, column=12).fill = row_fill

        # Columna M: RUTA ARCHIVO
        ruta_archivo_completa = linux_to_windows_path(evaluacion.ruta_completa)

        # Solo mostrar la ruta si el archivo es ejecutable por Windows
        if es_extension_ejecutable(evaluacion.nombre_archivo):
            cell_ruta = ws.cell(row=row, column=13, value=ruta_archivo_completa)
            cell_ruta.font = Font(bold=False, color="1E88E5", underline="single")
            cell_ruta.alignment = left_align
            cell_ruta.border = border
            cell_ruta.fill = row_fill
        else:
            cell_ruta = ws.cell(row=row, column=13, value="🚫 Requiere software específico")
            cell_ruta.font = Font(bold=False, color="757575", italic=True)
            cell_ruta.alignment = center_align
            cell_ruta.border = border
            cell_ruta.fill = row_fill

        row += 1

    # AJUSTAR DIMENSIONES - *** ACTUALIZADO PARA 13 COLUMNAS ***
    anchos_columnas = [15, 45, 40, 12, 18, 15, 12, 12, 12, 25, 35, 55, 60]  # 13 columnas A-M
    for i, ancho in enumerate(anchos_columnas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[3].height = 25
    ws.row_dimensions[5].height = 35
    
    for row_num in range(6, row):
        ws.row_dimensions[row_num].height = 25
    
    print(f"✅ PESTAÑA INVENTARIO MEJORADA:")
    print(f"   📊 {total_archivos} archivos ordenados alfabéticamente por subcarpeta")
    print(f"   🔗 {archivos_ejecutables} archivos ejecutables con rutas activas")
    print(f"   📑 13 columnas con CALIFICACIÓN (valor y texto)")
    print(f"   📝 Incluye valor numérico (0-1) y texto de calificación")



# AGREGAR al final de views.py - NO modificar nada existente

def obtener_concepto_calificacion_simple(evaluacion_id):
    """Función auxiliar independiente para obtener concepto de calificación"""
    if not evaluacion_id:
        return "SIN CALIFICACION (PENDIENTE)"
    
    try:
        calificacion = CalificacionInfoPost.objects.get(id=evaluacion_id)
        return f"{calificacion.concepto} ({calificacion.valor})"
    except CalificacionInfoPost.DoesNotExist:
        return "CALIFICACION NO ENCONTRADA"


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_datos_evaluacion_municipio(request, municipio_id):
    """
    🆕 NUEVO ENDPOINT - Obtiene datos de evaluación usando la MISMA lógica del Excel
    Ruta: /postoperacion/evaluacion-datos/{municipio_id}/
    """
    try:
        # Verificar permisos usando la misma lógica existente
        municipios_permitidos = get_municipios_permitidos(request.user)
        
        if municipios_permitidos != 'todos':
            if not municipios_permitidos or int(municipio_id) not in municipios_permitidos:
                return Response({
                    'error': 'No tiene permisos para acceder a este municipio'
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Obtener municipio usando la misma lógica del Excel
        try:
            municipio = Municipios.objects.get(cod_municipio=municipio_id)
        except Municipios.DoesNotExist:
            return Response({
                'error': 'Municipio no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # 🔄 USAR LA MISMA LÓGICA que generar_reporte_individual_postoperacion
        print(f"🎯 Obteniendo datos para {municipio.nom_municipio}")
        
        # 🔄 ENFOQUE CORREGIDO - Usar disposicion_post como tabla principal
        print(f"🎯 Obteniendo TODAS las disposiciones para municipio {municipio_id}")
        
        # PASO 1: OBTENER TODAS LAS DISPOSICIONES DEL MUNICIPIO
        disposiciones_municipio = DisposicionPost.objects.filter(
            cod_municipio=municipio_id
        ).select_related('cod_municipio')
        
        print(f"📁 Disposiciones encontradas: {disposiciones_municipio.count()}")
        
        if not disposiciones_municipio.exists():
            return Response({
                'version': 'POST COMPLETO',
                'municipio': {
                    'cod_municipio': municipio.cod_municipio,
                    'nom_municipio': municipio.nom_municipio,
                    'departamento': municipio.cod_depto.nom_depto if municipio.cod_depto else None
                },
                'mensaje': 'No se encontraron disposiciones para este municipio',
                'total_disposiciones': 0,
                'directorios': [],
                'archivos': []
            })
        
        # PASO 2: OBTENER TODAS LAS EVALUACIONES DEL MUNICIPIO
        # Filtrar archivos que el indexador marcó como inactivos (ya no existen en NAS)
        rutas_inactivas = set()
        try:
            from django.db import connection as db_conn
            with db_conn.cursor() as cur:
                cur.execute("""
                    SELECT ap.ruta_completa
                    FROM archivos_post ap
                    JOIN disposicion_post dp ON ap.id_disposicion = dp.id_disposicion
                    WHERE dp.cod_municipio = %s AND ap.activo = FALSE
                """, [municipio_id])
                rutas_inactivas = {row[0] for row in cur.fetchall()}
        except Exception as e:
            print(f"⚠️ No se pudo verificar archivos inactivos: {e}")

        evaluaciones_queryset = EvaluacionArchivosPost.objects.filter(
            id_disposicion__cod_municipio=municipio_id
        ).select_related('id_disposicion')

        # Excluir evaluaciones de archivos que ya no existen en NAS
        if rutas_inactivas:
            evaluaciones_queryset = evaluaciones_queryset.exclude(
                ruta_completa__in=rutas_inactivas
            )
            print(f"🗑️ Excluidas {len(rutas_inactivas)} evaluaciones de archivos inactivos")

        evaluaciones_municipio = evaluaciones_queryset

        print(f"📄 Evaluaciones encontradas: {evaluaciones_municipio.count()}")
        
        # PASO 3: CREAR DICCIONARIO DE EVALUACIONES POR DISPOSICIÓN
        evaluaciones_por_disposicion = defaultdict(list)
        for evaluacion in evaluaciones_municipio:
            evaluaciones_por_disposicion[evaluacion.id_disposicion.id_disposicion].append(evaluacion)
            
        # PASO 4: PROCESAR CADA DISPOSICIÓN (DIRECTORIO PADRE)
        directorios_data = []
        todos_archivos = []
        
        for disposicion in disposiciones_municipio:
            evaluaciones_disp = evaluaciones_por_disposicion.get(disposicion.id_disposicion, [])
            
            # Extraer nombre del directorio padre (después de 03_post)
            directorio_padre = "DIRECTORIO_DESCONOCIDO"
            jerarquia_completa = "Sin jerarquía"
            
            if disposicion.ruta_acceso:
                # Extraer directorio después de 03_post
                partes_ruta = disposicion.ruta_acceso.replace('\\', '/').split('/')
                try:
                    indice_post = next(i for i, parte in enumerate(partes_ruta) if '03_post' in parte)
                    if indice_post + 1 < len(partes_ruta):
                        directorio_padre = partes_ruta[indice_post + 1]
                        jerarquia_completa = '/'.join(partes_ruta[indice_post + 1:])
                except (StopIteration, IndexError):
                    # Si no encuentra 03_post, usar el último directorio
                    if len(partes_ruta) > 1:
                        directorio_padre = partes_ruta[-1]
                        jerarquia_completa = partes_ruta[-1]
            
            # Estadísticas del directorio
            stats_directorio = {
                'total_archivos': len(evaluaciones_disp),
                'pendientes': len([e for e in evaluaciones_disp if e.estado_archivo == 'PENDIENTE']),
                'evaluados': len([e for e in evaluaciones_disp if e.evaluado]),
                'aprobados': len([e for e in evaluaciones_disp if e.aprobado])
            }
            
            # Información del directorio
            directorio_info = {
                'id_disposicion': disposicion.id_disposicion,
                'nombre_directorio': directorio_padre,
                'ruta_acceso': disposicion.ruta_acceso,
                'jerarquia_completa': jerarquia_completa,
                'cod_municipio': disposicion.cod_municipio_id,  # ✅ USAR _id para obtener el valor
                'dispuesto': disposicion.dispuesto,
                'evaluado': disposicion.evaluado,
                'aprobado': disposicion.aprobado,
                'fecha_disposicion': disposicion.fecha_disposicion.isoformat() if disposicion.fecha_disposicion else None,
                'observaciones': disposicion.observaciones,
                'estadisticas': stats_directorio
            }
            
            directorios_data.append(directorio_info)
            
            # Procesar archivos de esta disposición
            archivos_disposicion = []
            for evaluacion in evaluaciones_disp:
                archivo_data = {
                    'id_evaluacion': evaluacion.id_evaluacion,
                    'id_archivo': evaluacion.id_archivo,
                    'id_disposicion': evaluacion.id_disposicion.id_disposicion,
                    'nombre_archivo': evaluacion.nombre_archivo,
                    'ruta_completa': evaluacion.ruta_completa,
                    'fecha_disposicion': evaluacion.fecha_disposicion.isoformat() if evaluacion.fecha_disposicion else None,
                    'observacion_original': evaluacion.observacion_original,
                    'hash_contenido': evaluacion.hash_contenido,
                    'usuario_windows': evaluacion.usuario_windows,
                    'peso_memoria': evaluacion.peso_memoria,
                    'evaluacion_archivo': evaluacion.evaluacion_archivo,
                    'estado_archivo': evaluacion.estado_archivo,
                    'observaciones_evaluacion': evaluacion.observaciones_evaluacion,
                    'fecha_creacion': evaluacion.fecha_creacion.isoformat() if evaluacion.fecha_creacion else None,
                    'fecha_actualizacion': evaluacion.fecha_actualizacion.isoformat() if evaluacion.fecha_actualizacion else None,
                    'usuario_evaluacion': evaluacion.usuario_evaluacion,
                    'evaluado': evaluacion.evaluado,
                    'aprobado': evaluacion.aprobado,
                    # Información adicional del directorio padre
                    'directorio_padre': directorio_padre,
                    'jerarquia_completa': jerarquia_completa,
                    'concepto_calificacion': obtener_concepto_calificacion_simple(evaluacion.evaluacion_archivo),
                    # Usuario que subió el archivo desde la plataforma web (para control de permisos)
                    'subido_por_plataforma': obtener_subido_por_plataforma(evaluacion.ruta_completa)
                }
                archivos_disposicion.append(archivo_data)
                todos_archivos.append(archivo_data)
        
        # RESPUESTA FINAL SIMPLIFICADA Y COMPLETA
        response_data = {
            'version': 'POST COMPLETO V2',
            'municipio': {
                'cod_municipio': municipio.cod_municipio,
                'nom_municipio': municipio.nom_municipio,
                'departamento': municipio.cod_depto.nom_depto if municipio.cod_depto else None
            },
            'fecha_consulta': datetime.now().isoformat(),
            'total_disposiciones': len(directorios_data),
            'total_archivos': len(todos_archivos),
            'directorios': directorios_data,
            'archivos': todos_archivos,
            # Estadísticas generales
            'estadisticas': {
                'total_disposiciones': len(directorios_data),
                'total_archivos': len(todos_archivos),
                'total_pendientes': len([a for a in todos_archivos if a['estado_archivo'] == 'PENDIENTE']),
                'total_evaluados': len([a for a in todos_archivos if a['evaluado']]),
                'total_aprobados': len([a for a in todos_archivos if a['aprobado']])
            }
        }
        
        print(f"✅ Datos generados COMPLETOS:")
        print(f"   📁 Total disposiciones: {len(directorios_data)}")
        print(f"   📄 Total archivos: {len(todos_archivos)}")
        print(f"   ⏳ Pendientes: {response_data['estadisticas']['total_pendientes']}")
        print(f"   ✅ Aprobados: {response_data['estadisticas']['total_aprobados']}")
        
        return Response(response_data)
        
    except Exception as e:
        print(f"❌ Error obteniendo datos de evaluación: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({
            'error': f'Error interno del servidor: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def actualizar_evaluacion_archivo(request, evaluacion_id):
    """
    🆕 NUEVO ENDPOINT - Actualizar evaluación de archivo individual
    Ruta: /postoperacion/evaluacion-actualizar/{evaluacion_id}/
    """
    try:
        # Obtener la evaluación
        try:
            evaluacion = EvaluacionArchivosPost.objects.get(id_evaluacion=evaluacion_id)
        except EvaluacionArchivosPost.DoesNotExist:
            return Response({
                'error': 'Evaluación no encontrada'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Verificar permisos
        municipios_permitidos = get_municipios_permitidos(request.user)
        if municipios_permitidos != 'todos':
            if not municipios_permitidos or evaluacion.id_disposicion.cod_municipio.cod_municipio not in municipios_permitidos:
                return Response({
                    'error': 'No tiene permisos para modificar este archivo'
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Obtener datos del request
        evaluacion_archivo_id = request.data.get('evaluacion_archivo')
        observaciones = request.data.get('observaciones_evaluacion', '')
        ruta_completa = request.data.get('ruta_completa')
        
        # LÓGICA AUTOMÁTICA (como especificaste)
        if evaluacion_archivo_id is not None:
            evaluacion.evaluacion_archivo = evaluacion_archivo_id
            
            # Aplicar lógica automática
            if evaluacion_archivo_id and evaluacion_archivo_id != 0:
                evaluacion.evaluado = True
                evaluacion.estado_archivo = 'EVALUADO'
                
                # Solo aprobado si es ID 7 (DOCUMENTO COMPLETO)
                if evaluacion_archivo_id == 7:
                    evaluacion.aprobado = True
                else:
                    evaluacion.aprobado = False
            else:
                # Si vuelve a 0 (SIN CALIFICACION)
                evaluacion.evaluado = False
                evaluacion.aprobado = False
                evaluacion.estado_archivo = 'PENDIENTE'
        
        # Actualizar otros campos si se proporcionan
        if observaciones is not None:
            evaluacion.observaciones_evaluacion = observaciones
        
        if ruta_completa is not None:
            evaluacion.ruta_completa = ruta_completa
        
        # Actualizar usuario y fecha
        evaluacion.usuario_evaluacion = request.user.username
        evaluacion.fecha_actualizacion = timezone.now()
        
        # Guardar
        evaluacion.save()
        
        # Respuesta con los datos actualizados
        return Response({
            'message': 'Evaluación actualizada correctamente',
            'evaluacion': {
                'id_evaluacion': evaluacion.id_evaluacion,
                'evaluacion_archivo': evaluacion.evaluacion_archivo,
                'estado_archivo': evaluacion.estado_archivo,
                'evaluado': evaluacion.evaluado,
                'aprobado': evaluacion.aprobado,
                'observaciones_evaluacion': evaluacion.observaciones_evaluacion,
                'ruta_completa': evaluacion.ruta_completa,
                'usuario_evaluacion': evaluacion.usuario_evaluacion,
                'fecha_actualizacion': evaluacion.fecha_actualizacion.isoformat()
            }
        })
        
    except Exception as e:
        print(f"❌ Error actualizando evaluación: {str(e)}")
        return Response({
            'error': f'Error interno del servidor: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def eliminar_evaluacion_archivo(request, evaluacion_id):
    """
    🆕 ENDPOINT MEJORADO - Eliminar evaluación de archivo (registro + físico opcional)
    Ruta: /postoperacion/evaluacion-eliminar/{evaluacion_id}/

    Query params:
        - eliminar_fisico: bool (default: false) - Si true, también elimina el archivo de la NAS
    """
    import os
    import re

    def normalizar_ruta_eliminacion(ruta):
        """
        Normaliza la ruta para soportar tanto Windows como Linux.
        Misma lógica que archivo_views.normalizar_ruta para consistencia.
        """
        if not ruta:
            return None

        # Primero convertir todas las barras invertidas a normales
        ruta_normalizada = ruta.replace('\\', '/')

        # Si la ruta ya está en formato Linux, devolverla tal cual
        if ruta_normalizada.startswith('/mnt/repositorio'):
            return ruta_normalizada

        # Remover prefijos comunes de Windows (varios formatos posibles)
        # Formato: //repositorio/DirGesCat/... (después de convertir barras)
        if ruta_normalizada.startswith('//repositorio/DirGesCat'):
            ruta_normalizada = ruta_normalizada.replace('//repositorio/DirGesCat', '/mnt/repositorio', 1)
        # Formato: repositorio/DirGesCat/... (sin barras iniciales)
        elif ruta_normalizada.startswith('repositorio/DirGesCat'):
            ruta_normalizada = ruta_normalizada.replace('repositorio/DirGesCat', '/mnt/repositorio', 1)

        return ruta_normalizada

    try:
        # Obtener la evaluación primero para verificar permisos
        try:
            evaluacion = EvaluacionArchivosPost.objects.get(id_evaluacion=evaluacion_id)
        except EvaluacionArchivosPost.DoesNotExist:
            return Response({
                'error': 'Evaluación no encontrada'
            }, status=status.HTTP_404_NOT_FOUND)

        # Verificar permisos
        es_admin = request.user.is_superuser or request.user.is_staff or request.user.groups.filter(name='Administradores').exists()

        if not es_admin:
            # Si no es admin, verificar si el usuario subió este archivo
            from postoperacion.models import AuditoriaArchivos

            # Buscar en auditoría si este usuario subió el archivo
            ruta_archivo = evaluacion.ruta_completa
            # Normalizar la ruta para buscar en auditoría
            ruta_buscar = ruta_archivo.replace('\\', '/') if ruta_archivo else ''
            if ruta_buscar.startswith('//repositorio/DirGesCat'):
                ruta_buscar = ruta_buscar.replace('//repositorio/DirGesCat', '/mnt/repositorio', 1)

            registro_upload = AuditoriaArchivos.objects.filter(
                ruta_completa=ruta_buscar,
                accion='UPLOAD',
                usuario=request.user.username
            ).first()

            if not registro_upload:
                print(f"⚠️ Usuario {request.user.username} intentó eliminar archivo que NO subió: {evaluacion.nombre_archivo}")
                return Response({
                    'error': 'Solo puedes eliminar archivos que tú mismo hayas subido'
                }, status=status.HTTP_403_FORBIDDEN)

            print(f"✅ Usuario {request.user.username} autorizado para eliminar archivo que subió el {registro_upload.fecha_accion}")

        # Obtener parámetro para eliminar físicamente
        eliminar_fisico = request.query_params.get('eliminar_fisico', 'false').lower() == 'true'

        # Guardar datos para respuesta y operaciones
        archivo_nombre = evaluacion.nombre_archivo
        ruta_windows = evaluacion.ruta_completa
        id_archivo = evaluacion.id_archivo
        id_disposicion = evaluacion.id_disposicion.id_disposicion if evaluacion.id_disposicion else None

        archivo_eliminado_fisicamente = False
        archivo_post_eliminado = False

        # Si se solicita eliminación física
        if eliminar_fisico and ruta_windows:
            # Logging detallado para debug
            print(f"🔍 DEBUG - Ruta original de BD: {repr(ruta_windows)}")

            # Usar la función de normalización robusta
            ruta_linux = normalizar_ruta_eliminacion(ruta_windows)

            print(f"🔍 DEBUG - Ruta normalizada: {repr(ruta_linux)}")

            print(f"🗑️ Intentando eliminar archivo físico: {ruta_linux}")
            print(f"🔍 DEBUG - ¿Existe el archivo? {os.path.exists(ruta_linux) if ruta_linux else 'ruta es None'}")

            if not ruta_linux:
                print(f"⚠️ No se pudo normalizar la ruta: {ruta_windows}")
                # Continuamos sin eliminar físicamente

            elif os.path.exists(ruta_linux):
                try:
                    os.remove(ruta_linux)
                    archivo_eliminado_fisicamente = True
                    print(f"✅ Archivo físico eliminado: {ruta_linux}")

                    # Registrar en auditoría
                    try:
                        from postoperacion.models import AuditoriaArchivos
                        AuditoriaArchivos.registrar_accion(
                            archivo=None,
                            nombre_archivo=archivo_nombre,
                            ruta_completa=ruta_linux,
                            accion='DELETE',
                            usuario=request.user.username,
                            usuario_email=request.user.email,
                            plataforma='WEB',
                            detalles={
                                'evaluacion_id': evaluacion_id,
                                'motivo': 'Eliminación solicitada por usuario'
                            }
                        )
                    except Exception as audit_err:
                        print(f"⚠️ Error registrando auditoría: {audit_err}")

                except PermissionError as perm_err:
                    print(f"❌ Error de permisos al eliminar: {perm_err}")
                    return Response({
                        'error': f'No se puede eliminar el archivo físico: permisos insuficientes'
                    }, status=status.HTTP_403_FORBIDDEN)
                except Exception as del_err:
                    print(f"❌ Error eliminando archivo físico: {del_err}")
                    return Response({
                        'error': f'Error al eliminar archivo físico: {str(del_err)}'
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                print(f"⚠️ Archivo físico no encontrado: {ruta_linux}")
                # Continuamos con la eliminación del registro aunque el archivo no exista

        # Eliminar registro de ArchivosPost si existe
        if id_archivo:
            try:
                archivo_post = ArchivosPost.objects.get(id_archivo=id_archivo)
                archivo_post.delete()
                archivo_post_eliminado = True
                print(f"✅ ArchivosPost eliminado: ID {id_archivo}")
            except ArchivosPost.DoesNotExist:
                print(f"⚠️ ArchivosPost no encontrado: ID {id_archivo}")
            except Exception as ap_err:
                print(f"⚠️ Error eliminando ArchivosPost: {ap_err}")

        # Eliminar EvaluacionArchivosPost
        evaluacion.delete()
        print(f"✅ EvaluacionArchivosPost eliminado: ID {evaluacion_id}")

        # Construir mensaje de respuesta
        mensaje = f'Archivo "{archivo_nombre}" eliminado'
        detalles = {
            'evaluacion_eliminada': True,
            'archivo_post_eliminado': archivo_post_eliminado
        }

        if eliminar_fisico:
            if archivo_eliminado_fisicamente:
                mensaje += ' (registro + archivo físico)'
                detalles['archivo_fisico_eliminado'] = True
            else:
                mensaje += ' (solo registro - archivo físico no encontrado)'
                detalles['archivo_fisico_eliminado'] = False
        else:
            mensaje += ' (solo registro)'
            detalles['archivo_fisico_eliminado'] = 'no solicitado'

        return Response({
            'message': mensaje,
            'detalles': detalles
        })

    except Exception as e:
        print(f"❌ Error eliminando evaluación: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({
            'error': f'Error interno del servidor: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===============================================
# MATRIZ RESUMEN POSTOPERACION
# ===============================================

def recopilar_datos_municipio_por_etapa(municipio, mecanismo, etapas_filtro=None):
    """
    Recopila datos de un municipio-mecanismo agrupados por etapa.
    Retorna diccionario con info por etapa y totales globales.
    """
    resultado = {
        'municipio': municipio,
        'mecanismo': mecanismo,
        'etapas': {},
        'totales': {
            'total_archivos': 0,
            'archivos_evaluados': 0,
            'archivos_aprobados': 0,
            'porcentaje_evaluado': 0.0,
            'porcentaje_aprobado': 0.0,
        }
    }

    # Obtener disposiciones del municipio
    disposiciones = DisposicionPost.objects.filter(
        cod_municipio=municipio.cod_municipio
    ).select_related('cod_municipio')

    # Filtrar por mecanismo y agrupar por etapa
    disposiciones_por_etapa = defaultdict(list)
    for disp in disposiciones:
        if not disp.ruta_acceso:
            continue
        mec = extraer_mecanismo_financiacion_desde_ruta(disp.ruta_acceso)
        if mec != mecanismo and not (mecanismo == "GENERAL" and mec in ["SIN_MECANISMO", "ERROR_MECANISMO"]):
            continue
        etapa = extraer_etapa_desde_ruta(disp.ruta_acceso)
        if not etapa:
            continue
        if etapas_filtro and etapa not in etapas_filtro:
            continue
        disposiciones_por_etapa[etapa].append(disp)

    # Obtener evaluaciones del municipio
    evaluaciones = EvaluacionArchivosPost.objects.filter(
        id_disposicion__cod_municipio=municipio.cod_municipio
    ).select_related('id_disposicion')

    eval_por_disposicion = defaultdict(list)
    for ev in evaluaciones:
        eval_por_disposicion[ev.id_disposicion.id_disposicion].append(ev)

    total_archivos_global = 0
    total_evaluados_global = 0
    total_aprobados_global = 0

    for prefijo_etapa, disps in disposiciones_por_etapa.items():
        nombre_etapa = ETAPAS_POSTOPERACION.get(prefijo_etapa, f'Etapa {prefijo_etapa}')

        total_archivos = 0
        archivos_evaluados = 0
        archivos_aprobados = 0
        calificaciones = []
        estados = defaultdict(int)
        observaciones = set()
        ruta_raiz = None

        for disp in disps:
            evals = eval_por_disposicion.get(disp.id_disposicion, [])
            total_archivos += len(evals)

            # Obtener ruta raiz de la etapa
            if ruta_raiz is None and disp.ruta_acceso:
                ruta_str = str(disp.ruta_acceso).replace('/', '\\')
                if '\\03_post\\' in ruta_str:
                    idx = ruta_str.find('\\03_post\\') + len('\\03_post\\')
                    resto = ruta_str[idx:]
                    if '\\' in resto:
                        etapa_dir = resto.split('\\')[0]
                        ruta_hasta_etapa = ruta_str[:idx] + etapa_dir
                    else:
                        ruta_hasta_etapa = ruta_str[:idx] + resto
                    ruta_raiz = linux_to_windows_path(ruta_hasta_etapa.replace('\\', '/'))
                    if ruta_raiz and not ruta_raiz.endswith('\\'):
                        ruta_raiz += '\\'

            for ev in evals:
                if ev.evaluado:
                    archivos_evaluados += 1
                if ev.aprobado:
                    archivos_aprobados += 1
                if ev.evaluacion_archivo:
                    val = obtener_valor_calificacion(ev.evaluacion_archivo)
                    if val is not None:
                        calificaciones.append(val)
                estado = getattr(ev, 'estado_archivo', 'PENDIENTE') or 'PENDIENTE'
                estados[estado] += 1
                if hasattr(ev, 'observaciones_evaluacion') and ev.observaciones_evaluacion:
                    observaciones.add(ev.observaciones_evaluacion)

            # Si no hay evaluaciones, contar archivos_relacionados
            if not evals:
                total_archivos += disp.archivos_relacionados.count()

        calificacion_promedio = sum(calificaciones) / len(calificaciones) if calificaciones else 0.0

        resultado['etapas'][prefijo_etapa] = {
            'nombre': nombre_etapa,
            'total_disposiciones': len(disps),
            'total_archivos': total_archivos,
            'archivos_evaluados': archivos_evaluados,
            'archivos_aprobados': archivos_aprobados,
            'calificacion_promedio': round(calificacion_promedio, 2),
            'estados': dict(estados),
            'ruta_raiz': ruta_raiz,
            'observaciones': list(observaciones)[:3],
        }

        total_archivos_global += total_archivos
        total_evaluados_global += archivos_evaluados
        total_aprobados_global += archivos_aprobados

    resultado['totales']['total_archivos'] = total_archivos_global
    resultado['totales']['archivos_evaluados'] = total_evaluados_global
    resultado['totales']['archivos_aprobados'] = total_aprobados_global
    resultado['totales']['porcentaje_evaluado'] = round(
        (total_evaluados_global / total_archivos_global * 100) if total_archivos_global > 0 else 0, 1
    )
    resultado['totales']['porcentaje_aprobado'] = round(
        (total_aprobados_global / total_archivos_global * 100) if total_archivos_global > 0 else 0, 1
    )

    return resultado


def _generar_hoja_matriz_resumen(ws, datos_todos, etapas_activas):
    """Genera Pestaña 1: Matriz Resumen - vista general de todos los municipios."""

    # Estilos
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Colores de estado
    fill_aprobado = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_parcial = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_sin_evaluar = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_vacio = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    num_etapas = len(etapas_activas)
    total_cols = 6 + num_etapas + 2  # 6 fijas + etapas + total + %eval

    # TITULO
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
    last_col_letter = openpyxl.utils.get_column_letter(total_cols)
    ws.merge_cells(f'A1:{last_col_letter}2')
    titulo_cell = ws['A1']
    titulo_cell.value = f"MATRIZ RESUMEN POST-OPERACIÓN (PRODUCTOS)\n{len(datos_todos)} COMBINACIONES MUNICIPIO-MECANISMO - Generado: {fecha_actual}"
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    for r in range(1, 3):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = border

    # LEYENDA DE COLORES (fila 3)
    ws.merge_cells(f'A3:{last_col_letter}3')
    leyenda_cell = ws['A3']
    leyenda_cell.value = "🟢 100% Aprobado  |  🟡 Parcialmente evaluado/aprobado  |  🔵 Sin evaluar  |  ⚪ Vacío/Sin datos  |  --- No existe"
    leyenda_cell.font = Font(bold=True, size=9, color="333333")
    leyenda_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    leyenda_cell.alignment = center_align
    for c in range(1, total_cols + 1):
        ws.cell(row=3, column=c).border = border

    # HEADERS (fila 5)
    headers_fijos = ['MUNICIPIO', 'CÓDIGO DANE', 'DEPARTAMENTO', 'MECANISMO', 'ALCANCE', 'TERRITORIAL']
    headers_etapas = [f"{pref}\n{nombre[:20]}" for pref, nombre in etapas_activas.items()]
    headers_totales = ['TOTAL\nARCHIVOS', '% EVALUACIÓN\nGLOBAL']
    all_headers = headers_fijos + headers_etapas + headers_totales

    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # DATOS
    row = 6
    for dato in datos_todos:
        mun = dato['municipio']
        departamento = mun.cod_depto.nom_depto if mun.cod_depto else "N/A"

        # Columnas fijas
        valores_fijos = [
            mun.nom_municipio,
            mun.cod_municipio,
            departamento,
            dato['mecanismo'],
            mun.alcance_operacion or "N/A",
            mun.nom_territorial or "N/A",
        ]
        for col, val in enumerate(valores_fijos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = left_align if col == 1 else center_align

        # Columnas de etapas
        col_offset = 7
        for prefijo in etapas_activas.keys():
            col_num = col_offset
            col_offset += 1
            cell = ws.cell(row=row, column=col_num)
            cell.border = border
            cell.alignment = center_align

            etapa_data = dato['etapas'].get(prefijo)
            if etapa_data is None:
                cell.value = "---"
                cell.font = Font(color='AAAAAA', size=9)
            elif etapa_data['total_archivos'] == 0:
                cell.value = f"0 archivos\n({etapa_data['total_disposiciones']} dirs)"
                cell.fill = fill_vacio
                cell.font = Font(color='666666', size=9)
                if etapa_data['ruta_raiz']:
                    cell.hyperlink = etapa_data['ruta_raiz']
            else:
                ta = etapa_data['total_archivos']
                ev = etapa_data['archivos_evaluados']
                ap = etapa_data['archivos_aprobados']

                if ap == ta and ta > 0:
                    cell.value = f"{ta} archivos\n(100% aprobado)"
                    cell.fill = fill_aprobado
                    cell.font = Font(color='006100', size=9, bold=True)
                elif ev > 0:
                    pct_ev = round(ev / ta * 100)
                    pct_ap = round(ap / ta * 100)
                    cell.value = f"{ta} archivos\n({pct_ev}% eval / {pct_ap}% aprob)"
                    cell.fill = fill_parcial
                    cell.font = Font(color='9C6500', size=9)
                else:
                    cell.value = f"{ta} archivos\n(sin evaluar)"
                    cell.fill = fill_sin_evaluar
                    cell.font = Font(color='003366', size=9)

                if etapa_data['ruta_raiz']:
                    cell.hyperlink = etapa_data['ruta_raiz']

        # Total archivos
        cell_total = ws.cell(row=row, column=col_offset, value=dato['totales']['total_archivos'])
        cell_total.border = border
        cell_total.alignment = center_align
        cell_total.font = Font(bold=True, size=10)

        # % Evaluacion global
        pct = dato['totales']['porcentaje_evaluado']
        cell_pct = ws.cell(row=row, column=col_offset + 1, value=f"{pct}%")
        cell_pct.border = border
        cell_pct.alignment = center_align
        if pct >= 80:
            cell_pct.font = Font(bold=True, color="006100", size=10)
        elif pct >= 50:
            cell_pct.font = Font(bold=True, color="9C6500", size=10)
        else:
            cell_pct.font = Font(bold=True, color="C00000", size=10)

        row += 1

    # AJUSTAR DIMENSIONES
    anchos_fijos = [25, 12, 20, 18, 15, 20]
    for i, ancho in enumerate(anchos_fijos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    for i in range(7, 7 + num_etapas):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22
    ws.column_dimensions[openpyxl.utils.get_column_letter(7 + num_etapas)].width = 14
    ws.column_dimensions[openpyxl.utils.get_column_letter(7 + num_etapas + 1)].width = 16

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[5].height = 45
    for r in range(6, row):
        ws.row_dimensions[r].height = 35

    # Congelar paneles
    ws.freeze_panes = 'G6'

    print(f"✅ HOJA 1 - Matriz Resumen: {len(datos_todos)} filas")


def _generar_hoja_detalle_evaluacion(ws, datos_todos, etapas_activas):
    """Genera Pestaña 2: Detalle Evaluación por municipio y etapa."""

    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Colores alternados por municipio (bicolor suave)
    mun_fill_a = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")  # Azul muy suave
    mun_fill_b = PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid")  # Melocotón muy suave

    # TITULO
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.merge_cells('A1:N2')
    titulo_cell = ws['A1']
    titulo_cell.value = f"DETALLE DE EVALUACIÓN POR ETAPA - POST-OPERACIÓN\nGenerado: {fecha_actual}"
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    for r in range(1, 3):
        for c in range(1, 15):
            ws.cell(row=r, column=c).border = border

    # HEADERS
    headers = [
        'MUNICIPIO', 'CÓDIGO\nDANE', 'MECANISMO', 'ETAPA',
        'TOTAL\nDIRS', 'TOTAL\nARCHIVOS', 'ARCHIVOS\nEVALUADOS',
        'ARCHIVOS\nAPROBADOS', '% EVALUADO', '% APROBADO',
        'CALIFICACIÓN\nPROMEDIO', 'RESUMEN ESTADOS', 'RUTA DIRECTORIO', 'OBSERVACIONES'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # DATOS
    row = 5
    municipio_anterior = None
    indice_municipio = 0  # Para alternar colores

    for dato in datos_todos:
        mun = dato['municipio']
        if municipio_anterior != mun.cod_municipio:
            indice_municipio += 1
            municipio_anterior = mun.cod_municipio

        # Alternar color por municipio
        row_fill = mun_fill_a if indice_municipio % 2 == 1 else mun_fill_b

        etapas_ordenadas = sorted(dato['etapas'].items(), key=lambda x: x[0])
        if not etapas_ordenadas:
            continue

        for prefijo, etapa_data in etapas_ordenadas:
            ta = etapa_data['total_archivos']
            ev = etapa_data['archivos_evaluados']
            ap = etapa_data['archivos_aprobados']
            pct_ev = round(ev / ta * 100, 1) if ta > 0 else 0
            pct_ap = round(ap / ta * 100, 1) if ta > 0 else 0

            # Resumen estados
            estados = etapa_data['estados']
            partes_estado = []
            for est_nombre, est_count in sorted(estados.items()):
                partes_estado.append(f"{est_count} {est_nombre.lower()}")
            resumen_estados = ", ".join(partes_estado) if partes_estado else "Sin datos"

            observaciones_texto = "; ".join(etapa_data['observaciones']) if etapa_data['observaciones'] else ""

            valores = [
                mun.nom_municipio,
                mun.cod_municipio,
                dato['mecanismo'],
                f"{prefijo} - {etapa_data['nombre']}",
                etapa_data['total_disposiciones'],
                ta,
                ev,
                ap,
                f"{pct_ev}%",
                f"{pct_ap}%",
                etapa_data['calificacion_promedio'],
                resumen_estados,
                etapa_data['ruta_raiz'] or "Sin ruta",
                observaciones_texto,
            ]

            for col, val in enumerate(valores, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.alignment = left_align if col in (1, 4, 12, 13, 14) else center_align
                cell.fill = row_fill

                # Color condicional en % evaluado y % aprobado
                if col == 9:  # % evaluado
                    if pct_ev >= 80:
                        cell.font = Font(bold=True, color="006100")
                    elif pct_ev >= 50:
                        cell.font = Font(bold=True, color="9C6500")
                    else:
                        cell.font = Font(bold=True, color="C00000")
                elif col == 10:  # % aprobado
                    if pct_ap >= 80:
                        cell.font = Font(bold=True, color="006100")
                    elif pct_ap >= 50:
                        cell.font = Font(bold=True, color="9C6500")
                    else:
                        cell.font = Font(bold=True, color="C00000")
                elif col == 13 and etapa_data['ruta_raiz']:
                    cell.hyperlink = etapa_data['ruta_raiz']
                    cell.font = Font(color='0000FF', underline='single', size=9)

            row += 1

    # AJUSTAR DIMENSIONES
    anchos = [25, 12, 18, 30, 8, 10, 10, 10, 10, 10, 12, 35, 55, 40]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[4].height = 35
    for r in range(5, row):
        ws.row_dimensions[r].height = 25

    ws.freeze_panes = 'A5'

    print(f"✅ HOJA 2 - Detalle Evaluación: {row - 5} filas")


def _generar_hoja_estadisticas(ws, datos_todos, etapas_activas):
    """Genera Pestaña 3: Estadísticas agregadas."""

    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
    seccion_font = Font(bold=True, size=13, color="FFFFFF")
    seccion_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
    stat_font = Font(bold=True, size=12, color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')

    # ========== SECCION A: ESTADÍSTICAS GENERALES ==========
    ws.merge_cells('A1:F2')
    titulo_cell = ws['A1']
    titulo_cell.value = f"ESTADÍSTICAS POST-OPERACIÓN (PRODUCTOS)\nGenerado: {fecha_actual}"
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    for r in range(1, 3):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = border

    # Calcular totales globales
    total_municipios = len(set(d['municipio'].cod_municipio for d in datos_todos))
    total_combinaciones = len(datos_todos)
    total_archivos = sum(d['totales']['total_archivos'] for d in datos_todos)
    total_evaluados = sum(d['totales']['archivos_evaluados'] for d in datos_todos)
    total_aprobados = sum(d['totales']['archivos_aprobados'] for d in datos_todos)
    pct_eval_global = round(total_evaluados / total_archivos * 100, 1) if total_archivos > 0 else 0
    pct_aprob_global = round(total_aprobados / total_archivos * 100, 1) if total_archivos > 0 else 0

    stats = [
        ('Municipios procesados', total_municipios),
        ('Combinaciones municipio-mecanismo', total_combinaciones),
        ('Total archivos', f"{total_archivos:,}".replace(',', '.')),
        ('Archivos evaluados', f"{total_evaluados:,}".replace(',', '.') + f" ({pct_eval_global}%)"),
        ('Archivos aprobados', f"{total_aprobados:,}".replace(',', '.') + f" ({pct_aprob_global}%)"),
    ]

    row = 4
    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=1).alignment = left_align
        ws.merge_cells(f'B{row}:C{row}')
        val_cell = ws.cell(row=row, column=2, value=str(value))
        val_cell.font = stat_font
        val_cell.border = border
        val_cell.alignment = center_align
        ws.cell(row=row, column=3).border = border
        row += 1

    # ========== SECCION B: COMPLETITUD POR ETAPA ==========
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    sec_cell = ws.cell(row=row, column=1, value="COMPLETITUD POR ETAPA")
    sec_cell.font = seccion_font
    sec_cell.fill = seccion_fill
    sec_cell.alignment = center_align
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = border
    row += 1

    # Headers de la tabla
    table_headers = ['ETAPA', 'MUNICIPIOS\nCON DATOS', 'TOTAL\nARCHIVOS', '% EVALUADO', '% APROBADO', 'CALIFICACIÓN\nPROMEDIO']
    for col, header in enumerate(table_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    row += 1

    # Datos por etapa
    for prefijo, nombre in etapas_activas.items():
        municipios_con_datos = 0
        archivos_etapa = 0
        evaluados_etapa = 0
        aprobados_etapa = 0
        calificaciones_etapa = []

        for dato in datos_todos:
            etapa_data = dato['etapas'].get(prefijo)
            if etapa_data and (etapa_data['total_archivos'] > 0 or etapa_data['total_disposiciones'] > 0):
                municipios_con_datos += 1
                archivos_etapa += etapa_data['total_archivos']
                evaluados_etapa += etapa_data['archivos_evaluados']
                aprobados_etapa += etapa_data['archivos_aprobados']
                if etapa_data['calificacion_promedio'] > 0:
                    calificaciones_etapa.append(etapa_data['calificacion_promedio'])

        pct_ev = round(evaluados_etapa / archivos_etapa * 100, 1) if archivos_etapa > 0 else 0
        pct_ap = round(aprobados_etapa / archivos_etapa * 100, 1) if archivos_etapa > 0 else 0
        calif_prom = round(sum(calificaciones_etapa) / len(calificaciones_etapa), 2) if calificaciones_etapa else 0

        valores = [
            f"{prefijo} - {nombre}",
            f"{municipios_con_datos}/{total_combinaciones}",
            archivos_etapa,
            f"{pct_ev}%",
            f"{pct_ap}%",
            calif_prom,
        ]

        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = center_align if col > 1 else left_align

            if col == 4:  # % evaluado
                if pct_ev >= 80:
                    cell.font = Font(bold=True, color="006100")
                elif pct_ev >= 50:
                    cell.font = Font(bold=True, color="9C6500")
                else:
                    cell.font = Font(color="C00000")
            elif col == 5:  # % aprobado
                if pct_ap >= 80:
                    cell.font = Font(bold=True, color="006100")
                elif pct_ap >= 50:
                    cell.font = Font(bold=True, color="9C6500")
                else:
                    cell.font = Font(color="C00000")

        row += 1

    # ========== SECCION C: RANKINGS ==========
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    rank_cell = ws.cell(row=row, column=1, value="RANKING DE MUNICIPIOS POR % APROBACIÓN")
    rank_cell.font = seccion_font
    rank_cell.fill = seccion_fill
    rank_cell.alignment = center_align
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = border
    row += 1

    # Ordenar por porcentaje de aprobación
    datos_ordenados = sorted(datos_todos, key=lambda d: d['totales']['porcentaje_aprobado'], reverse=True)
    datos_con_archivos = [d for d in datos_ordenados if d['totales']['total_archivos'] > 0]
    datos_sin_archivos = [d for d in datos_ordenados if d['totales']['total_archivos'] == 0]

    # Top 5
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="TOP 5 - Mayor % Aprobación").font = Font(bold=True, size=11, color="006100")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = border
    row += 1

    rank_headers = ['#', 'MUNICIPIO', 'MECANISMO', 'ARCHIVOS', '% EVALUADO', '% APROBADO']
    for col, h in enumerate(rank_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, size=9)
        cell.border = border
        cell.alignment = center_align
    row += 1

    for i, dato in enumerate(datos_con_archivos[:5], 1):
        vals = [
            i,
            dato['municipio'].nom_municipio,
            dato['mecanismo'],
            dato['totales']['total_archivos'],
            f"{dato['totales']['porcentaje_evaluado']}%",
            f"{dato['totales']['porcentaje_aprobado']}%",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = center_align if col != 2 else left_align
        row += 1

    # Bottom 5
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="BOTTOM 5 - Menor % Aprobación (con archivos)").font = Font(bold=True, size=11, color="C00000")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = border
    row += 1

    for col, h in enumerate(rank_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, size=9)
        cell.border = border
        cell.alignment = center_align
    row += 1

    bottom5 = datos_con_archivos[-5:] if len(datos_con_archivos) >= 5 else datos_con_archivos
    for i, dato in enumerate(reversed(bottom5), 1):
        vals = [
            i,
            dato['municipio'].nom_municipio,
            dato['mecanismo'],
            dato['totales']['total_archivos'],
            f"{dato['totales']['porcentaje_evaluado']}%",
            f"{dato['totales']['porcentaje_aprobado']}%",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = center_align if col != 2 else left_align
        row += 1

    # Municipios sin archivos
    if datos_sin_archivos:
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value=f"MUNICIPIOS SIN ARCHIVOS ({len(datos_sin_archivos)})").font = Font(bold=True, size=11, color="666666")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = border
        row += 1

        for dato in datos_sin_archivos[:10]:
            cell = ws.cell(row=row, column=1, value=dato['municipio'].nom_municipio)
            cell.border = border
            cell2 = ws.cell(row=row, column=2, value=dato['mecanismo'])
            cell2.border = border
            row += 1

    # AJUSTAR DIMENSIONES
    anchos = [35, 15, 15, 12, 12, 15]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    print(f"✅ HOJA 3 - Estadísticas generadas")


def generar_matriz_resumen_postoperacion(municipios_mecanismos_list, temp_dir, etapas_filtro=None):
    """
    Genera el Excel de Matriz Resumen Postoperación con 3 pestañas.
    """
    try:
        print(f"📋 Generando MATRIZ RESUMEN POST para {len(municipios_mecanismos_list)} combinaciones...")

        # Determinar etapas activas
        etapas_activas = OrderedDict()
        for prefijo, nombre in ETAPAS_POSTOPERACION.items():
            if etapas_filtro is None or prefijo in etapas_filtro:
                etapas_activas[prefijo] = nombre

        # Recopilar datos de todos los municipios
        datos_todos = []
        for i, (municipio_obj, mecanismo_str) in enumerate(municipios_mecanismos_list, 1):
            print(f"🔄 Recopilando datos {i}/{len(municipios_mecanismos_list)}: {municipio_obj.nom_municipio} - {mecanismo_str}")
            datos = recopilar_datos_municipio_por_etapa(municipio_obj, mecanismo_str, etapas_filtro)
            datos_todos.append(datos)

        # Crear workbook con 3 pestañas
        wb = openpyxl.Workbook()

        # Pestaña 1: Matriz Resumen
        ws1 = wb.active
        ws1.title = "Matriz Resumen"
        _generar_hoja_matriz_resumen(ws1, datos_todos, etapas_activas)

        # Pestaña 2: Detalle Evaluación
        ws2 = wb.create_sheet(title="Detalle Evaluacion")
        _generar_hoja_detalle_evaluacion(ws2, datos_todos, etapas_activas)

        # Pestaña 3: Estadísticas
        ws3 = wb.create_sheet(title="Estadisticas")
        _generar_hoja_estadisticas(ws3, datos_todos, etapas_activas)

        # Guardar
        fecha = datetime.now().strftime('%d_%m_%Y_%H%M%S')
        nombre = f"Matriz_Resumen_Postoperacion_{len(municipios_mecanismos_list)}_municipios_{fecha}.xlsx"
        archivo_path = os.path.join(temp_dir, nombre)
        wb.save(archivo_path)

        print(f"✅ MATRIZ RESUMEN POST generada: {nombre}")
        return archivo_path

    except Exception as e:
        print(f"❌ Error generando matriz resumen post: {str(e)}")
        import traceback
        traceback.print_exc()
        return None



