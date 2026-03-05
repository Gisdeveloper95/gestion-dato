"""
Vistas para generación de reportes en Excel
"""
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from datetime import datetime
from io import BytesIO

from .models import (
    ProfesionalesSeguimiento,
    ProfesionalTerritorial,
    ProfesionalMunicipio,
    Municipios
)


def aplicar_estilos_cabecera(ws, num_columnas):
    """Aplica estilos a la fila de cabecera"""
    # Colores corporativos
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, num_columnas + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def aplicar_estilos_datos(ws, fila_inicio, fila_fin, num_columnas):
    """Aplica estilos a las filas de datos"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    data_alignment = Alignment(vertical="center", wrap_text=True)
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row in range(fila_inicio, fila_fin + 1):
        for col in range(1, num_columnas + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = data_alignment

            # Filas alternadas
            if (row - fila_inicio) % 2 == 1:
                cell.fill = alt_fill


def ajustar_anchos_columnas(ws, anchos):
    """Ajusta el ancho de las columnas"""
    for idx, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reporte_asignaciones_profesionales(request):
    """
    Genera un reporte Excel con las asignaciones de los profesionales seleccionados.

    Body:
    {
        "profesionales": ["COD001", "COD002", ...]
    }

    Respuesta: Archivo Excel descargable
    """
    codigos_profesionales = request.data.get('profesionales', [])

    if not codigos_profesionales:
        return Response(
            {'error': 'Debe seleccionar al menos un profesional'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Obtener datos de profesionales
    profesionales = ProfesionalesSeguimiento.objects.filter(
        cod_profesional__in=codigos_profesionales
    ).order_by('nombre_profesional')

    if not profesionales.exists():
        return Response(
            {'error': 'No se encontraron profesionales con los códigos proporcionados'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Asignaciones"

    # Título del reporte
    ws.merge_cells('A1:E1')
    titulo_cell = ws['A1']
    titulo_cell.value = f"Reporte de Asignaciones - Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    titulo_cell.font = Font(bold=True, size=14, color="1F4E79")
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Cabeceras
    cabeceras = [
        "Nombre Completo",
        "Correo Electrónico",
        "Rol",
        "Territoriales Asignadas",
        "Municipios Asignados"
    ]

    for col, cabecera in enumerate(cabeceras, 1):
        ws.cell(row=2, column=col, value=cabecera)

    ws.row_dimensions[2].height = 25
    aplicar_estilos_cabecera(ws, len(cabeceras))

    # Datos
    fila = 3
    for profesional in profesionales:
        # Obtener territoriales asignadas (cod_profesional es ForeignKey, usar _id)
        territoriales = ProfesionalTerritorial.objects.filter(
            cod_profesional_id=profesional.cod_profesional
        ).values_list('territorial_seguimiento_id', flat=True)
        territoriales_str = ", ".join(filter(None, territoriales)) or "Sin asignación"

        # Obtener municipios asignados (cod_profesional es ForeignKey, usar _id)
        municipios_ids = ProfesionalMunicipio.objects.filter(
            cod_profesional_id=profesional.cod_profesional
        ).values_list('cod_municipio_id', flat=True)

        # Obtener nombres de municipios
        municipios_nombres = Municipios.objects.filter(
            cod_municipio__in=municipios_ids
        ).values_list('nom_municipio', flat=True).order_by('nom_municipio')

        municipios_str = ", ".join(municipios_nombres) if municipios_nombres else "Sin asignación"

        # Escribir fila (rol_profesional es ForeignKey, convertir a string)
        ws.cell(row=fila, column=1, value=profesional.nombre_profesional or "")
        ws.cell(row=fila, column=2, value=profesional.correo_profesional or "No registrado")
        ws.cell(row=fila, column=3, value=str(profesional.rol_profesional) if profesional.rol_profesional else "")
        ws.cell(row=fila, column=4, value=territoriales_str)
        ws.cell(row=fila, column=5, value=municipios_str)

        # Ajustar altura según contenido
        num_municipios = len(municipios_nombres) if municipios_nombres else 0
        altura = max(20, min(100, 15 + (num_municipios // 5) * 10))
        ws.row_dimensions[fila].height = altura

        fila += 1

    # Aplicar estilos a datos
    if fila > 3:
        aplicar_estilos_datos(ws, 3, fila - 1, len(cabeceras))

    # Ajustar anchos de columnas
    anchos = [35, 35, 20, 40, 80]
    ajustar_anchos_columnas(ws, anchos)

    # Habilitar filtros
    ws.auto_filter.ref = f"A2:E{fila - 1}"

    # Congelar panel (cabecera visible al hacer scroll)
    ws.freeze_panes = "A3"

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generar nombre de archivo
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"Reporte_Asignaciones_{fecha_str}.xlsx"

    # Crear respuesta HTTP
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
