

import re 
import statistics
import os
import zipfile
import tempfile
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.http import HttpResponse, StreamingHttpResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import json
from collections import defaultdict, OrderedDict
from django.contrib.auth.models import Group
from django.shortcuts import render
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from rest_framework import status,serializers
from collections import defaultdict
from django.contrib.auth.models import User
from .permissions import ReadPublicWriteAdminOnly, IsAdminUserOrReadOnly
from rest_framework import generics, viewsets, filters,status
from rest_framework.permissions import BasePermission, SAFE_METHODS
from rest_framework.views import APIView
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django_filters.rest_framework import DjangoFilterBackend
from django.core.cache import cache
from django.db.models import Prefetch

from pathlib import Path
import os.path
from postoperacion.models import ComponentesPost, DisposicionPost

# Importar utilidad para conversión de rutas Linux -> Windows
from backend.path_utils import linux_to_windows_path
try:
    from postoperacion.models import ArchivosPost
except ImportError:
    print("⚠️ Warning: No se pudo importar ArchivosPost. Post-operación será omitida.")
    ArchivosPost = None

from .serializers import (
    UserSerializer, InfoAdministrativaSerializer, InfoAdministrativaSimpleSerializer,
    CentrosPobladosSerializer, CentrosPobladosSimpleSerializer,
    DepartamentosSerializer, MunicipiosSerializer, MunicipiosSimpleSerializer, 
    UsuariosSerializer, TiposInsumosSerializer, CategoriasSerializer, 
    TiposFormatoSerializer, ConceptoSerializer, EntidadesSerializer,
    DetalleInsumoSerializer, DetalleInsumoSimpleSerializer, InsumosSerializer, ClasificacionInsumoSerializer,
    InsumosSimpleSerializer, ClasificacionInsumoSimpleSerializer, NotificacionesSerializer,
    MecanismoGeneralSerializer, MecanismoDetalleSerializer,  AlcanceOperacionSerializer, GrupoSerializer,
    ZonasSerializer,PathDirPreSerializer, PathDirPostSerializer,ListaArchivosPreSerializer,MecanismoOperacionSerializer,EstadosInsumoSerializer,
    RolesSeguimientoSerializer,TerritorialesIgacSerializer,ProfesionalesSeguimientoSerializer,ProfesionalTerritorialSerializer,ProfesionalMunicipioSerializer,
    AuditoriaSerializer, DetalleInsumoOptimizadoSerializer,
    SubClasificacionFuenteSecundariaSerializer
)
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from .models import (
    Departamentos, Municipios, Usuarios, TiposInsumos, Categorias,
    TiposFormato, Concepto, Entidades, DetalleInsumo, Insumos, ClasificacionInsumo,
    Notificaciones, MecanismoGeneral, MecanismoDetalle,  AlcanceOperacion, Grupo,  Zonas, PathDirPre, PathDirPost,ListaArchivosPre,
    TerritorialesIgac,MecanismoOperacion,EstadosInsumo,Auditoria,
    RolesSeguimiento,ProfesionalesSeguimiento,ProfesionalTerritorial, ProfesionalMunicipio,
    InfoAdministrativa, CentrosPoblados,
    SubClasificacionFuenteSecundaria
)
from .utils import registrar_auditoria
from .permissions import IsAuthenticatedOrReadOnly
from django.http import JsonResponse
from django.db.models import Count, Q, Sum, Max, Min, F,ProtectedError
from django.db import transaction, connection, IntegrityError
from datetime import datetime, timedelta
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)

# Vista para crear usuarios
class CreateUserView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]

# ViewSet para Zonas
class ZonasViewSet(viewsets.ModelViewSet):
    queryset = Zonas.objects.all()
    serializer_class = ZonasSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['zona']
    search_fields = ['zona']
    ordering_fields = ['zona']

# ViewSets para las tablas de dominio
class MecanismoGeneralViewSet(viewsets.ModelViewSet):
    queryset = MecanismoGeneral.objects.all()
    serializer_class = MecanismoGeneralSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_mecanismo']
    search_fields = ['cod_mecanismo', 'descripcion']
    ordering_fields = ['cod_mecanismo']

class MecanismoDetalleViewSet(viewsets.ModelViewSet):
    queryset = MecanismoDetalle.objects.all()
    serializer_class = MecanismoDetalleSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_mecanismo_detalle']
    search_fields = ['cod_mecanismo_detalle', 'descripcion']
    ordering_fields = ['cod_mecanismo_detalle']


class AlcanceOperacionViewSet(viewsets.ModelViewSet):
    queryset = AlcanceOperacion.objects.all()
    serializer_class = AlcanceOperacionSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_alcance']
    search_fields = ['cod_alcance']
    ordering_fields = ['cod_alcance']

class GrupoViewSet(viewsets.ModelViewSet):
    queryset = Grupo.objects.all()
    serializer_class = GrupoSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_grupo']
    search_fields = ['cod_grupo', 'descripcion']
    ordering_fields = ['cod_grupo']

# Departamentos ViewSet
class DepartamentosViewSet(viewsets.ModelViewSet):
    queryset = Departamentos.objects.all()
    serializer_class = DepartamentosSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_depto', 'nom_depto']
    search_fields = ['nom_depto']
    ordering_fields = ['cod_depto', 'nom_depto']

    @action(detail=True, methods=['get'])
    def municipios(self, request, pk=None):
        """Obtener todos los municipios de un departamento"""
        municipios = Municipios.objects.filter(cod_depto=pk)
        serializer = MunicipiosSerializer(municipios, many=True)
        return Response(serializer.data)
        
    @action(detail=True, methods=['get'])
    def estadisticas(self, request, pk=None):
        """Obtener estadísticas de municipios por departamento"""
        total_municipios = Municipios.objects.filter(cod_depto=pk).count()
        
        # Conteo por mecanismo general
        mecanismos_stats = Municipios.objects.filter(
            cod_depto=pk, mecanismo_general__isnull=False
        ).values('mecanismo_general').annotate(total=Count('cod_municipio'))
        
        # Conteo por grupo
        grupos_stats = Municipios.objects.filter(
            cod_depto=pk, grupo__isnull=False
        ).values('grupo').annotate(total=Count('cod_municipio'))
        
        # Conteo por alcance operación
        alcance_stats = Municipios.objects.filter(
            cod_depto=pk, alcance_operacion__isnull=False
        ).values('alcance_operacion').annotate(total=Count('cod_municipio'))
        
        # Conteo por operación directa
        operacion_stats = Municipios.objects.filter(
            cod_depto=pk, mecanismo_operacion__isnull=False
        ).values('mecanismo_operacion').annotate(total=Count('cod_municipio'))
        
        return Response({
            'total_municipios': total_municipios,
            'mecanismos': mecanismos_stats,
            'grupos': grupos_stats,
            'alcances': alcance_stats,
            'operaciones': operacion_stats
        })



#  MunicipioViewSet
class MunicipiosViewSet(viewsets.ModelViewSet):
    queryset = Municipios.objects.all()
    serializer_class = MunicipiosSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_municipio', 'nom_municipio', 'cod_depto', 'fecha_inicio', 
                        'mecanismo_general', 'mecanismo_detalle', 'alcance_operacion', 'grupo', 
                        'mecanismo_operacion','nom_territorial','area']
    search_fields = ['nom_municipio']
    ordering_fields = ['cod_municipio', 'nom_municipio', 'fecha_inicio']

    def get_serializer_class(self):
        if self.action == 'list':
            return MunicipiosSimpleSerializer
        return MunicipiosSerializer

    @action(detail=True, methods=['get'])
    def insumos(self, request, pk=None):
        """Obtener todos los insumos de un municipio"""
        insumos = Insumos.objects.filter(cod_municipio=pk)
        serializer = InsumosSerializer(insumos, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def clasificaciones(self, request, pk=None):
        """Obtener todas las clasificaciones de insumos de un municipio"""
        insumos_ids = Insumos.objects.filter(cod_municipio=pk).values_list('cod_insumo', flat=True)
        clasificaciones = ClasificacionInsumo.objects.filter(cod_insumo__in=insumos_ids)
        serializer = ClasificacionInsumoSerializer(clasificaciones, many=True)
        return Response(serializer.data)
        
    @action(detail=True, methods=['get'])
    def detalles(self, request, pk=None):
        """Obtener todos los detalles de insumos de un municipio"""
        insumos_ids = Insumos.objects.filter(cod_municipio=pk).values_list('cod_insumo', flat=True)
        clasificaciones_ids = ClasificacionInsumo.objects.filter(cod_insumo__in=insumos_ids).values_list('cod_clasificacion', flat=True)
        detalles = DetalleInsumo.objects.filter(cod_clasificacion__in=clasificaciones_ids)
        serializer = DetalleInsumoSerializer(detalles, many=True)
        return Response(serializer.data)
        
    @action(detail=False, methods=['get'])
    def por_mecanismo(self, request):
        """Obtener municipios por mecanismo general"""
        mecanismo = request.query_params.get('mecanismo', None)
        if mecanismo:
            municipios = Municipios.objects.filter(mecanismo_general=mecanismo)
            serializer = MunicipiosSerializer(municipios, many=True)
            return Response(serializer.data)
        return Response({"error": "Se requiere el parámetro 'mecanismo'"}, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=['get'])
    def por_operacion(self, request):
        """Obtener municipios por operación directa"""
        operacion = request.query_params.get('operacion', None)
        if operacion:
            municipios = Municipios.objects.filter(mecanismo_operacion=operacion)
            serializer = MunicipiosSerializer(municipios, many=True)
            return Response(serializer.data)
        return Response({"error": "Se requiere el parámetro 'operacion'"}, status=status.HTTP_400_BAD_REQUEST)
        

        
    @action(detail=False, methods=['get'])
    def por_grupo(self, request):
        """Obtener municipios por grupo"""
        grupo = request.query_params.get('grupo', None)
        if grupo:
            municipios = Municipios.objects.filter(grupo=grupo)
            serializer = MunicipiosSerializer(municipios, many=True)
            return Response(serializer.data)
        return Response({"error": "Se requiere el parámetro 'grupo'"}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['delete'])
    def eliminar_cascada(self, request, pk=None):
        """
        Elimina un municipio y todas sus dependencias en cascada.
        Este método debe usarse con precaución ya que eliminará todos los registros
        relacionados con el municipio en todas las tablas.
        """
        try:
            # Obtener el municipio que se va a eliminar
            municipio = self.get_object()
            print(f"Iniciando eliminación en cascada del municipio {municipio.cod_municipio} - {municipio.nom_municipio}")
            
            with transaction.atomic():
                # 1. Eliminar dependencias de postoperacion (si existe la app)
                try:
                    # Archivos Post (a través de Disposiciones)
                    with connection.cursor() as cursor:
                        # Primero obtener todas las disposiciones relacionadas
                        cursor.execute(
                            "SELECT id_disposicion FROM disposicion_post WHERE cod_municipio = %s",
                            [municipio.cod_municipio]
                        )
                        disposiciones_ids = [row[0] for row in cursor.fetchall()]
                        
                        # Eliminar los archivos asociados a esas disposiciones
                        if disposiciones_ids:
                            disposiciones_str = ','.join(str(id) for id in disposiciones_ids)
                            cursor.execute(
                                f"DELETE FROM archivos_post WHERE id_disposicion IN ({disposiciones_str})"
                            )
                            print(f"Eliminados {cursor.rowcount} archivos de postoperación")
                        
                        # Eliminar las disposiciones
                        cursor.execute(
                            "DELETE FROM disposicion_post WHERE cod_municipio = %s",
                            [municipio.cod_municipio]
                        )
                        print(f"Eliminadas {cursor.rowcount} disposiciones de postoperación")
                    
                    # Eliminar rutas de directorios post
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "DELETE FROM path_dir_post WHERE cod_municipio = %s",
                            [municipio.cod_municipio]
                        )
                        print(f"Eliminadas {cursor.rowcount} rutas de directorios postoperación")
                except Exception as e:
                    logger.error(f"Error eliminando dependencias de postoperación: {str(e)}")
                    # No levantamos la excepción para seguir intentando eliminar otras dependencias
                
                # 2. Eliminar dependencias de pre-operación
                try:
                    # Buscar todos los insumos del municipio
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "SELECT cod_insumo FROM insumos WHERE cod_municipio = %s",
                            [municipio.cod_municipio]
                        )
                        insumos_ids = [row[0] for row in cursor.fetchall()]
                    
                    # Para cada insumo, eliminar sus clasificaciones y dependencias
                    for insumo_id in insumos_ids:
                        print(f"Procesando insumo {insumo_id}")
                        
                        # Buscar clasificaciones de este insumo
                        with connection.cursor() as cursor:
                            cursor.execute(
                                "SELECT cod_clasificacion FROM clasificacion_insumo WHERE cod_insumo = %s",
                                [insumo_id]
                            )
                            clasificaciones_ids = [row[0] for row in cursor.fetchall()]
                        
                        # Para cada clasificación, eliminar detalles y archivos
                        for clasificacion_id in clasificaciones_ids:
                            print(f"Procesando clasificación {clasificacion_id}")
                            
                            # Eliminar detalles de insumo
                            with connection.cursor() as cursor:
                                cursor.execute(
                                    "DELETE FROM detalle_insumo WHERE cod_clasificacion = %s",
                                    [clasificacion_id]
                                )
                                print(f"Eliminados {cursor.rowcount} detalles de insumo")
                            
                            # Eliminar lista de archivos pre
                            with connection.cursor() as cursor:
                                cursor.execute(
                                    "DELETE FROM lista_archivos_preo WHERE cod_insumo = %s",
                                    [clasificacion_id]
                                )
                                print(f"Eliminados {cursor.rowcount} archivos pre-operación")
                            
                            # Eliminar notificaciones relacionadas con la clasificación
                            with connection.cursor() as cursor:
                                cursor.execute(
                                    "DELETE FROM notificaciones WHERE tipo_entidad = 'clasificacion' AND id_entidad = %s",
                                    [clasificacion_id]
                                )
                                print(f"Eliminadas {cursor.rowcount} notificaciones de clasificación")
                        
                        # Eliminar las clasificaciones
                        with connection.cursor() as cursor:
                            cursor.execute(
                                "DELETE FROM clasificacion_insumo WHERE cod_insumo = %s",
                                [insumo_id]
                            )
                            print(f"Eliminadas {cursor.rowcount} clasificaciones")
                        
                        # Eliminar notificaciones relacionadas con el insumo
                        with connection.cursor() as cursor:
                            cursor.execute(
                                "DELETE FROM notificaciones WHERE tipo_entidad = 'insumo' AND id_entidad = %s",
                                [insumo_id]
                            )
                            print(f"Eliminadas {cursor.rowcount} notificaciones de insumo")
                    
                    # Eliminar asignaciones de profesionales a municipio
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "DELETE FROM profesional_municipio WHERE cod_municipio = %s",
                            [municipio.cod_municipio]
                        )
                        print(f"Eliminadas {cursor.rowcount} asignaciones de profesionales")
                    
                    # Eliminar rutas de directorios pre
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "DELETE FROM path_dir_pre WHERE cod_municipio = %s",
                            [municipio.cod_municipio]
                        )
                        print(f"Eliminadas {cursor.rowcount} rutas de directorios pre-operación")
                    
                    # Eliminar los insumos
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "DELETE FROM insumos WHERE cod_municipio = %s",
                            [municipio.cod_municipio]
                        )
                        print(f"Eliminados {cursor.rowcount} insumos")
                    
                    # Eliminar notificaciones relacionadas con el municipio
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "DELETE FROM notificaciones WHERE tipo_entidad = 'municipio' AND id_entidad = %s",
                            [municipio.cod_municipio]
                        )
                        print(f"Eliminadas {cursor.rowcount} notificaciones de municipio")
                    
                    # 3. Finalmente, eliminar el municipio
                    municipio.delete()
                    print(f"Municipio {municipio.cod_municipio} eliminado exitosamente")
                    
                    return Response({'message': f'Municipio {municipio.nom_municipio} eliminado con éxito'}, status=status.HTTP_200_OK)
                
                except Exception as e:
                    logger.error(f"Error en eliminación cascada: {str(e)}")
                    return Response(
                        {'error': f'Error al eliminar el municipio: {str(e)}'},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
        
        except Exception as e:
            logger.error(f"Error general en eliminación cascada: {str(e)}")
            return Response(
                {'error': f'Error al eliminar el municipio: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# Usuarios ViewSet
class UsuariosViewSet(viewsets.ModelViewSet):
    queryset = Usuarios.objects.all()
    serializer_class = UsuariosSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_usuario', 'nombre', 'correo']
    search_fields = ['nombre', 'correo']
    ordering_fields = ['cod_usuario', 'nombre']
    
    @action(detail=True, methods=['get'])
    def detalles(self, request, pk=None):
        """Obtener todos los detalles de insumo asociados a un usuario"""
        detalles = DetalleInsumo.objects.filter(cod_usuario=pk)
        serializer = DetalleInsumoSerializer(detalles, many=True)
        return Response(serializer.data)

# TipoInsumo ViewSet
class TiposInsumosViewSet(viewsets.ModelViewSet):
    queryset = TiposInsumos.objects.all()
    serializer_class = TiposInsumosSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo_insumo']
    search_fields = ['tipo_insumo']
    ordering_fields = ['tipo_insumo']

# Categoria ViewSet
class CategoriasViewSet(viewsets.ModelViewSet):

    queryset = Categorias.objects.all()
    serializer_class = CategoriasSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_categoria', 'nom_categoria']
    search_fields = ['nom_categoria']
    ordering_fields = ['cod_categoria', 'nom_categoria']

    @action(detail=True, methods=['get'])
    def verificar_dependencias(self, request, pk=None):
        return Response({"mensaje": "verificar_dependencias funciona", "categoria_id": pk})

    @action(detail=True, methods=['get'])
    def dependencias_detalle(self, request, pk=None):
        return Response({"mensaje": "dependencias_detalle funciona", "categoria_id": pk})
    
    def destroy(self, request, *args, **kwargs):
        """
        Override del método destroy para manejar mejor los errores de integridad
        """
        try:
            return super().destroy(request, *args, **kwargs)
        except (IntegrityError, ProtectedError) as e:
            return Response({
                'error': 'No se puede eliminar esta categoría porque está siendo utilizada por otros datos',
                'detalle': str(e),
                'tipo_error': 'integridad',
                'sugerencia': 'Use la funcionalidad de verificar dependencias para ver qué registros la están usando'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            return Response({
                'error': 'Error interno del servidor',
                'detalle': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                   
# TipoFormato ViewSet
class TiposFormatoViewSet(viewsets.ModelViewSet):
    queryset = TiposFormato.objects.all()
    serializer_class = TiposFormatoSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_formato_tipo']
    search_fields = ['cod_formato_tipo']
    ordering_fields = ['cod_formato_tipo']

# Concepto ViewSet
class ConceptoViewSet(viewsets.ModelViewSet):
    queryset = Concepto.objects.all()
    serializer_class = ConceptoSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_concepto', 'concepto', 'fecha', 'evaluacion', 'pdf']
    search_fields = ['concepto', 'detalle_concepto', 'observacion', 'pdf']
    ordering_fields = ['cod_concepto', 'concepto', 'fecha']

    @action(detail=True, methods=['get'])
    def detalles(self, request, pk=None):
        """Obtener todos los detalles que usan este concepto"""
        detalles = DetalleInsumo.objects.filter(cod_concepto=pk)
        serializer = DetalleInsumoSerializer(detalles, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def buscar(self, request):
        """Buscar conceptos con filtros jerárquicos"""
        queryset = self.queryset
        
        # Si tenemos el filtro de detalle, es prioritario y anula los demás
        detalle_id = request.query_params.get('detalle')
        if detalle_id:
            # Buscar conceptos directamente por el detalle específico
            queryset = queryset.filter(cod_detalle=detalle_id)
        else:
            # Aplicar filtros jerárquicos normales
            departamento_id = request.query_params.get('departamento')
            municipio_id = request.query_params.get('municipio')
            insumo_id = request.query_params.get('insumo')
            clasificacion_id = request.query_params.get('clasificacion')
            
            # Construir la cadena de filtros
            if departamento_id:
                # Obtener municipios en este departamento
                municipios_ids = Municipios.objects.filter(cod_depto=departamento_id).values_list('cod_municipio', flat=True)
                
                # Si también hay un municipio específico, usarlo
                if municipio_id:
                    municipios_ids = [int(municipio_id)]
                
                # Encontrar insumos en estos municipios
                insumos_ids = Insumos.objects.filter(cod_municipio__in=municipios_ids).values_list('cod_insumo', flat=True)
                
                # Si hay un insumo específico, usarlo
                if insumo_id:
                    insumos_ids = [int(insumo_id)]
                
                # Encontrar clasificaciones de estos insumos
                clasificaciones_ids = ClasificacionInsumo.objects.filter(cod_insumo__in=insumos_ids).values_list('cod_clasificacion', flat=True)
                
                # Si hay una clasificación específica, usarla
                if clasificacion_id:
                    clasificaciones_ids = [int(clasificacion_id)]
                
                # Buscar los detalles relacionados con estas clasificaciones
                detalles_ids = DetalleInsumo.objects.filter(cod_clasificacion__in=clasificaciones_ids).values_list('cod_detalle', flat=True)
                
                # Filtrar conceptos por estos detalles
                queryset = queryset.filter(cod_detalle__in=detalles_ids)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

# Entidades ViewSet
class EntidadesViewSet(viewsets.ModelViewSet):
    queryset = Entidades.objects.all()
    serializer_class = EntidadesSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_entidad', 'nom_entidad']
    search_fields = ['nom_entidad']
    ordering_fields = ['cod_entidad', 'nom_entidad']

    @action(detail=True, methods=['get'])
    def detalles(self, request, pk=None):
        """Obtener todos los detalles asociados a una entidad"""
        detalles = DetalleInsumo.objects.filter(cod_entidad=pk)
        serializer = DetalleInsumoSerializer(detalles, many=True)
        return Response(serializer.data)

from rest_framework.pagination import PageNumberPagination

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 25  # ← Reducir de 50 a 25
    page_size_query_param = 'page_size'
    max_page_size = 50  # ← Límite máximo estricto
    
    def get_page_size(self, request):
        """Detectar consultas iniciales y limitarlas más"""
        
        # Si no hay filtros específicos, limitar a 10
        has_filters = any(param in request.query_params for param in [
            'cod_clasificacion', 'estado', 'zona', 'cod_entidad', 'search'
        ])
        
        if not has_filters:
            return 10  # ← LÍMITE INICIAL MUY BAJO
            
        # Con filtros, usar paginación normal
        page_size = super().get_page_size(request)
        return min(page_size, self.max_page_size)

# DetalleInsumo ViewSet
class DetalleInsumoViewSet(viewsets.ModelViewSet):
    # ✅ MANTENER el queryset base para el router
    queryset = DetalleInsumo.objects.all()
    
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        'cod_detalle', 'escala', 'estado', 'cubrimiento', 
        'fecha_entrega', 'fecha_disposicion', 'area', 
        'cod_entidad', 'vigencia', 'formato_tipo', 
        'cod_usuario', 'cod_clasificacion', 'zona', 'cod_centro_poblado'
    ]
    search_fields = ['escala', 'estado', 'cubrimiento', 'area', 'observacion', 'vigencia']
    ordering_fields = ['cod_detalle', 'fecha_entrega', 'fecha_disposicion']
    
    # 🚀 PAGINACIÓN MÁS RESTRICTIVA
    pagination_class = StandardResultsSetPagination  # ← Crear esta clase
    
    def get_serializer_class(self):
        if self.action == 'list':
            return DetalleInsumoSimpleSerializer
        return DetalleInsumoSerializer
    
    def get_queryset(self):
        """🔥 OPTIMIZACIÓN CRÍTICA: Filtros por departamento/municipio en backend"""

        request_params = self.request.query_params

        # 🆕 FILTROS JERÁRQUICOS OPTIMIZADOS (ejecutados en BD, no en frontend)
        cod_departamento = request_params.get('cod_departamento')
        cod_municipio = request_params.get('cod_municipio')
        cod_insumo = request_params.get('cod_insumo')

        # Base queryset con select_related optimizado
        queryset = DetalleInsumo.objects.select_related(
            'cod_clasificacion',
            'cod_clasificacion__cod_insumo',
            'cod_clasificacion__cod_insumo__cod_municipio',
            'cod_clasificacion__cod_insumo__cod_municipio__cod_depto',
            'cod_clasificacion__cod_insumo__cod_categoria',
            'cod_entidad',
            'cod_usuario',
            'cod_centro_poblado',
        )

        # 🎯 FILTRO POR INSUMO (más específico)
        if cod_insumo:
            clasificaciones_ids = ClasificacionInsumo.objects.filter(
                cod_insumo=cod_insumo
            ).values_list('cod_clasificacion', flat=True)
            queryset = queryset.filter(cod_clasificacion__in=clasificaciones_ids)

        # 🎯 FILTRO POR MUNICIPIO
        elif cod_municipio:
            insumos_ids = Insumos.objects.filter(
                cod_municipio=cod_municipio
            ).values_list('cod_insumo', flat=True)
            clasificaciones_ids = ClasificacionInsumo.objects.filter(
                cod_insumo__in=insumos_ids
            ).values_list('cod_clasificacion', flat=True)
            queryset = queryset.filter(cod_clasificacion__in=clasificaciones_ids)

        # 🎯 FILTRO POR DEPARTAMENTO
        elif cod_departamento:
            municipios_ids = Municipios.objects.filter(
                cod_depto=cod_departamento
            ).values_list('cod_municipio', flat=True)
            insumos_ids = Insumos.objects.filter(
                cod_municipio__in=municipios_ids
            ).values_list('cod_insumo', flat=True)
            clasificaciones_ids = ClasificacionInsumo.objects.filter(
                cod_insumo__in=insumos_ids
            ).values_list('cod_clasificacion', flat=True)
            queryset = queryset.filter(cod_clasificacion__in=clasificaciones_ids)

        # Detectar si hay filtros específicos
        has_specific_filters = any(param in request_params for param in [
            'cod_clasificacion', 'estado', 'zona', 'cod_entidad', 'search',
            'cod_departamento', 'cod_municipio', 'cod_insumo'
        ])

        if not has_specific_filters and self.action == 'list':
            # 🛡️ CONSULTA INICIAL: Límite estricto
            return queryset.order_by('-fecha_entrega')[:50]

        return queryset.order_by('-fecha_entrega')
    
    # 🆕 ENDPOINT PARA INFORMACIÓN BÁSICA (sin relaciones)
    @action(detail=False, methods=['get'])
    def basico(self, request):
        """
        Endpoint optimizado para consultas básicas sin relaciones costosas
        """
        queryset = DetalleInsumo.objects.only(
            'cod_detalle', 'estado', 'fecha_entrega', 'escala',
            'cod_clasificacion', 'cod_entidad', 'observacion'
        ).order_by('-fecha_entrega')
        
        # Aplicar filtros básicos
        estado = request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
            
        zona = request.query_params.get('zona')  
        if zona:
            queryset = queryset.filter(zona=zona)
            
        # Paginación estricta
        paginator = StandardResultsSetPagination()
        page = paginator.paginate_queryset(queryset, request)
        
        if page is not None:
            # Serialización mínima
            data = [{
                'cod_detalle': detalle.cod_detalle,
                'estado': detalle.estado,
                'fecha_entrega': detalle.fecha_entrega,
                'escala': detalle.escala,
                'cod_clasificacion': detalle.cod_clasificacion_id,
                'cod_entidad': detalle.cod_entidad_id,
                'observacion': detalle.observacion[:100] if detalle.observacion else None
            } for detalle in page]
            
            return paginator.get_paginated_response(data)
        
        return Response([])
    
    # 🆕 ENDPOINT PARA CARGAR NOMBRES EN LOTE
    @action(detail=False, methods=['post'])
    def nombres_lote(self, request):
        """
        Carga nombres de múltiples detalles en una sola petición
        POST: {'detalles_ids': [1, 2, 3, ...]}
        """
        detalles_ids = request.data.get('detalles_ids', [])
        
        if not detalles_ids or len(detalles_ids) > 50:  # Límite de seguridad
            return Response(
                {'error': 'Máximo 50 detalles por lote'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Cache key para este lote
        cache_key = f"detalle_nombres_{hash(tuple(sorted(detalles_ids)))}"
        cached_result = cache.get(cache_key)
        
        if cached_result:
            return Response(cached_result)
        
        # Consulta optimizada con prefetch
        detalles = DetalleInsumo.objects.filter(
            cod_detalle__in=detalles_ids
        ).select_related(
            'cod_clasificacion__cod_insumo__cod_municipio',
            'cod_entidad'
        ).only(
            'cod_detalle', 
            'cod_clasificacion__nombre',
            'cod_clasificacion__cod_insumo__cod_municipio__nom_municipio',
            'cod_entidad__nom_entidad'
        )
        
        result = {}
        for detalle in detalles:
            result[detalle.cod_detalle] = {
                'clasificacion_nombre': detalle.cod_clasificacion.nombre if detalle.cod_clasificacion else None,
                'municipio_nombre': (
                    detalle.cod_clasificacion.cod_insumo.cod_municipio.nom_municipio 
                    if detalle.cod_clasificacion and detalle.cod_clasificacion.cod_insumo 
                    and detalle.cod_clasificacion.cod_insumo.cod_municipio else None
                ),
                'entidad_nombre': detalle.cod_entidad.nom_entidad if detalle.cod_entidad else None
            }
        
        # Cache por 5 minutos
        cache.set(cache_key, result, 300)
        
        return Response(result)
    
    @action(detail=False, methods=['get'])
    def por_usuario(self, request):
        """Obtener detalles filtrados por usuario"""
        usuario_id = request.query_params.get('usuario', None)
        if not usuario_id:
            return Response({"error": "Se requiere el parámetro 'usuario'"}, status=status.HTTP_400_BAD_REQUEST)
            
        detalles = DetalleInsumo.objects.filter(cod_usuario=usuario_id)
        serializer = self.get_serializer(detalles, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def por_clasificacion(self, request):
        """Obtener detalles filtrados por clasificación"""
        clasificacion_id = request.query_params.get('clasificacion', None)
        if not clasificacion_id:
            return Response({"error": "Se requiere el parámetro 'clasificacion'"}, status=status.HTTP_400_BAD_REQUEST)
            
        detalles = DetalleInsumo.objects.filter(cod_clasificacion=clasificacion_id)
        serializer = self.get_serializer(detalles, many=True)
        return Response(serializer.data)
        
    @action(detail=False, methods=['get'])
    def por_zona(self, request):
        """Obtener detalles filtrados por zona"""
        zona = request.query_params.get('zona', None)
        if not zona:
            return Response({"error": "Se requiere el parámetro 'zona'"}, status=status.HTTP_400_BAD_REQUEST)
            
        detalles = DetalleInsumo.objects.filter(zona=zona)
        serializer = self.get_serializer(detalles, many=True)
        return Response(serializer.data)
        
    @action(detail=False, methods=['get'])
    def por_formato(self, request):
        """Obtener detalles filtrados por formato tipo"""
        formato = request.query_params.get('formato', None)
        if not formato:
            return Response({"error": "Se requiere el parámetro 'formato'"}, status=status.HTTP_400_BAD_REQUEST)
            
        detalles = DetalleInsumo.objects.filter(formato_tipo=formato)
        serializer = self.get_serializer(detalles, many=True)
        return Response(serializer.data)
        
    @action(detail=False, methods=['get'])
    def recientes(self, request):
        """Obtener detalles más recientes por fecha de entrega"""
        dias = request.query_params.get('dias', 30)
        try:
            dias = int(dias)
        except ValueError:
            dias = 30
            
        fecha_limite = datetime.now() - timedelta(days=dias)
        detalles = DetalleInsumo.objects.filter(fecha_entrega__gte=fecha_limite)
        serializer = self.get_serializer(detalles, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def por_centro_poblado(self, request):
        """Obtener detalles filtrados por centro poblado"""
        centro_poblado_id = request.query_params.get('centro_poblado', None)
        if not centro_poblado_id:
            return Response({"error": "Se requiere el parámetro 'centro_poblado'"}, status=status.HTTP_400_BAD_REQUEST)
            
        detalles = DetalleInsumo.objects.filter(cod_centro_poblado=centro_poblado_id)
        serializer = self.get_serializer(detalles, many=True)
        return Response(serializer.data)

    # 🆕 ENDPOINT OPTIMIZADO: Obtener siguiente código disponible
    @action(detail=False, methods=['get'])
    def siguiente_codigo(self, request):
        """
        Obtiene el siguiente código disponible para crear un nuevo detalle.
        Ejecuta MAX(cod_detalle) + 1 directamente en la BD.
        """
        try:
            # Cache por 5 segundos para evitar consultas repetidas
            cache_key = 'detalle_siguiente_codigo'
            cached_codigo = cache.get(cache_key)

            if cached_codigo:
                return Response({'siguiente_codigo': cached_codigo})

            # Consulta optimizada: solo obtiene el MAX
            resultado = DetalleInsumo.objects.aggregate(max_codigo=Max('cod_detalle'))
            max_codigo = resultado['max_codigo'] or 0
            siguiente = max_codigo + 1

            # Guardar en cache por 5 segundos
            cache.set(cache_key, siguiente, 5)

            return Response({
                'siguiente_codigo': siguiente,
                'max_actual': max_codigo
            })
        except Exception as e:
            logger.error(f"Error obteniendo siguiente código: {e}")
            # Fallback: usar timestamp
            fallback = int(datetime.now().timestamp()) % 100000
            return Response({
                'siguiente_codigo': fallback,
                'es_fallback': True,
                'error': str(e)
            })

    # 🆕 ENDPOINT: Estadísticas rápidas por filtros jerárquicos
    @action(detail=False, methods=['get'])
    def conteo_por_filtros(self, request):
        """
        Cuenta detalles por departamento/municipio/insumo sin cargar todos los datos.
        Útil para mostrar totales antes de cargar resultados.
        """
        cod_departamento = request.query_params.get('cod_departamento')
        cod_municipio = request.query_params.get('cod_municipio')
        cod_insumo = request.query_params.get('cod_insumo')

        try:
            if cod_insumo:
                clasificaciones_ids = ClasificacionInsumo.objects.filter(
                    cod_insumo=cod_insumo
                ).values_list('cod_clasificacion', flat=True)
                total = DetalleInsumo.objects.filter(
                    cod_clasificacion__in=clasificaciones_ids
                ).count()

            elif cod_municipio:
                insumos_ids = Insumos.objects.filter(
                    cod_municipio=cod_municipio
                ).values_list('cod_insumo', flat=True)
                clasificaciones_ids = ClasificacionInsumo.objects.filter(
                    cod_insumo__in=insumos_ids
                ).values_list('cod_clasificacion', flat=True)
                total = DetalleInsumo.objects.filter(
                    cod_clasificacion__in=clasificaciones_ids
                ).count()

            elif cod_departamento:
                municipios_ids = Municipios.objects.filter(
                    cod_depto=cod_departamento
                ).values_list('cod_municipio', flat=True)
                insumos_ids = Insumos.objects.filter(
                    cod_municipio__in=municipios_ids
                ).values_list('cod_insumo', flat=True)
                clasificaciones_ids = ClasificacionInsumo.objects.filter(
                    cod_insumo__in=insumos_ids
                ).values_list('cod_clasificacion', flat=True)
                total = DetalleInsumo.objects.filter(
                    cod_clasificacion__in=clasificaciones_ids
                ).count()
            else:
                total = DetalleInsumo.objects.count()

            return Response({
                'total': total,
                'filtros': {
                    'cod_departamento': cod_departamento,
                    'cod_municipio': cod_municipio,
                    'cod_insumo': cod_insumo
                }
            })
        except Exception as e:
            logger.error(f"Error en conteo_por_filtros: {e}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def retrieve(self, request, *args, **kwargs):
        """Añadir auditoría a la consulta detallada"""
        response = super().retrieve(request, *args, **kwargs)
        
        # Registrar la auditoría de consulta
        instance = self.get_object()
        registrar_auditoria(
            request=request,
            tipo_entidad='detalle',
            id_entidad=instance.cod_detalle,
            accion='consultar',
            detalles={
                'clasificacion_id': instance.cod_clasificacion_id
            }
        )
        
        return response
    
    def create(self, request, *args, **kwargs):
        """Añadir auditoría a la creación"""
        response = super().create(request, *args, **kwargs)
        
        # Registrar la auditoría de creación
        nuevos_datos = response.data
        registrar_auditoria(
            request=request,
            tipo_entidad='detalle',
            id_entidad=nuevos_datos['cod_detalle'],
            accion='crear',
            detalles={
                'clasificacion_id': nuevos_datos['cod_clasificacion'],
                'datos_creados': nuevos_datos
            }
        )
        
        return response
    
    def update(self, request, *args, **kwargs):
        """Añadir auditoría a la actualización"""
        # Obtener datos previos para comparación
        instance = self.get_object()
        datos_previos = DetalleInsumoSerializer(instance).data
        
        # Realizar la actualización
        response = super().update(request, *args, **kwargs)
        
        # Determinar campos modificados comparando datos previos y nuevos
        nuevos_datos = response.data
        campos_modificados = {}
        for campo, valor in nuevos_datos.items():
            if campo in datos_previos and datos_previos[campo] != valor:
                campos_modificados[campo] = {
                    'antes': datos_previos[campo],
                    'despues': valor
                }
        
        # Registrar la auditoría de actualización
        registrar_auditoria(
            request=request,
            tipo_entidad='detalle',
            id_entidad=instance.cod_detalle,
            accion='actualizar',
            detalles={
                'clasificacion_id': instance.cod_clasificacion_id,
                'campos_modificados': campos_modificados
            }
        )
        
        return response
    
    def destroy(self, request, *args, **kwargs):
        """Añadir auditoría a la eliminación"""
        # Guardar información de la instancia antes de eliminarla
        instance = self.get_object()
        datos_eliminados = DetalleInsumoSerializer(instance).data
        
        # Realizar la eliminación
        response = super().destroy(request, *args, **kwargs)
        
        # Registrar la auditoría de eliminación
        registrar_auditoria(
            request=request,
            tipo_entidad='detalle',
            id_entidad=instance.cod_detalle,
            accion='eliminar',
            detalles={
                'clasificacion_id': instance.cod_clasificacion_id,
                'datos_eliminados': datos_eliminados
            }
        )
        
        return response


# Insumos ViewSet
class InsumosViewSet(viewsets.ModelViewSet):
    queryset = Insumos.objects.all()
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_insumo', 'cod_municipio', 'cod_categoria', 'tipo_insumo']
    ordering_fields = ['cod_insumo']

    def get_serializer_class(self):
        # Usar serializer simple para mejorar rendimiento en lista
        if self.action == 'list':
            return InsumosSimpleSerializer
        return InsumosSerializer

    @action(detail=True, methods=['get'])
    def clasificaciones(self, request, pk=None):
        """Obtener todas las clasificaciones de un insumo"""
        clasificaciones = ClasificacionInsumo.objects.filter(cod_insumo=pk)
        serializer = ClasificacionInsumoSerializer(clasificaciones, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def detalles(self, request, pk=None):
        """Obtener todos los detalles asociados a un insumo a través de sus clasificaciones"""
        clasificaciones_ids = ClasificacionInsumo.objects.filter(cod_insumo=pk).values_list('cod_clasificacion', flat=True)
        detalles = DetalleInsumo.objects.filter(cod_clasificacion__in=clasificaciones_ids)
        serializer = DetalleInsumoSerializer(detalles, many=True)
        return Response(serializer.data)
        
    @action(detail=False, methods=['get'])
    def por_tipo(self, request):
        """Filtrar insumos por tipo"""
        tipo = request.query_params.get('tipo', None)
        if not tipo:
            return Response({"error": "Se requiere el parámetro 'tipo'"}, status=status.HTTP_400_BAD_REQUEST)
        
        insumos = Insumos.objects.filter(tipo_insumo=tipo)
        serializer = self.get_serializer(insumos, many=True)
        return Response(serializer.data)
        
    @action(detail=False, methods=['get'])
    def por_categoria(self, request):
        """Filtrar insumos por categoría"""
        categoria = request.query_params.get('categoria', None)
        if not categoria:
            return Response({"error": "Se requiere el parámetro 'categoria'"}, status=status.HTTP_400_BAD_REQUEST)
        
        insumos = Insumos.objects.filter(cod_categoria=categoria)
        serializer = self.get_serializer(insumos, many=True)
        return Response(serializer.data)


# ClasificacionInsumo ViewSet
class ClasificacionInsumoViewSet(viewsets.ModelViewSet):
    queryset = ClasificacionInsumo.objects.all()
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_clasificacion', 'cod_insumo', 'nombre']
    search_fields = ['nombre', 'observacion', 'ruta', 'descripcion']
    ordering_fields = ['cod_clasificacion', 'nombre']

    def get_serializer_class(self):
        # Usar serializer simple para mejorar rendimiento en lista
        if self.action == 'list':
            return ClasificacionInsumoSimpleSerializer
        return ClasificacionInsumoSerializer

    @action(detail=True, methods=['get'])
    def detalles(self, request, pk=None):
        """Obtener los detalles asociados a una clasificación"""
        detalles = DetalleInsumo.objects.filter(cod_clasificacion=pk)
        serializer = DetalleInsumoSerializer(detalles, many=True)
        return Response(serializer.data)
        
    @action(detail=False, methods=['get'])
    def sin_detalles(self, request):
        """Obtener clasificaciones que no tienen detalles asociados"""
        # Obtener IDs de clasificaciones que tienen detalles
        clasificaciones_con_detalles = DetalleInsumo.objects.values_list('cod_clasificacion', flat=True).distinct()
        
        # Filtrar clasificaciones que no están en la lista anterior
        clasificaciones_sin_detalles = ClasificacionInsumo.objects.exclude(
            cod_clasificacion__in=clasificaciones_con_detalles
        )
        
        serializer = self.get_serializer(clasificaciones_sin_detalles, many=True)
        return Response(serializer.data)


# Vista para obtener todos los insumos de un municipio específico
class InsumosByMunicipioView(generics.ListAPIView):
    serializer_class = InsumosSimpleSerializer
    permission_classes = [AllowAny]  # Acceso público para consulta

    def get_queryset(self):
        municipio_id = self.kwargs['municipio_id']
        return Insumos.objects.filter(cod_municipio=municipio_id)


# Vista para obtener todas las clasificaciones de un insumo específico
class ClasificacionesByInsumoView(generics.ListAPIView):
    serializer_class = ClasificacionInsumoSerializer
    permission_classes = [AllowAny]  # Acceso público para consulta

    def get_queryset(self):
        insumo_id = self.kwargs['insumo_id']
        return ClasificacionInsumo.objects.filter(cod_insumo=insumo_id)


# Vista para obtener todos los insumos de una categoría específica
class InsumosByCategoriaView(generics.ListAPIView):
    serializer_class = InsumosSimpleSerializer
    permission_classes = [AllowAny]  # Acceso público para consulta

    def get_queryset(self):
        categoria_id = self.kwargs['categoria_id']
        return Insumos.objects.filter(cod_categoria=categoria_id)


# Vista para obtener todos los insumos de un tipo de insumo específico
class InsumosByTipoView(generics.ListAPIView):
    serializer_class = InsumosSimpleSerializer
    permission_classes = [AllowAny]  # Acceso público para consulta

    def get_queryset(self):
        tipo_id = self.kwargs['tipo_id']
        return Insumos.objects.filter(tipo_insumo=tipo_id)


# Vista para obtener todos los detalles asociados a un usuario
class DetallesByUsuarioView(generics.ListAPIView):
    serializer_class = DetalleInsumoSerializer
    permission_classes = [AllowAny]  # Acceso público para consulta

    def get_queryset(self):
        usuario_id = self.kwargs['usuario_id']
        return DetalleInsumo.objects.filter(cod_usuario=usuario_id)


# Vista para obtener todos los detalles de una clasificación
class DetallesByClasificacionView(generics.ListAPIView):
    serializer_class = DetalleInsumoSerializer
    permission_classes = [AllowAny]  # Acceso público para consulta

    def get_queryset(self):
        clasificacion_id = self.kwargs['clasificacion_id']
        return DetalleInsumo.objects.filter(cod_clasificacion=clasificacion_id)


class NotificacionesViewSet(viewsets.ModelViewSet):
    queryset = Notificaciones.objects.all().order_by('-fecha_cambio')
    serializer_class = NotificacionesSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    # ACTIVAR filter_backends para soportar filtros
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    # Filtros básicos
    filterset_fields = ['tipo_entidad', 'accion', 'leido']
    search_fields = ['descripcion', 'tipo_entidad', 'accion']
    ordering_fields = ['fecha_cambio', 'id']
    
    def get_queryset(self):
        """
        Queryset personalizado con filtros avanzados
        """
        queryset = Notificaciones.objects.all()
        
        # Filtro por fechas
        fecha_desde = self.request.query_params.get('fecha_desde', None)
        fecha_hasta = self.request.query_params.get('fecha_hasta', None)
        
        if fecha_desde:
            try:
                queryset = queryset.filter(fecha_cambio__date__gte=fecha_desde)
            except:
                pass
                
        if fecha_hasta:
            try:
                queryset = queryset.filter(fecha_cambio__date__lte=fecha_hasta)
            except:
                pass
        
        # Filtro por rango de fechas con hora
        fecha_cambio_gte = self.request.query_params.get('fecha_cambio__gte', None)
        fecha_cambio_lte = self.request.query_params.get('fecha_cambio__lte', None)
        
        if fecha_cambio_gte:
            try:
                queryset = queryset.filter(fecha_cambio__gte=fecha_cambio_gte)
            except:
                pass
                
        if fecha_cambio_lte:
            try:
                queryset = queryset.filter(fecha_cambio__lte=fecha_cambio_lte)
            except:
                pass
        
        # Filtro por municipio (buscar en datos_contexto)
        municipio_id = self.request.query_params.get('municipio_id', None)
        if municipio_id:
            try:
                # Buscar en datos_contexto JSON
                from django.db.models import Q
                queryset = queryset.filter(
                    Q(datos_contexto__municipio_id=municipio_id) |
                    Q(datos_contexto__cod_municipio=municipio_id) |
                    Q(datos_contexto__municipio_id=int(municipio_id)) |
                    Q(datos_contexto__cod_municipio=int(municipio_id))
                )
            except:
                pass
        
        # Filtro por departamento (buscar municipios que empiecen con código de depto)
        departamento_id = self.request.query_params.get('departamento_id', None)
        if departamento_id:
            try:
                from django.db.models import Q
                depto_str = str(departamento_id).zfill(2)  # Asegurar 2 dígitos
                
                # Buscar municipios cuyo código empiece con el código del departamento
                # Por ejemplo, depto 25 incluye municipios 25001, 25019, etc.
                queryset = queryset.extra(
                    where=[
                        """(
                            datos_contexto->>'municipio_id' LIKE %s OR 
                            datos_contexto->>'cod_municipio' LIKE %s
                        )"""
                    ],
                    params=[f"{depto_str}%", f"{depto_str}%"]
                )
            except Exception as e:
                print(f"Error filtrando por departamento: {e}")
                pass
        
        # Filtro por usuario (buscar en datos_contexto)
        usuario_windows = self.request.query_params.get('usuario_windows', None)
        if usuario_windows:
            try:
                from django.db.models import Q
                queryset = queryset.filter(
                    Q(datos_contexto__usuario_windows__icontains=usuario_windows)
                )
            except:
                pass
        
        # Filtro por leído/no leído (convertir string a boolean)
        leido_param = self.request.query_params.get('leido', None)
        if leido_param is not None:
            if leido_param.lower() in ['true', '1']:
                queryset = queryset.filter(leido=True)
            elif leido_param.lower() in ['false', '0']:
                queryset = queryset.filter(leido=False)
        
        # Aplicar ordenación final
        ordering = self.request.query_params.get('ordering', '-fecha_cambio')
        try:
            queryset = queryset.order_by(ordering)
        except:
            queryset = queryset.order_by('-fecha_cambio')
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def no_leidas(self, request):
        try:
            queryset = self.get_queryset().filter(leido=False)[:100]
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        except Exception as e:
            print(f"❌ ERROR en no_leidas: {e}")
            return Response([])
    
    @action(detail=False, methods=['get'])
    def resumen(self, request):
        try:
            total_no_leidas = self.get_queryset().filter(leido=False).count()
            
            return Response({
                'total_no_leidas': total_no_leidas,
                'por_tipo': [],
                'ultimas': []
            })
        except Exception as e:
            print(f"❌ ERROR en resumen: {e}")
            return Response({
                'total_no_leidas': 0,
                'por_tipo': [],
                'ultimas': []
            })



class InfoAdministrativaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de información administrativa
    - GET: Acceso público (sin autenticación)
    - POST, PUT, PATCH, DELETE: Solo administradores autenticados
    """
    queryset = InfoAdministrativa.objects.all()
    
    # 🛡️ PERMISSION PERSONALIZADA
    permission_classes = [ReadPublicWriteAdminOnly]  # ← CAMBIO PRINCIPAL
    
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    filterset_fields = [
        'cod_info_admin', 'cod_municipio', 'id_gestor_catas', 
        'gestor_prestador_servicio', 'publicacion_year', 'vigencia_rural', 
        'vigencia_urbana', 'estado_rural', 'estado_urbano'
    ]
    
    search_fields = [
        'id_gestor_catas', 'gestor_prestador_servicio', 'observacion',
        'cod_municipio__nom_municipio', 'cod_municipio__cod_depto__nom_depto'
    ]
    
    ordering_fields = [
        'cod_info_admin', 'publicacion_year', 'total_predios', 
        'total_area_terreno_ha', 'total_avaluos'
    ]
    
    def get_serializer_class(self):
        """Usar serializer simple para listas, completo para detalles"""
        if self.action == 'list':
            return InfoAdministrativaSimpleSerializer
        return InfoAdministrativaSerializer
    
    def get_queryset(self):
        """Optimizar consultas con select_related"""
        queryset = super().get_queryset()
        return queryset.select_related('cod_municipio', 'cod_municipio__cod_depto')
    
    # 🔒 PERSONALIZAR PERMISOS POR ACTION (OPCIONAL)
    def get_permissions(self):
        """
        Permisos específicos por action si necesitas más control
        """
        if self.action in ['por_municipio', 'por_departamento', 'estadisticas']:
            # Actions de consulta: acceso libre
            permission_classes = [AllowAny]
        elif self.action in ['create', 'update', 'partial_update', 'destroy']:
            # Actions de escritura: solo administradores
            permission_classes = [IsAuthenticated, IsAdminUser]
        else:
            # Usar permission por defecto
            permission_classes = self.permission_classes
            
        return [permission() for permission in permission_classes]
    
    @action(detail=False, methods=['get'], permission_classes=[AllowAny])
    def por_municipio(self, request):
        """Obtener información administrativa de un municipio específico - ACCESO LIBRE"""
        municipio_id = request.query_params.get('municipio_id', None)
        if not municipio_id:
            return Response(
                {"error": "Se requiere el parámetro 'municipio_id'"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        info_admin = InfoAdministrativa.objects.filter(cod_municipio=municipio_id)
        serializer = self.get_serializer(info_admin, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], permission_classes=[AllowAny])
    def por_departamento(self, request):
        """Obtener información administrativa por departamento - ACCESO LIBRE"""
        departamento_id = request.query_params.get('departamento_id', None)
        if not departamento_id:
            return Response(
                {"error": "Se requiere el parámetro 'departamento_id'"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        info_admin = InfoAdministrativa.objects.filter(
            cod_municipio__cod_depto=departamento_id
        )
        serializer = self.get_serializer(info_admin, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], permission_classes=[AllowAny])
    def estadisticas(self, request):
        """Obtener estadísticas - ACCESO LIBRE"""
        from django.db.models import Count
        
        total_registros = InfoAdministrativa.objects.count()
        
        por_gestor = InfoAdministrativa.objects.filter(
            gestor_prestador_servicio__isnull=False
        ).values('gestor_prestador_servicio').annotate(
            total=Count('cod_info_admin')
        ).order_by('-total')[:10]
        
        por_año = InfoAdministrativa.objects.filter(
            publicacion_year__isnull=False
        ).values('publicacion_year').annotate(
            total=Count('cod_info_admin')
        ).order_by('-publicacion_year')
        
        con_estado_rural = InfoAdministrativa.objects.filter(
            estado_rural__isnull=False
        ).exclude(estado_rural='').count()
        
        con_estado_urbano = InfoAdministrativa.objects.filter(
            estado_urbano__isnull=False
        ).exclude(estado_urbano='').count()
        
        return Response({
            'total_registros': total_registros,
            'por_gestor': por_gestor,
            'por_año': por_año,
            'con_estado_rural': con_estado_rural,
            'con_estado_urbano': con_estado_urbano,
            'municipios_con_info': InfoAdministrativa.objects.values(
                'cod_municipio'
            ).distinct().count()
        })


class CentrosPobladosViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de centros poblados
    - GET: Acceso público (sin autenticación)
    - POST, PUT, PATCH, DELETE: Solo administradores autenticados
    """
    queryset = CentrosPoblados.objects.all()
    
    # 🛡️ PERMISSION PERSONALIZADA
    permission_classes = [ReadPublicWriteAdminOnly]  # ← CAMBIO PRINCIPAL
    
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    filterset_fields = ['cod_centro_poblado', 'cod_municipio']
    
    search_fields = [
        'nom_centro_poblado', 'cod_centro_poblado',
        'cod_municipio__nom_municipio', 'cod_municipio__cod_depto__nom_depto'
    ]
    
    ordering_fields = [
        'cod_centro_poblado', 'nom_centro_poblado', 'area_oficial_ha'
    ]
    
    def get_serializer_class(self):
        """Usar serializer simple para listas, completo para detalles"""
        if self.action == 'list':
            return CentrosPobladosSimpleSerializer
        return CentrosPobladosSerializer
    
    def get_queryset(self):
        """Optimizar consultas con select_related"""
        queryset = super().get_queryset()
        return queryset.select_related('cod_municipio', 'cod_municipio__cod_depto')

    def por_municipio(self, request, municipio_id=None):
        """Obtener centros poblados de un municipio específico - ACCESO LIBRE"""
        
        if not municipio_id:
            return Response(
                {"error": "Se requiere el parámetro 'municipio_id'"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Validar que municipio_id sea un entero
            municipio_id = int(municipio_id)
        except (ValueError, TypeError):
            return Response(
                {"error": f"El municipio_id '{municipio_id}' debe ser un número entero válido"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Filtrar centros poblados por municipio
            centros = CentrosPoblados.objects.filter(cod_municipio=municipio_id)
            
            # Logging para debug
            print(f"🔍 Buscando centros poblados para municipio {municipio_id}")
            print(f"📊 Encontrados {centros.count()} centros poblados")
            
            # Usar el serializer
            serializer = self.get_serializer(centros, many=True)
            
            return Response(serializer.data)
            
        except Exception as e:
            print(f"❌ Error al buscar centros poblados: {str(e)}")
            return Response(
                {
                    "error": "Error interno del servidor",
                    "message": str(e),
                    "municipio_id": municipio_id
                }, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        

    @action(detail=False, methods=['get'], permission_classes=[AllowAny])
    def por_departamento(self, request):
        """Obtener centros poblados por departamento - ACCESO LIBRE"""
        departamento_id = request.query_params.get('departamento_id', None)
        if not departamento_id:
            return Response(
                {"error": "Se requiere el parámetro 'departamento_id'"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        centros = CentrosPoblados.objects.filter(
            cod_municipio__cod_depto=departamento_id
        )
        serializer = self.get_serializer(centros, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], permission_classes=[AllowAny])
    def estadisticas(self, request):
        """Obtener estadísticas de centros poblados - ACCESO LIBRE"""
        from django.db.models import Count
        
        total_centros = CentrosPoblados.objects.count()
        
        por_municipio = CentrosPoblados.objects.values(
            'cod_municipio__nom_municipio', 
            'cod_municipio__cod_depto__nom_depto'
        ).annotate(
            total_centros=Count('cod_centro_poblado')
        ).order_by('-total_centros')[:15]
        
        por_departamento = CentrosPoblados.objects.values(
            'cod_municipio__cod_depto__nom_depto'
        ).annotate(
            total_centros=Count('cod_centro_poblado'),
            municipios_con_centros=Count('cod_municipio', distinct=True)
        ).order_by('-total_centros')
        
        return Response({
            'total_centros': total_centros,
            'municipios_con_centros': CentrosPoblados.objects.values(
                'cod_municipio'
            ).distinct().count(),
            'por_municipio': por_municipio,
            'por_departamento': por_departamento
        })
    
    @action(detail=True, methods=['get'], permission_classes=[AllowAny])
    def municipio_info(self, request, pk=None):
        """Obtener información del municipio del centro poblado - ACCESO LIBRE"""
        centro = self.get_object()
        municipio_serializer = MunicipiosSerializer(centro.cod_municipio)
        
        otros_centros = CentrosPoblados.objects.filter(
            cod_municipio=centro.cod_municipio
        ).exclude(cod_centro_poblado=pk)
        
        otros_centros_serializer = CentrosPobladosSimpleSerializer(
            otros_centros, many=True
        )
        
        return Response({
            'centro_poblado': self.get_serializer(centro).data,
            'municipio': municipio_serializer.data,
            'otros_centros_poblados': otros_centros_serializer.data
        })




    
# Vista para estadísticas generales
@api_view(['GET'])
@permission_classes([AllowAny])
def estadisticas_generales(request):
    """Retorna estadísticas generales de la base de datos"""
    try:
        # Contar registros en cada tabla principal
        cantidad_departamentos = Departamentos.objects.count()
        cantidad_municipios = Municipios.objects.count()
        cantidad_usuarios = Usuarios.objects.count()
        cantidad_insumos = Insumos.objects.count()
        cantidad_clasificaciones = ClasificacionInsumo.objects.count()
        cantidad_detalles = DetalleInsumo.objects.count()
        
        # Estadísticas por mecanismo
        mecanismos_stats = Municipios.objects.filter(
            mecanismo_general__isnull=False).values(
            'mecanismo_general').annotate(total=Count('cod_municipio'))
        
        # Estadísticas por alcance
        alcance_stats = Municipios.objects.filter(
            alcance_operacion__isnull=False).values(
            'alcance_operacion').annotate(total=Count('cod_municipio'))
        
        # Estadísticas por grupos
        grupos_stats = Municipios.objects.filter(
            grupo__isnull=False).values(
            'grupo').annotate(total=Count('cod_municipio'))
            
        # Estadísticas por operación directa
        operacion_stats = Municipios.objects.filter(
            mecanismo_operacion__isnull=False).values(
            'mecanismo_operacion').annotate(total=Count('cod_municipio'))
            
        # Estadísticas de detalles por zona
        zona_stats = DetalleInsumo.objects.filter(
            zona__isnull=False).values(
            'zona').annotate(total=Count('cod_detalle'))
            
        # Estadísticas de detalles por formato
        formato_stats = DetalleInsumo.objects.values(
            'formato_tipo').annotate(total=Count('cod_detalle'))
            
        # Estadísticas de detalles por usuario (top 10)
        usuario_stats = DetalleInsumo.objects.values(
            'cod_usuario').annotate(
            total=Count('cod_detalle')).order_by('-total')[:10]
        
        # Devolver datos
        return Response({
            'total_departamentos': cantidad_departamentos,
            'total_municipios': cantidad_municipios,
            'total_usuarios': cantidad_usuarios,
            'total_insumos': cantidad_insumos,
            'total_clasificaciones': cantidad_clasificaciones,
            'total_detalles': cantidad_detalles,
            'mecanismos': mecanismos_stats,
            'alcances': alcance_stats,
            'grupos': grupos_stats,
            'operaciones': operacion_stats,
            'zonas': zona_stats,
            'formatos': formato_stats,
            'top_usuarios': usuario_stats
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Estadísticas específicas de municipios
@api_view(['GET'])
@permission_classes([AllowAny])
def estadisticas_municipios(request):
    """Retorna estadísticas específicas sobre municipios"""
    try:
        # Total de municipios por departamento
        depto_stats = Municipios.objects.values(
            'cod_depto__cod_depto', 
            'cod_depto__nom_depto'
        ).annotate(
            total=Count('cod_municipio')
        ).order_by('-total')
        
        # Total de municipios por mecanismo general
        mecanismo_stats = Municipios.objects.filter(
            mecanismo_general__isnull=False
        ).values(
            'mecanismo_general'
        ).annotate(
            total=Count('cod_municipio')
        ).order_by('-total')
        
        # Total de municipios por grupo
        grupo_stats = Municipios.objects.filter(
            grupo__isnull=False
        ).values(
            'grupo'
        ).annotate(
            total=Count('cod_municipio')
        ).order_by('-total')
        
        # Total de municipios por operación directa
        operacion_stats = Municipios.objects.filter(
            mecanismo_operacion__isnull=False
        ).values(
            'mecanismo_operacion'
        ).annotate(
            total=Count('cod_municipio')
        ).order_by('-total')
        
        # Cantidad de municipios con fecha de inicio establecida vs sin fecha
        con_fecha = Municipios.objects.filter(fecha_inicio__isnull=False).count()
        sin_fecha = Municipios.objects.filter(fecha_inicio__isnull=True).count()
        
        return Response({
            'por_departamento': depto_stats,
            'por_mecanismo': mecanismo_stats,
            'por_grupo': grupo_stats,
            'por_operacion': operacion_stats,
            'con_fecha_inicio': con_fecha,
            'sin_fecha_inicio': sin_fecha
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Estadísticas específicas de insumos
@api_view(['GET'])
@permission_classes([AllowAny])
def estadisticas_insumos(request):
    """Retorna estadísticas específicas sobre insumos"""
    try:
        # Total de insumos por municipio (top 10)
        municipio_stats = Insumos.objects.values(
            'cod_municipio__cod_municipio', 
            'cod_municipio__nom_municipio'
        ).annotate(
            total=Count('cod_insumo')
        ).order_by('-total')[:10]
        
        # Total de insumos por categoría
        categoria_stats = Insumos.objects.values(
            'cod_categoria__cod_categoria', 
            'cod_categoria__nom_categoria'
        ).annotate(
            total=Count('cod_insumo')
        ).order_by('-total')
        
        # Total de insumos por tipo
        tipo_stats = Insumos.objects.values(
            'tipo_insumo'
        ).annotate(
            total=Count('cod_insumo')
        ).order_by('-total')
        
        # Promedio de clasificaciones por insumo
        from django.db.models import Avg, Subquery, OuterRef
        
        # Subconsulta para contar clasificaciones por insumo
        clasificaciones_por_insumo = ClasificacionInsumo.objects.filter(
            cod_insumo=OuterRef('pk')
        ).values(
            'cod_insumo'
        ).annotate(
            count=Count('*')
        ).values('count')
        
        # Promedio de clasificaciones por insumo
        promedio_clasificaciones = Insumos.objects.annotate(
            num_clasificaciones=Subquery(clasificaciones_por_insumo)
        ).aggregate(
            promedio=Avg('num_clasificaciones')
        )
        
        return Response({
            'top_municipios': municipio_stats,
            'por_categoria': categoria_stats,
            'por_tipo': tipo_stats,
            'promedio_clasificaciones_por_insumo': promedio_clasificaciones
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Estadísticas específicas de detalles
@api_view(['GET'])
@permission_classes([AllowAny])
def estadisticas_detalles(request):
    """Retorna estadísticas específicas sobre detalles de insumos"""
    try:
        # Total de detalles por formato
        formato_stats = DetalleInsumo.objects.values(
            'formato_tipo'
        ).annotate(
            total=Count('cod_detalle')
        ).order_by('-total')
        
        # Total de detalles por zona
        zona_stats = DetalleInsumo.objects.filter(
            zona__isnull=False
        ).values(
            'zona'
        ).annotate(
            total=Count('cod_detalle')
        ).order_by('-total')
        
        # Total de detalles por estado
        estado_stats = DetalleInsumo.objects.filter(
            estado__isnull=False
        ).values(
            'estado'
        ).annotate(
            total=Count('cod_detalle')
        ).order_by('-total')
        
        # Total de detalles por usuario (top 10)
        usuario_stats = DetalleInsumo.objects.values(
            'cod_usuario__cod_usuario', 
            'cod_usuario__nombre'
        ).annotate(
            total=Count('cod_detalle')
        ).order_by('-total')[:10]
        
        # Total de detalles por entidad
        entidad_stats = DetalleInsumo.objects.values(
            'cod_entidad__cod_entidad', 
            'cod_entidad__nom_entidad'
        ).annotate(
            total=Count('cod_detalle')
        ).order_by('-total')
        
        return Response({
            'por_formato': formato_stats,
            'por_zona': zona_stats,
            'por_estado': estado_stats,
            'top_usuarios': usuario_stats,
            'por_entidad': entidad_stats
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ReadOnlyOrAuthenticated(BasePermission):
    """
    Permite acceso de lectura a todos, pero exige autenticación para escritura.
    """
    def has_permission(self, request, view):
        # Permitir GET, HEAD, OPTIONS a todos
        if request.method in SAFE_METHODS:
            return True
            
        # Requerir autenticación para otros métodos
        return request.user and request.user.is_authenticated
    



class PathDirPreViewSet(viewsets.ModelViewSet):
    """ViewSet para rutas preoperativas"""
    queryset = PathDirPre.objects.all()
    serializer_class = PathDirPreSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_municipio']
    search_fields = ['path']
    ordering_fields = ['id', 'fecha_creacion']
    
    @action(detail=False, methods=['get'])
    def by_municipio(self, request):
        """Obtener rutas filtradas por código de municipio"""
        cod_municipio = request.query_params.get('cod_municipio', None)
        if cod_municipio:
            rutas = PathDirPre.objects.filter(cod_municipio=cod_municipio)
            serializer = self.get_serializer(rutas, many=True)
            return Response(serializer.data)
        return Response({"error": "Se requiere parámetro cod_municipio"}, status=400)


class PathDirPostViewSet(viewsets.ModelViewSet):
    """ViewSet para rutas postoperativas"""
    queryset = PathDirPost.objects.all()
    serializer_class = PathDirPostSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_municipio']
    search_fields = ['path']
    ordering_fields = ['id', 'fecha_creacion']
    
    @action(detail=False, methods=['get'])
    def by_municipio(self, request):
        """Obtener rutas filtradas por código de municipio"""
        cod_municipio = request.query_params.get('cod_municipio', None)
        if cod_municipio:
            rutas = PathDirPost.objects.filter(cod_municipio=cod_municipio)
            serializer = self.get_serializer(rutas, many=True)
            return Response(serializer.data)
        return Response({"error": "Se requiere parámetro cod_municipio"}, status=400)


def get_municipios_permitidos(user):
    if not user.is_authenticated:
        return []

    if user.groups.filter(name='Administradores').exists() or user.is_superuser or user.is_staff:
        return 'todos'

    # Buscar directamente si existe un profesional con este username
    # (no requiere grupo, busca por cod_profesional = username)
    try:
        profesional = ProfesionalesSeguimiento.objects.get(
            cod_profesional=user.username
        )

        municipios_ids = ProfesionalMunicipio.objects.filter(
            cod_profesional=profesional
        ).values_list('cod_municipio', flat=True)

        return list(municipios_ids)
    except ProfesionalesSeguimiento.DoesNotExist:
        pass

    return []

class ListaArchivosPreViewSet(viewsets.ModelViewSet):
    queryset = ListaArchivosPre.objects.all()
    serializer_class = ListaArchivosPreSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        municipios_permitidos = get_municipios_permitidos(self.request.user)
        
        if municipios_permitidos == 'todos':
            return queryset
        elif municipios_permitidos:
            # Filtrar archivos por municipios permitidos
            clasificaciones_permitidas = ClasificacionInsumo.objects.filter(
                cod_insumo__cod_municipio__in=municipios_permitidos
            ).values_list('cod_clasificacion', flat=True)
            
            return queryset.filter(cod_insumo__in=clasificaciones_permitidas)
        else:
            # Si no tiene municipios permitidos, no ve nada
            return queryset.none()
    
    def get_permissions(self):
        """
        Solo GET para profesionales, todo para administradores
        """
        if self.action in ['list', 'retrieve', 'por_municipio']:  # ← AGREGAR 'por_municipio'
            permission_classes = [IsAuthenticated]
        else:
            # Para POST, PUT, DELETE, etc.
            permission_classes = [IsAuthenticated, IsAdminUser]
        return [permission() for permission in permission_classes]

    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        """
        Obtiene todos los archivos de preoperación de un municipio específico de manera eficiente
        """
        municipio_id = request.query_params.get('municipio_id')
        if not municipio_id:
            return Response({"error": "Se requiere el parámetro municipio_id"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Verificar permisos
            municipios_permitidos = get_municipios_permitidos(request.user)
            
            if municipios_permitidos != 'todos':
                if not municipios_permitidos or int(municipio_id) not in municipios_permitidos:
                    return Response({"error": "No tiene permisos para acceder a este municipio"}, status=status.HTTP_403_FORBIDDEN)
            
            # Obtener archivos del municipio con información de clasificación e insumo incluida
            archivos = ListaArchivosPre.objects.filter(
                cod_insumo__cod_insumo__cod_municipio=municipio_id
            ).select_related(
                'cod_insumo',
                'cod_insumo__cod_insumo',
                'cod_insumo__cod_insumo__cod_municipio'
            )
            
            serializer = self.get_serializer(archivos, many=True)
            return Response(serializer.data)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ListaArchivosPreByClasificacionView(generics.ListAPIView):
    serializer_class = ListaArchivosPreSerializer

    def get_queryset(self):
        clasificacion_id = self.kwargs['clasificacion_id']
        return ListaArchivosPre.objects.filter(cod_insumo=clasificacion_id)
    
# Agregar estos ViewSets a preoperacion/views.py

class RolesSeguimientoViewSet(viewsets.ModelViewSet):
    queryset = RolesSeguimiento.objects.all()
    serializer_class = RolesSeguimientoSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['rol_profesional']
    search_fields = ['rol_profesional']
    ordering_fields = ['rol_profesional']

class TerritorialesIgacViewSet(viewsets.ModelViewSet):
    queryset = TerritorialesIgac.objects.all()
    serializer_class = TerritorialesIgacSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['nom_territorial']
    search_fields = ['nom_territorial']
    ordering_fields = ['nom_territorial']

class ProfesionalesSeguimientoViewSet(viewsets.ModelViewSet):
    queryset = ProfesionalesSeguimiento.objects.all()
    serializer_class = ProfesionalesSeguimientoSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_profesional', 'rol_profesional']
    search_fields = ['nombre_profesional', 'correo_profesional']
    ordering_fields = ['cod_profesional', 'nombre_profesional']
    # Permitir puntos en cod_profesional (ej: julio.gonzales)
    lookup_value_regex = '[^/]+'
    
    def destroy(self, request, *args, **kwargs):
        """Eliminar profesional + usuario completamente (cascada total)"""
        profesional = self.get_object()
        cod = profesional.cod_profesional
        nombre = profesional.nombre_profesional
        correo = profesional.correo_profesional
        usuario_django = profesional.usuario_django
        try:
            from django.db import transaction
            with transaction.atomic():
                # 1. Eliminar asignaciones dependientes
                n_terr = ProfesionalTerritorial.objects.filter(cod_profesional=cod).delete()[0]
                n_muni = ProfesionalMunicipio.objects.filter(cod_profesional=cod).delete()[0]
                print(f"🗑️ Asignaciones eliminadas: {n_terr} territoriales, {n_muni} municipios")

                # 2. Eliminar registro de profesional
                profesional.delete()
                print(f"✅ Profesional {cod} eliminado")

                # 3. Eliminar usuario custom (tabla usuarios) por correo
                if correo:
                    usuario_custom = Usuarios.objects.filter(correo=correo).first()
                    if usuario_custom:
                        # Reasignar detalles a usuario sistema antes de eliminar
                        detalles = DetalleInsumo.objects.filter(cod_usuario=usuario_custom)
                        if detalles.exists():
                            usuario_sistema, _ = Usuarios.objects.get_or_create(
                                cod_usuario=9999,
                                defaults={'nombre': 'Usuario Eliminado', 'correo': 'sistema@igac.gov.co'}
                            )
                            detalles.update(cod_usuario=usuario_sistema)
                            print(f"✅ {detalles.count()} detalles reasignados a usuario sistema")
                        usuario_custom.delete()
                        print(f"✅ Usuario custom eliminado (correo: {correo})")

                # 4. Eliminar usuario Django (auth_user)
                if usuario_django:
                    username = usuario_django.username
                    usuario_django.delete()
                    print(f"✅ Usuario Django eliminado: {username}")

            return Response(
                {'message': f'Profesional {nombre} y su usuario eliminados completamente'},
                status=status.HTTP_200_OK
            )
        except Exception as e:
            print(f"❌ Error eliminando profesional {cod}: {e}")
            return Response(
                {'error': f'Error al eliminar: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def territoriales(self, request, pk=None):
        """Obtener territoriales asignadas a un profesional"""
        territoriales = ProfesionalTerritorial.objects.filter(cod_profesional=pk)
        serializer = ProfesionalTerritorialSerializer(territoriales, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def municipios(self, request, pk=None):
        """Obtener municipios asignados a un profesional"""
        municipios = ProfesionalMunicipio.objects.filter(cod_profesional=pk)
        serializer = ProfesionalMunicipioSerializer(municipios, many=True)
        return Response(serializer.data)

class ProfesionalTerritorialViewSet(viewsets.ModelViewSet):
    queryset = ProfesionalTerritorial.objects.all()
    serializer_class = ProfesionalTerritorialSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_profesional', 'territorial_seguimiento']
    ordering_fields = ['id']

class ProfesionalMunicipioViewSet(viewsets.ModelViewSet):
    queryset = ProfesionalMunicipio.objects.all()
    serializer_class = ProfesionalMunicipioSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_profesional', 'cod_municipio']
    ordering_fields = ['id']


# ============================================================================
# ASIGNACIONES MASIVAS DE MUNICIPIOS A PROFESIONALES
# ============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def asignacion_masiva_municipios(request):
    """
    Asigna múltiples municipios a múltiples profesionales de manera masiva.
    Opcionalmente asigna también la territorial si se especifica.

    Body:
    {
        "profesionales": ["COD001", "COD002"],
        "municipios": [12345, 12346, 12347],
        "territorial": "TERRITORIAL NORTE"  // opcional
    }

    Response:
    {
        "creados": 6,
        "duplicados": 2,
        "total_intentos": 8,
        "territoriales_asignadas": 2
    }
    """
    profesionales = request.data.get('profesionales', [])
    municipios = request.data.get('municipios', [])
    territorial = request.data.get('territorial', None)

    if not profesionales or not municipios:
        return Response({
            'error': 'Debe proporcionar al menos un profesional y un municipio'
        }, status=status.HTTP_400_BAD_REQUEST)

    creados = 0
    duplicados = 0
    errores = []
    territoriales_asignadas = 0

    # Si se especifica territorial, asignarla a cada profesional
    if territorial:
        for cod_prof in profesionales:
            try:
                obj, created = ProfesionalTerritorial.objects.get_or_create(
                    cod_profesional_id=cod_prof,
                    territorial_seguimiento_id=territorial
                )
                if created:
                    territoriales_asignadas += 1
            except Exception as e:
                errores.append(f"Territorial {cod_prof}-{territorial}: {str(e)}")

    # Asignar municipios
    for cod_prof in profesionales:
        for cod_mun in municipios:
            try:
                # Usar _id suffix para ForeignKeys
                obj, created = ProfesionalMunicipio.objects.get_or_create(
                    cod_profesional_id=cod_prof,
                    cod_municipio_id=cod_mun
                )
                if created:
                    creados += 1
                else:
                    duplicados += 1
            except Exception as e:
                errores.append(f"{cod_prof}-{cod_mun}: {str(e)}")

    # Registrar auditoría
    registrar_auditoria(
        request,
        tipo_entidad='asignacion_masiva',
        id_entidad=0,
        accion='crear',
        detalles={
            'profesionales': profesionales,
            'municipios': municipios,
            'territorial': territorial,
            'creados': creados,
            'duplicados': duplicados,
            'territoriales_asignadas': territoriales_asignadas,
            'errores': errores
        }
    )

    return Response({
        'creados': creados,
        'duplicados': duplicados,
        'errores': errores,
        'territoriales_asignadas': territoriales_asignadas,
        'total_intentos': len(profesionales) * len(municipios)
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def municipios_por_territorial(request, territorial):
    """
    Retorna todos los códigos de municipios que pertenecen a una territorial.

    URL: /preoperacion/municipios-por-territorial/<territorial>/

    Response:
    {
        "territorial": "TERRITORIAL CENTRO",
        "municipios": [12345, 12346, 12347],
        "total": 3
    }
    """
    municipios = Municipios.objects.filter(
        nom_territorial=territorial
    ).values_list('cod_municipio', flat=True)

    return Response({
        'territorial': territorial,
        'municipios': list(municipios),
        'total': len(municipios)
    })


# ============================================================================
# FIN ASIGNACIONES MASIVAS
# ============================================================================

# Agregar estas vistas a preoperacion/views.py

class ProfesionalesByTerritorialView(generics.ListAPIView):
    serializer_class = ProfesionalesSeguimientoSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        territorial_id = self.kwargs['territorial_id']
        profesionales_ids = ProfesionalTerritorial.objects.filter(
            territorial_seguimiento=territorial_id
        ).values_list('cod_profesional', flat=True)
        return ProfesionalesSeguimiento.objects.filter(cod_profesional__in=profesionales_ids)

class ProfesionalesByMunicipioView(generics.ListAPIView):
    serializer_class = ProfesionalesSeguimientoSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        municipio_id = self.kwargs['municipio_id']
        profesionales_ids = ProfesionalMunicipio.objects.filter(
            cod_municipio=municipio_id
        ).values_list('cod_profesional', flat=True)
        return ProfesionalesSeguimiento.objects.filter(cod_profesional__in=profesionales_ids)
    

class MecanismoOperacionViewSet(viewsets.ModelViewSet):
    queryset = MecanismoOperacion.objects.all()
    serializer_class = MecanismoOperacionSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['cod_operacion']
    search_fields = ['cod_operacion', 'descripcion']
    ordering_fields = ['cod_operacion']


class EstadosInsumoViewSet(viewsets.ModelViewSet):
    queryset = EstadosInsumo.objects.all()
    serializer_class = EstadosInsumoSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['estado']
    search_fields = ['estado']
    ordering_fields = ['estado']



# En preoperacion/views.py
# En preoperacion/views.py - VERSIÓN CORREGIDA
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def usuario_actual(request):
    try:
        user = request.user
        
        # 🎯 LÓGICA CLARA DE ROLES
        rol_tipo = 'publico'
        municipios_asignados = []
        
        # SUPER ADMINISTRADORES: is_superuser = True
        if user.is_superuser:
            rol_tipo = 'super_admin'
            print(f"✅ Usuario identificado como SUPER ADMINISTRADOR")
            
        # ADMINISTRADORES NORMALES: is_staff = True pero is_superuser = False
        elif user.is_staff:
            rol_tipo = 'admin'
            print(f"✅ Usuario identificado como ADMINISTRADOR NORMAL")
            
        # PROFESIONALES: Buscar directamente en tabla profesionales_seguimiento por username
        else:
            try:
                profesional = ProfesionalesSeguimiento.objects.get(cod_profesional=user.username)
                rol_tipo = 'profesional'
                municipios_asignados = get_municipios_permitidos(user)
                print(f"✅ Usuario identificado como PROFESIONAL (cod_profesional={user.username})")
            except ProfesionalesSeguimiento.DoesNotExist:
                print(f"⚠️ Usuario {user.username} no es profesional ni admin")
        
        # Buscar usuario custom
        usuario_custom = None
        try:
            if user.email:
                usuario_custom = Usuarios.objects.get(correo=user.email)
        except Usuarios.DoesNotExist:
            pass
        
        # 🔧 RESPUESTA CLARA Y ESPECÍFICA
        response_data = {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'firstName': user.first_name,
            'lastName': user.last_name,
            'isActive': user.is_active,
            'isStaff': user.is_staff,
            'isSuperUser': user.is_superuser,  # ✅ Más claro
            'isAdmin': user.is_staff and not user.is_superuser,  # ✅ Admin normal
            'isSuperAdmin': user.is_superuser,  # ✅ Super admin
            'rol_tipo': rol_tipo,  # ✅ Explícito: super_admin, admin, profesional, publico
            'municipios_asignados': municipios_asignados,
            'groups': list(user.groups.values_list('name', flat=True)),
            'cod_usuario': usuario_custom.cod_usuario if usuario_custom else None,
            'nombre': usuario_custom.nombre if usuario_custom else f"{user.first_name} {user.last_name}".strip()
        }
        
        print(f"📤 Enviando - rol_tipo: {rol_tipo}")
        
        return Response(response_data)
        
    except Exception as e:
        print(f"❌ Error en usuario_actual: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)
       
@api_view(['GET'])
@permission_classes([AllowAny])
def estadisticas_dashboard(request):
    """Retorna estadísticas EXACTAS para el dashboard"""
    try:
        # Consultas directas y simples a cada tabla
        total_municipios_pre = Municipios.objects.count()
        
        from postoperacion.models import DisposicionPost
        total_municipios_post = DisposicionPost.objects.values('cod_municipio').distinct().count()
        
        total_archivos_pre = ListaArchivosPre.objects.count()
        
        from postoperacion.models import ArchivosPost
        total_archivos_post = ArchivosPost.objects.count()
        
        total_territoriales = TerritorialesIgac.objects.count()
        
        # ✅ CONTAR ENLACES (NUEVA LÓGICA)
        total_enlaces = 0
        total_profesionales_las = 0  # Mantener para debug
        total_profesionales_pas = 0  # Mantener para debug
        
        # Contar profesionales por rol
        for prof in ProfesionalesSeguimiento.objects.all():
            if hasattr(prof, 'rol_profesional'):
                rol = str(prof.rol_profesional).upper()
                
                # ✅ BUSCAR EL NUEVO ROL
                if "ENLACE" in rol:
                    total_enlaces += 1
                
                # 🔍 Mantener conteo de roles antiguos para debug
                if "L.A.S" in rol:
                    total_profesionales_las += 1
                elif "P.A.S" in rol:
                    total_profesionales_pas += 1
        
        # 🧪 DEBUG: Imprimir roles para verificar
        print("=== VERIFICACIÓN DE ROLES ===")
        roles_encontrados = set()
        for prof in ProfesionalesSeguimiento.objects.all():
            if hasattr(prof, 'rol_profesional'):
                roles_encontrados.add(str(prof.rol_profesional))
        
        for rol in sorted(roles_encontrados):
            print(f"  - '{rol}'")
        
        print(f"CONTEOS:")
        print(f"  L.A.S antiguos: {total_profesionales_las}")
        print(f"  P.A.S antiguos: {total_profesionales_pas}")
        print(f"  ENLACES nuevos: {total_enlaces}")
        print("============================")
        
        # Municipios sin insumos
        municipios_con_insumos = set()
        for insumo in Insumos.objects.all():
            municipios_con_insumos.add(insumo.cod_municipio_id)
        
        total_municipios_sin_insumos = total_municipios_pre - len(municipios_con_insumos)
        
        # ✅ RESPUESTA ACTUALIZADA
        return Response({
            'total_municipios': total_municipios_pre,
            'total_municipios_post': total_municipios_post, 
            'total_archivos_pre': total_archivos_pre,
            'total_archivos_post': total_archivos_post,
            'total_territoriales': total_territoriales,
            'total_enlaces': total_enlaces,   
            # Mantener campos antiguos temporalmente:
            'total_profesionales_las': total_profesionales_las,
            'total_profesionales_pas': total_profesionales_pas,
            'total_municipios_sin_insumos': total_municipios_sin_insumos
        })
    except Exception as e:
        import traceback
        print("ERROR EN ESTADÍSTICAS DASHBOARD:")
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AuditoriaViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para consultar registros de auditoría (solo lectura)"""
    queryset = Auditoria.objects.all()
    serializer_class = AuditoriaSerializer
    permission_classes = [IsAuthenticated]  # Solo usuarios autenticados pueden ver la auditoría
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['usuario', 'tipo_entidad', 'accion', 'fecha_hora']
    search_fields = ['usuario__nombre', 'tipo_entidad', 'accion']
    ordering_fields = ['fecha_hora', 'usuario', 'tipo_entidad', 'accion']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtrar por rango de fechas si se proporciona
        fecha_desde = self.request.query_params.get('fecha_desde', None)
        fecha_hasta = self.request.query_params.get('fecha_hasta', None)
        
        if fecha_desde:
            try:
                fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d')
                queryset = queryset.filter(fecha_hora__gte=fecha_desde)
            except ValueError:
                pass
                
        if fecha_hasta:
            try:
                fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d')
                fecha_hasta = fecha_hasta.replace(hour=23, minute=59, second=59)
                queryset = queryset.filter(fecha_hora__lte=fecha_hasta)
            except ValueError:
                pass
        
        # Filtrar por entidad específica si se proporciona
        id_entidad = self.request.query_params.get('id_entidad', None)
        if id_entidad:
            queryset = queryset.filter(id_entidad=id_entidad)
        
        return queryset

class VerifyTokenView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Si llega hasta aquí, está autenticado
        return Response({"valid": True})
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def actualizar_perfil(request):
    """
    Actualiza información del perfil del usuario autenticado en auth_user y también en la tabla Usuarios.
    """
    try:
        # Obtener el usuario de Django
        user = request.user
        
        # Obtener datos del request
        data = request.data
        first_name = data.get('firstName', None)
        last_name = data.get('lastName', None)
        email = data.get('email', None)
        
        # Verificar si hay que actualizar el email
        if email and email != user.email:
            # Verificar si el email ya existe
            if User.objects.filter(email=email).exclude(id=user.id).exists():
                return Response(
                    {'error': 'Este correo electrónico ya está registrado por otro usuario.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            user.email = email
        
        # Actualizar nombres si se proporcionaron
        if first_name is not None:
            user.first_name = first_name
        
        if last_name is not None:
            user.last_name = last_name
        
        # Guardar cambios
        user.save()
        
        # También actualizar en la tabla Usuarios si existe un registro
        try:
            from .models import Usuarios
            usuario_custom = Usuarios.objects.filter(correo=user.email).first()
            if usuario_custom:
                # Actualizar nombre combinando first_name y last_name
                nombre_completo = f"{user.first_name} {user.last_name}".strip()
                if nombre_completo:
                    usuario_custom.nombre = nombre_completo
                    usuario_custom.correo = user.email
                    usuario_custom.save()
        except Exception as e:
            # No interrumpir el proceso si falla la actualización en Usuarios
            print(f"Error actualizando la tabla Usuarios: {str(e)}")
        
        # Devolver datos actualizados
        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'firstName': user.first_name,
            'lastName': user.last_name,
            'isActive': user.is_active,
            'isStaff': user.is_staff,
            'isAdmin': user.is_superuser
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cambiar_password(request):
    """
    Cambia la contraseña del usuario autenticado.
    """
    try:
        user = request.user
        data = request.data
        
        # Obtener contraseñas del request
        old_password = data.get('old_password')
        new_password = data.get('new_password')
        
        # Validar datos de entrada
        if not old_password or not new_password:
            return Response(
                {'error': 'Se requieren ambas contraseñas, antigua y nueva.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verificar contraseña actual
        if not user.check_password(old_password):
            return Response(
                {'error': 'La contraseña actual es incorrecta.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar nueva contraseña con validadores de Django
        try:
            validate_password(new_password, user)
        except ValidationError as e:
            return Response(
                {'error': ' '.join(e.messages)},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Cambiar la contraseña
        user.set_password(new_password)
        user.save()
        
        # Mantener la sesión activa
        update_session_auth_hash(request, user)
        
        return Response({'message': 'Contraseña actualizada con éxito.'})
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

# Reemplazar COMPLETAMENTE la función generar_reportes_preoperacion en preoperacion/views.py

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generar_reportes_preoperacion(request):
    """
    ⚡ VERSIÓN OPTIMIZADA CON THREADING (Threading = Más simple y funciona)
    """
    try:
        data = json.loads(request.body) if request.body else request.data
        municipios_ids = data.get('municipios', [])
        generar_individuales = data.get('generar_individuales', True)
        generar_resumen = data.get('generar_resumen', True)
        
        if not municipios_ids:
            return Response(
                {'error': 'No se proporcionaron municipios'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with tempfile.TemporaryDirectory() as temp_dir:
            archivos_generados = []
            municipios_mecanismos_procesados = []
            
            if generar_individuales:
                # Preparar tareas
                tareas = []
                for municipio_id in municipios_ids:
                    try:
                        mecanismos = obtener_mecanismos_municipio(municipio_id)
                        tareas.append((municipio_id, temp_dir, mecanismos))
                    except:
                        continue
                
                # ⚡ THREADING - Número de workers (20 es un buen número)
                max_workers = min(20, len(tareas))
                
                # ⚡ EJECUTAR EN PARALELO CON THREADS
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    # Enviar todas las tareas
                    futuros = {
                        executor.submit(generar_reporte_municipio_thread, *tarea): tarea 
                        for tarea in tareas
                    }
                    
                    # Recopilar resultados conforme terminan
                    for futuro in as_completed(futuros):
                        try:
                            resultado = futuro.result()
                            for archivo_path, municipio, mecanismo in resultado:
                                archivos_generados.append(archivo_path)
                                municipios_mecanismos_procesados.append((municipio, mecanismo))
                        except Exception as e:
                            pass  # Silencioso
                
            # Generar resumen
            if generar_resumen and municipios_mecanismos_procesados:
                archivo_resumen = generar_matriz_resumen_preoperacion(
                    municipios_mecanismos_procesados, 
                    temp_dir
                )
                if archivo_resumen:
                    archivos_generados.append(archivo_resumen)
            
            if not archivos_generados:
                return Response(
                    {'error': 'No se pudieron generar reportes'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            # Crear ZIP en archivo temporal para streaming
            fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_zip = f'reportes_preoperacion_por_mecanismos_{fecha_actual}.zip'

            temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
            temp_zip_path = temp_zip.name
            temp_zip.close()

            with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=1) as zip_file:
                for archivo_path in archivos_generados:
                    nombre_archivo = os.path.basename(archivo_path)
                    zip_file.write(archivo_path, nombre_archivo)

            # Streaming del ZIP
            def zip_iterator():
                CHUNK_SIZE = 8 * 1024 * 1024  # 8MB chunks
                try:
                    with open(temp_zip_path, 'rb') as f:
                        while True:
                            chunk = f.read(CHUNK_SIZE)
                            if not chunk:
                                break
                            yield chunk
                finally:
                    try:
                        os.unlink(temp_zip_path)
                    except:
                        pass

            response = StreamingHttpResponse(
                zip_iterator(),
                content_type='application/zip'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_zip}"'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['X-Accel-Buffering'] = 'no'  # Desactivar buffering en nginx

            return response
            
    except Exception as e:
        return Response(
            {'error': f'Error interno: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

import json
import os
import tempfile
import zipfile
from io import BytesIO
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework import status

def generar_reporte_municipio_thread(municipio_id, temp_dir, mecanismos):
    """
    Worker para procesar 1 municipio con threading
    """
    archivos_generados = []
    
    try:
        municipio = Municipios.objects.get(cod_municipio=municipio_id)
        
        if not mecanismos:
            archivo_path = generar_reporte_individual_preoperacion_completo(
                municipio, temp_dir, "SIN_MECANISMO"
            )
            if archivo_path:
                archivos_generados.append((archivo_path, municipio, "SIN_MECANISMO"))
        else:
            for mecanismo in mecanismos:
                archivo_path = generar_reporte_individual_preoperacion_completo(
                    municipio, temp_dir, mecanismo
                )
                if archivo_path:
                    archivos_generados.append((archivo_path, municipio, mecanismo))
        
        return archivos_generados
        
    except Exception as e:
        return []



def obtener_mecanismos_municipio(municipio_id):
    """
    🆕 Obtiene todos los mecanismos de financiación de un municipio
    """
    try:
        # Obtener todas las rutas del municipio
        rutas_municipio = PathDirPre.objects.filter(cod_municipio=municipio_id)
        
        mecanismos = set()
        
        for ruta in rutas_municipio:
            mecanismo = extraer_mecanismo_financiacion(ruta.path, municipio_id)
            if mecanismo and mecanismo != "SIN_MECANISMO":
                mecanismos.add(mecanismo)
        
        mecanismos_list = sorted(list(mecanismos))
        print(f"🔍 Municipio {municipio_id} - Mecanismos encontrados: {mecanismos_list}")
        
        return mecanismos_list
        
    except Exception as e:
        print(f"❌ Error obteniendo mecanismos para municipio {municipio_id}: {e}")
        return []
    
def extraer_mecanismo_financiacion(ruta_archivo, cod_municipio):
    """
    🆕 Extrae el mecanismo de financiación de la ruta
    Ejemplo: \\repositorio\\...\\25\\175\\PGN-IGAC\\01_preo\\ → "PGN-IGAC"
    """
    if not ruta_archivo or not cod_municipio:
        return "SIN_MECANISMO"
    
    try:
        ruta_str = str(ruta_archivo).replace('/', '\\')
        cod_str = str(cod_municipio)
        
        # Crear patrón del código del municipio (\25\175\)
        if len(cod_str) == 5:  # Código completo como 25175
            depto = cod_str[:2]  # 25
            mun = cod_str[2:]    # 175
            patron_municipio = f"\\{depto}\\{mun}\\"
        else:
            return "SIN_MECANISMO"
        
        # Buscar el patrón en la ruta
        if patron_municipio in ruta_str:
            # Encontrar la posición después del patrón del municipio
            inicio_mecanismo = ruta_str.find(patron_municipio) + len(patron_municipio)
            resto_ruta = ruta_str[inicio_mecanismo:]
            
            # El mecanismo es el primer directorio después del código del municipio
            if '\\' in resto_ruta:
                mecanismo = resto_ruta.split('\\')[0]
                if mecanismo and mecanismo.strip():
                    return mecanismo.strip()
        
        return "SIN_MECANISMO"
        
    except Exception as e:
        print(f"⚠️ Error extrayendo mecanismo de {ruta_archivo}: {e}")
        return "SIN_MECANISMO"
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mecanismos_preoperacion_municipio(request, municipio_id):
    """
    Endpoint para obtener los mecanismos de financiación de un municipio en preoperación.
    GET /preoperacion/mecanismos-preoperacion/<municipio_id>/
    """
    try:
        try:
            municipio = Municipios.objects.get(cod_municipio=municipio_id)
        except Municipios.DoesNotExist:
            return Response({
                'success': False,
                'error': f'Municipio {municipio_id} no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)

        directorios = DirectorioPreoperacion.objects.filter(cod_mpio=municipio_id)

        mecanismos_stats = {}
        for directorio in directorios:
            if directorio.ruta_directorio:
                mecanismo = extraer_mecanismo_financiacion(directorio.ruta_directorio, municipio_id)
                if mecanismo and mecanismo not in ["SIN_MECANISMO", "ERROR_MECANISMO"]:
                    if mecanismo not in mecanismos_stats:
                        mecanismos_stats[mecanismo] = {
                            'codigo': mecanismo,
                            'total_directorios': 0,
                            'total_archivos': 0
                        }
                    mecanismos_stats[mecanismo]['total_directorios'] += 1
                    archivos_count = ArchivoPreoperacion.objects.filter(cod_directorio=directorio).count()
                    mecanismos_stats[mecanismo]['total_archivos'] += archivos_count

        mecanismos_list = sorted(mecanismos_stats.values(), key=lambda x: x['codigo'])

        return Response({
            'success': True,
            'municipio': {
                'cod_municipio': municipio.cod_municipio,
                'nom_municipio': municipio.nom_municipio
            },
            'mecanismos': mecanismos_list,
            'total_mecanismos': len(mecanismos_list),
            'tiene_multiples': len(mecanismos_list) > 1
        })

    except Exception as e:
        print(f"Error obteniendo mecanismos preoperacion para {municipio_id}: {e}")
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def obtener_nombre_directorio_final(ruta_directorio):
    """
    Obtiene el nombre del último directorio de una ruta
    Ejemplo: \\server\\folder\\subfolder → subfolder
    """
    if not ruta_directorio:
        return "Sin nombre"
    
    try:
        parts = ruta_directorio.replace('/', '\\').split('\\')
        # Filtrar partes vacías
        parts = [p for p in parts if p.strip()]
        return parts[-1] if parts else "Sin nombre"
    except:
        return "Sin nombre"


def generar_pestana_detalles_categorias(ws, municipio, todas_categorias, archivos_por_categoria, mecanismo_financiacion):
    """
    Genera la segunda pestaña con detalles de archivos por categoría CON MECANISMO
    ✅ CORREGIDA: archivos_por_categoria valores son listas
    """
    #print("🧠 Aplicando lógica inteligente para categorías con mecanismo...")
    
    # ✅ LÓGICA INTELIGENTE: Filtrar categorías con archivos
    categorias_con_archivos = [
        cat for cat in todas_categorias 
        if cat.cod_categoria in archivos_por_categoria and len(archivos_por_categoria[cat.cod_categoria]) > 0
    ]
    
    # 🎯 DECISIÓN INTELIGENTE
    if len(categorias_con_archivos) > 0:
        categorias_a_mostrar = categorias_con_archivos
        mostrar_mensaje = f"📊 Mostrando {len(categorias_a_mostrar)} categorías CON archivos (de {len(todas_categorias)} totales) para {mecanismo_financiacion}"
        tipo_vista = "FILTRADA"
        titulo_color = "2E7D32"
        header_color = "4CAF50"
        subheader_color = "81C784"
        icono_titulo = "🎯 VISTA FILTRADA"
    else:
        categorias_a_mostrar = todas_categorias[:13]
        mostrar_mensaje = f"📋 Mostrando las 13 categorías principales (ninguna tiene archivos para {mecanismo_financiacion})"
        tipo_vista = "COMPLETA"
        titulo_color = "616161"
        header_color = "9E9E9E"
        subheader_color = "BDBDBD"
        icono_titulo = "📋 VISTA COMPLETA"
    
    print(f"🎯 {mostrar_mensaje}")
    
    # 🎨 ESTILOS DINÁMICOS (mantener existentes)
    titulo_font = Font(bold=True, size=14, color="FFFFFF")
    titulo_fill = PatternFill(start_color=titulo_color, end_color=titulo_color, fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    
    subheader_font = Font(bold=True, color="FFFFFF", size=9)
    subheader_fill = PatternFill(start_color=subheader_color, end_color=subheader_color, fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 🏛️ TÍTULO PRINCIPAL CON MECANISMO
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
    num_columnas = len(categorias_a_mostrar) * 2
    
    from openpyxl.utils import get_column_letter
    ultima_columna = get_column_letter(max(num_columnas, 10))
    
    ws.merge_cells(f'A1:{ultima_columna}2')
    titulo_cell = ws['A1']
    titulo_cell.value = (
        f"{icono_titulo} - DETALLES DE ARCHIVOS POR CATEGORÍA\n"
        f"🏛️ {municipio.nom_municipio.upper()} | MECANISMO: {mecanismo_financiacion} | {mostrar_mensaje} | 📅 {fecha_actual}"
    )
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    titulo_cell.border = border
    
    # 📊 HEADERS PRINCIPALES (ROW 4) - Nombres de categorías
    row_header = 4
    col = 1
    
    for numero, categoria in enumerate(categorias_a_mostrar, 1):
        # Merge 2 columnas para cada categoría
        start_col = col
        end_col = col + 1
        
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        
        ws.merge_cells(f'{start_letter}{row_header}:{end_letter}{row_header}')
        
        # Información de la categoría
        num_archivos = len(archivos_por_categoria.get(categoria.cod_categoria, []))
        icono_estado = f"📁({num_archivos})" if num_archivos > 0 else "📭"
        
        # Nombre de la categoría (truncado)
        nombre_cat = categoria.nom_categoria[:25] + "..." if len(categoria.nom_categoria) > 25 else categoria.nom_categoria
        
        header_cell = ws.cell(row=row_header, column=start_col)
        header_cell.value = f"{numero:02d}. {icono_estado} {nombre_cat}"
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = center_align
        header_cell.border = border
        
        col += 2
    
    # 📊 SUB-HEADERS (ROW 5)
    row_subheader = 5
    col = 1
    
    for categoria in categorias_a_mostrar:
        # Columna 1: Nombre Archivo
        cell_nombre = ws.cell(row=row_subheader, column=col, value="📄 Nombre Archivo")
        cell_nombre.font = subheader_font
        cell_nombre.fill = subheader_fill
        cell_nombre.alignment = center_align
        cell_nombre.border = border
        
        # Columna 2: Ruta
        cell_ruta = ws.cell(row=row_subheader, column=col+1, value="📁 Ruta de Acceso")
        cell_ruta.font = subheader_font
        cell_ruta.fill = subheader_fill
        cell_ruta.alignment = center_align
        cell_ruta.border = border
        
        col += 2
    
    # 📋 DATOS - Archivos por categoría
    max_archivos = max([len(archivos_por_categoria.get(cat.cod_categoria, [])) 
                       for cat in categorias_a_mostrar], default=3)
    
    if max_archivos == 0:
        max_archivos = 3
    
    for fila_archivo in range(max_archivos):
        row_data = 6 + fila_archivo
        col = 1
        
        # Alternar colores
        if fila_archivo % 2 == 0:
            row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for categoria in categorias_a_mostrar:
            archivos_cat = archivos_por_categoria.get(categoria.cod_categoria, [])
            
            if fila_archivo < len(archivos_cat):
                archivo = archivos_cat[fila_archivo]
                
                # Nombre del archivo
                nombre_archivo = archivo.nombre_insumo or obtener_nombre_directorio_final(archivo.path_file) or "Sin nombre"
                cell_nombre = ws.cell(row=row_data, column=col, value=f"📄 {nombre_archivo}")
                cell_nombre.alignment = left_align
                cell_nombre.border = border
                cell_nombre.font = Font(size=9, color="2E7D32")
                cell_nombre.fill = row_fill
                
                # Ruta del directorio (no del archivo)
                directorio = extraer_directorio_desde_ruta(archivo.path_file)
                cell_ruta = ws.cell(row=row_data, column=col+1, value=directorio or "Sin ruta")
                cell_ruta.alignment = left_align
                cell_ruta.border = border
                cell_ruta.font = Font(size=8, color="1565C0")
                cell_ruta.fill = row_fill
                
                # Agregar hipervínculo si hay directorio
                if directorio:
                    cell_ruta.hyperlink = directorio
                    cell_ruta.font = Font(size=8, color="0000FF", underline='single')
            else:
                # Celdas vacías
                cell_nombre = ws.cell(row=row_data, column=col, value="")
                cell_nombre.border = border
                cell_nombre.fill = row_fill
                
                cell_ruta = ws.cell(row=row_data, column=col+1, value="")
                cell_ruta.border = border
                cell_ruta.fill = row_fill
            
            col += 2
    
    # 📏 AJUSTAR DIMENSIONES (mantener existentes)
    for i in range(1, len(categorias_a_mostrar) * 2 + 1):
        column_letter = get_column_letter(i)
        
        if i % 2 == 1:  # Nombres de archivo
            ws.column_dimensions[column_letter].width = 35
        else:  # Rutas
            ws.column_dimensions[column_letter].width = 50
    
    # Altura de filas
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 25
    
    for row_num in range(6, 6 + max_archivos):
        ws.row_dimensions[row_num].height = 28
    
    print(f"✅ Matriz de detalles por categorías creada para mecanismo {mecanismo_financiacion}: {len(categorias_a_mostrar)} categorías")
# ===============================================
# FUNCIONES COMPLETAS CORREGIDAS PARA REPORTES PRE-OPERACIÓN
# ===============================================

# NO OLVIDES AGREGAR AL INICIO DEL ARCHIVO:
# from pathlib import Path
# import os.path

# =============================================
# FUNCIÓN 1: EXTRACCIÓN DE DIRECTORIO INTELIGENTE
# =============================================

def extraer_directorio_desde_ruta(ruta_archivo):
    """
    ✅ VERSIÓN ACTUALIZADA: SIEMPRE extrae el directorio padre del archivo
    NO abre archivos directamente, TODOS van al directorio padre
    CONVIERTE de formato Linux a formato Windows para usuarios
    """
    if not ruta_archivo:
        return None

    try:
        ruta_path = Path(ruta_archivo)

        # 🎯 SIEMPRE OBTENER EL DIRECTORIO PADRE
        # Sin excepciones para PDFs, Excel, etc.

        # Detectar si es un archivo por la presencia de extensión
        if '.' in ruta_path.name and len(ruta_path.suffix) <= 5:
            # ES UN ARCHIVO - Devolver el directorio padre
            directorio_padre = str(ruta_path.parent)
            #print(f"📁 ARCHIVO DETECTADO - Directorio padre: {directorio_padre}")
            resultado = directorio_padre
        else:
            # YA ES UN DIRECTORIO - Mantenerlo
            #print(f"📂 DIRECTORIO - Mantener: {str(ruta_archivo)}")
            resultado = str(ruta_archivo)

        # 🔄 Convertir de Linux a Windows para que los usuarios puedan acceder
        resultado = linux_to_windows_path(resultado)
        return resultado

    except Exception as e:
        print(f"⚠️ Error procesando ruta {ruta_archivo}: {str(e)}")

        # FALLBACK usando manipulación de strings
        try:
            ruta_str = str(ruta_archivo)

            # Verificar si tiene extensión (es archivo)
            if '.' in ruta_str:
                ultimo_punto = ruta_str.rfind('.')
                posible_extension = ruta_str[ultimo_punto:]

                # Si parece una extensión válida (máximo 5 caracteres)
                if len(posible_extension) <= 5:
                    # Obtener directorio padre
                    if '\\' in ruta_str:
                        resultado = '\\'.join(ruta_str.split('\\')[:-1])
                    elif '/' in ruta_str:
                        resultado = '/'.join(ruta_str.split('/')[:-1])
                    else:
                        resultado = ruta_str
                    # 🔄 Convertir de Linux a Windows
                    return linux_to_windows_path(resultado)

            # Si no es archivo o no se puede determinar, convertir y devolver
            return linux_to_windows_path(ruta_str)
        except:
            # Último recurso: intentar convertir la ruta original
            return linux_to_windows_path(str(ruta_archivo)) if ruta_archivo else None

# =============================================
# FUNCIÓN 2: MATRIZ PRIMARIOS CON RUTAS INTELIGENTES
# =============================================

def generar_pestana_matriz_primarios(ws, municipio, insumos_municipio, clasificaciones_municipio, archivos_municipio, mecanismo_financiacion):
    """
    ✅ VERSIÓN ACTUALIZADA con columna CATEGORÍA agrupadora
    """
    # 🎨 ESTILOS PROFESIONALES (mantener existentes)
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    
    info_font = Font(bold=True, size=12, color="1F4E79")
    info_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    
    # Nuevo estilo para categorías
    categoria_font = Font(bold=True, size=10, color="FFFFFF")
    categoria_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thick_border = Border(
        left=Side(style='thick', color='1F4E79'),
        right=Side(style='thick', color='1F4E79'),
        top=Side(style='thick', color='1F4E79'),
        bottom=Side(style='thick', color='1F4E79')
    )
    
    thin_border = Border(
        left=Side(style='thin', color='5D6D7E'),
        right=Side(style='thin', color='5D6D7E'),
        top=Side(style='thin', color='5D6D7E'),
        bottom=Side(style='thin', color='5D6D7E')
    )
    
    hyperlink_font = Font(color='0000FF', underline='single', size=10)
    normal_font = Font(color='000000', size=10)
    
    # 🏛️ ENCABEZADO PRINCIPAL CON MECANISMO
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    # Fila 1-2: Título principal CON MECANISMO (ahora con 8 columnas por la nueva)
    ws.merge_cells('A1:H2')
    titulo_cell = ws['A1']
    titulo_cell.value = f"MATRIZ DE INSUMOS PRIMARIOS\nMUNICIPIO DE {municipio.nom_municipio.upper()}\nMECANISMO: {mecanismo_financiacion}"
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    titulo_cell.border = thick_border
    
    # 🏛️ INFORMACIÓN DEL MUNICIPIO
    ws.merge_cells('A3:H3')
    departamento = municipio.cod_depto.nom_depto if municipio.cod_depto else "No definido"
    territorial = municipio.nom_territorial or "No asignada"
    alcance = municipio.alcance_operacion or "No definido"
    
    info_cell = ws['A3']
    info_cell.value = f"📍 Código DANE: {municipio.cod_municipio} | 🏛️ Departamento: {departamento} | 🌍 Territorial: {territorial} | 📋 Alcance: {alcance} | 📅 Generado: {fecha_actual}"
    info_cell.font = info_font
    info_cell.fill = info_fill
    info_cell.alignment = center_align
    info_cell.border = thin_border
    
    # 📊 ESTADÍSTICAS RÁPIDAS
    total_archivos = len(archivos_municipio)
    total_clasificaciones = clasificaciones_municipio.count()
    total_insumos = insumos_municipio.count()
    
    ws.merge_cells('A4:H4')
    stats_cell = ws['A4']
    stats_cell.value = f"📊 RESUMEN MECANISMO {mecanismo_financiacion}: {total_insumos} Insumos | {total_clasificaciones} Clasificaciones | {total_archivos} Archivos Filtrados"
    stats_cell.font = Font(bold=True, size=11, color="2E86AB")
    stats_cell.fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    stats_cell.alignment = center_align
    stats_cell.border = thin_border
    
    # 📊 HEADERS PRINCIPALES CON NUEVA COLUMNA CATEGORÍA
    row = 6
    headers = [
        'CATEGORÍA',  # ✅ NUEVA COLUMNA
        'ENTIDAD\nRESPONSABLE', 
        'TIPO DE INSUMO\nREQUERIDO', 
        'FUENTE\nDE DATOS', 
        'FECHA DE\nDISPOSICIÓN', 
        'RUTA DE ACCESO\nAL REPOSITORIO', 
        'OBSERVACIONES\nGENERALES', 
        'USUARIO\nDILIGENCIA'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thick_border
    
    # 📋 CONFIGURACIÓN DE CATEGORÍAS E INSUMOS AGRUPADOS
    categorias_insumos = [
        {
            'categoria': '📂 INSUMOS DE ORDENAMIENTO TERRITORIAL',
            'categoria_color': '16A085',
            'insumos': [
                {
                    'entidad': '🏛️ Alcaldías Municipales',
                    'insumo': 'Instrumentos de Ordenamiento Territorial\n(POT, PBOT, EOT)',
                    'buscar_en_ruta': '08_inst_ord_terri',
                    'fuente': 'PRIMARIA'
                }
            ]
        },
        {
            'categoria': '🗄️ INFORMACIÓN CATASTRAL',
            'categoria_color': '2874A6',
            'insumos': [
                {
                    'entidad': '🗺️ IGAC',
                    'insumo': 'Base de Datos Catastral\n(Registros R1 y R2)',
                    'buscar_en_ruta': '03_info_catas\\01_r1_r2',
                    'fuente': 'PRIMARIA'
                },
                {
                    'entidad': '🗺️ IGAC',
                    'insumo': 'Base Catastral Gráfica\n(Información Geoespacial)',
                    'buscar_en_ruta': '03_info_catas\\02_gdb',
                    'fuente': 'PRIMARIA'
                },
                {
                    'entidad': '🗺️ IGAC',
                    'insumo': 'Información de Construcciones\n(Cartografía Vectorial)',
                    'buscar_en_ruta': '03_info_catas\\03_tab_terr_constr',
                    'fuente': 'PRIMARIA'
                },
                {
                    'entidad': '🗺️ IGAC',
                    'insumo': 'Estudio de Zonas Homogeneas Fisicas y Geoeconomicas',
                    'buscar_en_ruta': '03_info_catas\\04_estu_zhf_zhg',
                    'fuente': 'PRIMARIA'
                }
            ]
        },
        {
            'categoria': '🗺️ CARTOGRAFÍA BÁSICA',
            'categoria_color': 'B7950B',
            'insumos': [
                {
                    'entidad': '🗺️ IGAC',
                    'insumo': 'Ortoimágenes de Alta Resolución\n(Zonas Urbana y Rural)',
                    'buscar_en_ruta': '01_carto_basic\\01_rast\\01_orto',
                    'fuente': 'PRIMARIA'
                },
                {
                    'entidad': '🗺️ IGAC',
                    'insumo': 'Cartografía Base Vectorial\n(Elementos Geográficos)',
                    'buscar_en_ruta': '01_carto_basic\\02_vect',
                    'fuente': 'PRIMARIA'
                },
                {
                    'entidad': '🗺️ IGAC',
                    'insumo': 'Modelo Digital de Elevación\n(DEM y Curvas de Nivel)',
                    'buscar_en_ruta': '01_carto_basic\\01_rast\\02_dtm',
                    'fuente': 'PRIMARIA'
                }
            ]
        },
        {
            'categoria': '📍 DESLINDES',
            'categoria_color': '884EA0',
            'insumos': [
                {
                    'entidad': '🗺️ IGAC',
                    'insumo': 'Delimitación Territorial\n(Límites Municipales)',
                    'buscar_en_ruta': '04_deslin\\01_shp',
                    'fuente': 'PRIMARIA'
                }
            ]
        },
        {
            'categoria': '📋 INSUMOS REGISTRALES',
            'categoria_color': 'DC7633',
            'insumos': [
                {
                    'entidad': '📋 SNR (Superintendencia\nde Notariado y Registro)',
                    'insumo': 'Histórico de Anotaciones\n(Base Registral Municipal y XTE)',
                    'buscar_en_ruta': '06_insu_regis',
                    'fuente': 'PRIMARIA'
                }
            ]
        }
    ]
    
    # Procesar cada categoría y sus insumos
    row = 7
    categoria_actual = None
    primera_fila_categoria = {}
    filas_por_categoria = {}
    
    for grupo in categorias_insumos:
        categoria = grupo['categoria']
        categoria_color = grupo['categoria_color']
        primera_fila_categoria[categoria] = row
        filas_categoria = 0
        
        for config in grupo['insumos']:
            # Buscar archivos para este tipo de insumo
            archivos_tipo = []
            # 🔧 NORMALIZAR: Convertir patrón de búsqueda a formato con / para coincidir con rutas Linux en BD
            patron_normalizado = config['buscar_en_ruta'].replace('\\', '/')
            for archivo in archivos_municipio:
                if archivo.path_file:
                    # Normalizar también la ruta del archivo para comparación
                    ruta_normalizada = archivo.path_file.replace('\\', '/')
                    if patron_normalizado in ruta_normalizada:
                        archivos_tipo.append(archivo)
            
            # Alternar colores de fila
            row_num = row - 6
            if row_num % 2 == 0:
                row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            else:
                row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # Columna A: CATEGORÍA (solo en la primera fila de cada grupo)
            if filas_categoria == 0:
                cell_categoria = ws.cell(row=row, column=1, value=categoria)
                cell_categoria.font = categoria_font
                cell_categoria.fill = PatternFill(start_color=categoria_color, end_color=categoria_color, fill_type="solid")
                cell_categoria.alignment = center_align
                cell_categoria.border = thick_border
            else:
                cell_categoria = ws.cell(row=row, column=1, value="")
                cell_categoria.fill = PatternFill(start_color=categoria_color, end_color=categoria_color, fill_type="solid")
                cell_categoria.border = thin_border
            
            # Columna B: ENTIDAD
            cell_entidad = ws.cell(row=row, column=2, value=config['entidad'])
            cell_entidad.font = Font(bold=True, size=10, color="1F4E79")
            cell_entidad.alignment = center_align
            cell_entidad.border = thin_border
            cell_entidad.fill = row_fill
            
            # Columna C: TIPO DE INSUMO
            cell_insumo = ws.cell(row=row, column=3, value=config['insumo'])
            cell_insumo.font = Font(size=10, color="2C3E50")
            cell_insumo.alignment = left_align
            cell_insumo.border = thin_border
            cell_insumo.fill = row_fill
            
            # Columna D: FUENTE
            cell_fuente = ws.cell(row=row, column=4, value=config['fuente'])
            cell_fuente.font = Font(bold=True, size=10, color="27AE60")
            cell_fuente.alignment = center_align
            cell_fuente.border = thin_border
            cell_fuente.fill = row_fill
            
            if archivos_tipo:
                primer_archivo = archivos_tipo[0]
                
                # Columna E: FECHA
                fecha_disp = primer_archivo.fecha_disposicion.strftime('%d/%m/%Y') if primer_archivo.fecha_disposicion else "Sin fecha"
                cell_fecha = ws.cell(row=row, column=5, value=fecha_disp)
                cell_fecha.font = Font(size=10, color="E67E22")
                cell_fecha.alignment = center_align
                cell_fecha.border = thin_border
                cell_fecha.fill = row_fill
                
                # Columna F: RUTA (SIEMPRE al directorio padre)
                directorio = extraer_directorio_desde_ruta(primer_archivo.path_file)
                
                if directorio:
                    nombre_dir = obtener_nombre_directorio_final(directorio)
                    cell_value = f"📁 {nombre_dir}"
                    cell_ruta = ws.cell(row=row, column=6, value=cell_value)
                    cell_ruta.hyperlink = directorio
                    cell_ruta.font = Font(color="0000FF", underline='single', size=10)
                    cell_ruta.alignment = center_align
                    cell_ruta.border = thin_border
                    cell_ruta.fill = row_fill
                else:
                    cell_ruta = ws.cell(row=row, column=6, value="❌ Sin ruta")
                    cell_ruta.font = Font(size=10, color="E74C3C")
                    cell_ruta.alignment = center_align
                    cell_ruta.border = thin_border
                    cell_ruta.fill = row_fill
                
                # Columna G: OBSERVACIONES
                observacion = primer_archivo.observacion or f"✅ {len(archivos_tipo)} archivo(s) disponible(s)"
                cell_obs = ws.cell(row=row, column=7, value=observacion)
                cell_obs.font = Font(size=9, color="34495E")
                cell_obs.alignment = left_align
                cell_obs.border = thin_border
                cell_obs.fill = row_fill
            else:
                # Sin archivos
                cell_fecha = ws.cell(row=row, column=5, value="⏳ Pendiente")
                cell_fecha.font = Font(size=10, color="95A5A6", italic=True)
                cell_fecha.alignment = center_align
                cell_fecha.border = thin_border
                cell_fecha.fill = row_fill
                
                cell_ruta = ws.cell(row=row, column=6, value="📭 Vacío")
                cell_ruta.font = Font(size=10, color="95A5A6", italic=True)
                cell_ruta.alignment = center_align
                cell_ruta.border = thin_border
                cell_ruta.fill = row_fill
                
                cell_obs = ws.cell(row=row, column=7, value=f"⚠️ No se encontraron archivos")
                cell_obs.font = Font(size=9, color="E74C3C", italic=True)
                cell_obs.alignment = left_align
                cell_obs.border = thin_border
                cell_obs.fill = row_fill
            
            # Columna H: USUARIO
            cell_usuario = ws.cell(row=row, column=8, value="🤖 Sistema Automatizado")
            cell_usuario.font = Font(size=10, color="8E44AD")
            cell_usuario.alignment = center_align
            cell_usuario.border = thin_border
            cell_usuario.fill = row_fill
            
            filas_categoria += 1
            row += 1
        
        filas_por_categoria[categoria] = filas_categoria
        
        # Combinar celdas de categoría si hay más de una fila
        if filas_categoria > 1:
            ws.merge_cells(f'A{primera_fila_categoria[categoria]}:A{primera_fila_categoria[categoria] + filas_categoria - 1}')
    
    # 📐 AJUSTAR DIMENSIONES (actualizado para 8 columnas)
    anchos_columnas = [25, 20, 45, 15, 18, 30, 40, 18]
    
    for i, ancho in enumerate(anchos_columnas):
        column_letter = get_column_letter(i + 1)
        ws.column_dimensions[column_letter].width = ancho
    
    # Altura de filas
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[6].height = 40
    
    for row_num in range(7, row):
        ws.row_dimensions[row_num].height = 35
    
    print(f"✅ Pestaña Matriz Primarios generada con columna CATEGORÍA para {mecanismo_financiacion}")

# =============================================
# FUNCIÓN 3: MATRIZ RESUMEN CON RUTAS INTELIGENTES
# =============================================

def generar_matriz_resumen_preoperacion(municipios_mecanismos_list, temp_dir):
    """
    Genera el REPORTE MATRIZ RESUMEN con MECANISMOS DE FINANCIACIÓN
    ✅ CORREGIDA: Manejo correcto de tuplas (municipio, mecanismo)
    """
    try:
        #print(f"📋 Generando MATRIZ RESUMEN con mecanismos para {len(municipios_mecanismos_list)} combinaciones...")
        
        # Crear workbook para resumen
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matriz Resumen"
        
        # 🎨 ESTILOS (mantener los existentes)
        titulo_font = Font(bold=True, size=16, color="FFFFFF")
        titulo_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="5BA0F2", end_color="5BA0F2", fill_type="solid")
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        hyperlink_font = Font(color='0000FF', underline='single')
        ruta_generica_font = Font(color='0066CC', underline='single')
        
        # 🏛️ TÍTULO
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        ws.merge_cells('A1:Q2')
        titulo_cell = ws['A1']
        titulo_cell.value = f"MATRIZ RESUMEN PRE-OPERACIÓN CON MECANISMOS DE FINANCIACIÓN\n{len(municipios_mecanismos_list)} COMBINACIONES MUNICIPIO-MECANISMO - Generado: {fecha_actual}"
        titulo_cell.font = titulo_font
        titulo_cell.fill = titulo_fill
        titulo_cell.alignment = center_align
        titulo_cell.border = border
        
        # 📊 HEADERS - CON NUEVA COLUMNA DE MECANISMO
        row = 4
        headers = [
            'MUNICIPIO', 'CÓDIGO DANE', 'MECANISMO FINANCIACIÓN',
            'ALCANCE OPERACIÓN', 
            'INSUMOS DE ORDENAMIENTO TERRITORIAL\nPOT, PBOT, EOT',
            'BASE DE DATOS CATASTRAL\n(R1y R2)',
            'BASE CATASTRAL DEL MUNICIPIO\nGRÁFICA', 
            'ORTOIMAGENES\n(URBANA Y RURAL)',
            'BASE CARTOGRÁFICA\n(URBANA Y RURAL)',
            'CONSTRUCCIONES\n(CARTOGRAFÍA VECTORIAL)',
            'CONSTRUCCIONES (CLASIFICACIÓN SUPERVISADA)/\nCONSTRUCCIONES (IMÁGENES DE BING MAPS)',
            'DEM / CURVAS DE NIVEL',
            'LIMITES MUNICIPALES',
            'HISTORICO DE ANOTACIONES DE LA BASE\nREGISTRAL DEL MUNICIPIO',
            'INSUMOS SECUNDARIOS',
            'OBSERVACIÓN',
            'DILIGENCIO'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # 📋 DATOS DE CADA COMBINACIÓN MUNICIPIO-MECANISMO
        row = 5
        
        # Patrones de búsqueda (mantener los existentes)
        patrones_busqueda_config = [
            {'patron': '08_inst_ord_terri', 'es_generico': False, 'descripcion': 'POT/PBOT/EOT'},
            {'patron': '03_info_catas\\01_r1_r2', 'es_generico': False, 'descripcion': 'R1 y R2'},
            {'patron': '03_info_catas\\02_gdb', 'es_generico': True, 'descripcion': 'Base catastral gráfica'},
            {'patron': '01_carto_basic\\01_rast\\01_orto', 'es_generico': True, 'descripcion': 'Ortoimágenes (genérico urbana/rural)'},
            {'patron': '01_carto_basic\\02_vect', 'es_generico': True, 'descripcion': 'Base cartográfica (genérico urbana/rural)'},
            {'patron': '03_info_catas\\03_tab_terr_constr', 'es_generico': False, 'descripcion': 'Construcciones vectorial'},
            {'patron': '03_info_catas\\04_estu_zhf_zhg', 'es_generico': False, 'descripcion': 'Construcciones clasificadas'},
            {'patron': '01_carto_basic\\01_rast\\02_dtm', 'es_generico': False, 'descripcion': 'DEM'},
            {'patron': '04_deslin\\01_shp', 'es_generico': False, 'descripcion': 'Límites municipales'},
            {'patron': '06_insu_regis', 'es_generico': False, 'descripcion': 'Histórico anotaciones'},
            {'patron': '07_insu_fuente_secun', 'es_generico': False, 'descripcion': 'Insumos secundarios'}
        ]
        
        # ✅ CORREGIDO: Iterar correctamente sobre las tuplas
        for numero, tupla_municipio_mecanismo in enumerate(municipios_mecanismos_list, 1):
            # ✅ DESEMPAQUETAR CORRECTAMENTE LA TUPLA
            municipio_obj, mecanismo_str = tupla_municipio_mecanismo
            
            #print(f"🔄 Procesando fila {numero}: {municipio_obj.nom_municipio} - {mecanismo_str}")
            
            # Columna A: Municipio
            ws.cell(row=row, column=1, value=municipio_obj.nom_municipio).border = border
            
            # Columna B: Código DANE
            ws.cell(row=row, column=2, value=municipio_obj.cod_municipio).border = border
            
            # Columna C: Mecanismo de Financiación
            ws.cell(row=row, column=3, value=mecanismo_str).border = border
            
            # Columna D: Alcance Operación
            ws.cell(row=row, column=4, value=municipio_obj.alcance_operacion or "N/A").border = border
            
            # 🔥 Obtener archivos FILTRADOS por mecanismo
            archivos_municipio = ListaArchivosPre.objects.filter(
                cod_insumo__cod_insumo__cod_municipio=municipio_obj.cod_municipio
            )
            
            # Filtrar por mecanismo
            archivos_filtrados = []
            for archivo in archivos_municipio:
                mecanismo_archivo = extraer_mecanismo_financiacion(archivo.path_file, municipio_obj.cod_municipio)
                if mecanismo_archivo == mecanismo_str:
                    archivos_filtrados.append(archivo)
            
            #print(f"   📊 Archivos filtrados para {mecanismo_str}: {len(archivos_filtrados)}")
            
            # Procesar patrones con archivos filtrados (columnas E a O)
            for i, patron_config in enumerate(patrones_busqueda_config):
                col_num = 5 + i  # Empezar en columna E (5)
                patron = patron_config['patron']
                es_generico = patron_config['es_generico']

                # 🔧 NORMALIZAR: Convertir patrón y rutas a formato / para coincidir con rutas Linux en BD
                patron_normalizado = patron.replace('\\', '/')
                archivos_patron = [a for a in archivos_filtrados if a.path_file and patron_normalizado in a.path_file.replace('\\', '/')]
                
                cell = ws.cell(row=row, column=col_num)
                cell.border = border
                cell.alignment = center_align
                
                if archivos_patron:
                    primer_archivo = archivos_patron[0]
                    directorio = extraer_directorio_desde_ruta(primer_archivo.path_file)
                    
                    if directorio:
                        nombre_dir = obtener_nombre_directorio_final(directorio)
                        
                        if es_generico:
                            cell.value = f"🌐 {nombre_dir}"
                            cell.font = ruta_generica_font
                        else:
                            cell.value = nombre_dir
                            cell.font = hyperlink_font
                        
                        cell.hyperlink = directorio
                        
                        #print(f"   ✅ {patron}: {nombre_dir}")
                    else:
                        cell.value = "Sin ruta"
                        cell.font = Font(color='E74C3C')
                        #print(f"   ⚠️ {patron}: Sin ruta válida")
                else:
                    cell.value = "Vacío"
                    cell.font = Font(color='95A5A6')
                    #print(f"   📭 {patron}: Vacío")
            
            # Columna P: Observación
            observacion_cell = ws.cell(row=row, column=16, value=f"Mecanismo: {mecanismo_str}")
            observacion_cell.border = border
            
            # Columna Q: Diligencio
            diligencio_cell = ws.cell(row=row, column=17, value="sistema_reportes_web")
            diligencio_cell.border = border
            
            row += 1
            
            if numero % 5 == 0:
                print(f"Procesadas {numero} combinaciones municipio-mecanismo...")
        
        # 📏 AJUSTAR DIMENSIONES (ampliado para nueva columna)
        anchos = [25, 12, 20, 15, 20, 20, 20, 25, 25, 25, 35, 20, 20, 30, 20, 15, 20]
        
        for i, ancho in enumerate(anchos, 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = ancho
        
        # Altura de filas
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[4].height = 40
        
        for row_num in range(5, row):
            ws.row_dimensions[row_num].height = 25
        
        # 💾 GUARDAR ARCHIVO RESUMEN
        nombre_resumen = f"Matriz_Resumen_Preoperacion_CON_MECANISMOS_{datetime.now().strftime('%d_%m_%Y_%H%M%S')}.xlsx"
        archivo_path = os.path.join(temp_dir, nombre_resumen)
        wb.save(archivo_path)
        
        print(f"✅ MATRIZ RESUMEN CON MECANISMOS generado: {nombre_resumen}")
        return archivo_path
        
    except Exception as e:
        print(f"Error generando matriz resumen con mecanismos: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


# ============================================================================
# REPORTE PRE-OPERACION COMPLETO (TODOS LOS 8 DIRECTORIOS)
# ============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generar_reporte_preoperacion_completo(request):
    """
    Genera un reporte Excel completo de Pre-Operacion con todos los 8 directorios:
    - 01_prop (Propuesta)
    - 02_carta_acept (Carta de Aceptacion)
    - 03_cto_modif (Contrato/Modificacion)
    - 04_acta_ini (Acta de Inicio)
    - 05_plan_gest_proy (Plan de Gestion del Proyecto)
    - 06_precono (Pre-Conocimiento)
    - 07_insu (Insumos)
    - 08_contr_pers (Control de Personal)
    """
    try:
        data = json.loads(request.body) if request.body else request.data
        municipios_ids = data.get('municipios', [])

        if not municipios_ids:
            return Response(
                {'error': 'No se proporcionaron municipios'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar permisos
        municipios_permitidos = get_municipios_permitidos(request.user)
        if municipios_permitidos != 'todos':
            municipios_ids = [m for m in municipios_ids if m in municipios_permitidos]
            if not municipios_ids:
                return Response(
                    {'error': 'No tiene permisos para acceder a estos municipios'},
                    status=status.HTTP_403_FORBIDDEN
                )

        with tempfile.TemporaryDirectory() as temp_dir:
            archivos_generados = []

            # Generar reporte individual por municipio
            for municipio_id in municipios_ids:
                try:
                    archivo_path = _generar_excel_preoperacion_completo(municipio_id, temp_dir)
                    if archivo_path:
                        archivos_generados.append(archivo_path)
                except Exception as e:
                    print(f"Error procesando municipio {municipio_id}: {str(e)}")
                    continue

            # Generar resumen consolidado si hay mas de un municipio
            if len(municipios_ids) > 1:
                try:
                    resumen_path = _generar_resumen_preoperacion_completo(municipios_ids, temp_dir)
                    if resumen_path:
                        archivos_generados.append(resumen_path)
                except Exception as e:
                    print(f"Error generando resumen: {str(e)}")

            if not archivos_generados:
                return Response(
                    {'error': 'No se pudieron generar reportes'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            # Crear ZIP en archivo temporal para streaming
            fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_zip = f'reportes_preoperacion_completo_{fecha_actual}.zip'

            temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
            temp_zip_path = temp_zip.name
            temp_zip.close()

            with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=1) as zip_file:
                for archivo_path in archivos_generados:
                    nombre_archivo = os.path.basename(archivo_path)
                    zip_file.write(archivo_path, nombre_archivo)

            # Streaming del ZIP
            def zip_iterator():
                CHUNK_SIZE = 8 * 1024 * 1024  # 8MB chunks
                try:
                    with open(temp_zip_path, 'rb') as f:
                        while True:
                            chunk = f.read(CHUNK_SIZE)
                            if not chunk:
                                break
                            yield chunk
                finally:
                    try:
                        os.unlink(temp_zip_path)
                    except:
                        pass

            response = StreamingHttpResponse(
                zip_iterator(),
                content_type='application/zip'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_zip}"'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['X-Accel-Buffering'] = 'no'  # Desactivar buffering en nginx

            return response

    except Exception as e:
        return Response(
            {'error': f'Error interno: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def _extraer_directorio_nivel1_de_ruta(ruta):
    """
    Extrae el directorio de nivel 1 (hijo directo de 01_preo) de una ruta.
    Ej: .../01_preo/05_plan_gest_proy/01_plan_opera/archivo.pdf → 05_plan_gest_proy
    Ej: .../01_preo/01_prop/01_prop/archivo.pdf → 01_prop
    """
    if not ruta:
        return 'N/A'
    try:
        ruta_str = str(ruta).replace('\\', '/')
        partes = [p for p in ruta_str.split('/') if p]
        # Buscar '01_preo' o variantes en las partes
        for i, parte in enumerate(partes):
            if parte.lower() in ('01_preo', '01_pre'):
                # El siguiente elemento es el directorio nivel 1
                if i + 1 < len(partes) - 1:  # -1 para no tomar el archivo final
                    return partes[i + 1]
        return 'N/A'
    except:
        return 'N/A'


def _generar_excel_preoperacion_completo(municipio_id, temp_dir):
    """
    Genera un Excel profesional con inventario detallado de Pre-Operacion.
    Formato identico al de Operacion/Transversal.
    """
    try:
        municipio = Municipios.objects.select_related('cod_depto').get(cod_municipio=municipio_id)
        print(f"Generando reporte PRE-OPERACION COMPLETO para {municipio.nom_municipio}...")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario Pre-Operacion"

        # ESTILOS PROFESIONALES EN TONOS CYAN/TEAL (PRE-OPERACION)
        titulo_font = Font(bold=True, size=16, color="FFFFFF")
        titulo_fill = PatternFill(start_color="006064", end_color="006064", fill_type="solid")  # Cyan oscuro

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="00838F", end_color="00838F", fill_type="solid")  # Cyan medio
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # TITULO PRINCIPAL
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        depto_nombre = municipio.cod_depto.nom_depto if municipio.cod_depto else 'N/A'

        ws.merge_cells('A1:I2')
        titulo_cell = ws['A1']
        titulo_cell.value = (
            f"INVENTARIO DETALLADO DE ARCHIVOS PRE-OPERACION\n"
            f"{municipio.nom_municipio.upper()} | {depto_nombre} | {fecha_actual}"
        )
        titulo_cell.font = titulo_font
        titulo_cell.fill = titulo_fill
        titulo_cell.alignment = center_align
        titulo_cell.border = border

        # Recolectar TODOS los archivos de los 8 directorios
        todos_archivos = []

        # ============================================================
        # PASO 1: Obtener TODOS los archivos de ArchivoPreoperacion
        # para el municipio (directorios 01-06, 08)
        # ============================================================
        archivos_preo_todos = ArchivoPreoperacion.objects.filter(
            cod_directorio__cod_mpio=municipio_id
        ).select_related('cod_directorio')

        for archivo in archivos_preo_todos:
            ruta_str = archivo.ruta_archivo or ''
            # Extraer el directorio padre de nivel 1 desde la ruta
            # Ej: .../01_preo/05_plan_gest_proy/01_plan_opera/archivo.pdf → 05_plan_gest_proy
            directorio_padre = _extraer_directorio_nivel1_de_ruta(ruta_str)

            ruta_archivo_win = linux_to_windows_path(ruta_str) if ruta_str else 'N/A'
            ruta_directorio_win = linux_to_windows_path(archivo.cod_directorio.ruta_directorio) if archivo.cod_directorio and archivo.cod_directorio.ruta_directorio else 'N/A'
            todos_archivos.append({
                'etapa': directorio_padre,
                'subcarpeta': archivo.cod_directorio.nom_directorio if archivo.cod_directorio else 'N/A',
                'nombre': archivo.nom_archivo or 'N/A',
                'ruta_archivo': ruta_archivo_win,
                'ruta_directorio': ruta_directorio_win,
                'extension': (archivo.extension or '').upper().replace('.', ''),
                'fecha': archivo.fecha_modificacion.strftime('%d/%m/%Y %H:%M:%S') if archivo.fecha_modificacion else 'Sin fecha',
                'peso': f"{int(archivo.tamano_bytes):,}".replace(',', '.') + " bytes" if archivo.tamano_bytes else "0 bytes",
            })

        # ============================================================
        # PASO 2: Obtener archivos de insumos (07_insu) desde ListaArchivosPre
        # ============================================================
        insumos = Insumos.objects.filter(cod_municipio=municipio_id)
        clasificaciones = ClasificacionInsumo.objects.filter(cod_insumo__in=insumos)
        archivos_insu = ListaArchivosPre.objects.filter(cod_insumo__in=clasificaciones).select_related('cod_insumo')

        for archivo in archivos_insu:
            nombre_archivo = os.path.basename(archivo.path_file) if archivo.path_file else archivo.nombre_insumo or 'N/A'
            extension = os.path.splitext(archivo.path_file)[1] if archivo.path_file else ''
            ruta_archivo_win = linux_to_windows_path(archivo.path_file) if archivo.path_file else 'N/A'
            ruta_directorio_win = linux_to_windows_path(os.path.dirname(archivo.path_file)) if archivo.path_file else 'N/A'
            # Extraer directorio nivel 1 de la ruta del insumo
            directorio_padre = _extraer_directorio_nivel1_de_ruta(archivo.path_file or '')
            todos_archivos.append({
                'etapa': directorio_padre if directorio_padre != 'N/A' else '07_insu',
                'subcarpeta': archivo.cod_insumo.nombre if archivo.cod_insumo else 'N/A',
                'nombre': nombre_archivo,
                'ruta_archivo': ruta_archivo_win,
                'ruta_directorio': ruta_directorio_win,
                'extension': extension.upper().replace('.', '') if extension else 'N/A',
                'fecha': archivo.fecha_disposicion.strftime('%d/%m/%Y %H:%M:%S') if archivo.fecha_disposicion else 'Sin fecha',
                'peso': f"{int(archivo.peso_memoria):,}".replace(',', '.') + " bytes" if archivo.peso_memoria else "0 bytes",
            })

        # ESTADISTICAS RAPIDAS
        ws.merge_cells('A3:I3')
        stats_cell = ws['A3']
        stats_cell.value = f"RESUMEN: {len(todos_archivos)} archivos totales de Pre-Operacion"
        stats_cell.font = Font(bold=True, size=11, color="004D40")  # Cyan texto oscuro
        stats_cell.fill = PatternFill(start_color="B2EBF2", end_color="B2EBF2", fill_type="solid")  # Cyan claro
        stats_cell.alignment = center_align
        stats_cell.border = border

        # HEADERS DE LA TABLA (ROW 5)
        row = 5
        headers = [
            'ETAPA',
            'SUBCARPETA',
            'NOMBRE DOCUMENTO',
            'RUTA COMPLETA ARCHIVO',
            'RUTA DIRECTORIO',
            'FORMATO\nDOCUMENTO',
            'FECHA DE CREACION\nO MODIFICACION',
            'TAMANO',
            'OBSERVACIONES'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # DATOS - TODOS LOS ARCHIVOS (ordenados por directorio y nombre)
        todos_archivos.sort(key=lambda x: (x['etapa'], x['nombre']))
        row = 6
        for numero, archivo in enumerate(todos_archivos, 1):
            # Alternar colores de fila en tonos cyan
            if numero % 2 == 0:
                row_fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")  # Cyan muy claro
            else:
                row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Col A: ETAPA
            cell = ws.cell(row=row, column=1, value=archivo['etapa'])
            cell.alignment = center_align
            cell.border = border
            cell.fill = row_fill
            cell.font = Font(size=9, bold=True, color="006064")  # Cyan oscuro

            # Col B: SUBCARPETA
            cell = ws.cell(row=row, column=2, value=archivo['subcarpeta'])
            cell.alignment = left_align
            cell.border = border
            cell.fill = row_fill
            cell.font = Font(size=9)

            # Col C: NOMBRE DOCUMENTO
            cell = ws.cell(row=row, column=3, value=archivo['nombre'])
            cell.alignment = left_align
            cell.border = border
            cell.fill = row_fill
            cell.font = Font(size=9)

            # Col D: RUTA COMPLETA ARCHIVO
            cell = ws.cell(row=row, column=4, value=archivo['ruta_archivo'])
            cell.alignment = left_align
            cell.border = border
            cell.fill = row_fill
            cell.font = Font(size=8, color="00838F")  # Cyan medio

            # Col E: RUTA DIRECTORIO
            cell = ws.cell(row=row, column=5, value=archivo['ruta_directorio'])
            cell.alignment = left_align
            cell.border = border
            cell.fill = row_fill
            cell.font = Font(size=8, color="00838F")  # Cyan medio

            # Col F: FORMATO
            cell = ws.cell(row=row, column=6, value=archivo['extension'])
            cell.alignment = center_align
            cell.border = border
            cell.fill = row_fill
            cell.font = Font(size=9, bold=True, color="004D40")  # Cyan oscuro

            # Col G: FECHA
            cell = ws.cell(row=row, column=7, value=archivo['fecha'])
            cell.alignment = center_align
            cell.border = border
            cell.fill = row_fill
            cell.font = Font(size=9)

            # Col H: TAMANO
            cell = ws.cell(row=row, column=8, value=archivo['peso'])
            cell.alignment = center_align
            cell.border = border
            cell.fill = row_fill
            cell.font = Font(size=9, color="00838F")  # Cyan medio

            # Col I: OBSERVACIONES (vacio)
            cell = ws.cell(row=row, column=9, value="")
            cell.alignment = left_align
            cell.border = border
            cell.fill = row_fill

            row += 1

        # AJUSTAR ANCHOS DE COLUMNAS
        from openpyxl.utils import get_column_letter
        anchos = [18, 50, 45, 80, 80, 12, 22, 18, 30]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        # Ajustar alturas
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[5].height = 35

        # Guardar archivo
        fecha_str = datetime.now().strftime('%d_%m_%Y')
        municipio_limpio = municipio.nom_municipio.replace(' ', '_').replace('/', '_').upper()
        nombre_archivo = f"{municipio.cod_municipio}_{municipio_limpio}_PreOperacion_Completo_{fecha_str}.xlsx"
        archivo_path = os.path.join(temp_dir, nombre_archivo)
        wb.save(archivo_path)

        print(f"Reporte PRE-OPERACION generado: {nombre_archivo} con {len(todos_archivos)} archivos")
        return archivo_path

    except Exception as e:
        print(f"Error generando Excel para municipio {municipio_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def _crear_hoja_directorio(wb, municipio_id, codigo, descripcion, header_font, header_fill, border, center_align, left_align, alt_fill):
    """
    Crea una hoja en el Excel con el detalle de archivos de un directorio.
    """
    # Nombre de hoja (max 31 caracteres)
    nombre_hoja = codigo[:31]
    ws = wb.create_sheet(title=nombre_hoja)

    ws['A1'] = f"Directorio: {codigo} - {descripcion}"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:E1')

    row = 3
    if codigo == '07_insu':
        # Cabeceras para insumos
        headers = ['Clasificacion', 'Nombre Archivo', 'Extension', 'Ruta']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        row += 1

        insumos = Insumos.objects.filter(cod_municipio=municipio_id)
        clasificaciones = ClasificacionInsumo.objects.filter(cod_insumo__in=insumos)
        archivos = ListaArchivosPre.objects.filter(cod_insumo__in=clasificaciones).select_related('cod_insumo')

        for archivo in archivos[:1000]:  # Limitar a 1000 archivos por hoja
            clasificacion_nombre = archivo.cod_insumo.nombre if archivo.cod_insumo else 'N/A'
            # Extraer nombre del archivo desde path_file
            nombre_archivo = os.path.basename(archivo.path_file) if archivo.path_file else archivo.nombre_insumo or 'N/A'
            # Extraer extension desde path_file
            extension = os.path.splitext(archivo.path_file)[1] if archivo.path_file else 'N/A'
            ws.cell(row=row, column=1, value=clasificacion_nombre).alignment = left_align
            ws.cell(row=row, column=2, value=nombre_archivo).alignment = left_align
            ws.cell(row=row, column=3, value=extension).alignment = center_align
            ws.cell(row=row, column=4, value=archivo.path_file or 'N/A').alignment = left_align

            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
                if row % 2 == 0:
                    ws.cell(row=row, column=col).fill = alt_fill

            row += 1
    else:
        # Cabeceras para directorios indexados
        headers = ['Subdirectorio', 'Nombre Archivo', 'Extension', 'Tamano', 'Ruta']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        row += 1

        directorios = DirectorioPreoperacion.objects.filter(
            cod_mpio=municipio_id,
            nom_directorio__startswith=codigo
        )
        archivos = ArchivoPreoperacion.objects.filter(
            cod_directorio__in=directorios
        ).select_related('cod_directorio')[:1000]

        for archivo in archivos:
            subdir_nombre = archivo.cod_directorio.nom_directorio if archivo.cod_directorio else 'N/A'
            tamano = f"{archivo.tamano_bytes / 1024:.1f} KB" if archivo.tamano_bytes else 'N/A'

            ws.cell(row=row, column=1, value=subdir_nombre).alignment = left_align
            ws.cell(row=row, column=2, value=archivo.nom_archivo or 'N/A').alignment = left_align
            ws.cell(row=row, column=3, value=archivo.extension or 'N/A').alignment = center_align
            ws.cell(row=row, column=4, value=tamano).alignment = center_align
            ws.cell(row=row, column=5, value=archivo.ruta_relativa or 'N/A').alignment = left_align

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
                if row % 2 == 0:
                    ws.cell(row=row, column=col).fill = alt_fill

            row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50


def _generar_resumen_preoperacion_completo(municipios_ids, temp_dir):
    """
    Genera un Excel de resumen consolidado de multiples municipios.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen Consolidado"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alt_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")

        # Encabezado
        ws.merge_cells('A1:K1')
        ws['A1'] = f"RESUMEN CONSOLIDADO PRE-OPERACION - {len(municipios_ids)} MUNICIPIOS"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_align

        ws['A2'] = f"Fecha de generacion: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10)

        # Cabeceras
        row = 4
        headers = ['Codigo', 'Municipio', 'Departamento', '01_prop', '02_carta', '03_cto', '04_acta', '05_plan', '06_precono', '07_insu', '08_contr']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        row += 1

        for mpio_id in municipios_ids:
            try:
                municipio = Municipios.objects.select_related('cod_depto').get(cod_municipio=mpio_id)
                depto_nombre = municipio.cod_depto.nom_depto if municipio.cod_depto else 'N/A'

                ws.cell(row=row, column=1, value=municipio.cod_municipio)
                ws.cell(row=row, column=2, value=municipio.nom_municipio)
                ws.cell(row=row, column=3, value=depto_nombre)

                # Contar archivos por directorio
                directorios_codes = ['01_prop', '02_carta_acept', '03_cto_modif', '04_acta_ini', '05_plan_gest_proy', '06_precono', '07_insu', '08_contr_pers']

                for col_idx, codigo in enumerate(directorios_codes, 4):
                    if codigo == '07_insu':
                        insumos = Insumos.objects.filter(cod_municipio=mpio_id)
                        clasificaciones = ClasificacionInsumo.objects.filter(cod_insumo__in=insumos)
                        count = ListaArchivosPre.objects.filter(cod_insumo__in=clasificaciones).count()
                    else:
                        directorios = DirectorioPreoperacion.objects.filter(
                            cod_mpio=mpio_id,
                            nom_directorio__startswith=codigo
                        )
                        count = ArchivoPreoperacion.objects.filter(cod_directorio__in=directorios).count()

                    ws.cell(row=row, column=col_idx, value=count)

                for col in range(1, 12):
                    ws.cell(row=row, column=col).border = border
                    ws.cell(row=row, column=col).alignment = center_align
                    if row % 2 == 0:
                        ws.cell(row=row, column=col).fill = alt_fill

                row += 1

            except Municipios.DoesNotExist:
                continue

        # Ajustar anchos
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col].width = 10

        # Guardar
        nombre_archivo = f"Resumen_PreOperacion_Consolidado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        archivo_path = os.path.join(temp_dir, nombre_archivo)
        wb.save(archivo_path)

        return archivo_path

    except Exception as e:
        print(f"Error generando resumen consolidado: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


# FUNCION DEBUG OPCIONAL
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def debug_datos_preoperacion(request):
    """Vista temporal para verificar estructura de datos de preoperación"""
    
    municipio_id = request.GET.get('municipio', 17174)
    
    try:
        municipio = Municipios.objects.get(cod_municipio=municipio_id)
        
        # Estadísticas
        insumos = Insumos.objects.filter(cod_municipio=municipio_id)
        clasificaciones = ClasificacionInsumo.objects.filter(cod_insumo__in=insumos)
        archivos = ListaArchivosPre.objects.filter(cod_insumo__in=clasificaciones)
        
        debug_info = {
            'municipio': {
                'codigo': municipio.cod_municipio,
                'nombre': municipio.nom_municipio
            },
            'totales': {
                'insumos': insumos.count(),
                'clasificaciones': clasificaciones.count(),
                'archivos': archivos.count()
            },
            'insumos_muestra': [
                {
                    'id': ins.cod_insumo,
                    'categoria': ins.cod_categoria.nom_categoria if ins.cod_categoria else None,
                    'tipo': ins.tipo_insumo
                } for ins in insumos[:3]
            ],
            'clasificaciones_muestra': [
                {
                    'id': clas.cod_clasificacion,
                    'nombre': clas.nombre,
                    'insumo_id': clas.cod_insumo.cod_insumo
                } for clas in clasificaciones[:3]
            ]
        }
        
        return Response(debug_info)
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)
    




# ✅ FUNCIÓN DEBUG PARA VER CAMPOS DISPONIBLES EN ARCHIVOSPOST
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def debug_campos_archivos_post(request):
    """
    Función debug para ver qué campos tiene el modelo ArchivosPost
    """
    try:
        # Obtener información del modelo
        campos_modelo = [field.name for field in ArchivosPost._meta.get_fields()]
        
        # Intentar obtener un registro de muestra
        archivo_muestra = ArchivosPost.objects.first()
        campos_con_datos = {}
        
        if archivo_muestra:
            for campo in campos_modelo:
                try:
                    valor = getattr(archivo_muestra, campo)
                    campos_con_datos[campo] = {
                        'tipo': str(type(valor).__name__),
                        'valor_muestra': str(valor)[:50] if valor else 'None'
                    }
                except Exception as e:
                    campos_con_datos[campo] = {'error': str(e)}
        
        return Response({
            'modelo': 'ArchivosPost',
            'total_registros': ArchivosPost.objects.count(),
            'campos_disponibles': campos_modelo,
            'campos_con_datos': campos_con_datos
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)


# REEMPLAZAR COMPLETAMENTE las funciones en preoperacion/views.py

# ✅ LISTA DE USUARIOS ESPECÍFICOS (NOMBRES REALES DE TU BD)
USUARIOS_INFORME_ESPECIAL = [
    'felipe.vargas',
    'elizabeth.rosas',
    'elizabeth.eraso',

]

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generar_informe_usuarios_especificos(request):
    """
    Genera informe PROFESIONAL EJECUTIVO con dashboard analítico y pestañas detalladas
    """
    try:
        # Validación de usuario autorizado
        usuario_actual = request.user.username
        usuarios_autorizados = ['andres.osorio', 'elizabeth.eraso']
        
        if usuario_actual not in usuarios_autorizados:
            return Response(
                {'error': 'No tiene permisos para generar este informe'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Obtener parámetros
        try:
            data = json.loads(request.body) if request.body else request.data
        except json.JSONDecodeError:
            data = request.data
            
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')
        
        if not fecha_desde or not fecha_hasta:
            return Response(
                {'error': 'Se requieren fecha_desde y fecha_hasta'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        print(f"📊 GENERANDO DASHBOARD EJECUTIVO COMPLETO")
        print(f"📅 Rango: {fecha_desde} a {fecha_hasta}")
        
        # Generar informe completo
        with tempfile.TemporaryDirectory() as temp_dir:
            archivo_excel = generar_dashboard_ejecutivo_completo(fecha_desde, fecha_hasta, temp_dir)
            
            if not archivo_excel:
                return Response(
                    {'error': 'No se pudo generar el informe'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            # Leer y retornar archivo
            with open(archivo_excel, 'rb') as file:
                excel_data = file.read()
            
            response = HttpResponse(
                excel_data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            nombre_archivo = f'Dashboard_Ejecutivo_{fecha_desde}_{fecha_hasta}_{usuario_actual}_{datetime.now().strftime("%Y%m%dT%H%M")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            response['Content-Length'] = len(excel_data)
            
            print(f"✅ DASHBOARD EJECUTIVO generado: {nombre_archivo}")
            return response
                
    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response(
            {'error': f'Error interno: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

def generar_dashboard_ejecutivo_completo(fecha_desde, fecha_hasta, temp_dir):
    """
    Genera el DASHBOARD EJECUTIVO COMPLETO
    """
    try:
        print(f"🎯 Iniciando generación de DASHBOARD EJECUTIVO...")
        
        # Recopilar datos
        datos_completos = recopilar_datos_completos(fecha_desde, fecha_hasta)
        
        if not datos_completos['tiene_datos']:
            print("⚠️ No se encontraron datos")
            return None
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Dashboard principal
        ws_dashboard = wb.active
        ws_dashboard.title = "Dashboard Ejecutivo"
        generar_dashboard_principal(ws_dashboard, datos_completos, fecha_desde, fecha_hasta)
        
        # Pestañas individuales
        for usuario_data in datos_completos['usuarios_con_datos']:
            usuario = usuario_data['usuario']
            #print(f"🎨 Generando pestaña para {usuario}...")
            
            nombre_pestaña = usuario.replace('.', '_')[:31]
            ws_usuario = wb.create_sheet(title=nombre_pestaña)
            generar_analisis_usuario(ws_usuario, usuario_data, fecha_desde, fecha_hasta)
        
        # Guardar archivo
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"Dashboard_Ejecutivo_{fecha_str}.xlsx"
        archivo_path = os.path.join(temp_dir, nombre_archivo)
        wb.save(archivo_path)
        
        print(f"✅ Dashboard completo generado: {nombre_archivo}")
        return archivo_path
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def recopilar_datos_completos(fecha_desde, fecha_hasta):
    """
    Recopila TODOS los datos necesarios
    """
    try:
        datos = {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'usuarios_con_datos': [],
            'resumen_global': {
                'total_archivos_pre': 0,
                'total_archivos_post': 0,
                'total_municipios': set(),
                'total_categorias': set(),
                'total_componentes': set(),
                'actividad_diaria_global': defaultdict(int),
                'municipios_frecuencia': defaultdict(int),
                'categorias_frecuencia': defaultdict(int),
                'componentes_frecuencia': defaultdict(int),
                'usuarios_productividad': []
            },
            'kpis': {
                'promedio_archivos_usuario': 0,
                'cobertura_territorial': 0,
                'eficiencia_operacional': 0
            },
            'tiene_datos': False
        }
        
        # Procesar cada usuario
        for usuario in USUARIOS_INFORME_ESPECIAL:
            usuario_data = procesar_usuario_completo(usuario, fecha_desde, fecha_hasta)
            
            if usuario_data['tiene_actividad']:
                datos['usuarios_con_datos'].append(usuario_data)
                datos['tiene_datos'] = True
                
                # Acumular estadísticas
                datos['resumen_global']['total_archivos_pre'] += usuario_data['stats']['total_archivos_pre']
                datos['resumen_global']['total_archivos_post'] += usuario_data['stats']['total_archivos_post']
                
                # Municipios, categorías, componentes
                datos['resumen_global']['total_municipios'].update(usuario_data['municipios_trabajados'].keys())
                datos['resumen_global']['total_categorias'].update(usuario_data['categorias_trabajadas'].keys())
                datos['resumen_global']['total_componentes'].update(usuario_data['componentes_trabajados'].keys())
                
                # Actividad diaria global
                for fecha, cantidad in usuario_data['actividad_diaria'].items():
                    datos['resumen_global']['actividad_diaria_global'][fecha] += cantidad
                
                # Frecuencias para gráficas
                for mun_id, mun_data in usuario_data['municipios_trabajados'].items():
                    total_archivos_mun = mun_data['archivos_pre'] + mun_data['archivos_post']
                    datos['resumen_global']['municipios_frecuencia'][mun_id] += total_archivos_mun
                
                for cat_id, cantidad in usuario_data['categorias_detalle'].items():
                    datos['resumen_global']['categorias_frecuencia'][cat_id] += cantidad
                
                for comp_id, cantidad in usuario_data['componentes_detalle'].items():
                    datos['resumen_global']['componentes_frecuencia'][comp_id] += cantidad
                
                # Productividad
                total_archivos = usuario_data['stats']['total_archivos_pre'] + usuario_data['stats']['total_archivos_post']
                datos['resumen_global']['usuarios_productividad'].append({
                    'usuario': usuario,
                    'total_archivos': total_archivos,
                    'dias_activos': usuario_data['stats']['dias_activos'],
                    'municipios': usuario_data['stats']['municipios_unicos'],
                    'promedio_diario': usuario_data['stats']['promedio_archivos_dia']
                })
        
        # Calcular KPIs
        if datos['usuarios_con_datos']:
            total_usuarios = len(datos['usuarios_con_datos'])
            total_archivos_global = datos['resumen_global']['total_archivos_pre'] + datos['resumen_global']['total_archivos_post']
            
            datos['kpis']['promedio_archivos_usuario'] = round(total_archivos_global / total_usuarios, 2)
            datos['kpis']['cobertura_territorial'] = len(datos['resumen_global']['total_municipios'])
            
            total_dias_activos = sum([u['stats']['dias_activos'] for u in datos['usuarios_con_datos']])
            if total_dias_activos > 0:
                datos['kpis']['eficiencia_operacional'] = round(total_archivos_global / total_dias_activos, 2)
        
        # Convertir sets a números
        datos['resumen_global']['total_municipios'] = len(datos['resumen_global']['total_municipios'])
        datos['resumen_global']['total_categorias'] = len(datos['resumen_global']['total_categorias'])
        datos['resumen_global']['total_componentes'] = len(datos['resumen_global']['total_componentes'])
        
        return datos
        
    except Exception as e:
        print(f"❌ Error recopilando datos: {str(e)}")
        raise

def procesar_usuario_completo(usuario, fecha_desde, fecha_hasta):
    """
    Procesa datos completos de un usuario
    """
    try:
        usuario_data = {
            'usuario': usuario,
            'tiene_actividad': False,
            'archivos_preoperacion': [],
            'archivos_postoperacion': [],
            'municipios_trabajados': {},
            'categorias_trabajadas': {},
            'componentes_trabajados': {},
            'categorias_detalle': defaultdict(int),
            'componentes_detalle': defaultdict(int),
            'actividad_diaria': defaultdict(int),
            'stats': {
                'total_archivos_pre': 0,
                'total_archivos_post': 0,
                'municipios_unicos': 0,
                'categorias_unicas': 0,
                'componentes_unicos': 0,
                'dias_activos': 0,
                'promedio_archivos_dia': 0,
                'categoria_principal': None,
                'componente_principal': None,
                'municipio_principal': None,
                'consistency_score': 0
            }
        }
        
        # Archivos de preoperación
        archivos_pre = ListaArchivosPre.objects.filter(
            usuario_windows=usuario,
            fecha_disposicion__range=[fecha_desde, fecha_hasta]
        ).select_related(
            'cod_insumo__cod_insumo__cod_municipio',
            'cod_insumo__cod_insumo__cod_categoria'
        ).order_by('fecha_disposicion')
        
        for archivo in archivos_pre:
            if archivo.cod_insumo and archivo.cod_insumo.cod_insumo:
                municipio = archivo.cod_insumo.cod_insumo.cod_municipio
                categoria = archivo.cod_insumo.cod_insumo.cod_categoria
                
                # Registrar municipio
                if municipio:
                    mun_id = municipio.cod_municipio
                    if mun_id not in usuario_data['municipios_trabajados']:
                        usuario_data['municipios_trabajados'][mun_id] = {
                            'nombre': municipio.nom_municipio,
                            'archivos_pre': 0,
                            'archivos_post': 0
                        }
                    usuario_data['municipios_trabajados'][mun_id]['archivos_pre'] += 1
                
                # Registrar categoría
                if categoria:
                    cat_id = categoria.cod_categoria
                    usuario_data['categorias_trabajadas'][cat_id] = categoria.nom_categoria
                    usuario_data['categorias_detalle'][cat_id] += 1
                
                # Actividad diaria
                if archivo.fecha_disposicion:
                    fecha_str = archivo.fecha_disposicion.strftime('%Y-%m-%d')
                    usuario_data['actividad_diaria'][fecha_str] += 1
                
                # Agregar archivo
                usuario_data['archivos_preoperacion'].append({
                    'nombre': archivo.nombre_insumo or 'Sin nombre',
                    'fecha': archivo.fecha_disposicion,
                    'municipio': municipio.nom_municipio if municipio else 'Sin municipio',
                    'categoria': categoria.nom_categoria if categoria else 'Sin categoría',
                    'ruta': archivo.path_file or 'Sin ruta',
                    'observacion': archivo.observacion or ''
                })
        
        # Archivos de postoperación
        archivos_post = ArchivosPost.objects.filter(
            usuario_windows=usuario,
            fecha_disposicion__range=[fecha_desde, fecha_hasta]
        ).select_related(
            'id_disposicion__cod_municipio',
            'id_disposicion__id_componente'
        ).order_by('fecha_disposicion')
        
        for archivo in archivos_post:
            if archivo.id_disposicion:
                municipio = archivo.id_disposicion.cod_municipio
                componente = archivo.id_disposicion.id_componente
                
                # Registrar municipio
                if municipio:
                    mun_id = municipio.cod_municipio
                    if mun_id not in usuario_data['municipios_trabajados']:
                        usuario_data['municipios_trabajados'][mun_id] = {
                            'nombre': municipio.nom_municipio,
                            'archivos_pre': 0,
                            'archivos_post': 0
                        }
                    usuario_data['municipios_trabajados'][mun_id]['archivos_post'] += 1
                
                # Registrar componente
                if componente:
                    comp_id = componente.id_componente
                    usuario_data['componentes_trabajados'][comp_id] = componente.nombre_componente
                    usuario_data['componentes_detalle'][comp_id] += 1
                
                # Actividad diaria
                if archivo.fecha_disposicion:
                    fecha_str = archivo.fecha_disposicion.strftime('%Y-%m-%d')
                    usuario_data['actividad_diaria'][fecha_str] += 1
                
                # Agregar archivo
                usuario_data['archivos_postoperacion'].append({
                    'nombre': archivo.nombre_archivo,
                    'fecha': archivo.fecha_disposicion,
                    'municipio': municipio.nom_municipio if municipio else 'Sin municipio',
                    'componente': componente.nombre_componente if componente else 'Sin componente',
                    'ruta': archivo.ruta_completa,
                    'observacion': archivo.observacion or ''
                })
        
        # Calcular estadísticas
        usuario_data['stats']['total_archivos_pre'] = len(usuario_data['archivos_preoperacion'])
        usuario_data['stats']['total_archivos_post'] = len(usuario_data['archivos_postoperacion'])
        usuario_data['stats']['municipios_unicos'] = len(usuario_data['municipios_trabajados'])
        usuario_data['stats']['categorias_unicas'] = len(usuario_data['categorias_trabajadas'])
        usuario_data['stats']['componentes_unicos'] = len(usuario_data['componentes_trabajados'])
        usuario_data['stats']['dias_activos'] = len(usuario_data['actividad_diaria'])
        
        total_archivos = usuario_data['stats']['total_archivos_pre'] + usuario_data['stats']['total_archivos_post']
        if usuario_data['stats']['dias_activos'] > 0:
            usuario_data['stats']['promedio_archivos_dia'] = round(total_archivos / usuario_data['stats']['dias_activos'], 2)
        
        # Principales
        if usuario_data['categorias_detalle']:
            top_cat = max(usuario_data['categorias_detalle'].items(), key=lambda x: x[1])
            usuario_data['stats']['categoria_principal'] = {
                'nombre': usuario_data['categorias_trabajadas'].get(top_cat[0], 'Sin nombre'),
                'cantidad': top_cat[1]
            }
        
        if usuario_data['componentes_detalle']:
            top_comp = max(usuario_data['componentes_detalle'].items(), key=lambda x: x[1])
            usuario_data['stats']['componente_principal'] = {
                'nombre': usuario_data['componentes_trabajados'].get(top_comp[0], 'Sin nombre'),
                'cantidad': top_comp[1]
            }
        
        if usuario_data['municipios_trabajados']:
            municipio_counts = {
                mun_id: data['archivos_pre'] + data['archivos_post'] 
                for mun_id, data in usuario_data['municipios_trabajados'].items()
            }
            top_mun_id = max(municipio_counts.items(), key=lambda x: x[1])
            usuario_data['stats']['municipio_principal'] = {
                'nombre': usuario_data['municipios_trabajados'][top_mun_id[0]]['nombre'],
                'cantidad': top_mun_id[1]
            }
        
        # Consistencia
        if usuario_data['actividad_diaria']:
            actividades = list(usuario_data['actividad_diaria'].values())
            if len(actividades) > 1:
                promedio = statistics.mean(actividades)
                if promedio > 0:
                    desv_std = statistics.stdev(actividades)
                    usuario_data['stats']['consistency_score'] = max(0, round(100 - (desv_std / promedio * 100), 1))
        
        # Determinar si tiene actividad
        usuario_data['tiene_actividad'] = total_archivos > 0
        
        return usuario_data
        
    except Exception as e:
        print(f"❌ Error procesando {usuario}: {str(e)}")
        raise

def generar_dashboard_principal(ws, datos_completos, fecha_desde, fecha_hasta):
    """
    Genera el dashboard principal con 6 GRÁFICAS COMPLETAS
    """
    print("📊 Generando dashboard principal con 6 gráficas...")
    
    # Estilos
    titulo_font = Font(bold=True, size=18, color="FFFFFF")
    titulo_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    seccion_font = Font(bold=True, size=12, color="FFFFFF")
    seccion_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # TÍTULO PRINCIPAL
    ws.merge_cells('A1:Z3')
    titulo_cell = ws['A1']
    titulo_cell.value = f"📊 DASHBOARD EJECUTIVO - ANÁLISIS DE PRODUCTIVIDAD\n📅 Período: {fecha_desde} al {fecha_hasta}\n⚙️ Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    titulo_cell.border = border
    
    # KPIs PRINCIPALES
    ws.merge_cells('A5:Z5')
    ws['A5'] = "📈 INDICADORES CLAVE DE RENDIMIENTO"
    ws['A5'].font = seccion_font
    ws['A5'].fill = seccion_fill
    ws['A5'].alignment = center_align
    
    kpis = [
        ("📁 Promedio Archivos/Usuario", datos_completos['kpis']['promedio_archivos_usuario']),
        ("🏛️ Cobertura Territorial", f"{datos_completos['kpis']['cobertura_territorial']} municipios"),
        ("⚡ Eficiencia Operacional", f"{datos_completos['kpis']['eficiencia_operacional']} arch/día"),
        ("👥 Usuarios Activos", len(datos_completos['usuarios_con_datos']))
    ]
    
    for i, (titulo, valor) in enumerate(kpis):
        col = i * 6 + 1
        ws.cell(row=6, column=col, value=titulo).font = Font(bold=True, color="1F4E79")
        ws.merge_cells(f'{get_column_letter(col+1)}6:{get_column_letter(col+4)}6')
        ws.cell(row=6, column=col+1, value=valor).font = Font(bold=True, size=14, color="C5504B")
        ws.cell(row=6, column=col+1).alignment = center_align
    
    current_row = 10
    
    # GRÁFICA 1: PRODUCTIVIDAD POR USUARIO
    ws.merge_cells(f'A{current_row}:L{current_row}')
    ws[f'A{current_row}'] = "📊 PRODUCTIVIDAD POR USUARIO"
    ws[f'A{current_row}'].font = seccion_font
    ws[f'A{current_row}'].fill = seccion_fill
    ws[f'A{current_row}'].alignment = center_align
    
    # Datos de productividad
    usuarios_ordenados = sorted(datos_completos['resumen_global']['usuarios_productividad'], key=lambda x: x['total_archivos'], reverse=True)
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="Usuario").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="Total Archivos").font = Font(bold=True)
    
    for i, usuario_info in enumerate(usuarios_ordenados):
        ws.cell(row=current_row+1+i, column=1, value=usuario_info['usuario'])
        ws.cell(row=current_row+1+i, column=2, value=usuario_info['total_archivos'])
    
    # Crear gráfica de barras
    chart1 = BarChart()
    chart1.title = "Productividad por Usuario"
    chart1.style = 10
    chart1.width = 15
    chart1.height = 10
    
    data_ref = Reference(ws, min_col=2, min_row=current_row, max_col=2, max_row=current_row+len(usuarios_ordenados))
    cats_ref = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+len(usuarios_ordenados))
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    
    ws.add_chart(chart1, f"N{current_row}")
    
    current_row += len(usuarios_ordenados) + 8
    
    # GRÁFICA 2: DISTRIBUCIÓN PRE vs POST
    ws.merge_cells(f'A{current_row}:L{current_row}')
    ws[f'A{current_row}'] = "🥧 DISTRIBUCIÓN PRE-OPERACIÓN vs POST-OPERACIÓN"
    ws[f'A{current_row}'].font = seccion_font
    ws[f'A{current_row}'].fill = seccion_fill
    ws[f'A{current_row}'].alignment = center_align
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="Tipo").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="Cantidad").font = Font(bold=True)
    ws.cell(row=current_row+1, column=1, value="Pre-Operación")
    ws.cell(row=current_row+1, column=2, value=datos_completos['resumen_global']['total_archivos_pre'])
    ws.cell(row=current_row+2, column=1, value="Post-Operación")
    ws.cell(row=current_row+2, column=2, value=datos_completos['resumen_global']['total_archivos_post'])
    
    # Crear gráfica de pastel
    pie_chart = PieChart()
    pie_chart.title = "Distribución Pre vs Post"
    pie_chart.width = 15
    pie_chart.height = 10
    
    data_pie = Reference(ws, min_col=2, min_row=current_row, max_col=2, max_row=current_row+2)
    cats_pie = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+2)
    pie_chart.add_data(data_pie, titles_from_data=True)
    pie_chart.set_categories(cats_pie)
    
    ws.add_chart(pie_chart, f"N{current_row}")
    
    current_row += 10
    
    # GRÁFICA 3: TOP CATEGORÍAS
    ws.merge_cells(f'A{current_row}:L{current_row}')
    ws[f'A{current_row}'] = "📋 TOP 10 CATEGORÍAS MÁS TRABAJADAS"
    ws[f'A{current_row}'].font = seccion_font
    ws[f'A{current_row}'].fill = seccion_fill
    ws[f'A{current_row}'].alignment = center_align
    
    top_categorias = sorted(datos_completos['resumen_global']['categorias_frecuencia'].items(), key=lambda x: x[1], reverse=True)[:10]
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="Categoría").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="Archivos").font = Font(bold=True)
    
    for i, (cat_id, cat_count) in enumerate(top_categorias):
        try:
            categoria = Categorias.objects.get(cod_categoria=cat_id)
            cat_nombre = categoria.nom_categoria
        except:
            cat_nombre = f"Categoría {cat_id}"
        
        ws.cell(row=current_row+1+i, column=1, value=cat_nombre)
        ws.cell(row=current_row+1+i, column=2, value=cat_count)
    
    # Gráfica de barras horizontales para categorías
    chart_cat = BarChart()
    chart_cat.type = "bar"
    chart_cat.title = "Top 10 Categorías"
    chart_cat.width = 15
    chart_cat.height = 12
    
    data_cat = Reference(ws, min_col=2, min_row=current_row, max_col=2, max_row=current_row+len(top_categorias))
    cats_cat = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+len(top_categorias))
    chart_cat.add_data(data_cat, titles_from_data=True)
    chart_cat.set_categories(cats_cat)
    
    ws.add_chart(chart_cat, f"N{current_row}")
    
    current_row += len(top_categorias) + 8
    
    # GRÁFICA 4: TOP COMPONENTES
    ws.merge_cells(f'A{current_row}:L{current_row}')
    ws[f'A{current_row}'] = "🎯 TOP 10 COMPONENTES MÁS TRABAJADOS"
    ws[f'A{current_row}'].font = seccion_font
    ws[f'A{current_row}'].fill = seccion_fill
    ws[f'A{current_row}'].alignment = center_align
    
    top_componentes = sorted(datos_completos['resumen_global']['componentes_frecuencia'].items(), key=lambda x: x[1], reverse=True)[:10]
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="Componente").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="Archivos").font = Font(bold=True)
    
    for i, (comp_id, comp_count) in enumerate(top_componentes):
        try:
            componente = ComponentesPost.objects.get(id_componente=comp_id)
            comp_nombre = componente.nombre_componente
        except:
            comp_nombre = f"Componente {comp_id}"
        
        ws.cell(row=current_row+1+i, column=1, value=comp_nombre)
        ws.cell(row=current_row+1+i, column=2, value=comp_count)
    
    # Gráfica de barras horizontales para componentes
    chart_comp = BarChart()
    chart_comp.type = "bar"
    chart_comp.title = "Top 10 Componentes"
    chart_comp.width = 15
    chart_comp.height = 12
    
    data_comp = Reference(ws, min_col=2, min_row=current_row, max_col=2, max_row=current_row+len(top_componentes))
    cats_comp = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+len(top_componentes))
    chart_comp.add_data(data_comp, titles_from_data=True)
    chart_comp.set_categories(cats_comp)
    
    ws.add_chart(chart_comp, f"N{current_row}")
    
    current_row += len(top_componentes) + 8
    
    # GRÁFICA 5: TOP MUNICIPIOS
    ws.merge_cells(f'A{current_row}:L{current_row}')
    ws[f'A{current_row}'] = "🏛️ TOP 15 MUNICIPIOS MÁS TRABAJADOS"
    ws[f'A{current_row}'].font = seccion_font
    ws[f'A{current_row}'].fill = seccion_fill
    ws[f'A{current_row}'].alignment = center_align
    
    top_municipios = sorted(datos_completos['resumen_global']['municipios_frecuencia'].items(), key=lambda x: x[1], reverse=True)[:15]
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="Municipio").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="Archivos").font = Font(bold=True)
    
    for i, (mun_id, mun_count) in enumerate(top_municipios):
        try:
            municipio = Municipios.objects.get(cod_municipio=mun_id)
            mun_nombre = municipio.nom_municipio
        except:
            mun_nombre = f"Municipio {mun_id}"
        
        ws.cell(row=current_row+1+i, column=1, value=mun_nombre)
        ws.cell(row=current_row+1+i, column=2, value=mun_count)
    
    # Gráfica de barras horizontales para municipios
    chart_mun = BarChart()
    chart_mun.type = "bar"
    chart_mun.title = "Top 15 Municipios"
    chart_mun.width = 15
    chart_mun.height = 15
    
    data_mun = Reference(ws, min_col=2, min_row=current_row, max_col=2, max_row=current_row+len(top_municipios))
    cats_mun = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+len(top_municipios))
    chart_mun.add_data(data_mun, titles_from_data=True)
    chart_mun.set_categories(cats_mun)
    
    ws.add_chart(chart_mun, f"N{current_row}")
    
    current_row += len(top_municipios) + 8
    
    # GRÁFICA 6: ACTIVIDAD TEMPORAL (CORREGIDA - TODOS LOS DÍAS)
    ws.merge_cells(f'A{current_row}:L{current_row}')
    ws[f'A{current_row}'] = "📈 ACTIVIDAD DIARIA DEL EQUIPO (PERÍODO COMPLETO)"
    ws[f'A{current_row}'].font = seccion_font
    ws[f'A{current_row}'].fill = seccion_fill
    ws[f'A{current_row}'].alignment = center_align
    
    # GENERAR TODOS LOS DÍAS DEL PERÍODO
    fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d')
    fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d')
    
    dias_completos = []
    fecha_actual_iter = fecha_inicio
    while fecha_actual_iter <= fecha_fin:
        fecha_str = fecha_actual_iter.strftime('%Y-%m-%d')
        cantidad = datos_completos['resumen_global']['actividad_diaria_global'].get(fecha_str, 0)
        dias_completos.append((fecha_str, cantidad))
        fecha_actual_iter += timedelta(days=1)
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="Fecha").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="Archivos").font = Font(bold=True)
    
    for i, (fecha, cantidad) in enumerate(dias_completos):
        ws.cell(row=current_row+1+i, column=1, value=fecha)
        ws.cell(row=current_row+1+i, column=2, value=cantidad)
    
    # Gráfica de líneas
    line_chart = LineChart()
    line_chart.title = "Actividad Diaria del Equipo"
    line_chart.style = 13
    line_chart.width = 18
    line_chart.height = 12
    
    data_line = Reference(ws, min_col=2, min_row=current_row, max_col=2, max_row=current_row+len(dias_completos))
    cats_line = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+len(dias_completos))
    line_chart.add_data(data_line, titles_from_data=True)
    line_chart.set_categories(cats_line)
    
    ws.add_chart(line_chart, f"N{current_row}")
    
    # AJUSTAR ANCHOS DE COLUMNAS (MUY IMPORTANTE)
    anchos = [30, 25, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    
    print("✅ Dashboard principal con 6 gráficas completas generado")

def generar_analisis_usuario(ws, usuario_data, fecha_desde, fecha_hasta):
    """
    Genera análisis INDIVIDUAL con gráficas y orden correcto
    """
    usuario = usuario_data['usuario']
    print(f"🎨 Generando análisis completo para {usuario}...")
    
    # Estilos
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="0F4C75", end_color="0F4C75", fill_type="solid")
    seccion_font = Font(bold=True, size=12, color="FFFFFF")
    seccion_fill = PatternFill(start_color="3282B8", end_color="3282B8", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # TÍTULO INDIVIDUAL
    stats = usuario_data['stats']
    total_archivos = stats['total_archivos_pre'] + stats['total_archivos_post']
    
    ws.merge_cells('A1:Z3')
    titulo_cell = ws['A1']
    titulo_cell.value = f"👤 ANÁLISIS DETALLADO - {usuario.upper()}\n📅 Período: {fecha_desde} al {fecha_hasta}\n📊 Rendimiento: {total_archivos} archivos | {stats['municipios_unicos']} municipios | {stats['dias_activos']} días activos"
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    titulo_cell.border = border
    
    # ESTADÍSTICAS PERSONALES
    ws.merge_cells('A5:Z5')
    ws['A5'] = f"📊 MÉTRICAS DE RENDIMIENTO - {usuario.upper()}"
    ws['A5'].font = seccion_font
    ws['A5'].fill = seccion_fill
    ws['A5'].alignment = center_align
    
    stats_personales = [
        ("📁 Archivos Pre", stats['total_archivos_pre']),
        ("📁 Archivos Post", stats['total_archivos_post']),
        ("🏛️ Municipios", stats['municipios_unicos']),
        ("📋 Categorías", stats['categorias_unicas']),
        ("🎯 Componentes", stats['componentes_unicos']),
        ("📅 Días Activos", stats['dias_activos']),
        ("⚡ Promedio Diario", f"{stats['promedio_archivos_dia']} arch/día"),
        ("📈 Consistencia", f"{stats['consistency_score']}%")
    ]
    
    for i, (titulo, valor) in enumerate(stats_personales):
        col = (i % 4) * 6 + 1
        row = 6 if i < 4 else 7
        
        ws.cell(row=row, column=col, value=titulo).font = Font(bold=True, color="0F4C75")
        ws.merge_cells(f'{get_column_letter(col+1)}{row}:{get_column_letter(col+4)}{row}')
        ws.cell(row=row, column=col+1, value=valor).font = Font(bold=True, size=12, color="C5504B")
        ws.cell(row=row, column=col+1).alignment = center_align
    
    current_row = 10
    
    # GRÁFICA 1: DISTRIBUCIÓN PRE vs POST INDIVIDUAL
    if stats['total_archivos_pre'] > 0 or stats['total_archivos_post'] > 0:
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'] = f"🥧 DISTRIBUCIÓN DE TRABAJO - {usuario.upper()}"
        ws[f'A{current_row}'].font = seccion_font
        ws[f'A{current_row}'].fill = seccion_fill
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        ws.cell(row=current_row, column=1, value="Tipo").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value="Cantidad").font = Font(bold=True)
        ws.cell(row=current_row+1, column=1, value="Pre-Operación")
        ws.cell(row=current_row+1, column=2, value=stats['total_archivos_pre'])
        ws.cell(row=current_row+2, column=1, value="Post-Operación")
        ws.cell(row=current_row+2, column=2, value=stats['total_archivos_post'])
        
        # Gráfica de pastel individual
        pie_chart_user = PieChart()
        pie_chart_user.title = f"Distribución - {usuario}"
        pie_chart_user.width = 15
        pie_chart_user.height = 10
        
        data_pie_user = Reference(ws, min_col=2, min_row=current_row, max_col=2, max_row=current_row+2)
        cats_pie_user = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+2)
        pie_chart_user.add_data(data_pie_user, titles_from_data=True)
        pie_chart_user.set_categories(cats_pie_user)
        
        ws.add_chart(pie_chart_user, f"N{current_row}")
        
        current_row += 12
    
    # GRÁFICA 2: MUNICIPIOS DEL USUARIO
    if usuario_data['municipios_trabajados']:
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'] = f"🏛️ TOP 10 MUNICIPIOS - {usuario.upper()}"
        ws[f'A{current_row}'].font = seccion_font
        ws[f'A{current_row}'].fill = seccion_fill
        ws[f'A{current_row}'].alignment = center_align
        
        municipios_ordenados = sorted(
            usuario_data['municipios_trabajados'].items(),
            key=lambda x: x[1]['archivos_pre'] + x[1]['archivos_post'],
            reverse=True
        )[:10]
        
        current_row += 1
        ws.cell(row=current_row, column=1, value="Municipio").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value="Total Archivos").font = Font(bold=True)
        
        for i, (mun_id, mun_data) in enumerate(municipios_ordenados):
            total_mun = mun_data['archivos_pre'] + mun_data['archivos_post']
            ws.cell(row=current_row+1+i, column=1, value=mun_data['nombre'])
            ws.cell(row=current_row+1+i, column=2, value=total_mun)
        
        # Gráfica de barras para municipios del usuario
        chart_mun_user = BarChart()
        chart_mun_user.type = "bar"
        chart_mun_user.title = f"Top 10 Municipios - {usuario}"
        chart_mun_user.width = 16
        chart_mun_user.height = 12
        
        data_mun_user = Reference(ws, min_col=2, min_row=current_row, max_col=2, max_row=current_row+len(municipios_ordenados))
        cats_mun_user = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+len(municipios_ordenados))
        chart_mun_user.add_data(data_mun_user, titles_from_data=True)
        chart_mun_user.set_categories(cats_mun_user)
        
        ws.add_chart(chart_mun_user, f"N{current_row}")
        
        current_row += len(municipios_ordenados) + 10
    
    # GRÁFICA 3: ACTIVIDAD DIARIA INDIVIDUAL (CORREGIDA)
    if usuario_data['actividad_diaria']:
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'] = f"📈 ACTIVIDAD DIARIA - {usuario.upper()}"
        ws[f'A{current_row}'].font = seccion_font
        ws[f'A{current_row}'].fill = seccion_fill
        ws[f'A{current_row}'].alignment = center_align
        
        # GENERAR TODOS LOS DÍAS DEL PERÍODO PARA EL USUARIO
        fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d')
        fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        
        dias_completos_usuario = []
        fecha_actual_iter = fecha_inicio
        while fecha_actual_iter <= fecha_fin:
            fecha_str = fecha_actual_iter.strftime('%Y-%m-%d')
            cantidad = usuario_data['actividad_diaria'].get(fecha_str, 0)
            dias_completos_usuario.append((fecha_str, cantidad))
            fecha_actual_iter += timedelta(days=1)
        
        current_row += 1
        ws.cell(row=current_row, column=1, value="Fecha").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value="Archivos").font = Font(bold=True)
        
        for i, (fecha, cantidad) in enumerate(dias_completos_usuario):
            ws.cell(row=current_row+1+i, column=1, value=fecha)
            ws.cell(row=current_row+1+i, column=2, value=cantidad)
        
        # Gráfica de líneas individual
        line_chart_user = LineChart()
        line_chart_user.title = f"Actividad Diaria - {usuario}"
        line_chart_user.style = 13
        line_chart_user.width = 18
        line_chart_user.height = 12
        
        data_line_user = Reference(ws, min_col=2, min_row=current_row, max_col=2, max_row=current_row+len(dias_completos_usuario))
        cats_line_user = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+len(dias_completos_usuario))
        line_chart_user.add_data(data_line_user, titles_from_data=True)
        line_chart_user.set_categories(cats_line_user)
        
        ws.add_chart(line_chart_user, f"N{current_row}")
        
        current_row += len(dias_completos_usuario) + 8
    
    # TABLA RESUMEN POR MUNICIPIOS (PRIMERO - COMO PEDISTE)
    if usuario_data['municipios_trabajados']:
        ws.merge_cells(f'A{current_row}:Z{current_row}')
        ws[f'A{current_row}'] = f"🏛️ RESUMEN COMPLETO POR MUNICIPIOS - {usuario.upper()}"
        ws[f'A{current_row}'].font = seccion_font
        ws[f'A{current_row}'].fill = seccion_fill
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        headers_mun = ['CÓDIGO', 'MUNICIPIO', 'ARCHIVOS PRE', 'ARCHIVOS POST', 'TOTAL', '% DEL TOTAL']
        for col, header in enumerate(headers_mun, 1):
            ws.cell(row=current_row, column=col, value=header).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="0F4C75", end_color="0F4C75", fill_type="solid")
            ws.cell(row=current_row, column=col).alignment = center_align
        
        municipios_completos = sorted(
            usuario_data['municipios_trabajados'].items(),
            key=lambda x: x[1]['archivos_pre'] + x[1]['archivos_post'],
            reverse=True
        )
        
        current_row += 1
        for mun_id, mun_data in municipios_completos:
            total_mun = mun_data['archivos_pre'] + mun_data['archivos_post']
            porcentaje = round((total_mun / total_archivos) * 100, 1) if total_archivos > 0 else 0
            
            datos_municipio = [mun_id, mun_data['nombre'], mun_data['archivos_pre'], mun_data['archivos_post'], total_mun, f"{porcentaje}%"]
            
            for col, valor in enumerate(datos_municipio, 1):
                ws.cell(row=current_row, column=col, value=valor).border = border
                if col >= 3:
                    ws.cell(row=current_row, column=col).font = Font(bold=True, color="C5504B")
            
            current_row += 1
        
        current_row += 3
    
    # TABLA ARCHIVOS PRE-OPERACIÓN
    if usuario_data['archivos_preoperacion']:
        ws.merge_cells(f'A{current_row}:Z{current_row}')
        ws[f'A{current_row}'] = f"📋 ARCHIVOS PRE-OPERACIÓN - {usuario.upper()} ({len(usuario_data['archivos_preoperacion'])} archivos)"
        ws[f'A{current_row}'].font = seccion_font
        ws[f'A{current_row}'].fill = seccion_fill
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        headers_pre = ['#', 'NOMBRE ARCHIVO', 'FECHA', 'MUNICIPIO', 'CATEGORÍA', 'RUTA', 'OBSERVACIÓN']
        for col, header in enumerate(headers_pre, 1):
            ws.cell(row=current_row, column=col, value=header).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="0F4C75", end_color="0F4C75", fill_type="solid")
        
        current_row += 1
        for i, archivo in enumerate(usuario_data['archivos_preoperacion'], 1):
            datos_archivo = [
                i,
                archivo['nombre'],
                archivo['fecha'].strftime('%d/%m/%Y') if archivo['fecha'] else 'Sin fecha',
                archivo['municipio'],
                archivo['categoria'],
                archivo['ruta'][:70] + '...' if len(archivo['ruta']) > 70 else archivo['ruta'],
                archivo['observacion'][:50] + '...' if len(archivo['observacion']) > 50 else archivo['observacion']
            ]
            
            for col, valor in enumerate(datos_archivo, 1):
                ws.cell(row=current_row, column=col, value=valor).border = border
                ws.cell(row=current_row, column=col).alignment = center_align if col == 1 else left_align
            
            current_row += 1
        
        current_row += 3
    
    # TABLA ARCHIVOS POST-OPERACIÓN
    if usuario_data['archivos_postoperacion']:
        ws.merge_cells(f'A{current_row}:Z{current_row}')
        ws[f'A{current_row}'] = f"🎯 ARCHIVOS POST-OPERACIÓN - {usuario.upper()} ({len(usuario_data['archivos_postoperacion'])} archivos)"
        ws[f'A{current_row}'].font = seccion_font
        ws[f'A{current_row}'].fill = seccion_fill
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        headers_post = ['#', 'NOMBRE ARCHIVO', 'FECHA', 'MUNICIPIO', 'COMPONENTE', 'RUTA', 'OBSERVACIÓN']
        for col, header in enumerate(headers_post, 1):
            ws.cell(row=current_row, column=col, value=header).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="0F4C75", end_color="0F4C75", fill_type="solid")
        
        current_row += 1
        for i, archivo in enumerate(usuario_data['archivos_postoperacion'], 1):
            datos_archivo = [
                i,
                archivo['nombre'],
                archivo['fecha'].strftime('%d/%m/%Y') if archivo['fecha'] else 'Sin fecha',
                archivo['municipio'],
                archivo['componente'],
                archivo['ruta'][:70] + '...' if len(archivo['ruta']) > 70 else archivo['ruta'],
                archivo['observacion'][:50] + '...' if len(archivo['observacion']) > 50 else archivo['observacion']
            ]
            
            for col, valor in enumerate(datos_archivo, 1):
                ws.cell(row=current_row, column=col, value=valor).border = border
                ws.cell(row=current_row, column=col).alignment = center_align if col == 1 else left_align
            
            current_row += 1
    
    # AJUSTAR ANCHOS DE COLUMNAS (MUY ANCHO PARA QUE NO SE CORTEN)
    anchos = [8, 50, 15, 35, 40, 80, 60, 20, 20, 20, 20, 20, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    
    print(f"✅ Análisis completo para {usuario} generado con orden correcto")




class UsuarioCreateSerializer(serializers.ModelSerializer):
    # Sobrescribir username para quitar validador unico por defecto de Django
    # y usar nuestro validador personalizado con mensaje en espanol
    username = serializers.CharField(max_length=150, validators=[])
    password = serializers.CharField(write_only=True, min_length=4)
    confirm_password = serializers.CharField(write_only=True)
    rol_tipo = serializers.ChoiceField(
        choices=[('profesional', 'Profesional'), ('admin', 'Administrador')],
        default='profesional',
        write_only=True
    )

    class Meta:
        model = User
        fields = ['username', 'email', 'first_name', 'last_name', 'password', 'confirm_password', 'rol_tipo']
    
    def validate_username(self, value):
        """Validar que el username sea único"""
        if User.objects.filter(username=value).exists():
            raise serializers.ValidationError("Este nombre de usuario ya existe.")
        
        # Validar formato del username
        if not re.match(r'^[a-zA-Z0-9._]+$', value):
            raise serializers.ValidationError("El nombre de usuario solo puede contener letras, números, puntos y guiones bajos.")
        
        return value
    
    def validate_email(self, value):
        """Validar que el email sea único"""
        if User.objects.filter(email=value).exists():
            raise serializers.ValidationError("Este correo electrónico ya está registrado.")
        
        return value
    
    def validate(self, data):
        """Validar que las contraseñas coincidan"""
        if data['password'] != data['confirm_password']:
            raise serializers.ValidationError("Las contraseñas no coinciden.")
        
        password = data['password']
        if len(password) < 4:
            raise serializers.ValidationError("La contraseña debe tener al menos 4 caracteres.")
        
        return data
    
    def create(self, validated_data):
        """Crear usuario y su registro en ProfesionalesSeguimiento"""
        validated_data.pop('confirm_password')
        password = validated_data.pop('password')
        rol_tipo = validated_data.pop('rol_tipo', 'profesional')
        
        # Crear usuario Django
        user = User.objects.create_user(
            username=validated_data['username'],
            email=validated_data['email'],
            first_name=validated_data.get('first_name', ''),
            last_name=validated_data.get('last_name', ''),
            password=password
        )
        
        # Configurar permisos según rol
        if rol_tipo == 'admin':
            user.is_staff = True
            user.is_superuser = False
            
            # Agregar a grupo de administradores
            try:
                grupo_admin = Group.objects.get(name='Administradores')
                user.groups.add(grupo_admin)
            except Group.DoesNotExist:
                print("⚠️ Grupo 'Administradores' no existe")
                
        else:  # profesional
            user.is_staff = False
            user.is_superuser = False
            
            # Agregar a grupo de profesionales
            try:
                grupo_profesionales = Group.objects.get(name='Profesionales_Seguimiento')
                user.groups.add(grupo_profesionales)
            except Group.DoesNotExist:
                print("⚠️ Grupo 'Profesionales_Seguimiento' no existe")
            
            # ✅ CREAR REGISTRO EN ProfesionalesSeguimiento
            try:
                # Obtener o crear rol por defecto
                rol_default, _ = RolesSeguimiento.objects.get_or_create(
                    rol_profesional='Profesional de Seguimiento'
                )
                
                # Crear profesional de seguimiento
                profesional = ProfesionalesSeguimiento.objects.create(
                    cod_profesional=user.username,  # ✅ Usar username como código
                    nombre_profesional=f"{user.first_name} {user.last_name}".strip() or user.username,
                    correo_profesional=user.email,
                    rol_profesional=rol_default,
                    usuario_django=user  # ✅ Vincular con usuario Django
                )
                
                print(f"✅ Profesional creado: {profesional.cod_profesional}")
                
            except Exception as e:
                print(f"❌ Error creando profesional: {str(e)}")
                # No fallar la creación del usuario por esto
        
        user.save()
        return user
    
class UsuarioDetailSerializer(serializers.ModelSerializer):
    rol_tipo = serializers.SerializerMethodField()
    fecha_registro = serializers.DateTimeField(source='date_joined', read_only=True)
    ultimo_login = serializers.DateTimeField(source='last_login', read_only=True)
    
    class Meta:
        model = User
        fields = [
            'id', 'username', 'email', 'first_name', 'last_name', 
            'is_active', 'is_staff', 'is_superuser', 'rol_tipo',
            'fecha_registro', 'ultimo_login'
        ]
        read_only_fields = ['id', 'username', 'fecha_registro', 'ultimo_login']
    
    def get_rol_tipo(self, obj):
        """Determinar el tipo de rol"""
        if obj.is_superuser and obj.is_staff:
            return 'super_admin'
        elif obj.is_staff and not obj.is_superuser:
            return 'admin'
        else:
            return 'profesional'

class UsuarioUpdateSerializer(serializers.ModelSerializer):
    rol_tipo = serializers.ChoiceField(
        choices=[('profesional', 'Profesional'), ('admin', 'Administrador'), ('super_admin', 'Super Administrador')],
        write_only=True
    )

    class Meta:
        model = User
        fields = ['first_name', 'last_name', 'email', 'is_active', 'rol_tipo']

    def update(self, instance, validated_data):
        """Actualizar usuario con lógica de roles"""
        rol_tipo = validated_data.pop('rol_tipo', None)
        request = self.context.get('request')

        # =============== USUARIOS PROTEGIDOS ===============
        # Estos usuarios NO pueden ser modificados en rol, estado ni contraseña
        USUARIOS_PROTEGIDOS = ['andres.osorio', 'elizabeth.eraso']

        if instance.username in USUARIOS_PROTEGIDOS:
            # Solo permitir modificar nombre y apellido, NO el rol ni estado
            instance.first_name = validated_data.get('first_name', instance.first_name)
            instance.last_name = validated_data.get('last_name', instance.last_name)
            instance.email = validated_data.get('email', instance.email)
            # NO modificar is_active ni rol para usuarios protegidos
            instance.save()
            return instance

        # Actualizar campos básicos
        instance.first_name = validated_data.get('first_name', instance.first_name)
        instance.last_name = validated_data.get('last_name', instance.last_name)
        instance.email = validated_data.get('email', instance.email)
        instance.is_active = validated_data.get('is_active', instance.is_active)

        # Aplicar lógica de roles
        # Solo un super_admin puede dar rol de super_admin a otro usuario
        if rol_tipo == 'super_admin':
            if request and request.user.is_superuser:
                instance.is_staff = True
                instance.is_superuser = True
            # Si no es superuser quien hace la petición, ignorar el intento
        elif rol_tipo == 'admin':
            instance.is_staff = True
            instance.is_superuser = False
        elif rol_tipo == 'profesional':
            instance.is_staff = False
            instance.is_superuser = False

        instance.save()
        return instance

# =============== VISTAS DE GESTIÓN DE USUARIOS ===============

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def crear_usuario(request):
    """
    Crear un nuevo usuario - Solo para administradores
    """
    try:
        # Verificar que el usuario actual es admin o super admin
        if not (request.user.is_staff or request.user.is_superuser):
            return Response(
                {'error': 'No tiene permisos para crear usuarios'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = UsuarioCreateSerializer(data=request.data)
        if serializer.is_valid():
            usuario = serializer.save()
            
            # ✅ CREAR PROFESIONAL SI ES NECESARIO
            rol_tipo = request.data.get('rol_tipo', 'profesional')
            
            if rol_tipo == 'profesional':
                try:
                    rol_default, _ = RolesSeguimiento.objects.get_or_create(
                        rol_profesional='Profesional de Seguimiento'
                    )
                    
                    ProfesionalesSeguimiento.objects.get_or_create(
                        cod_profesional=usuario.username,
                        defaults={
                            'nombre_profesional': f"{usuario.first_name} {usuario.last_name}".strip() or usuario.username,
                            'correo_profesional': usuario.email,
                            'rol_profesional': rol_default,
                            'usuario_django': usuario
                        }
                    )
                except Exception as e:
                    print(f"❌ Error creando profesional: {str(e)}")
            
            # Respuesta con datos del usuario creado
            response_serializer = UsuarioDetailSerializer(usuario)
            return Response({
                'message': 'Usuario creado exitosamente',
                'usuario': response_serializer.data
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
    except Exception as e:
        return Response(
            {'error': f'Error al crear usuario: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def listar_usuarios(request):
    """
    Listar usuarios agrupados por rol - Solo para administradores
    """
    try:
        # Obtener todos los usuarios (excepto el usuario actual para evitar auto-modificación)
        usuarios = User.objects.exclude(id=request.user.id).order_by('-date_joined')
        
        # Agrupar por roles
        super_admins = []
        admins = []
        profesionales = []
        
        for usuario in usuarios:
            usuario_data = UsuarioDetailSerializer(usuario).data
            
            if usuario.is_superuser and usuario.is_staff:
                super_admins.append(usuario_data)
            elif usuario.is_staff and not usuario.is_superuser:
                admins.append(usuario_data)
            else:
                profesionales.append(usuario_data)
        
        return Response({
            'super_admins': super_admins,
            'admins': admins,
            'profesionales': profesionales,
            'total': len(usuarios),
            'estadisticas': {
                'total_super_admins': len(super_admins),
                'total_admins': len(admins),
                'total_profesionales': len(profesionales)
            }
        })
        
    except Exception as e:
        return Response(
            {'error': f'Error al listar usuarios: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def obtener_usuario(request, user_id):
    """
    Obtener un usuario específico - Solo para administradores
    """
    try:
        usuario = User.objects.get(id=user_id)
        
        # No permitir ver/editar super admins si no eres super admin
        if usuario.is_superuser and not request.user.is_superuser:
            return Response(
                {'error': 'No tiene permisos para ver este usuario'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = UsuarioDetailSerializer(usuario)
        return Response(serializer.data)
        
    except User.DoesNotExist:
        return Response(
            {'error': 'Usuario no encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Error al obtener usuario: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def actualizar_usuario(request, user_id):
    """
    Actualizar un usuario - Solo para administradores
    """
    try:
        # =============== USUARIOS PROTEGIDOS ===============
        USUARIOS_PROTEGIDOS = ['andres.osorio', 'elizabeth.eraso']

        usuario = User.objects.get(id=user_id)

        # Verificar si es usuario protegido y se intenta cambiar rol o estado
        if usuario.username in USUARIOS_PROTEGIDOS:
            rol_tipo = request.data.get('rol_tipo')
            is_active = request.data.get('is_active')

            # No permitir cambiar rol de usuarios protegidos
            if rol_tipo and rol_tipo != 'super_admin':
                return Response(
                    {'error': f'No se puede modificar el rol del usuario protegido {usuario.username}'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # No permitir desactivar usuarios protegidos
            if is_active is False:
                return Response(
                    {'error': f'No se puede desactivar al usuario protegido {usuario.username}'},
                    status=status.HTTP_403_FORBIDDEN
                )

        # Validaciones de permisos
        if usuario.is_superuser and not request.user.is_superuser:
            return Response(
                {'error': 'No puede modificar super administradores'},
                status=status.HTTP_403_FORBIDDEN
            )

        if usuario.id == request.user.id:
            return Response(
                {'error': 'No puede modificar su propio usuario desde aquí'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Pasar el request como contexto para verificar permisos de super_admin
        serializer = UsuarioUpdateSerializer(
            usuario,
            data=request.data,
            partial=True,
            context={'request': request}
        )
        if serializer.is_valid():
            usuario_actualizado = serializer.save()

            response_serializer = UsuarioDetailSerializer(usuario_actualizado)
            return Response({
                'message': 'Usuario actualizado exitosamente',
                'usuario': response_serializer.data
            })

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    except User.DoesNotExist:
        return Response(
            {'error': 'Usuario no encontrado'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Error al actualizar usuario: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def eliminar_usuario(request, user_id):
    """
    Eliminar usuario en CASCADA - Solo para SUPER ADMINISTRADORES
    """
    try:
        # =============== USUARIOS PROTEGIDOS ===============
        USUARIOS_PROTEGIDOS = ['andres.osorio', 'elizabeth.eraso']

        # Solo super admins pueden eliminar usuarios
        if not request.user.is_superuser:
            return Response(
                {'error': 'Solo los super administradores pueden eliminar usuarios'},
                status=status.HTTP_403_FORBIDDEN
            )

        usuario = User.objects.get(id=user_id)

        # Validar usuarios protegidos (estos NUNCA pueden ser eliminados)
        if usuario.username in USUARIOS_PROTEGIDOS:
            return Response(
                {'error': f'El usuario {usuario.username} está protegido y no puede ser eliminado'},
                status=status.HTTP_403_FORBIDDEN
            )

        # No puede eliminarse a sí mismo
        if usuario.id == request.user.id:
            return Response(
                {'error': 'No puede eliminar su propio usuario'},
                status=status.HTTP_403_FORBIDDEN
            )

        username = usuario.username

        # ELIMINACIÓN EN CASCADA
        with transaction.atomic():
            print(f"🗑️ Iniciando eliminación en cascada del usuario {username}...")
            
            # 1. Eliminar de Usuarios (tabla custom)
            try:
                usuario_custom = Usuarios.objects.filter(correo=usuario.email).first()
                if usuario_custom:
                    usuario_custom.delete()
                    print(f"✅ Usuario custom eliminado")
            except Exception as e:
                print(f"⚠️ Error eliminando usuario custom: {e}")
            
            # 2. Eliminar profesional de seguimiento si existe
            try:
                profesional = ProfesionalesSeguimiento.objects.filter(
                    usuario_django=usuario
                ).first()
                if profesional:
                    # Eliminar asignaciones de territoriales y municipios
                    ProfesionalTerritorial.objects.filter(cod_profesional=profesional).delete()
                    ProfesionalMunicipio.objects.filter(cod_profesional=profesional).delete()
                    profesional.delete()
                    print(f"✅ Profesional de seguimiento eliminado")
            except Exception as e:
                print(f"⚠️ Error eliminando profesional: {e}")
            
            # 3. Actualizar registros que referencian al usuario (en lugar de eliminar)
            try:
                # Actualizar detalles de insumo
                detalles_usuario = DetalleInsumo.objects.filter(cod_usuario__correo=usuario.email)
                if detalles_usuario.exists():
                    # Buscar usuario "sistema" o crear uno genérico
                    usuario_sistema, created = Usuarios.objects.get_or_create(
                        cod_usuario=9999,
                        defaults={
                            'nombre': 'Usuario Eliminado',
                            'correo': 'sistema@igac.gov.co'
                        }
                    )
                    detalles_usuario.update(cod_usuario=usuario_sistema)
                    print(f"✅ {detalles_usuario.count()} detalles reasignados")
            except Exception as e:
                print(f"⚠️ Error reasignando detalles: {e}")
            
            # 4. Eliminar el usuario de Django
            usuario.delete()
            print(f"✅ Usuario Django eliminado")
        
        return Response({
            'message': f'Usuario {username} eliminado exitosamente',
            'username': username
        })
        
    except User.DoesNotExist:
        return Response(
            {'error': 'Usuario no encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        print(f"❌ Error en eliminación: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response(
            {'error': f'Error al eliminar usuario: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def cambiar_password_usuario(request, user_id):
    """
    Cambiar contraseña de un usuario - Solo para administradores
    """
    try:
        # =============== USUARIOS PROTEGIDOS ===============
        USUARIOS_PROTEGIDOS = ['andres.osorio', 'elizabeth.eraso']

        usuario = User.objects.get(id=user_id)

        # Verificar si es usuario protegido
        if usuario.username in USUARIOS_PROTEGIDOS:
            return Response(
                {'error': f'No se puede cambiar la contraseña del usuario protegido {usuario.username}'},
                status=status.HTTP_403_FORBIDDEN
            )

        nueva_password = request.data.get('nueva_password')

        if not nueva_password or len(nueva_password) < 4:
            return Response(
                {'error': 'La contraseña debe tener al menos 4 caracteres'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validaciones de permisos
        if usuario.is_superuser and not request.user.is_superuser:
            return Response(
                {'error': 'No puede cambiar la contraseña de super administradores'},
                status=status.HTTP_403_FORBIDDEN
            )

        usuario.set_password(nueva_password)
        usuario.save()

        return Response({
            'message': f'Contraseña actualizada para {usuario.username}'
        })
        
    except User.DoesNotExist:
        return Response(
            {'error': 'Usuario no encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': f'Error al cambiar contraseña: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def estadisticas_usuarios(request):
    """
    Obtener estadísticas de usuarios del sistema
    """
    try:
        total_usuarios = User.objects.count()
        usuarios_activos = User.objects.filter(is_active=True).count()
        super_admins = User.objects.filter(is_superuser=True, is_staff=True).count()
        admins = User.objects.filter(is_staff=True, is_superuser=False).count()
        profesionales = User.objects.filter(is_staff=False, is_superuser=False).count()
        
        # Usuarios recientes (últimos 30 días)
        from datetime import datetime, timedelta
        hace_30_dias = datetime.now() - timedelta(days=30)
        usuarios_recientes = User.objects.filter(date_joined__gte=hace_30_dias).count()
        
        return Response({
            'total_usuarios': total_usuarios,
            'usuarios_activos': usuarios_activos,
            'usuarios_inactivos': total_usuarios - usuarios_activos,
            'super_admins': super_admins,
            'admins': admins,
            'profesionales': profesionales,
            'usuarios_recientes': usuarios_recientes,
            'porcentaje_activos': round((usuarios_activos / total_usuarios) * 100, 2) if total_usuarios > 0 else 0
        })
        
    except Exception as e:
        return Response(
            {'error': f'Error al obtener estadísticas: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    




# Mapeo de directorios de preoperación a categorías legibles para reportes
# Incluye todas las variantes encontradas en la BD
DIRECTORIOS_PREOPERACION_MAP = {
    '00_capacitaciones': 'Capacitaciones',
    '01_prop': 'Propuesta',
    '02_carta_acept': 'Carta de Aceptacion',
    '03_cnt_modif': 'Contrato/Modificacion',
    '03_constr_modif': 'Contrato/Modificacion',
    '03_cont_modif': 'Contrato/Modificacion',
    '03_cont_y_modif': 'Contrato/Modificacion',
    '03_conte_modif': 'Contrato/Modificacion',
    '03_contr_modif': 'Contrato/Modificacion',
    '03_contrac_modif': 'Contrato/Modificacion',
    '03_contrl_modif': 'Contrato/Modificacion',
    '03_cto_modif': 'Contrato/Modificacion',
    '04_acta_ini': 'Acta de Inicio',
    '05_plan_gest_proy': 'Plan de Gestion del Proyecto',
    '05_plan_trab': 'Plan de Gestion del Proyecto',
    '06_insu': 'Insumos',
    '06_pre_recono': 'Pre-Conocimiento',
    '06_precono': 'Pre-Conocimiento',
    '06_prerecono': 'Pre-Conocimiento',
    '06_recono': 'Pre-Conocimiento',
    '07_contr_pers': 'Control de Personal',
    '07_cto_pers': 'Control de Personal',
    '08_contr_pers': 'Control de Personal',
    '08_contrac_pers': 'Control de Personal',
    '08_cto_pers': 'Control de Personal',
    '08_corres': 'Correspondencia',
    '09_asis': 'Asistencias',
    '10_ppt': 'Presentaciones',
    '10_presen': 'Presentaciones',
}


def extraer_subcarpeta_desde_ruta_pre(path_file):
    """
    Extrae la subcarpeta principal de una ruta completa para PRE-OPERACIÓN
    Ejemplo: \\repositorio\\...\\25\\175\\PGN-IGAC\\01_preo\\01_carto_basic\\archivo.shp 
    → "01_preo\\01_carto_basic"
    """
    if not path_file:
        return "Sin subcarpeta"
    
    try:
        ruta_str = str(path_file).replace('/', '\\')
        partes = ruta_str.split('\\')
        
        # Buscar el índice donde aparece '01_preo' o similar
        indice_preo = -1
        for i, parte in enumerate(partes):
            if 'preo' in parte.lower() or 'pre' in parte.lower() or '01_' in parte:
                indice_preo = i
                break
        
        if indice_preo != -1 and indice_preo < len(partes) - 1:
            # Tomar desde 01_preo hasta antes del archivo final
            subcarpeta_partes = partes[indice_preo:-1]  # Excluir el nombre del archivo
            if subcarpeta_partes:
                return '\\'.join(subcarpeta_partes)
        
        # Fallback: tomar las últimas 2-3 carpetas antes del archivo
        if len(partes) >= 3:
            return '\\'.join(partes[-3:-1])
        elif len(partes) >= 2:
            return partes[-2]
        
        return "Sin subcarpeta"
        
    except Exception as e:
        print(f"⚠️ Error extrayendo subcarpeta PRE: {e}")
        return "Error subcarpeta"

def extraer_formato_documento_pre(nombre_insumo):
    """
    Extrae la extensión/formato del archivo para PRE-OPERACIÓN
    Ejemplo: "cartografia_base.shp" → "SHP"
    """
    if not nombre_insumo:
        return "SIN FORMATO"
    
    try:
        # Si hay punto, extraer extensión
        if '.' in nombre_insumo:
            extension = nombre_insumo.split('.')[-1].upper()
            return extension
        else:
            # Si no hay extensión, podría ser una carpeta o archivo sin extensión
            return "CARPETA"
    except:
        return "ERROR FORMATO"

def formatear_tamaño_bytes_pre(peso_archivos_str):
    """
    Formatea el tamaño en bytes agregando puntos como separadores de miles para PRE-OPERACIÓN
    Ejemplo: "2500000" → "2.500.000"
    """
    if not peso_archivos_str:
        return "Sin datos"
    
    try:
        # Limpiar el string y extraer solo números
        peso_limpio = ''.join(filter(str.isdigit, str(peso_archivos_str)))
        
        if peso_limpio:
            peso_int = int(peso_limpio)
            # Formatear con puntos como separadores de miles
            peso_formateado = f"{peso_int:,}".replace(',', '.')
            return f"{peso_formateado} bytes"
        else:
            return "0 bytes"
    except:
        return "Error tamaño"

def determinar_etapa_archivo_pre(path_file):
    """
    Determina la etapa basada en la ruta del archivo para PRE-OPERACIÓN
    """
    if not path_file:
        return "NO DEFINIDA"
    
    ruta_lower = str(path_file).lower()
    
    if 'preo' in ruta_lower or 'pre' in ruta_lower:
        return "PRE-OPERACIÓN"
    elif 'post' in ruta_lower or 'postop' in ruta_lower:
        return "POST-OPERACIÓN"
    else:
        return "PRE-OPERACIÓN"  # Por defecto, ya que estamos en pre-operación

def obtener_categoria_desde_clasificacion(cod_insumo_obj):
    """
    Obtiene el nombre de la categoría desde una clasificación de insumo
    """
    try:
        if cod_insumo_obj and cod_insumo_obj.cod_insumo:
            insumo = cod_insumo_obj.cod_insumo
            if insumo.cod_categoria:
                return insumo.cod_categoria.nom_categoria
        return "SIN CATEGORÍA"
    except:
        return "ERROR CATEGORÍA"

def obtener_categoria_desde_ruta_preoperacion(ruta_archivo):
    """
    Determina la categoría de un ArchivoPreoperacion basándose en su ruta.
    Busca nombres de directorio como 01_prop, 02_carta_acept, etc.
    """
    if not ruta_archivo:
        return "SIN CATEGORIA"

    try:
        ruta_str = str(ruta_archivo).replace('\\', '/').lower()
        partes = [p for p in ruta_str.split('/') if p]

        # Buscar coincidencia en las partes de la ruta
        for parte in partes:
            for codigo, nombre in DIRECTORIOS_PREOPERACION_MAP.items():
                if parte == codigo.lower():
                    return nombre

        return "Otros Pre-Operacion"
    except:
        return "ERROR CATEGORIA"


def generar_pestana_inventario_archivos_pre(ws, municipio, archivos_filtrados, mecanismo_financiacion, archivos_preoperacion=None):
    """
    Genera la tercera pestaña con el inventario detallado de archivos PRE-OPERACIÓN.
    Combina datos de ListaArchivosPre (insumos) + ArchivoPreoperacion (demás directorios).
    """
    print(f"📋 Generando pestaña de Inventario de Archivos PRE para {mecanismo_financiacion}...")
    
    # 🎨 ESTILOS PROFESIONALES EN TONOS AZULES (PRE-OPERACIÓN)
    titulo_font = Font(bold=True, size=16, color="FFFFFF")
    titulo_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")  # Azul oscuro
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")  # Azul medio
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 🏛️ TÍTULO PRINCIPAL
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    ws.merge_cells('A1:H2')
    titulo_cell = ws['A1']
    titulo_cell.value = (
        f"📋 INVENTARIO DETALLADO DE ARCHIVOS PRE-OPERACIÓN\n"
        f"🏛️ {municipio.nom_municipio.upper()} | MECANISMO: {mecanismo_financiacion} | 📅 {fecha_actual}"
    )
    titulo_cell.font = titulo_font
    titulo_cell.fill = titulo_fill
    titulo_cell.alignment = center_align
    titulo_cell.border = border
    
    # ============================================================
    # CONSTRUIR LISTA UNIFICADA DE ARCHIVOS
    # ============================================================
    archivos_unificados = []

    # Fuente 1: ListaArchivosPre (insumos / 07_insu)
    for archivo in archivos_filtrados:
        archivos_unificados.append({
            'etapa': determinar_etapa_archivo_pre(archivo.path_file),
            'subcarpeta': extraer_subcarpeta_desde_ruta_pre(archivo.path_file),
            'nombre': archivo.nombre_insumo or "Sin nombre",
            'formato': extraer_formato_documento_pre(archivo.nombre_insumo),
            'fecha': archivo.fecha_disposicion.strftime('%d/%m/%Y') if archivo.fecha_disposicion else "Sin fecha",
            'tamano': formatear_tamaño_bytes_pre(archivo.peso_memoria),
            'categoria': obtener_categoria_desde_clasificacion(archivo.cod_insumo),
        })

    # Fuente 2: ArchivoPreoperacion (directorios 01-06, 08)
    if archivos_preoperacion:
        for archivo_preo in archivos_preoperacion:
            fecha_doc = None
            if archivo_preo.fecha_modificacion:
                fecha_doc = archivo_preo.fecha_modificacion.strftime('%d/%m/%Y')
            elif archivo_preo.fecha_creacion:
                fecha_doc = archivo_preo.fecha_creacion.strftime('%d/%m/%Y')

            archivos_unificados.append({
                'etapa': determinar_etapa_archivo_pre(archivo_preo.ruta_archivo),
                'subcarpeta': extraer_subcarpeta_desde_ruta_pre(archivo_preo.ruta_archivo),
                'nombre': archivo_preo.nom_archivo or "Sin nombre",
                'formato': (archivo_preo.extension or "").upper().replace('.', '') or extraer_formato_documento_pre(archivo_preo.nom_archivo),
                'fecha': fecha_doc or "Sin fecha",
                'tamano': formatear_tamaño_bytes_pre(str(archivo_preo.tamano_bytes)) if archivo_preo.tamano_bytes else "Sin datos",
                'categoria': obtener_categoria_desde_ruta_preoperacion(archivo_preo.ruta_archivo),
            })

    # 📊 ESTADÍSTICAS RÁPIDAS
    total_archivos = len(archivos_unificados)
    categorias_con_archivos = len(set([arch['categoria'] for arch in archivos_unificados]))
    archivos_insumos = len(archivos_filtrados)
    archivos_directorios = len(archivos_preoperacion) if archivos_preoperacion else 0

    ws.merge_cells('A3:H3')
    stats_cell = ws['A3']
    stats_cell.value = f"RESUMEN: {total_archivos} archivos ({archivos_insumos} insumos + {archivos_directorios} otros dirs) | {categorias_con_archivos} categorias | Mecanismo: {mecanismo_financiacion}"
    stats_cell.font = Font(bold=True, size=11, color="1565C0")  # Azul oscuro
    stats_cell.fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")  # Azul claro
    stats_cell.alignment = center_align
    stats_cell.border = border
    
    # 📊 HEADERS DE LA TABLA (ROW 5)
    row = 5
    headers = [
        'ETAPA',
        'SUBCARPETA', 
        'NOMBRE DOCUMENTO',
        'FORMATO\nDOCUMENTO',
        'FECHA DE CREACIÓN\nO MODIFICACIÓN',
        'TAMAÑO (bytes)',
        'OBSERVACIONES',
        'CATEGORÍA EN\nREPORTE WEB'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # 📋 DATOS - TODOS LOS ARCHIVOS UNIFICADOS
    row = 6

    # Ordenar por categoría y luego por nombre
    archivos_ordenados = sorted(archivos_unificados, key=lambda x: (
        x['categoria'],
        x['nombre']
    ))

    for numero, arch_data in enumerate(archivos_ordenados, 1):
        if numero % 2 == 0:
            row_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Columna A: ETAPA
        cell_etapa = ws.cell(row=row, column=1, value=arch_data['etapa'])
        cell_etapa.alignment = center_align
        cell_etapa.border = border
        cell_etapa.fill = row_fill
        cell_etapa.font = Font(size=9, bold=True, color="1565C0")

        # Columna B: SUBCARPETA
        cell_subcarpeta = ws.cell(row=row, column=2, value=arch_data['subcarpeta'])
        cell_subcarpeta.alignment = left_align
        cell_subcarpeta.border = border
        cell_subcarpeta.fill = row_fill
        cell_subcarpeta.font = Font(size=9)

        # Columna C: NOMBRE DOCUMENTO
        cell_nombre = ws.cell(row=row, column=3, value=arch_data['nombre'])
        cell_nombre.alignment = left_align
        cell_nombre.border = border
        cell_nombre.fill = row_fill
        cell_nombre.font = Font(size=9)

        # Columna D: FORMATO DOCUMENTO
        cell_formato = ws.cell(row=row, column=4, value=arch_data['formato'])
        cell_formato.alignment = center_align
        cell_formato.border = border
        cell_formato.fill = row_fill
        cell_formato.font = Font(size=9, bold=True, color="1976D2")

        # Columna E: FECHA
        cell_fecha = ws.cell(row=row, column=5, value=arch_data['fecha'])
        cell_fecha.alignment = center_align
        cell_fecha.border = border
        cell_fecha.fill = row_fill
        cell_fecha.font = Font(size=9)

        # Columna F: TAMAÑO
        cell_tamano = ws.cell(row=row, column=6, value=arch_data['tamano'])
        cell_tamano.alignment = center_align
        cell_tamano.border = border
        cell_tamano.fill = row_fill
        cell_tamano.font = Font(size=9, color="42A5F5")

        # Columna G: OBSERVACIONES
        cell_observaciones = ws.cell(row=row, column=7, value="")
        cell_observaciones.alignment = left_align
        cell_observaciones.border = border
        cell_observaciones.fill = row_fill

        # Columna H: CATEGORÍA
        cell_categoria = ws.cell(row=row, column=8, value=arch_data['categoria'])
        cell_categoria.alignment = left_align
        cell_categoria.border = border
        cell_categoria.fill = row_fill
        cell_categoria.font = Font(size=9, bold=True, color="1565C0")

        row += 1
    
    # 📏 AJUSTAR DIMENSIONES DE COLUMNAS
    anchos_columnas = [15, 60, 50, 12, 18, 15, 30, 50]
    
    for i, ancho in enumerate(anchos_columnas, 1):
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = ancho
    
    # Ajustar altura de filas
    ws.row_dimensions[1].height = 35  # Título
    ws.row_dimensions[3].height = 25  # Estadísticas
    ws.row_dimensions[5].height = 35  # Headers
    
    for row_num in range(6, row):
        ws.row_dimensions[row_num].height = 20  # Datos
    
    print(f"✅ Inventario de archivos PRE-OPERACIÓN generado: {len(archivos_ordenados)} archivos ({archivos_insumos} insumos + {archivos_directorios} otros dirs)")

# ===============================================
# 🔄 MODIFICACIÓN DE LA FUNCIÓN PRINCIPAL PRE-OPERACIÓN
# ===============================================

def generar_reporte_individual_preoperacion_completo(municipio, temp_dir, mecanismo_financiacion):
    """
    🔄 VERSIÓN MEJORADA: Genera reporte Excel PRE-OPERACIÓN CON TERCERA PESTAÑA de inventario
    """
    try:
        #print(f"🎯 Generando reporte PRE-OPERACIÓN COMPLETO (3 pestañas) para {municipio.nom_municipio} - Mecanismo: {mecanismo_financiacion}...")
        
        # Crear nuevo workbook
        wb = openpyxl.Workbook()
        
        # ✅ PASO 1: RECOPILAR DATOS (adaptado para pre-operación)
        #print(f"📊 Recopilando datos de {municipio.nom_municipio}...")
        
        # Obtener insumos del municipio
        insumos_municipio = Insumos.objects.filter(
            cod_municipio=municipio.cod_municipio
        ).select_related('cod_categoria', 'tipo_insumo')
        
        # Obtener clasificaciones de estos insumos
        clasificaciones_municipio = ClasificacionInsumo.objects.filter(
            cod_insumo__in=insumos_municipio
        ).select_related('cod_insumo')
        
        # 🔥 FILTRAR ARCHIVOS POR MECANISMO DE FINANCIACIÓN
        archivos_municipio = ListaArchivosPre.objects.filter(
            cod_insumo__in=clasificaciones_municipio
        ).select_related('cod_insumo__cod_insumo')
        
        # Filtrar archivos que pertenezcan al mecanismo específico
        archivos_filtrados = []
        archivos_sin_mecanismo = []
        
        for archivo in archivos_municipio:
            # Usar la función de extracción de mecanismo que ya existe en views.py
            mecanismo_archivo = extraer_mecanismo_financiacion(archivo.path_file, municipio.cod_municipio)
            
            if mecanismo_archivo == mecanismo_financiacion:
                archivos_filtrados.append(archivo)
            elif mecanismo_archivo == "SIN_MECANISMO":
                archivos_sin_mecanismo.append(archivo)
        
        # 🎯 LÓGICA ESPECIAL: Si el mecanismo es "GENERAL" y no hay archivos específicos,
        # incluir archivos sin mecanismo definido
        if mecanismo_financiacion == "GENERAL" and len(archivos_filtrados) == 0 and len(archivos_sin_mecanismo) > 0:
            archivos_filtrados = archivos_sin_mecanismo
            print(f"🔄 Usando archivos sin mecanismo específico para reporte GENERAL")
        
        print(f"📋 Datos filtrados por mecanismo {mecanismo_financiacion}:")
        print(f"   - Archivos insumos filtrados: {len(archivos_filtrados)} de {archivos_municipio.count()} totales")

        # ============================================================
        # QUERY ArchivoPreoperacion (directorios 01-06, 08 excl. 07_insu)
        # ============================================================
        archivos_preo_todos = ArchivoPreoperacion.objects.filter(
            cod_directorio__cod_mpio=municipio.cod_municipio
        ).select_related('cod_directorio')

        archivos_preo_filtrados = []
        archivos_preo_sin_mecanismo = []

        for archivo_preo in archivos_preo_todos:
            mecanismo_archivo = extraer_mecanismo_financiacion(
                archivo_preo.ruta_archivo, municipio.cod_municipio
            )
            if mecanismo_archivo == mecanismo_financiacion:
                archivos_preo_filtrados.append(archivo_preo)
            elif mecanismo_archivo == "SIN_MECANISMO":
                archivos_preo_sin_mecanismo.append(archivo_preo)

        if mecanismo_financiacion == "GENERAL" and len(archivos_preo_filtrados) == 0 and len(archivos_preo_sin_mecanismo) > 0:
            archivos_preo_filtrados = archivos_preo_sin_mecanismo

        print(f"   - Archivos directorios filtrados: {len(archivos_preo_filtrados)} de {archivos_preo_todos.count()} totales")

        # 🚀 GENERAR REPORTE INCLUSO SI NO HAY ARCHIVOS PARA ESTE MECANISMO
        total_archivos_combinados = len(archivos_filtrados) + len(archivos_preo_filtrados)
        if total_archivos_combinados == 0:
            print(f"⚠️ ADVERTENCIA: No hay archivos para el mecanismo {mecanismo_financiacion}")
            print(f"   📋 Generando reporte vacío con estructura completa...")
        else:
            print(f"✅ Generando reporte con {total_archivos_combinados} archivos ({len(archivos_filtrados)} insumos + {len(archivos_preo_filtrados)} directorios)")
        
        # Agrupar archivos por categoría
        archivos_por_categoria = defaultdict(list)
        for archivo in archivos_filtrados:
            if archivo.cod_insumo and archivo.cod_insumo.cod_insumo:
                insumo = archivo.cod_insumo.cod_insumo
                if insumo.cod_categoria:
                    categoria_id = insumo.cod_categoria.cod_categoria
                    archivos_por_categoria[categoria_id].append(archivo)
        
        # ✅ PASO 2: GENERAR PRIMERA PESTAÑA - MATRIZ PRIMARIOS
        print("🎨 Generando Pestaña 1: Matriz Primarios...")
        ws1 = wb.active
        ws1.title = "Matriz Primarios"
        generar_pestana_matriz_primarios(ws1, municipio, insumos_municipio, clasificaciones_municipio, archivos_filtrados, mecanismo_financiacion)
        
        # ✅ PASO 3: GENERAR SEGUNDA PESTAÑA - DETALLES DE CATEGORÍAS
        print("🎨 Generando Pestaña 2: Detalles de Categorias...")
        todas_categorias = Categorias.objects.all().order_by('cod_categoria')
        ws2 = wb.create_sheet(title="Detalles de Categorias")
        generar_pestana_detalles_categorias(ws2, municipio, todas_categorias, archivos_por_categoria, mecanismo_financiacion)
        
        # ✅ PASO 4: 🆕 GENERAR TERCERA PESTAÑA - INVENTARIO DE ARCHIVOS
        print("🎨 Generando Pestaña 3: Inventario de Archivos...")
        ws3 = wb.create_sheet(title="Inventario de Archivos")
        generar_pestana_inventario_archivos_pre(ws3, municipio, archivos_filtrados, mecanismo_financiacion, archivos_preo_filtrados)
        
        # ✅ PASO 5: GUARDAR ARCHIVO
        fecha_actual = datetime.now().strftime('%d_%m_%Y')
        
        # Limpiar nombre del mecanismo para el archivo
        mecanismo_limpio = mecanismo_financiacion.replace('\\', '_').replace('/', '_').replace(':', '_').replace('-', '_').replace(' ', '_')
        municipio_limpio = municipio.nom_municipio.replace(' ', '_').replace('/', '_').replace('-', '_').upper()
        
        nombre_archivo = f"{municipio.cod_municipio}_{municipio_limpio}_{mecanismo_limpio}_Preoperacion_COMPLETO_{fecha_actual}.xlsx"
        archivo_path = os.path.join(temp_dir, nombre_archivo)
        wb.save(archivo_path)
        
        #print(f"✅ Reporte PRE-OPERACIÓN COMPLETO generado exitosamente:")
        #print(f"   📁 Archivo: {nombre_archivo}")
        #print(f"   💰 Mecanismo: {mecanismo_financiacion}")
        #print(f"   📊 Archivos incluidos: {len(archivos_filtrados)}")
        #print(f"   🎯 Categorías con datos: {len(archivos_por_categoria)}")
        #print(f"   📋 Pestañas: 3 (Matriz + Detalles + Inventario)")
        
        return archivo_path
        
    except Exception as e:
        print(f"❌ Error generando reporte PRE-OPERACIÓN COMPLETO para {municipio.nom_municipio} - {mecanismo_financiacion}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


# ============================================================================
# VIEWSETS PARA INDEXACIÓN PRE-OPERACIÓN (EXCLUYENDO 07_insu)
# ============================================================================
# Estos viewsets exponen los datos indexados por Script_PREO_RESTO_Linux.py
# NO incluyen datos de 07_insu (que se manejan con los viewsets de insumos)
# ============================================================================

from .models import DirectorioPreoperacion, ArchivoPreoperacion
from .serializers import (
    DirectorioPreoperacionSerializer, DirectorioPreoperacionSimpleSerializer,
    DirectorioPreoperacionArbolSerializer, ArchivoPreoperacionSerializer,
    ArchivoPreoperacionSimpleSerializer
)


class DirectorioPreoperacionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para consultar directorios de pre-operación (excluyendo 07_insu).

    Endpoints:
    - GET /api/directorios-preoperacion/ - Lista todos los directorios
    - GET /api/directorios-preoperacion/{id}/ - Detalle de un directorio
    - GET /api/directorios-preoperacion/?cod_mpio=17380 - Filtrar por municipio
    - GET /api/directorios-preoperacion/?nivel=0 - Filtrar por nivel (0=raíz)
    """
    queryset = DirectorioPreoperacion.objects.all()
    serializer_class = DirectorioPreoperacionSerializer
    filterset_fields = ['cod_mpio', 'nivel', 'parent']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Optimizar consultas
        queryset = queryset.select_related('cod_mpio', 'parent')
        queryset = queryset.prefetch_related('archivos', 'subdirectorios')

        # Ordenar por nivel y nombre
        return queryset.order_by('nivel', 'nom_directorio')

    @action(detail=False, methods=['get'])
    def por_municipio(self, request):
        """Obtiene todos los directorios raíz (nivel 0) de un municipio"""
        cod_mpio = request.query_params.get('cod_mpio')
        if not cod_mpio:
            return Response(
                {'error': 'Parámetro cod_mpio es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        directorios = self.get_queryset().filter(cod_mpio=cod_mpio, nivel=0)
        serializer = DirectorioPreoperacionSimpleSerializer(directorios, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def archivos(self, request, pk=None):
        """Obtiene los archivos de un directorio específico"""
        directorio = self.get_object()
        archivos = directorio.archivos.all()
        serializer = ArchivoPreoperacionSimpleSerializer(archivos, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def subdirectorios(self, request, pk=None):
        """Obtiene los subdirectorios de un directorio específico"""
        directorio = self.get_object()
        subdirs = directorio.subdirectorios.all()
        serializer = DirectorioPreoperacionSimpleSerializer(subdirs, many=True)
        return Response(serializer.data)


class ArchivoPreoperacionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para consultar archivos de pre-operación (excluyendo 07_insu).

    Endpoints:
    - GET /api/archivos-preoperacion/ - Lista todos los archivos
    - GET /api/archivos-preoperacion/{id}/ - Detalle de un archivo
    - GET /api/archivos-preoperacion/?cod_directorio=5 - Filtrar por directorio
    - GET /api/archivos-preoperacion/?extension=.pdf - Filtrar por extensión
    """
    queryset = ArchivoPreoperacion.objects.all()
    serializer_class = ArchivoPreoperacionSerializer
    filterset_fields = ['cod_directorio', 'extension', 'propietario']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Optimizar consultas
        queryset = queryset.select_related('cod_directorio', 'cod_directorio__cod_mpio')

        # Filtro adicional por municipio (a través del directorio)
        cod_mpio = self.request.query_params.get('cod_mpio')
        if cod_mpio:
            queryset = queryset.filter(cod_directorio__cod_mpio=cod_mpio)

        return queryset.order_by('nom_archivo')

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Obtiene estadísticas de archivos por municipio"""
        from django.db.models import Count, Sum

        cod_mpio = request.query_params.get('cod_mpio')
        queryset = self.get_queryset()

        if cod_mpio:
            queryset = queryset.filter(cod_directorio__cod_mpio=cod_mpio)

        stats = queryset.aggregate(
            total_archivos=Count('cod_archivo'),
            tamano_total=Sum('tamano_bytes')
        )

        # Por extensión
        por_extension = queryset.values('extension').annotate(
            cantidad=Count('cod_archivo'),
            tamano=Sum('tamano_bytes')
        ).order_by('-cantidad')[:10]

        return Response({
            'total_archivos': stats['total_archivos'] or 0,
            'tamano_total_bytes': stats['tamano_total'] or 0,
            'por_extension': list(por_extension)
        })


# ============================================================================
# VISTA ÁRBOL COMBINADA PRE-OPERACIÓN COMPLETA
# ============================================================================

class PreoperacionArbolCombinadoViewSet(viewsets.ViewSet):
    """
    ViewSet para obtener la vista de árbol COMPLETA de pre-operación.

    Combina:
    - Directorios indexados en directorios_preoperacion (01_prop, 02_carta_acept, etc.)
    - Estructura de insumos en 07_insu (desde insumos/clasificacion_insumo)

    Endpoints:
    - GET /api/preoperacion-arbol/arbol_municipio/?cod_mpio=17380
    - GET /api/preoperacion-arbol/estadisticas_municipio/?cod_mpio=17380

    Permisos:
    - Administradores y Super Administradores: acceso a todos los municipios
    - Profesionales: solo municipios asignados
    """
    permission_classes = [IsAuthenticated]

    def _verificar_acceso_municipio(self, user, cod_mpio):
        """
        Verifica si el usuario tiene acceso al municipio solicitado.
        Retorna True si tiene acceso, False si no.
        """
        municipios_permitidos = get_municipios_permitidos(user)

        # Administradores tienen acceso a todo
        if municipios_permitidos == 'todos':
            return True

        # Profesionales solo a sus municipios asignados
        if isinstance(municipios_permitidos, list):
            return cod_mpio in municipios_permitidos

        return False

    @action(detail=False, methods=['get'])
    def arbol_municipio(self, request):
        """Obtiene el árbol combinado completo de pre-operación"""
        cod_mpio = request.query_params.get('cod_mpio')

        if not cod_mpio:
            return Response(
                {'error': 'Parámetro cod_mpio es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            cod_mpio = int(cod_mpio)
        except ValueError:
            return Response(
                {'error': 'cod_mpio debe ser un número entero'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar permisos de acceso al municipio
        if not self._verificar_acceso_municipio(request.user, cod_mpio):
            return Response(
                {'error': 'No tiene permisos para acceder a este municipio'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            arbol = self._construir_arbol_completo(cod_mpio)
            return Response(arbol)
        except Exception as e:
            return Response(
                {'error': f'Error construyendo árbol: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _construir_arbol_completo(self, cod_mpio):
        """Construye el árbol combinado de pre-operación"""
        arbol = []

        # 1. Obtener directorios de directorios_preoperacion (nivel 0)
        directorios_resto = DirectorioPreoperacion.objects.filter(
            cod_mpio=cod_mpio,
            nivel=0
        ).prefetch_related('archivos', 'subdirectorios').order_by('nom_directorio')

        # 2. Construir nodos para cada directorio
        for directorio in directorios_resto:
            nodo = {
                'id': f'dir_{directorio.cod_directorio}',
                'nombre': directorio.nom_directorio,
                'tipo': 'directorio',
                'fuente': 'preoperacion',
                'nivel': directorio.nivel,
                'archivos_count': directorio.get_archivos_count(),
                'hijos': self._construir_hijos_directorio(directorio)
            }
            arbol.append(nodo)

        # 3. Obtener estructura de 07_insu desde insumos/clasificacion_insumo
        # Un municipio puede tener múltiples insumos, obtenemos todos
        insumos = Insumos.objects.filter(cod_municipio=cod_mpio)

        if insumos.exists():
            # Obtener todas las clasificaciones de todos los insumos del municipio
            clasificaciones = ClasificacionInsumo.objects.filter(
                cod_insumo__in=insumos
            ).order_by('cod_clasificacion')

            nodo_insu = {
                'id': 'dir_insu',
                'nombre': '07_insu',
                'tipo': 'directorio',
                'fuente': 'insumos',
                'nivel': 0,
                'archivos_count': self._contar_archivos_insumos(clasificaciones),
                'hijos': self._construir_hijos_insumos(clasificaciones)
            }
            arbol.append(nodo_insu)
        else:
            nodo_insu = {
                'id': 'dir_insu',
                'nombre': '07_insu',
                'tipo': 'directorio',
                'fuente': 'insumos',
                'nivel': 0,
                'archivos_count': 0,
                'hijos': [],
                'mensaje': 'Sin datos de insumos indexados'
            }
            arbol.append(nodo_insu)

        # 4. Ordenar todo por nombre
        arbol.sort(key=lambda x: x['nombre'])

        return {
            'cod_mpio': cod_mpio,
            'estructura': arbol,
            'total_directorios': len(arbol),
            'fuentes': {
                'directorios_preoperacion': len(directorios_resto),
                'insumos': 1
            }
        }

    def _construir_hijos_directorio(self, directorio, max_nivel=10):
        """Construye los hijos de un directorio recursivamente (hasta max_nivel)"""
        hijos = []

        # Agregar subdirectorios
        for subdir in directorio.subdirectorios.all().order_by('nom_directorio'):
            if subdir.nivel <= max_nivel:
                hijo = {
                    'id': f'dir_{subdir.cod_directorio}',
                    'nombre': subdir.nom_directorio,
                    'tipo': 'directorio',
                    'fuente': 'preoperacion',
                    'nivel': subdir.nivel,
                    'archivos_count': subdir.get_archivos_count(),
                    'ruta_windows': linux_to_windows_path(subdir.ruta_directorio),
                    'hijos': self._construir_hijos_directorio(subdir, max_nivel)
                }
                hijos.append(hijo)

        # Agregar archivos del directorio actual
        for archivo in directorio.archivos.all().order_by('nom_archivo'):
            hijo_archivo = {
                'id': f'file_{archivo.cod_archivo}',
                'nombre': archivo.nom_archivo,
                'tipo': 'archivo',
                'fuente': 'preoperacion',
                'nivel': directorio.nivel + 1,
                'archivos_count': 0,
                'ruta_windows': linux_to_windows_path(archivo.ruta_archivo),
                'extension': archivo.extension,
                'tamano_legible': archivo.get_tamano_legible(),
                'propietario': archivo.propietario,
                'fecha_modificacion': archivo.fecha_modificacion.isoformat() if archivo.fecha_modificacion else None,
                'hijos': []
            }
            hijos.append(hijo_archivo)

        return hijos

    def _construir_hijos_insumos(self, clasificaciones):
        """Construye los hijos del nodo 07_insu con archivos"""
        hijos = []
        for clasificacion in clasificaciones:
            archivos = ListaArchivosPre.objects.filter(
                cod_insumo=clasificacion
            ).order_by('nombre_insumo')

            # Construir nodos de archivos
            archivos_hijos = []
            for archivo in archivos:
                # Extraer extensión del nombre
                nombre = archivo.nombre_insumo or ''
                extension = None
                if '.' in nombre:
                    extension = '.' + nombre.rsplit('.', 1)[-1].lower()

                archivo_hijo = {
                    'id': f'insu_file_{archivo.id_lista_archivo}',
                    'nombre': nombre,
                    'tipo': 'archivo',
                    'fuente': 'insumos',
                    'nivel': 2,
                    'archivos_count': 0,
                    'ruta_windows': linux_to_windows_path(archivo.path_file) if archivo.path_file else None,
                    'extension': extension,
                    'tamano_legible': self._formatear_tamano(archivo.peso_memoria) if archivo.peso_memoria else None,
                    'propietario': archivo.usuario_windows,
                    'fecha_modificacion': archivo.fecha_disposicion.isoformat() if archivo.fecha_disposicion else None,
                    'hijos': []
                }
                archivos_hijos.append(archivo_hijo)

            hijo = {
                'id': f'clas_{clasificacion.cod_clasificacion}',
                'nombre': clasificacion.nombre,
                'tipo': 'categoria',
                'fuente': 'insumos',
                'nivel': 1,
                'archivos_count': len(archivos_hijos),
                'ruta_windows': linux_to_windows_path(clasificacion.ruta) if clasificacion.ruta else None,
                'hijos': archivos_hijos
            }
            hijos.append(hijo)
        return hijos

    def _formatear_tamano(self, tamano_bytes):
        """Formatea el tamaño en bytes a formato legible"""
        if not tamano_bytes:
            return "0 B"
        tamano = float(tamano_bytes)
        for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
            if tamano < 1024.0:
                return f"{tamano:.2f} {unit}"
            tamano /= 1024.0
        return f"{tamano:.2f} PB"

    def _contar_archivos_insumos(self, clasificaciones):
        """Cuenta el total de archivos en insumos"""
        total = 0
        for clasificacion in clasificaciones:
            total += ListaArchivosPre.objects.filter(
                cod_insumo=clasificacion
            ).count()
        return total

    @action(detail=False, methods=['get'])
    def estadisticas_municipio(self, request):
        """Obtiene estadísticas combinadas de pre-operación"""
        cod_mpio = request.query_params.get('cod_mpio')

        if not cod_mpio:
            return Response(
                {'error': 'Parámetro cod_mpio es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            cod_mpio = int(cod_mpio)
        except ValueError:
            return Response(
                {'error': 'cod_mpio debe ser un número entero'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar permisos de acceso al municipio
        if not self._verificar_acceso_municipio(request.user, cod_mpio):
            return Response(
                {'error': 'No tiene permisos para acceder a este municipio'},
                status=status.HTTP_403_FORBIDDEN
            )

        from django.db.models import Count, Sum

        try:

            stats_directorios = DirectorioPreoperacion.objects.filter(
                cod_mpio=cod_mpio
            ).aggregate(total_directorios=Count('cod_directorio'))

            stats_archivos_preo = ArchivoPreoperacion.objects.filter(
                cod_directorio__cod_mpio=cod_mpio
            ).aggregate(
                total_archivos=Count('cod_archivo'),
                tamano_total=Sum('tamano_bytes')
            )

            # Un municipio puede tener múltiples insumos
            insumos = Insumos.objects.filter(cod_municipio=cod_mpio)
            if insumos.exists():
                clasificaciones = ClasificacionInsumo.objects.filter(cod_insumo__in=insumos)
                total_archivos_insu = ListaArchivosPre.objects.filter(
                    cod_insumo__in=clasificaciones
                ).count()
                total_categorias = clasificaciones.count()
            else:
                total_archivos_insu = 0
                total_categorias = 0

            return Response({
                'cod_mpio': cod_mpio,
                'directorios_preoperacion': {
                    'total_directorios': stats_directorios['total_directorios'] or 0,
                    'total_archivos': stats_archivos_preo['total_archivos'] or 0,
                    'tamano_total_bytes': stats_archivos_preo['tamano_total'] or 0
                },
                'insumos_07_insu': {
                    'total_categorias': total_categorias,
                    'total_archivos': total_archivos_insu
                },
                'totales': {
                    'total_archivos': (stats_archivos_preo['total_archivos'] or 0) + total_archivos_insu
                }
            })

        except Exception as e:
            return Response(
                {'error': f'Error obteniendo estadísticas: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], permission_classes=[AllowAny])
    def municipios_con_preoperacion(self, request):
        """
        Lista todos los municipios que tienen directorios de pre-operación indexados.
        Incluye estadísticas de directorios, archivos y tamaño.

        NOTA: Endpoint público (como postoperacion) para compatibilidad con frontend.
        El control de acceso por rol se maneja en el frontend.
        """
        from django.db.models import Count, Sum, F, Value
        from django.db.models.functions import Coalesce

        try:
            # Obtener municipios que tienen directorios indexados
            municipios_con_dirs = DirectorioPreoperacion.objects.values(
                'cod_mpio'
            ).annotate(
                total_directorios=Count('cod_directorio')
            ).values_list('cod_mpio', flat=True).distinct()

            # Obtener información de municipios
            municipios = Municipios.objects.filter(
                cod_municipio__in=municipios_con_dirs
            ).select_related('cod_depto').annotate(
                total_directorios=Coalesce(
                    Count('directorios_preoperacion', distinct=True),
                    Value(0)
                ),
                total_archivos=Coalesce(
                    Count('directorios_preoperacion__archivos', distinct=True),
                    Value(0)
                ),
                tamano_bytes=Coalesce(
                    Sum('directorios_preoperacion__archivos__tamano_bytes'),
                    Value(0)
                )
            ).order_by('nom_municipio')

            # Formatear respuesta
            resultado = []
            for mun in municipios:
                # Formatear tamaño
                tamano = mun.tamano_bytes or 0
                if tamano >= 1024 * 1024 * 1024:
                    tamano_str = f"{tamano / (1024*1024*1024):.2f} GB"
                elif tamano >= 1024 * 1024:
                    tamano_str = f"{tamano / (1024*1024):.2f} MB"
                elif tamano >= 1024:
                    tamano_str = f"{tamano / 1024:.2f} KB"
                else:
                    tamano_str = f"{tamano} B"

                resultado.append({
                    'cod_municipio': mun.cod_municipio,
                    'nom_municipio': mun.nom_municipio,
                    'cod_depto': mun.cod_depto.cod_depto if mun.cod_depto else None,
                    'nom_territorial': mun.nom_territorial,
                    'total_directorios': mun.total_directorios,
                    'total_archivos': mun.total_archivos,
                    'tamano_total': tamano_str
                })

            return Response(resultado)

        except Exception as e:
            return Response(
                {'error': f'Error listando municipios: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def arbol_realtime(self, request):
        """
        Obtiene el árbol de pre-operación leyendo DIRECTAMENTE del sistema de archivos.
        NO depende de datos indexados - acceso en tiempo real al NAS.

        Parámetros:
        - cod_mpio: Código del municipio (requerido)
        - path: Subdirectorio a explorar (opcional, relativo a 01_preo)

        Permisos:
        - Administradores y Super Administradores: acceso a todos los municipios
        - Profesionales: solo municipios asignados
        """
        cod_mpio = request.query_params.get('cod_mpio')
        subpath = request.query_params.get('path', '')
        mecanismo_param = request.query_params.get('mecanismo', '')

        if not cod_mpio:
            return Response(
                {'error': 'Parámetro cod_mpio es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            cod_mpio = int(cod_mpio)
        except ValueError:
            return Response(
                {'error': 'cod_mpio debe ser un número entero'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar permisos de acceso al municipio
        if not self._verificar_acceso_municipio(request.user, cod_mpio):
            return Response(
                {'error': 'No tiene permisos para acceder a este municipio'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            # Obtener información del municipio
            municipio = Municipios.objects.select_related('cod_depto').filter(
                cod_municipio=cod_mpio
            ).first()

            if not municipio:
                return Response(
                    {'error': f'Municipio {cod_mpio} no encontrado'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Construir la ruta base al filesystem
            resultado_ruta = self._construir_ruta_preoperacion(municipio, mecanismo=mecanismo_param)

            if not resultado_ruta or not resultado_ruta.get('path'):
                from backend.path_utils import linux_to_windows_path
                error_msg = resultado_ruta.get('error', 'No se encontró directorio de pre-operación') if resultado_ruta else 'No se encontró directorio de pre-operación'
                ruta_esperada = resultado_ruta.get('ruta_esperada') if resultado_ruta else None
                # Convertir ruta a formato Windows para mostrar al usuario
                ruta_windows = linux_to_windows_path(ruta_esperada) if ruta_esperada else None
                # También convertir el mensaje de error si contiene la ruta
                if ruta_esperada and ruta_esperada in error_msg:
                    error_msg = error_msg.replace(ruta_esperada, ruta_windows)
                return Response({
                    'cod_mpio': cod_mpio,
                    'nom_municipio': municipio.nom_municipio,
                    'estructura': [],
                    'mensaje': error_msg,
                    'ruta_buscada': ruta_windows
                })

            base_path = resultado_ruta['path']
            mecanismo = resultado_ruta.get('mecanismo', 'desconocido')

            # Si hay subpath, añadirlo
            if subpath:
                full_path = os.path.join(base_path, subpath)
            else:
                full_path = base_path

            # Verificar que el directorio existe
            if not os.path.isdir(full_path):
                return Response({
                    'cod_mpio': cod_mpio,
                    'nom_municipio': municipio.nom_municipio,
                    'estructura': [],
                    'mensaje': f'Directorio no encontrado: {full_path}',
                    'ruta_buscada': full_path
                })

            # Leer el directorio recursivamente
            estructura = self._leer_directorio_recursivo(full_path, base_path, max_depth=10)

            return Response({
                'cod_mpio': cod_mpio,
                'nom_municipio': municipio.nom_municipio,
                'nom_depto': municipio.cod_depto.nom_depto if municipio.cod_depto else None,
                'estructura': estructura,
                'ruta_base': base_path,
                'total_elementos': len(estructura)
            })

        except Exception as e:
            import traceback
            return Response(
                {'error': f'Error leyendo filesystem: {str(e)}', 'trace': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _construir_ruta_preoperacion(self, municipio, mecanismo=''):
        """
        Construye la ruta al directorio 01_preo del municipio.
        Busca en el filesystem el mecanismo (PGN, SGP, etc.)

        Retorna:
        - dict con 'path' y 'error' si hubo problema
        - dict con 'path' si encontró el directorio
        - None si no existe
        """
        import os

        # Ruta base del repositorio
        repo_base = '/mnt/repositorio/2510SP/H_Informacion_Consulta/Sub_Proy/01_actualiz_catas'

        # Obtener código de departamento (2 dígitos)
        cod_depto = str(municipio.cod_depto.cod_depto).zfill(2) if municipio.cod_depto else None
        if not cod_depto:
            return {'path': None, 'error': 'Municipio sin departamento asignado'}

        # Obtener código de municipio (3 dígitos, sin el depto)
        cod_mpio_str = str(municipio.cod_municipio)
        if len(cod_mpio_str) == 5:
            # Es el código completo (depto + mpio), tomar los últimos 3
            cod_mpio_short = cod_mpio_str[2:]
        else:
            cod_mpio_short = cod_mpio_str.zfill(3)

        # Construir ruta al directorio del municipio
        mpio_base = os.path.join(repo_base, cod_depto, cod_mpio_short)

        try:
            if not os.path.isdir(mpio_base):
                return {'path': None, 'error': f'Directorio del municipio no existe: {mpio_base}', 'ruta_esperada': mpio_base}
        except OSError as e:
            return {'path': None, 'error': f'Error accediendo al repositorio: {str(e)}. Verifique el montaje del NAS.', 'ruta_esperada': mpio_base}

        # Buscar el mecanismo (PGN, SGP, CONPES, etc.)
        mecanismos_encontrados = []
        try:
            for item in os.listdir(mpio_base):
                item_path = os.path.join(mpio_base, item)
                try:
                    if os.path.isdir(item_path):
                        # Verificar si tiene 01_preo dentro
                        preo_path = os.path.join(item_path, '01_preo')
                        if os.path.isdir(preo_path):
                            mecanismos_encontrados.append({
                                'mecanismo': item,
                                'path': preo_path
                            })
                except OSError:
                    continue
        except PermissionError:
            return {'path': None, 'error': f'Permiso denegado para leer: {mpio_base}', 'ruta_esperada': mpio_base}
        except OSError as e:
            return {'path': None, 'error': f'Error de sistema de archivos: {str(e)}. Verifique conexión al NAS.', 'ruta_esperada': mpio_base}

        if not mecanismos_encontrados:
            return {'path': None, 'error': f'No se encontró carpeta 01_preo en ningún mecanismo dentro de {mpio_base}', 'ruta_esperada': mpio_base}

        # Si se especificó un mecanismo, buscar ese específicamente
        if mecanismo:
            for m in mecanismos_encontrados:
                if m['mecanismo'].upper() == mecanismo.upper():
                    return {'path': m['path'], 'mecanismo': m['mecanismo']}
            # Si no se encontró el mecanismo solicitado, informar
            disponibles = [m['mecanismo'] for m in mecanismos_encontrados]
            return {'path': None, 'error': f'Mecanismo {mecanismo} no encontrado. Disponibles: {", ".join(disponibles)}'}

        # Si hay múltiples mecanismos y no se especificó, usar el primero
        return {'path': mecanismos_encontrados[0]['path'], 'mecanismo': mecanismos_encontrados[0]['mecanismo']}

    def _leer_directorio_recursivo(self, dir_path, base_path, current_depth=0, max_depth=10):
        """
        Lee un directorio y sus contenidos recursivamente.
        Retorna estructura de árbol con archivos y carpetas.

        NOTA: Los directorios .gdb (geodatabases) se tratan como archivos descargables,
        no se itera dentro de ellos.
        """
        import os
        from datetime import datetime

        # Extensiones de directorios que deben tratarse como archivos (geodatabases, etc.)
        DIRECTORY_AS_FILE_EXTENSIONS = ['.gdb', '.mdb', '.geodatabase']

        if current_depth > max_depth:
            return []

        items = []

        try:
            entries = sorted(os.listdir(dir_path))
        except PermissionError:
            return [{
                'id': f'error_{current_depth}',
                'nombre': '[Sin acceso]',
                'tipo': 'error',
                'mensaje': 'Permiso denegado'
            }]
        except Exception as e:
            return [{
                'id': f'error_{current_depth}',
                'nombre': f'[Error: {str(e)}]',
                'tipo': 'error'
            }]

        for entry in entries:
            entry_path = os.path.join(dir_path, entry)
            relative_path = os.path.relpath(entry_path, base_path)

            try:
                stat_info = os.stat(entry_path)
                is_dir = os.path.isdir(entry_path)

                # Verificar si es un directorio que debe tratarse como archivo (.gdb, etc.)
                _, ext = os.path.splitext(entry)
                is_gdb_or_special = ext.lower() in DIRECTORY_AS_FILE_EXTENSIONS

                # Si es un .gdb o similar, tratarlo como archivo descargable
                if is_dir and is_gdb_or_special:
                    # Calcular tamaño total del directorio .gdb
                    total_size = 0
                    try:
                        for dirpath, dirnames, filenames in os.walk(entry_path):
                            for f in filenames:
                                fp = os.path.join(dirpath, f)
                                total_size += os.path.getsize(fp)
                    except:
                        total_size = 0

                    item = {
                        'id': f'gdb_{relative_path.replace("/", "_").replace(" ", "_")}',
                        'nombre': entry,
                        'tipo': 'archivo',  # Tratar como archivo
                        'subtipo': 'geodatabase',
                        'ruta_relativa': relative_path,
                        'ruta_absoluta': entry_path,
                        'nivel': current_depth,
                        'fecha_modificacion': datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                        'extension': ext.lower(),
                        'tamano_bytes': total_size,
                        'tamano_legible': self._formatear_tamano(total_size),
                        'archivos_count': 0,
                        'hijos': [],
                        'es_geodatabase': True
                    }
                elif is_dir:
                    # Directorio normal - iterar dentro
                    try:
                        hijos_count = len(os.listdir(entry_path))
                    except:
                        hijos_count = 0

                    item = {
                        'id': f'dir_{relative_path.replace("/", "_").replace(" ", "_")}',
                        'nombre': entry,
                        'tipo': 'directorio',
                        'ruta_relativa': relative_path,
                        'ruta_absoluta': entry_path,
                        'nivel': current_depth,
                        'fecha_modificacion': datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                        'archivos_count': hijos_count,
                        'hijos': self._leer_directorio_recursivo(
                            entry_path, base_path, current_depth + 1, max_depth
                        )
                    }
                else:
                    # Es un archivo normal
                    _, extension = os.path.splitext(entry)
                    item = {
                        'id': f'file_{relative_path.replace("/", "_").replace(" ", "_")}',
                        'nombre': entry,
                        'tipo': 'archivo',
                        'ruta_relativa': relative_path,
                        'ruta_absoluta': entry_path,
                        'nivel': current_depth,
                        'fecha_modificacion': datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                        'extension': extension.lower() if extension else None,
                        'tamano_bytes': stat_info.st_size,
                        'tamano_legible': self._formatear_tamano(stat_info.st_size),
                        'archivos_count': 0,
                        'hijos': []
                    }

                items.append(item)

            except Exception as e:
                items.append({
                    'id': f'error_{entry}',
                    'nombre': entry,
                    'tipo': 'error',
                    'mensaje': str(e)
                })

        return items


class SubClasificacionFuenteSecundariaViewSet(viewsets.ReadOnlyModelViewSet):
    """Sub-clasificaciones (sub-dominios) para Insumos Fuentes Secundarias, filtrable por dominio"""
    queryset = SubClasificacionFuenteSecundaria.objects.all()
    serializer_class = SubClasificacionFuenteSecundariaSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['dominio']
    ordering_fields = ['orden', 'nombre']
    ordering = ['orden']